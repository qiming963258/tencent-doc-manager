#!/usr/bin/env python3
"""
智能Excel标记系统 - 精确匹配和条纹涂色
实现从详细打分JSON到Excel文件的精确涂色标记
"""

import json
import os
import hashlib
import zipfile
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import glob
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DetailedScoreGenerator:
    """生成详细打分JSON"""
    
    @staticmethod
    def generate_score_json(baseline_file: str, target_file: str, output_dir: str) -> str:
        """
        生成详细打分JSON文件
        
        Args:
            baseline_file: 基准文件路径
            target_file: 目标文件路径
            output_dir: 输出目录
            
        Returns:
            生成的JSON文件路径
        """
        # 生成唯一标识符
        file_hash = hashlib.md5(f"{baseline_file}_{target_file}".encode()).hexdigest()[:8]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        doc_name = os.path.basename(target_file).replace('.xlsx', '').replace('.csv', '').replace('_fixed', '')
        
        # 根据文件格式加载数据
        def load_file_data(file_path):
            """加载CSV或Excel文件，返回数据矩阵"""
            if file_path.endswith('.csv'):
                # CSV文件处理（不依赖pandas）
                import csv
                data = []
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        data.append(row)
                return data, len(data), len(data[0]) if data else 0
            else:
                # Excel文件处理
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                data = []
                max_row = ws.max_row
                max_col = ws.max_column
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
                wb.close()
                return data, max_row, max_col
        
        # 加载两个文件的数据
        baseline_data, baseline_rows, baseline_cols = load_file_data(baseline_file)
        target_data, target_rows, target_cols = load_file_data(target_file)
        
        # 生成详细打分数据
        score_data = {
            "metadata": {
                "comparison_id": f"comp_{timestamp}_{file_hash}",
                "baseline_file": os.path.basename(baseline_file),
                "target_file": os.path.basename(target_file),
                "doc_name": doc_name,
                "doc_id": file_hash,
                "timestamp": timestamp,
                "week_number": DetailedScoreGenerator._extract_week_number(baseline_file)
            },
            "statistics": {
                "total_cells": 0,
                "changed_cells": 0,
                "high_risk_count": 0,
                "medium_risk_count": 0,
                "low_risk_count": 0
            },
            "cell_scores": {}
        }
        
        # 遍历所有单元格进行对比
        max_row = max(baseline_rows, target_rows)
        max_col = max(baseline_cols, target_cols)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_ref = openpyxl.utils.get_column_letter(col) + str(row)
                
                # 从数据矩阵获取值（注意索引是从0开始的）
                baseline_value = None
                target_value = None
                
                if row <= len(baseline_data) and col <= len(baseline_data[row-1] if baseline_data else []):
                    baseline_value = baseline_data[row-1][col-1]
                
                if row <= len(target_data) and col <= len(target_data[row-1] if target_data else []):
                    target_value = target_data[row-1][col-1]
                
                # 转换为字符串进行比较
                baseline_str = str(baseline_value) if baseline_value is not None else ""
                target_str = str(target_value) if target_value is not None else ""
                
                score_data["statistics"]["total_cells"] += 1
                
                if baseline_str != target_str:
                    score_data["statistics"]["changed_cells"] += 1
                    
                    # 计算变更评分
                    score, risk_level, change_type = DetailedScoreGenerator._calculate_score(
                        baseline_str, target_str, row, col
                    )
                    
                    # 更新风险统计
                    if risk_level == "high":
                        score_data["statistics"]["high_risk_count"] += 1
                    elif risk_level == "medium":
                        score_data["statistics"]["medium_risk_count"] += 1
                    else:
                        score_data["statistics"]["low_risk_count"] += 1
                    
                    # 记录详细信息
                    score_data["cell_scores"][cell_ref] = {
                        "row": row,
                        "column": col,
                        "old_value": baseline_str,
                        "new_value": target_str,
                        "change_type": change_type,
                        "risk_level": risk_level,
                        "score": score,
                        "color_code": DetailedScoreGenerator._get_color_code(score)
                    }
        
        # 保存JSON文件
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, f"detailed_scores_{doc_name}_{timestamp}.json")
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(score_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✓ 详细打分JSON已生成: {output_file}")
        logger.info(f"  - 总单元格: {score_data['statistics']['total_cells']}")
        logger.info(f"  - 变更单元格: {score_data['statistics']['changed_cells']}")
        
        return output_file
    
    @staticmethod
    def _extract_week_number(filename: str) -> str:
        """从文件名中提取周数"""
        import re
        match = re.search(r'W(\d+)', filename)
        return match.group(1) if match else "00"
    
    @staticmethod
    def _calculate_score(old_value: str, new_value: str, row: int, col: int) -> Tuple[int, str, str]:
        """
        计算变更评分
        
        Returns:
            (score, risk_level, change_type)
        """
        # 判断变更类型
        if old_value == "" and new_value != "":
            change_type = "addition"
            base_score = 60
        elif old_value != "" and new_value == "":
            change_type = "deletion"
            base_score = 30  # 删除通常风险较高
        else:
            # 尝试数值比较
            try:
                old_num = float(old_value)
                new_num = float(new_value)
                
                change_rate = abs((new_num - old_num) / old_num) if old_num != 0 else 1.0
                
                if change_rate < 0.1:
                    change_type = "minor_change"
                    base_score = 80
                elif change_rate < 0.5:
                    change_type = "moderate_change"
                    base_score = 50
                else:
                    change_type = "major_change"
                    base_score = 20
                    
                if new_num > old_num:
                    change_type = f"numeric_increase_{change_type}"
                else:
                    change_type = f"numeric_decrease_{change_type}"
                    
            except:
                # 文本变更
                change_type = "text_change"
                base_score = 40
        
        # 根据位置调整分数（前几行和前几列通常更重要）
        if row <= 3 or col <= 2:
            base_score = max(0, base_score - 20)  # 标题行/列变更风险更高
        
        # 确定风险等级
        if base_score < 30:
            risk_level = "high"
        elif base_score < 70:
            risk_level = "medium"
        else:
            risk_level = "low"
        
        return base_score, risk_level, change_type
    
    @staticmethod
    def _get_color_code(score: int) -> str:
        """根据分数返回颜色代码"""
        if score < 30:
            return "FF0000"  # 红色
        elif score < 70:
            return "FFFF00"  # 黄色
        else:
            return "00FF00"  # 绿色


class IntelligentExcelMarker:
    """智能Excel标记器 - 精确匹配和条纹涂色"""
    
    def __init__(self):
        """初始化标记器"""
        self.score_dir = "/root/projects/tencent-doc-manager/scoring_results/detailed/"
        self.output_dir = "/root/projects/tencent-doc-manager/excel_outputs/marked/"
        
        # 使用纯色填充以兼容腾讯文档
        # 根据列级别(L1/L2/L3)进行颜色映射，而不是风险等级
        # 使用更明显的颜色，避免在腾讯文档中看不清
        self.column_level_colors = {
            "L1": "FF6666",  # L1列(最重要) - 明显红色
            "L2": "FFB366",  # L2列(重要) - 明显橙色
            "L3": "66FF66"   # L3列(一般) - 明显绿色
        }

        # 备用：如果需要根据风险等级涂色（符合规范文档）
        self.risk_level_colors = {
            "EXTREME_HIGH": "FF3333",  # 极高风险 - 深红色
            "HIGH": "FF6666",          # 高风险 - 红色
            "MEDIUM": "FFB366",        # 中风险 - 橙色
            "LOW": "66FF66"            # 低风险 - 绿色
        }
        
        # 使用列级别涂色模式
        self.use_column_level_coloring = True
    
    def find_matching_score_file(self, excel_file: str) -> Optional[str]:
        """
        查找与Excel文件匹配的详细打分JSON
        
        匹配逻辑：
        1. 提取Excel文件的基础名称
        2. 在打分目录中查找包含该名称的JSON
        3. 如果有多个匹配，选择最新的
        """
        # 提取文档标识信息
        base_name = os.path.basename(excel_file)
        doc_name = base_name.replace('.xlsx', '').replace('_fixed', '')
        
        logger.info(f"查找匹配的打分文件: {doc_name}")
        
        # 方法1：通过文档名称匹配
        pattern1 = os.path.join(self.score_dir, f"detailed_scores_{doc_name}_*.json")
        matches = glob.glob(pattern1)
        
        # 方法2：通过文档ID匹配（如果文件名包含ID）
        if not matches:
            # 尝试提取可能的文档ID
            import re
            id_match = re.search(r'([a-f0-9]{8})', doc_name)
            if id_match:
                doc_id = id_match.group(1)
                pattern2 = os.path.join(self.score_dir, f"*{doc_id}*.json")
                matches = glob.glob(pattern2)
        
        # 方法3：通过时间戳查找最近的打分文件
        if not matches:
            # 获取所有打分文件
            all_scores = glob.glob(os.path.join(self.score_dir, "detailed_scores_*.json"))
            
            # 检查每个文件的metadata中是否包含目标文件名
            for score_file in all_scores:
                try:
                    with open(score_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        if 'metadata' in data:
                            target = data['metadata'].get('target_file', '')
                            if doc_name in target or base_name in target:
                                matches.append(score_file)
                except:
                    continue
        
        if matches:
            # 选择最新的文件
            latest_file = max(matches, key=os.path.getmtime)
            logger.info(f"✓ 找到匹配的打分文件: {latest_file}")
            return latest_file
        else:
            logger.warning(f"✗ 未找到匹配的打分文件")
            return None
    
    def apply_striped_coloring(self, excel_file: str, score_file: str, output_file: str = None) -> str:
        """
        应用纯色涂色到Excel文件（腾讯文档兼容版本）
        
        Args:
            excel_file: 要涂色的Excel文件（已修复格式）
            score_file: 详细打分JSON文件
            output_file: 输出文件路径（可选）
            
        Returns:
            输出文件路径
        """
        logger.info(f"开始应用纯色涂色（腾讯文档兼容模式）")
        logger.info(f"  Excel文件: {excel_file}")
        logger.info(f"  打分文件: {score_file}")
        
        # 加载打分数据
        with open(score_file, 'r', encoding='utf-8') as f:
            score_data = json.load(f)
        
        # 加载Excel文件
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 统计涂色情况
        if self.use_column_level_coloring:
            color_stats = {"L1": 0, "L2": 0, "L3": 0}
        else:
            color_stats = {"EXTREME_HIGH": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
        
        # 兼容两种数据格式
        if 'cell_scores' in score_data:
            # 旧格式：cell_scores字典
            scores_to_process = score_data['cell_scores'].items()
        elif 'scores' in score_data:
            # 新格式：scores数组，需要转换
            scores_to_process = []
            for item in score_data['scores']:
                cell_ref = item.get('cell')
                if cell_ref:
                    cell_data = {
                        'risk_level': item.get('risk_assessment', {}).get('risk_level', 'LOW'),
                        'column_level': item.get('column_level', 'L3'),  # 添加列级别
                        'score': item.get('scoring_details', {}).get('final_score', 0),
                        'old_value': item.get('old_value'),
                        'new_value': item.get('new_value'),
                        'column_name': item.get('column_name'),
                        'ai_analysis': item.get('ai_analysis', {})
                    }
                    scores_to_process.append((cell_ref, cell_data))
        else:
            logger.error("打分数据格式不正确：缺少cell_scores或scores字段")
            raise ValueError("打分数据格式不正确")
        
        # 遍历所有变更的单元格
        for cell_ref, cell_data in scores_to_process:
            try:
                # 获取单元格
                cell = ws[cell_ref]
                
                # 根据配置决定使用哪种涂色模式
                if self.use_column_level_coloring:
                    # 使用列级别涂色（L1红、L2黄、L3绿）
                    column_level = str(cell_data.get('column_level', 'L3')).upper()

                    # 获取对应的颜色
                    if column_level in self.column_level_colors:
                        colors = self.column_level_colors[column_level]
                        # 用于统计的键
                        stat_key = column_level
                    else:
                        logger.warning(f"未知的列级别: {column_level}，跳过单元格 {cell_ref}")
                        continue
                else:
                    # 使用风险等级涂色（备用模式）
                    risk_level = str(cell_data.get('risk_level', 'LOW')).upper()

                    if risk_level in self.risk_level_colors:
                        colors = self.risk_level_colors[risk_level]
                        stat_key = risk_level
                    else:
                        logger.warning(f"未知的风险等级: {risk_level}，跳过单元格 {cell_ref}")
                        continue
                
                # 创建纯色填充（使用新语法，腾讯文档兼容）
                fill = PatternFill(
                    start_color=colors,
                    end_color=colors,
                    fill_type='solid'  # 必须使用solid，腾讯文档唯一支持
                )
                
                # 应用填充
                cell.fill = fill
                
                # 腾讯文档不支持Excel注释，已移除注释功能
                # 仅使用颜色标记风险等级
                
                color_stats[stat_key] += 1
                
            except Exception as e:
                logger.warning(f"无法涂色单元格 {cell_ref}: {e}")
        
        # 生成输出文件名
        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            week_num = score_data['metadata'].get('week_number', '00')
            base_name = os.path.basename(excel_file).replace('_fixed.xlsx', '')
            output_file = os.path.join(
                self.output_dir, 
                f"{base_name}_marked_{timestamp}_W{week_num}.xlsx"
            )
        
        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        # 保存文件
        wb.save(output_file)
        wb.close()
        
        # 输出统计信息
        logger.info(f"✓ 涂色完成（纯色模式，腾讯文档兼容）！")
        logger.info(f"  输出文件: {output_file}")
        logger.info(f"  涂色统计:")
        
        if self.use_column_level_coloring:
            logger.info(f"    - L1列（红色）: {color_stats.get('L1', 0)} 个单元格")
            logger.info(f"    - L2列（橙色）: {color_stats.get('L2', 0)} 个单元格")
            logger.info(f"    - L3列（绿色）: {color_stats.get('L3', 0)} 个单元格")
        else:
            logger.info(f"    - 极高风险（红色）: {color_stats.get('EXTREME_HIGH', 0)} 个单元格")
            logger.info(f"    - 高风险（浅红色）: {color_stats.get('HIGH', 0)} 个单元格")
            logger.info(f"    - 中风险（浅橙色）: {color_stats.get('MEDIUM', 0)} 个单元格")
            logger.info(f"    - 低风险（黄色）: {color_stats.get('LOW', 0)} 个单元格")
        
        return output_file
    
    def process_excel_with_auto_match(self, excel_file: str) -> Optional[str]:
        """
        自动匹配打分文件并处理Excel
        
        完整处理流程：
        1. 修复Excel格式问题
        2. 查找匹配的打分文件
        3. 如果没有打分文件，生成新的
        4. 应用条纹涂色
        """
        logger.info("="*60)
        logger.info("开始智能Excel处理流程")
        logger.info("="*60)
        
        # 步骤1：修复格式（如果需要）
        if "_fixed" not in excel_file:
            from fix_tencent_excel import fix_tencent_excel
            fixed_file = excel_file.replace('.xlsx', '_fixed.xlsx')
            if not os.path.exists(fixed_file):
                logger.info("修复Excel格式...")
                fix_tencent_excel(excel_file, fixed_file)
                excel_file = fixed_file
            else:
                excel_file = fixed_file
        
        # 步骤2：查找或生成打分文件
        score_file = self.find_matching_score_file(excel_file)
        
        if not score_file:
            logger.info("未找到打分文件，尝试生成新的...")
            
            # 查找基准文件
            from production.core_modules.week_time_manager import WeekTimeManager
            wtm = WeekTimeManager()
            week_info = wtm.get_current_week_info()
            current_week = week_info['week_number']
            
            baseline_dir = f"/root/projects/tencent-doc-manager/csv_versions/2025_W{current_week}/baseline/"
            baseline_files = glob.glob(os.path.join(baseline_dir, "*baseline*.csv"))
            
            if baseline_files:
                # 将CSV转换为临时Excel（如果需要）
                baseline_file = baseline_files[0]
                logger.info(f"使用基准文件: {baseline_file}")
                
                # 生成打分
                os.makedirs(self.score_dir, exist_ok=True)
                generator = DetailedScoreGenerator()
                score_file = generator.generate_score_json(
                    baseline_file,
                    excel_file,
                    self.score_dir
                )
            else:
                logger.error("未找到基准文件，无法生成打分")
                return None
        
        # 步骤3：应用涂色
        if score_file:
            output_file = self.apply_striped_coloring(excel_file, score_file)
            return output_file
        
        return None


def main():
    """主函数 - 演示完整流程"""
    
    # 初始化标记器
    marker = IntelligentExcelMarker()
    
    # 查找最新下载的Excel文件
    download_dir = "/root/projects/tencent-doc-manager/downloads/"
    excel_files = glob.glob(os.path.join(download_dir, "*.xlsx"))
    
    # 过滤掉已处理的文件
    excel_files = [f for f in excel_files if "_fixed" not in f and "_marked" not in f]
    
    if excel_files:
        # 选择最新的文件
        latest_file = max(excel_files, key=os.path.getmtime)
        logger.info(f"处理文件: {latest_file}")
        
        # 执行智能处理
        result = marker.process_excel_with_auto_match(latest_file)
        
        if result:
            logger.info(f"\n✅ 处理成功！")
            logger.info(f"   最终文件: {result}")
            
            # 测试打开文件
            try:
                wb = openpyxl.load_workbook(result)
                logger.info(f"   验证: 文件可正常打开")
                wb.close()
            except Exception as e:
                logger.error(f"   验证失败: {e}")
        else:
            logger.error("处理失败")
    else:
        logger.info("未找到待处理的Excel文件")
        
        # 演示模式：使用测试文件
        test_file = "downloads/副本-副本-测试版本-出国销售计划表_fixed.xlsx"
        if os.path.exists(test_file):
            logger.info(f"\n使用测试文件: {test_file}")
            result = marker.process_excel_with_auto_match(test_file)


if __name__ == "__main__":
    main()