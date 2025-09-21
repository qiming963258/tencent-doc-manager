#!/usr/bin/env python3
"""
智能Excel标记器 V3 - 基于配置中心的统一版本
使用配置中心的标准列定义和风险分级
创建日期：2025-09-20
"""

import json
import os
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.comments import Comment

# 添加项目路径
sys.path.append('/root/projects/tencent-doc-manager')

# 从配置中心导入（单一真相源）
from production.config import (
    STANDARD_COLUMNS,
    L1_COLUMNS,
    L2_COLUMNS,
    L3_COLUMNS,
    get_column_risk_level,
    get_column_weight
)

class IntelligentExcelMarkerV3:
    """
    智能Excel标记器V3
    基于配置中心的统一涂色方案
    """

    def __init__(self):
        """初始化标记器"""
        self.output_dir = Path("/root/projects/tencent-doc-manager/excel_outputs/marked")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # 风险级别对应的颜色（纯色填充，兼容腾讯文档）
        self.risk_colors = {
            'HIGH': {
                'bg_color': 'FFCCCC',    # 浅红色背景
                'font_color': 'CC0000',   # 深红色字体
                'border': True
            },
            'MEDIUM': {
                'bg_color': 'FFE5CC',    # 浅橙色背景
                'font_color': 'FF6600',   # 橙色字体
                'border': True
            },
            'LOW': {
                'bg_color': 'FFFFCC',    # 浅黄色背景
                'font_color': 'CC9900',   # 深黄色字体
                'border': False
            }
        }

    def fix_tencent_excel(self, input_file: str) -> str:
        """
        修复腾讯文档导出的Excel文件
        处理空fill标签问题
        """
        print(f"🔧 修复腾讯Excel格式...")

        fixed_file = str(Path(input_file).with_suffix('.fixed.xlsx'))

        with tempfile.TemporaryDirectory() as temp_dir:
            # 解压Excel文件
            with zipfile.ZipFile(input_file, 'r') as zf:
                zf.extractall(temp_dir)

            # 修复styles.xml
            styles_path = Path(temp_dir) / 'xl' / 'styles.xml'
            if styles_path.exists():
                content = styles_path.read_text(encoding='utf-8')
                # 替换空fill标签
                content = content.replace('<fill/>', '<fill><patternFill patternType="none"/></fill>')
                styles_path.write_text(content, encoding='utf-8')

            # 重新打包
            with zipfile.ZipFile(fixed_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = Path(root) / file
                        arc_name = file_path.relative_to(temp_dir)
                        zf.write(file_path, arc_name)

        print(f"  ✅ 格式修复完成")
        return fixed_file

    def get_risk_level_from_score(self, score: float) -> str:
        """
        根据分数判断风险级别
        分数范围：0-100 或 0-1
        """
        # 标准化到0-1范围
        if score > 1:
            score = score / 100

        if score >= 0.6:
            return 'HIGH'
        elif score >= 0.3:
            return 'MEDIUM'
        else:
            return 'LOW'

    def apply_cell_marking(self, cell, risk_level: str, cell_data: dict):
        """
        对单个单元格应用标记

        Args:
            cell: openpyxl单元格对象
            risk_level: 风险级别(HIGH/MEDIUM/LOW)
            cell_data: 单元格数据（包含原值、新值等）
        """
        colors = self.risk_colors[risk_level]

        # 应用背景色
        cell.fill = PatternFill(
            patternType='solid',
            fgColor=colors['bg_color'],
            bgColor=colors['bg_color']
        )

        # 应用字体颜色
        cell.font = Font(
            color=colors['font_color'],
            bold=(risk_level == 'HIGH')
        )

        # 添加边框（高、中风险）
        if colors['border']:
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

        # 添加批注
        comment_text = self.build_comment_text(risk_level, cell_data)
        cell.comment = Comment(comment_text, "智能监控系统V3")

    def build_comment_text(self, risk_level: str, cell_data: dict) -> str:
        """构建批注文本"""
        old_value = cell_data.get('old_value', 'N/A')
        new_value = cell_data.get('new_value', 'N/A')
        change_type = cell_data.get('change_type', 'unknown')
        score = cell_data.get('score', 0)
        column_level = cell_data.get('column_level', 'unknown')

        # 获取列名（从单元格位置推断）
        col_index = cell_data.get('column_index')
        col_name = STANDARD_COLUMNS[col_index] if col_index and col_index < len(STANDARD_COLUMNS) else 'unknown'

        comment = (
            f"🎯 风险等级: {risk_level}\n"
            f"📊 列名: {col_name}\n"
            f"📈 列级别: {column_level}\n"
            f"🔄 变更类型: {change_type}\n"
            f"💯 评分: {score}\n"
            f"📝 原值: {old_value}\n"
            f"✏️ 新值: {new_value}"
        )

        # 添加风险说明
        if risk_level == 'HIGH':
            comment += "\n⚠️ 高风险变更，需要重点审核"
        elif risk_level == 'MEDIUM':
            comment += "\n⚡ 中等风险，建议复查"

        return comment

    def mark_excel_with_scores(self, excel_file: str, score_file: str, output_file: str = None) -> str:
        """
        基于打分结果对Excel进行智能标记

        Args:
            excel_file: 要标记的Excel文件路径
            score_file: 详细打分JSON文件路径
            output_file: 输出文件路径（可选）

        Returns:
            输出文件路径
        """
        print("\n" + "="*60)
        print("🎨 智能Excel标记器 V3 - 基于配置中心")
        print("="*60)

        # 加载打分数据
        print(f"\n📊 加载打分数据...")
        with open(score_file, 'r', encoding='utf-8') as f:
            score_data = json.load(f)

        stats = score_data.get('statistics', {})
        print(f"  - 总单元格: {stats.get('total_cells', 0)}")
        print(f"  - 变更单元格: {stats.get('changed_cells', 0)}")

        # 修复Excel格式（如果需要）
        if 'tencent' in excel_file.lower():
            fixed_file = self.fix_tencent_excel(excel_file)
        else:
            fixed_file = excel_file

        # 加载Excel
        print(f"\n📄 加载Excel文件...")
        wb = openpyxl.load_workbook(fixed_file)
        ws = wb.active

        # 标记统计
        mark_stats = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
        marked_cells = []

        # 获取单元格打分数据
        cell_scores = score_data.get('cell_scores', {})
        print(f"\n🎯 开始标记 {len(cell_scores)} 个变更单元格...")

        for cell_ref, cell_data in cell_scores.items():
            try:
                # 获取单元格
                cell = ws[cell_ref]

                # 获取风险等级
                if 'risk_level' in cell_data:
                    risk_level = cell_data['risk_level']
                else:
                    # 从分数推断风险等级
                    score = cell_data.get('score', 0)
                    risk_level = self.get_risk_level_from_score(score)

                # 应用标记
                self.apply_cell_marking(cell, risk_level, cell_data)

                # 更新统计
                mark_stats[risk_level] += 1
                marked_cells.append(cell_ref)

            except Exception as e:
                print(f"  ⚠️ 标记单元格 {cell_ref} 失败: {e}")

        # 生成输出文件名
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_name = Path(excel_file).stem
            output_file = str(self.output_dir / f"{base_name}_marked_v3_{timestamp}.xlsx")

        # 保存文件
        wb.save(output_file)
        wb.close()

        # 清理临时文件
        if fixed_file != excel_file and os.path.exists(fixed_file):
            os.remove(fixed_file)

        # 输出统计
        print(f"\n✅ 标记完成！")
        print(f"📁 输出文件: {output_file}")
        print(f"\n📊 标记统计:")
        print(f"  - 高风险（红色）: {mark_stats['HIGH']} 个")
        print(f"  - 中风险（橙色）: {mark_stats['MEDIUM']} 个")
        print(f"  - 低风险（黄色）: {mark_stats['LOW']} 个")
        print(f"  - 总计: {sum(mark_stats.values())} 个")

        if marked_cells[:10]:
            print(f"\n📍 标记的单元格示例:")
            print(f"  {', '.join(marked_cells[:10])}")

        return output_file

    def validate_configuration(self):
        """验证配置中心设置"""
        print("\n🔍 验证配置中心...")
        print(f"  - 标准列数: {len(STANDARD_COLUMNS)}")
        print(f"  - L1列数: {len(L1_COLUMNS)}")
        print(f"  - L2列数: {len(L2_COLUMNS)}")
        print(f"  - L3列数: {len(L3_COLUMNS)}")

        # 验证总数
        total = len(L1_COLUMNS) + len(L2_COLUMNS) + len(L3_COLUMNS)
        assert total == 19, f"列分级总数错误: {total} != 19"

        print(f"  ✅ 配置验证通过")


def main():
    """主函数 - 演示用法"""
    marker = IntelligentExcelMarkerV3()

    # 验证配置
    marker.validate_configuration()

    # 文件路径
    excel_file = "/root/projects/tencent-doc-manager/csv_versions/2025_W38/midweek/tencent_111测试版本-小红书部门_20250919_2255_midweek_W38.xlsx"
    score_file = "/root/projects/tencent-doc-manager/scoring_results/detailed/detailed_scores_tencent_111测试版本-小红书部门_20250919_2255_midweek_W38_20250920_010627.json"

    # 检查文件是否存在
    if not os.path.exists(excel_file):
        print(f"❌ Excel文件不存在: {excel_file}")
        return

    if not os.path.exists(score_file):
        print(f"❌ 打分文件不存在: {score_file}")
        return

    # 执行标记
    output_file = marker.mark_excel_with_scores(excel_file, score_file)

    print("\n" + "="*60)
    print("🎉 智能标记完成！基于配置中心的统一涂色")
    print("="*60)

    return output_file


if __name__ == "__main__":
    main()