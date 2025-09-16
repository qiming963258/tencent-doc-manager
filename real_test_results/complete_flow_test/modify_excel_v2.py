#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Complete Flow Test - 修改Excel文件 (v2)
使用更强健的方法处理Excel文件
"""

import os
import openpyxl
import json
from datetime import datetime

def safe_read_excel(source_file):
    """
    安全地读取Excel文件内容
    """
    try:
        print(f"尝试读取Excel文件: {source_file}")
        
        # 第一种方法：尝试正常加载
        try:
            workbook = openpyxl.load_workbook(source_file, data_only=True)
            worksheet = workbook.active
            print(f"✅ 成功使用data_only模式读取文件")
            return workbook, worksheet, "normal"
        except Exception as e1:
            print(f"data_only模式失败: {e1}")
        
        # 第二种方法：尝试不读取公式
        try:
            workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
            worksheet = workbook.active
            print(f"✅ 成功使用read_only模式读取文件")
            return workbook, worksheet, "read_only"
        except Exception as e2:
            print(f"read_only模式失败: {e2}")
        
        # 如果都失败了，返回错误
        raise Exception(f"无法读取Excel文件: {e1}, {e2}")
        
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")
        return None, None, "failed"

def create_modified_excel(original_data, output_file, original_a1_value):
    """
    创建修改后的Excel文件
    """
    try:
        # 创建新的工作簿
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        new_worksheet.title = "工作表2"  # 使用原始工作表名称
        
        # 设置A1单元格的新值
        if original_a1_value:
            new_a1_value = f"[已测试]{original_a1_value}"
        else:
            new_a1_value = "[已测试]"
        
        new_worksheet['A1'] = new_a1_value
        
        # 添加修改信息到B1
        modification_info = f"修改时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        new_worksheet['B1'] = modification_info
        
        # 复制其他数据（如果有的话）
        if original_data:
            for row_idx, row_data in enumerate(original_data, 1):
                if row_idx == 1:  # 跳过第一行，因为我们已经处理了A1
                    continue
                for col_idx, cell_value in enumerate(row_data, 1):
                    if cell_value is not None:
                        new_worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # 保存新文件
        new_workbook.save(output_file)
        print(f"✅ 成功创建修改后的Excel文件: {output_file}")
        
        return {
            "success": True,
            "new_a1_value": new_a1_value,
            "modification_info": modification_info
        }
        
    except Exception as e:
        print(f"❌ 创建修改后Excel文件失败: {e}")
        return {"success": False, "error": str(e)}

def modify_excel_file_v2(source_file, output_file):
    """
    修改Excel文件的强健版本
    """
    try:
        print(f"开始修改Excel文件 (v2): {source_file}")
        
        # 安全读取原始文件
        workbook, worksheet, read_mode = safe_read_excel(source_file)
        
        if not workbook or not worksheet:
            return {"success": False, "error": "无法读取源Excel文件"}
        
        # 读取原始A1值
        original_a1_value = None
        original_data = []
        
        try:
            original_a1_value = worksheet['A1'].value
            print(f"原始A1值: {original_a1_value}")
            
            # 如果是只读模式，我们需要复制数据
            if read_mode == "read_only":
                print("只读模式，复制所有数据...")
                for row in worksheet.iter_rows(values_only=True):
                    original_data.append(row)
                    if len(original_data) > 100:  # 限制行数，避免内存问题
                        break
                print(f"复制了 {len(original_data)} 行数据")
        
        except Exception as e:
            print(f"读取原始数据时出错: {e}")
            original_a1_value = None
        
        # 关闭原始文件（如果是只读模式）
        if read_mode == "read_only":
            workbook.close()
        
        # 创建修改后的文件
        result = create_modified_excel(original_data, output_file, original_a1_value)
        
        if result["success"]:
            # 验证修改结果
            try:
                verify_workbook = openpyxl.load_workbook(output_file, data_only=True)
                verify_worksheet = verify_workbook.active
                verified_a1 = verify_worksheet['A1'].value
                verified_b1 = verify_worksheet['B1'].value
                verify_workbook.close()
                
                # 获取文件大小
                source_size = os.path.getsize(source_file)
                output_size = os.path.getsize(output_file)
                
                return {
                    "success": True,
                    "source_file": source_file,
                    "output_file": output_file,
                    "original_a1_value": original_a1_value,
                    "modified_a1_value": verified_a1,
                    "modification_info": verified_b1,
                    "source_file_size": source_size,
                    "output_file_size": output_size,
                    "read_mode": read_mode,
                    "modification_timestamp": datetime.now().isoformat()
                }
                
            except Exception as e:
                print(f"验证修改结果时出错: {e}")
                return {
                    "success": True,  # 创建成功，但验证失败
                    "source_file": source_file,
                    "output_file": output_file,
                    "original_a1_value": original_a1_value,
                    "modified_a1_value": result.get("new_a1_value"),
                    "modification_info": result.get("modification_info"),
                    "read_mode": read_mode,
                    "verification_error": str(e),
                    "modification_timestamp": datetime.now().isoformat()
                }
        else:
            return result
        
    except Exception as e:
        error_msg = f"修改Excel文件v2失败: {e}"
        print(f"❌ {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "source_file": source_file,
            "output_file": output_file
        }

def main():
    print("=== 腾讯文档完整流程测试 - Excel修改阶段 (v2) ===")
    
    # 文件路径
    base_dir = "/root/projects/tencent-doc-manager/real_test_results/complete_flow_test"
    source_file = os.path.join(base_dir, "测试版本-小红书部门.xlsx")
    output_file = os.path.join(base_dir, "测试版本-小红书部门_修改标记.xlsx")
    
    # 检查源文件是否存在
    if not os.path.exists(source_file):
        error_msg = f"源文件不存在: {source_file}"
        print(f"❌ {error_msg}")
        return {"success": False, "error": error_msg}
    
    # 执行修改
    result = modify_excel_file_v2(source_file, output_file)
    
    # 保存测试结果
    test_result = {
        "timestamp": datetime.now().isoformat(),
        "test_phase": "excel_modification_v2",
        "source_file": source_file,
        "output_file": output_file,
        "modification_result": result
    }
    
    result_file = os.path.join(base_dir, "excel_modification_v2_result.json")
    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(test_result, f, ensure_ascii=False, indent=2)
    
    if result['success']:
        print(f"✅ Excel修改成功!")
        print(f"原始文件: {result['source_file']}")
        print(f"修改后文件: {result['output_file']}")
        print(f"原始A1值: {result.get('original_a1_value')}")
        print(f"修改后A1值: {result.get('modified_a1_value')}")
        print(f"读取模式: {result.get('read_mode', '未知')}")
        
        if result.get('verification_error'):
            print(f"⚠️ 验证时出现问题: {result['verification_error']}")
    else:
        print(f"❌ Excel修改失败: {result['error']}")
    
    return result

if __name__ == "__main__":
    main()