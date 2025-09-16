#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化Excel修改方案 - 避免样式问题
仅修改数据内容，保持原始格式
"""

import sys
import os
from datetime import datetime

def simple_excel_modify(input_file):
    """简化的Excel修改方案"""
    print(f"🛠️ 简化Excel修改")
    print(f"📂 输入文件: {input_file}")
    
    if not os.path.exists(input_file):
        print("❌ 输入文件不存在")
        return None
    
    original_size = os.path.getsize(input_file)
    print(f"📏 原始大小: {original_size} bytes")
    
    try:
        import openpyxl
        
        # 加载工作簿时忽略样式
        print("📖 加载Excel工作簿（忽略样式）...")
        wb = openpyxl.load_workbook(input_file, data_only=True)
        
        # 获取第一个工作表
        ws_names = wb.sheetnames
        print(f"📊 工作表: {ws_names}")
        
        ws = wb[ws_names[0]]  # 使用第一个工作表
        
        # 读取A1原始内容
        original_a1 = ws['A1'].value
        print(f"📝 A1原始内容: {original_a1}")
        
        # 修改A1内容
        if original_a1:
            new_a1_value = f"[已测试]{str(original_a1)}"
        else:
            new_a1_value = "[已测试]新内容"
        
        ws['A1'] = new_a1_value
        print(f"✏️ A1修改后: {new_a1_value}")
        
        # 在空白位置添加修改标记
        empty_cell = None
        for row in range(1, 10):
            for col in range(20, 26):  # T-Z列
                cell = ws.cell(row=row, column=col)
                if not cell.value:
                    empty_cell = cell
                    break
            if empty_cell:
                break
        
        if empty_cell:
            modification_mark = f"修改: {datetime.now().strftime('%H:%M:%S')}"
            empty_cell.value = modification_mark
            print(f"🏷️ 修改标记: {empty_cell.coordinate} = {modification_mark}")
        
        # 保存修改后的文件
        output_file = input_file.replace('.xlsx', '_简化修改.xlsx')
        
        # 保存时不保留样式
        print("💾 保存修改后的文件...")
        wb.save(output_file)
        wb.close()
        
        # 验证修改结果
        if os.path.exists(output_file):
            modified_size = os.path.getsize(output_file)
            print(f"✅ 修改成功!")
            print(f"📁 输出文件: {output_file}")
            print(f"📏 修改后大小: {modified_size} bytes")
            
            # 大小差异分析
            size_diff = abs(modified_size - original_size)
            size_change = (size_diff / original_size) * 100
            print(f"📊 大小变化: {size_change:.2f}%")
            
            # 验证修改内容
            wb_verify = openpyxl.load_workbook(output_file, data_only=True)
            ws_verify = wb_verify[ws_names[0]]
            verified_a1 = ws_verify['A1'].value
            wb_verify.close()
            
            print(f"🔍 验证A1内容: {verified_a1}")
            
            return {
                'success': True,
                'output_file': output_file,
                'original_size': original_size,
                'modified_size': modified_size,
                'size_change_percent': size_change,
                'original_a1': original_a1,
                'modified_a1': verified_a1,
                'modification_timestamp': datetime.now().isoformat()
            }
        else:
            print("❌ 输出文件创建失败")
            return {'success': False, 'error': '输出文件创建失败'}
            
    except Exception as e:
        print(f"❌ Excel修改失败: {e}")
        return {'success': False, 'error': str(e)}

if __name__ == "__main__":
    # 查找最新下载的Excel文件
    test_dir = "/root/projects/tencent-doc-manager/real_test_results/verified_test_163708"
    excel_file = os.path.join(test_dir, "测试版本-小红书部门.xlsx")
    
    print("🎯 Excel文件简化修改测试")
    print("=" * 50)
    
    if os.path.exists(excel_file):
        result = simple_excel_modify(excel_file)
        
        print("\n" + "=" * 50)
        print("📊 修改测试结果")
        print("=" * 50)
        
        if result and result['success']:
            print("✅ Excel修改测试成功!")
            print(f"📁 原始文件: {excel_file}")
            print(f"📁 修改文件: {result['output_file']}")
            print(f"📏 大小变化: {result['size_change_percent']:.2f}%")
            print(f"📝 A1修改: {result['original_a1']} → {result['modified_a1']}")
        else:
            print("❌ Excel修改测试失败")
            if result:
                print(f"错误: {result.get('error', '未知错误')}")
    else:
        print(f"❌ 找不到Excel文件: {excel_file}")