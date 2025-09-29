#!/usr/bin/env python3
"""
直接测试批量处理并验证Excel涂色
"""

import requests
import json
import time
import os
from datetime import datetime
from pathlib import Path


def trigger_batch_processing():
    """直接触发8093批量处理"""

    print("📋 直接调用8093批量处理API...")

    # 读取Cookie
    cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
    with open(cookie_file) as f:
        cookie_data = json.load(f)
        cookie = cookie_data.get('current_cookies', '')

    # 准备批量处理请求
    data = {
        'cookie': cookie,
        'advanced_settings': {
            'task_type': 'comprehensive',
            'auto_download': True,
            'force_download': False,  # 不强制下载，使用现有文件
            'enable_ai_analysis': True,
            'enable_excel_marking': True,
            'enable_upload': True,
            'use_existing_baseline': True
        }
    }

    try:
        response = requests.post(
            'http://localhost:8093/api/start-batch',
            json=data,
            timeout=10
        )

        if response.status_code == 200:
            result = response.json()
            print(f"✅ 批量处理已启动")
            print(f"   • 执行ID: {result.get('execution_id')}")
            return result.get('execution_id')
        else:
            print(f"❌ 启动失败: HTTP {response.status_code}")
            return None

    except Exception as e:
        print(f"❌ 请求失败: {e}")
        return None


def monitor_processing():
    """监控处理进度"""

    print("\n📡 监控处理进度...")
    print("-" * 50)

    max_wait = 120
    wait_time = 0

    while wait_time < max_wait:
        try:
            response = requests.get('http://localhost:8093/api/status', timeout=5)
            if response.status_code == 200:
                data = response.json()
                status = data.get('status', 'unknown')
                task = data.get('current_task', '')
                progress = data.get('progress', 0)

                if wait_time % 5 == 0:  # 每5秒输出一次
                    timestamp = datetime.now().strftime('%H:%M:%S')
                    print(f"[{timestamp}] {status} | {task} | {progress}%")

                if status in ['completed', 'error']:
                    return status, data

        except Exception:
            pass

        time.sleep(1)
        wait_time += 1

    return 'timeout', None


def validate_excel_files():
    """验证Excel文件涂色"""

    print("\n📋 验证Excel文件涂色...")

    excel_dir = Path('/root/projects/tencent-doc-manager/excel_outputs/marked')
    if not excel_dir.exists():
        print("❌ Excel目录不存在")
        return False

    # 查找最新的3个Excel文件
    excel_files = sorted(excel_dir.glob('*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)[:3]

    if not excel_files:
        print("❌ 未找到Excel文件")
        return False

    import openpyxl

    all_valid = True

    for excel_file in excel_files:
        print(f"\n🔍 检查: {excel_file.name}")

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            solid_count = 0
            lightup_count = 0

            # 检查前50行的涂色
            for row in range(1, min(51, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)

                    if cell.fill and cell.fill.patternType:
                        if cell.fill.patternType == "solid":
                            solid_count += 1
                        elif cell.fill.patternType == "lightUp":
                            lightup_count += 1

            if lightup_count > 0:
                print(f"   ❌ 发现lightUp填充: {lightup_count}个")
                all_valid = False
            elif solid_count > 0:
                print(f"   ✅ 正确使用solid填充: {solid_count}个")
            else:
                print(f"   ⚠️ 未发现涂色")

        except Exception as e:
            print(f"   ❌ 读取失败: {e}")
            all_valid = False

    return all_valid


def main():
    print("=" * 60)
    print("🚀 Excel涂色功能验证测试")
    print("=" * 60)

    # 启动批量处理
    execution_id = trigger_batch_processing()
    if not execution_id:
        print("❌ 无法启动批量处理")
        return False

    # 监控处理
    status, data = monitor_processing()

    print(f"\n处理结果: {status}")

    # 验证Excel
    valid = validate_excel_files()

    print("\n" + "=" * 60)
    print("📊 测试结果")
    print("=" * 60)

    if valid:
        print("✅ Excel涂色验证通过！")
        print("• 使用了正确的solid填充")
        print("• 未发现不兼容的lightUp填充")
        print("• 可以上传到腾讯文档")
    else:
        print("❌ Excel涂色验证失败")
        print("• 可能使用了lightUp填充")
        print("• 需要修复fill_type参数")

    return valid


if __name__ == "__main__":
    import sys
    success = main()
    sys.exit(0 if success else 1)