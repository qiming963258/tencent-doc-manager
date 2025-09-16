#!/usr/bin/env python3
"""
腾讯文档Excel转换器 - 将腾讯文档导出的Excel转换为标准格式
"""

import os
import sys
from pathlib import Path

def install_pandas():
    """安装pandas库"""
    print("正在安装pandas库...")
    os.system("pip install pandas openpyxl xlsxwriter --quiet")

try:
    import pandas as pd
except ImportError:
    install_pandas()
    import pandas as pd

def convert_tencent_excel(input_file, output_file=None):
    """
    转换腾讯文档Excel文件为标准格式
    
    Args:
        input_file: 腾讯文档导出的Excel文件路径
        output_file: 输出文件路径（可选）
    
    Returns:
        转换后的文件路径
    """
    if not output_file:
        path = Path(input_file)
        output_file = path.parent / f"{path.stem}_converted.xlsx"
    
    print(f"📥 正在转换: {input_file}")
    
    try:
        # 方法1：使用pandas读取并重新保存
        print("   尝试方法1: pandas转换...")
        
        # pandas有更好的容错性，可以处理大多数格式问题
        df = pd.read_excel(input_file, engine=None)  # 自动选择引擎
        
        # 保存为新的Excel文件
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        print(f"✅ 转换成功！")
        print(f"📄 输出文件: {output_file}")
        
        # 显示数据预览
        print(f"\n📊 数据预览:")
        print(f"   行数: {len(df)}")
        print(f"   列数: {len(df.columns)}")
        print(f"   列名: {', '.join(df.columns[:5])}...")
        
        return output_file
        
    except Exception as e:
        print(f"❌ pandas转换失败: {e}")
        
        # 方法2：尝试只提取数据，忽略格式
        try:
            print("\n   尝试方法2: 只提取数据...")
            
            # 使用openpyxl的只读模式
            from openpyxl import load_workbook
            wb = load_workbook(input_file, read_only=True, data_only=True)
            ws = wb.active
            
            # 提取所有数据
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            wb.close()
            
            # 转换为DataFrame并保存
            df = pd.DataFrame(data[1:], columns=data[0] if data else [])
            df.to_excel(output_file, index=False)
            
            print(f"✅ 数据提取成功！")
            return output_file
            
        except Exception as e2:
            print(f"❌ 数据提取也失败: {e2}")
            
            # 方法3：建议使用CSV格式
            print("\n💡 建议：")
            print("   1. 在腾讯文档中选择'导出为CSV'格式")
            print("   2. CSV格式没有样式问题，可以直接处理")
            print("   3. 或者用Microsoft Excel打开后另存为")
            
            return None

def batch_convert(directory, pattern="*tencent*.xlsx"):
    """
    批量转换目录下的腾讯文档Excel文件
    
    Args:
        directory: 目录路径
        pattern: 文件匹配模式
    """
    dir_path = Path(directory)
    files = list(dir_path.glob(pattern))
    
    print(f"📁 找到 {len(files)} 个文件")
    print("="*50)
    
    success_count = 0
    failed_files = []
    
    for file in files:
        print(f"\n处理文件 {success_count + len(failed_files) + 1}/{len(files)}")
        result = convert_tencent_excel(file)
        if result:
            success_count += 1
        else:
            failed_files.append(file)
        print("-"*50)
    
    # 显示结果
    print(f"\n📊 转换结果:")
    print(f"   ✅ 成功: {success_count} 个")
    print(f"   ❌ 失败: {len(failed_files)} 个")
    
    if failed_files:
        print(f"\n失败的文件:")
        for f in failed_files:
            print(f"   - {f}")

def convert_to_csv(input_file, output_file=None):
    """
    将Excel文件转换为CSV格式（最稳定的方案）
    """
    if not output_file:
        path = Path(input_file)
        output_file = path.parent / f"{path.stem}.csv"
    
    try:
        # 尝试读取Excel
        df = pd.read_excel(input_file)
        # 保存为CSV
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        print(f"✅ 已转换为CSV: {output_file}")
        return output_file
    except Exception as e:
        print(f"❌ 转换失败: {e}")
        return None

if __name__ == "__main__":
    print("🔧 腾讯文档Excel转换器")
    print("="*50)
    
    if len(sys.argv) < 2:
        print("用法:")
        print("  转换单个文件: python tencent_excel_converter.py <文件路径>")
        print("  批量转换: python tencent_excel_converter.py --batch <目录路径>")
        print("  转为CSV: python tencent_excel_converter.py --csv <文件路径>")
        sys.exit(1)
    
    if sys.argv[1] == "--batch" and len(sys.argv) > 2:
        batch_convert(sys.argv[2])
    elif sys.argv[1] == "--csv" and len(sys.argv) > 2:
        convert_to_csv(sys.argv[2])
    else:
        input_file = sys.argv[1]
        if not os.path.exists(input_file):
            print(f"错误: 文件不存在 - {input_file}")
            sys.exit(1)
        convert_tencent_excel(input_file)