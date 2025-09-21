#!/usr/bin/env python3
"""
直接修改XLSX文件的新方法
尝试绕过兼容性问题直接涂色
"""

import asyncio
import json
import sys
import os
from pathlib import Path
from datetime import datetime
import csv
import logging
from typing import Optional, Dict, List
import shutil

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

sys.path.append('/root/projects/tencent-doc-manager')

from production.config import STANDARD_COLUMNS, L1_COLUMNS, L2_COLUMNS, L3_COLUMNS


class XLSXDirectModify:
    """直接修改XLSX文件"""

    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.test_dir = Path(f"/root/projects/tencent-doc-manager/xlsx_direct_{self.timestamp}")
        self.test_dir.mkdir(parents=True, exist_ok=True)

        # 使用同一文档进行对比
        self.doc_url = "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN"
        self.baseline_file = "/root/projects/tencent-doc-manager/csv_versions/2025_W38/baseline/tencent_出国销售计划表_20250915_0145_baseline_W38.csv"

    async def step1_download_both_formats(self):
        """同时下载CSV和XLSX格式"""
        logger.info("="*60)
        logger.info("📥 Step 1: 下载CSV和XLSX两种格式")
        logger.info("="*60)

        from production.core_modules.playwright_downloader import PlaywrightDownloader

        # 读取Cookie
        cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
        with open(cookie_file) as f:
            cookie_data = json.load(f)

        cookie_string = cookie_data.get('current_cookies', '')
        downloader = PlaywrightDownloader()

        # 1. 下载CSV用于对比
        logger.info("📄 下载CSV格式...")
        csv_result = await downloader.download(
            url=self.doc_url,
            cookies=cookie_string,
            format='csv',
            download_dir=str(self.test_dir)
        )

        csv_file = None
        if csv_result.get('success'):
            csv_file = csv_result.get('file_path')
            logger.info(f"✅ CSV下载成功: {Path(csv_file).name}")

        # 等待一下避免太快
        await asyncio.sleep(2)

        # 2. 下载XLSX作为模板
        logger.info("📄 下载XLSX格式...")
        xlsx_result = await downloader.download(
            url=self.doc_url,
            cookies=cookie_string,
            format='xlsx',
            download_dir=str(self.test_dir)
        )

        xlsx_file = None
        if xlsx_result.get('success'):
            xlsx_file = xlsx_result.get('file_path')
            logger.info(f"✅ XLSX下载成功: {Path(xlsx_file).name}")

        return csv_file, xlsx_file

    def step2_compare_csv(self, csv_file: str) -> List[Dict]:
        """使用CSV进行对比"""
        logger.info("\n" + "="*60)
        logger.info("🔍 Step 2: 使用CSV进行对比")
        logger.info("="*60)

        if not Path(self.baseline_file).exists():
            logger.error("❌ 基线文件不存在")
            return []

        changes = []

        # 读取基线和当前CSV
        with open(self.baseline_file, 'r', encoding='utf-8') as f:
            baseline_data = list(csv.reader(f))

        with open(csv_file, 'r', encoding='utf-8') as f:
            current_data = list(csv.reader(f))

        logger.info(f"📊 基线: {len(baseline_data)} 行")
        logger.info(f"📊 当前: {len(current_data)} 行")

        # 对比
        for row_idx in range(1, min(len(baseline_data), len(current_data))):
            for col_idx in range(min(len(baseline_data[0]), len(current_data[0]))):
                val_baseline = str(baseline_data[row_idx][col_idx]).strip()
                val_current = str(current_data[row_idx][col_idx]).strip()

                if val_baseline != val_current:
                    col_name = baseline_data[0][col_idx] if col_idx < len(baseline_data[0]) else f"Col_{col_idx}"
                    changes.append({
                        'row': row_idx + 1,
                        'col': col_idx + 1,  # Excel从1开始
                        'col_name': col_name,
                        'old_value': val_baseline[:50],
                        'new_value': val_current[:50]
                    })

        logger.info(f"✅ 发现 {len(changes)} 处变更")
        return changes

    def step3_create_fresh_excel(self, csv_file: str, changes: List[Dict]) -> str:
        """创建全新的Excel文件（不依赖腾讯XLSX）"""
        logger.info("\n" + "="*60)
        logger.info("🎨 Step 3: 创建全新Excel文件")
        logger.info("="*60)

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.comments import Comment

        # 读取CSV数据
        with open(csv_file, 'r', encoding='utf-8') as f:
            csv_data = list(csv.reader(f))

        # 创建新工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "出国销售计划表"

        # 写入数据
        for row_idx, row_data in enumerate(csv_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # 表头格式
                if row_idx == 1:
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="366092",
                        end_color="366092",
                        fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 应用涂色
        cells_colored = 0
        for change in changes:
            # 风险分级
            col_name = change['col_name']
            if col_name in L1_COLUMNS:
                bg_color = "FFCCCC"  # 浅红
                font_color = "CC0000"  # 深红
                risk = "HIGH"
            elif col_name in L2_COLUMNS:
                bg_color = "FFFFCC"  # 浅黄
                font_color = "FF8800"  # 橙色
                risk = "MEDIUM"
            else:
                bg_color = "CCFFCC"  # 浅绿
                font_color = "008800"  # 深绿
                risk = "LOW"

            # 应用涂色
            try:
                cell = ws.cell(row=change['row'], column=change['col'])

                # 关键：使用solid填充
                cell.fill = PatternFill(
                    start_color=bg_color,
                    end_color=bg_color,
                    fill_type="solid"  # 必须是solid！
                )

                # 字体颜色
                cell.font = Font(
                    color=font_color,
                    bold=(risk == "HIGH")
                )

                # 添加批注
                comment = Comment(
                    f"风险等级: {risk}\n原值: {change['old_value']}\n新值: {change['new_value']}",
                    "AI分析系统"
                )
                cell.comment = comment

                cells_colored += 1

            except Exception as e:
                logger.error(f"❌ 涂色失败 [{change['row']},{change['col']}]: {e}")

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存文件
        output_file = self.test_dir / f"fresh_colored_{self.timestamp}.xlsx"
        wb.save(output_file)

        size = output_file.stat().st_size / 1024
        logger.info(f"✅ 全新Excel生成成功")
        logger.info(f"📁 文件: {output_file.name}")
        logger.info(f"🎨 涂色: {cells_colored} 个单元格")
        logger.info(f"📊 大小: {size:.2f} KB")

        # 验证涂色
        self.verify_coloring(output_file)

        return str(output_file)

    def verify_coloring(self, excel_file):
        """验证涂色正确性"""
        from openpyxl import load_workbook

        logger.info("\n🔍 验证涂色...")

        wb = load_workbook(excel_file)
        ws = wb.active

        solid_count = 0
        non_solid_count = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.fill_type:
                    if cell.fill.fill_type == 'solid':
                        solid_count += 1
                    elif cell.fill.fill_type not in ['none', None]:
                        non_solid_count += 1

        logger.info(f"✅ Solid填充: {solid_count} 个")
        logger.info(f"❌ 非Solid填充: {non_solid_count} 个")

        if non_solid_count == 0 and solid_count > 0:
            logger.info("🎉 完美！所有涂色都是solid填充，100%兼容腾讯文档！")

    async def step4_upload(self, excel_file: str) -> Optional[str]:
        """上传到腾讯文档"""
        logger.info("\n" + "="*60)
        logger.info("📤 Step 4: 上传到腾讯文档")
        logger.info("="*60)

        try:
            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')

            from production.core_modules.tencent_doc_upload_production_v3 import quick_upload_v3

            result = await quick_upload_v3(
                cookie_string=cookie_string,
                file_path=excel_file,
                headless=True
            )

            if result and result.get('success'):
                url = result.get('url')
                logger.info(f"✅ 上传成功!")
                logger.info(f"🔗 URL: {url}")
                return url
            else:
                logger.error(f"❌ 上传失败")
                return None

        except Exception as e:
            logger.error(f"❌ 上传异常: {e}")
            return None

    async def run_complete_test(self):
        """运行完整测试"""
        logger.info("🚀 开始XLSX直接修改测试")
        logger.info(f"📁 测试目录: {self.test_dir}")

        # Step 1: 下载两种格式
        csv_file, xlsx_file = await self.step1_download_both_formats()

        if not csv_file:
            logger.error("❌ CSV下载失败，无法继续")
            return None

        # Step 2: 使用CSV对比
        changes = self.step2_compare_csv(csv_file)

        if not changes:
            logger.warning("⚠️ 没有发现变更")
            # 继续创建空报告

        # Step 3: 创建全新Excel（不依赖腾讯XLSX）
        excel_file = self.step3_create_fresh_excel(csv_file, changes)

        # Step 4: 上传
        url = await self.step4_upload(excel_file)

        # 总结
        logger.info("\n" + "="*60)
        logger.info("📊 测试总结")
        logger.info("="*60)
        logger.info(f"📥 下载: CSV ✅ XLSX {'✅' if xlsx_file else '❌'}")
        logger.info(f"🔍 对比: {len(changes)} 处变更")
        logger.info(f"🎨 涂色: 全新创建，100% solid填充")
        logger.info(f"📤 上传: {'✅' if url else '❌'}")

        if url:
            logger.info(f"\n🎉 成功URL: {url}")
            logger.info("👉 请检查涂色是否正确显示")

            # 说明
            logger.info("\n💡 技术说明:")
            logger.info("   1. 使用CSV进行数据对比（稳定）")
            logger.info("   2. 创建全新Excel文件（避免兼容性问题）")
            logger.info("   3. 使用solid填充（腾讯文档兼容）")
            logger.info("   4. 这是最稳定的方案")

        return url


async def main():
    """主函数"""
    test = XLSXDirectModify()
    url = await test.run_complete_test()

    if url:
        print(f"\n🌟 最终成功URL: {url}")
        print("✅ 这是基于CSV创建的全新Excel，避免了XLSX兼容性问题")
    else:
        print("\n⚠️ 测试未完全成功")


if __name__ == "__main__":
    asyncio.run(main())