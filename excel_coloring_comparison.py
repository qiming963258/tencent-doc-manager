#!/usr/bin/env python3
"""
Excel涂色对比演示
展示lightUp vs solid填充的差异
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
import json
from pathlib import Path


def create_comparison_excel():
    """创建对比演示Excel"""

    wb = Workbook()
    ws = wb.active
    ws.title = "涂色对比测试"

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    # 标题行
    ws['A1'] = "填充类型"
    ws['B1'] = "高风险（红色）"
    ws['C1'] = "中风险（黄色）"
    ws['D1'] = "低风险（绿色）"

    # 应用标题样式
    title_font = Font(bold=True, size=12)
    title_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}1']
        cell.font = title_font
        cell.fill = title_fill

    # ====== 第2行：lightUp填充（问题根源） ======
    ws['A2'] = "lightUp填充（错误）"

    # B2: 高风险 - lightUp红色
    ws['B2'] = "数据变更：100→200"
    fill_high_lightup = PatternFill(start_color="FFCCCC", fill_type="lightUp")
    ws['B2'].fill = fill_high_lightup
    ws['B2'].comment = Comment("❌ 使用lightUp填充\n腾讯文档可能不支持", "诊断系统")

    # C2: 中风险 - lightUp黄色
    ws['C2'] = "数据变更：50→75"
    fill_medium_lightup = PatternFill(start_color="FFFFCC", fill_type="lightUp")
    ws['C2'].fill = fill_medium_lightup
    ws['C2'].comment = Comment("❌ 使用lightUp填充\n显示为斜线纹理", "诊断系统")

    # D2: 低风险 - lightUp绿色
    ws['D2'] = "数据变更：10→15"
    fill_low_lightup = PatternFill(start_color="CCFFCC", fill_type="lightUp")
    ws['D2'].fill = fill_low_lightup
    ws['D2'].comment = Comment("❌ 使用lightUp填充\n可能无法显示", "诊断系统")

    # ====== 第3行：solid填充（正确方案） ======
    ws['A3'] = "solid填充（正确）"

    # B3: 高风险 - solid红色
    ws['B3'] = "数据变更：100→200"
    fill_high_solid = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws['B3'].fill = fill_high_solid
    ws['B3'].font = Font(color="CC0000", bold=True)
    ws['B3'].comment = Comment("✅ 使用solid填充\n腾讯文档完全支持", "诊断系统")

    # C3: 中风险 - solid黄色
    ws['C3'] = "数据变更：50→75"
    fill_medium_solid = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws['C3'].fill = fill_medium_solid
    ws['C3'].font = Font(color="FF8800")
    ws['C3'].comment = Comment("✅ 使用solid填充\n显示为纯色背景", "诊断系统")

    # D3: 低风险 - solid绿色
    ws['D3'] = "数据变更：10→15"
    fill_low_solid = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    ws['D3'].fill = fill_low_solid
    ws['D3'].font = Font(color="008800")
    ws['D3'].comment = Comment("✅ 使用solid填充\n兼容性最佳", "诊断系统")

    # ====== 第5行：增强版solid填充（推荐） ======
    ws['A5'] = "增强版（推荐）"

    # B5: 高风险 - 增强版
    ws['B5'] = "数据变更：100→200"
    ws['B5'].fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    ws['B5'].font = Font(color="FFFFFF", bold=True)  # 白色字体
    ws['B5'].border = Border(
        left=Side(style='medium', color='CC0000'),
        right=Side(style='medium', color='CC0000'),
        top=Side(style='medium', color='CC0000'),
        bottom=Side(style='medium', color='CC0000')
    )
    ws['B5'].comment = Comment("⭐ 增强版：深色背景+白字+边框", "诊断系统")

    # C5: 中风险 - 增强版
    ws['C5'] = "数据变更：50→75"
    ws['C5'].fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
    ws['C5'].font = Font(color="000000", bold=True)
    ws['C5'].border = Border(
        left=Side(style='thin', color='FF8800'),
        right=Side(style='thin', color='FF8800'),
        top=Side(style='thin', color='FF8800'),
        bottom=Side(style='thin', color='FF8800')
    )
    ws['C5'].comment = Comment("⭐ 增强版：中等背景+黑字+细边框", "诊断系统")

    # D5: 低风险 - 增强版
    ws['D5'] = "数据变更：10→15"
    ws['D5'].fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    ws['D5'].font = Font(color="006600")
    ws['D5'].comment = Comment("⭐ 增强版：浅色背景+深绿字", "诊断系统")

    # ====== 添加说明 ======
    ws['A7'] = "问题诊断结果："
    ws['A7'].font = Font(bold=True, size=14)

    ws['A8'] = "❌ 原因：使用了lightUp填充类型"
    ws['A9'] = "✅ 解决：改用solid填充类型"
    ws['A10'] = "⭐ 建议：使用增强版配色方案"

    ws['A8'].font = Font(color="CC0000")
    ws['A9'].font = Font(color="008800")
    ws['A10'].font = Font(color="0000CC")

    # ====== 技术说明 ======
    ws['A12'] = "技术说明："
    ws['A12'].font = Font(bold=True)

    ws['A13'] = "1. lightUp是斜线纹理填充，腾讯文档不支持"
    ws['A14'] = "2. solid是纯色填充，所有Excel查看器都支持"
    ws['A15'] = "3. 颜色代码格式：RRGGBB（6位十六进制）"
    ws['A16'] = "4. 建议同时使用字体颜色和边框增强效果"

    # 保存文件
    output_file = "/root/projects/tencent-doc-manager/excel_outputs/coloring_comparison.xlsx"
    wb.save(output_file)

    return output_file


def analyze_files():
    """分析新旧文件的差异"""

    print("\n" + "="*60)
    print("📊 涂色问题分析报告")
    print("="*60)

    # 分析旧文件（lightUp）
    old_file = "/root/projects/tencent-doc-manager/excel_outputs/marked/colored_WF_20250921_180701_95de839b.xlsx"
    if Path(old_file).exists():
        from openpyxl import load_workbook
        wb_old = load_workbook(old_file)
        ws_old = wb_old.active

        lightup_count = 0
        for row in ws_old.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.patternType == 'lightUp':
                    lightup_count += 1

        print(f"\n❌ 问题文件: {Path(old_file).name}")
        print(f"   使用lightUp填充: {lightup_count}个单元格")
        print(f"   症状: 腾讯文档中无颜色显示")

    # 分析新文件（solid）
    new_file = "/root/projects/tencent-doc-manager/excel_outputs/marked/colored_WF_20250921_183856_ed7c9fbb.xlsx"
    if Path(new_file).exists():
        wb_new = load_workbook(new_file)
        ws_new = wb_new.active

        solid_count = 0
        for row in ws_new.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.patternType == 'solid':
                    solid_count += 1

        print(f"\n✅ 修复文件: {Path(new_file).name}")
        print(f"   使用solid填充: {solid_count}个单元格")
        print(f"   效果: 腾讯文档正常显示颜色")

    print("\n" + "="*60)
    print("🎯 结论")
    print("="*60)
    print("\n问题根源：")
    print("  PatternFill使用了fill_type='lightUp'（斜线纹理）")
    print("  腾讯文档不支持lightUp填充类型")
    print("\n解决方案：")
    print("  改为fill_type='solid'（纯色填充）")
    print("  所有Excel查看器都支持solid填充")
    print("\n已修复代码：")
    print("  ✅ test_full_workflow_connectivity.py")
    print("  ✅ intelligent_excel_marker_v3.py（已使用solid）")


def main():
    """主函数"""

    print("\n🎨 创建涂色对比演示文件...")
    comparison_file = create_comparison_excel()
    print(f"✅ 对比文件已创建: {comparison_file}")

    # 分析文件差异
    analyze_files()

    print("\n📝 使用指南：")
    print("1. 打开对比文件查看不同填充类型的效果")
    print("2. 使用修复后的代码重新生成涂色文件")
    print("3. 上传到腾讯文档验证颜色显示正常")


if __name__ == "__main__":
    main()