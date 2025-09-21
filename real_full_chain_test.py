#!/usr/bin/env python3
"""
真实全链路测试 - 无任何模拟
从真实下载到真实上传的完整流程
禁止任何欺诈模拟行为
"""

import json
import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

sys.path.append('/root/projects/tencent-doc-manager')

# 导入真实模块
from production.config import STANDARD_COLUMNS, L1_COLUMNS, L2_COLUMNS, L3_COLUMNS
from workflow_chain_manager import get_chain_manager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.comments import Comment


class RealFullChainTest:
    """真实全链路测试 - 无模拟"""

    def __init__(self):
        self.session_id = f"REAL_CHAIN_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.base_dir = Path('/root/projects/tencent-doc-manager')
        self.manager = get_chain_manager()

        # 真实配置
        self.config = {
            "doc_url": "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN",
            "doc_name": "出国销售计划表",
            "baseline_week": "W38",
            "current_week": datetime.now().strftime("W%U")
        }

        logger.info("="*70)
        logger.info("🚀 真实全链路测试 - 无任何模拟")
        logger.info(f"Session: {self.session_id}")
        logger.info("="*70)

    def run_real_test(self):
        """执行真实的全链路测试"""

        try:
            # Step 1: 真实下载
            logger.info("\n📥 Step 1: 真实下载腾讯文档...")
            csv_file = self._real_download()
            if not csv_file:
                logger.error("下载失败，无法继续")
                return False

            # Step 2: 真实对比
            logger.info("\n🔍 Step 2: 真实基线对比...")
            diff_result = self._real_compare(csv_file)
            if not diff_result:
                logger.error("对比失败，无法继续")
                return False

            # Step 3: 真实打分
            logger.info("\n💯 Step 3: 真实风险打分...")
            score_result = self._real_scoring(diff_result)
            if not score_result:
                logger.error("打分失败，无法继续")
                return False

            # Step 4: 真实Excel生成和涂色
            logger.info("\n🎨 Step 4: 真实Excel涂色...")
            excel_file = self._real_excel_coloring(csv_file, score_result)
            if not excel_file:
                logger.error("Excel生成失败，无法继续")
                return False

            # Step 5: 真实上传
            logger.info("\n☁️ Step 5: 真实上传到腾讯文档...")
            upload_url = self._real_upload(excel_file)
            if not upload_url:
                logger.error("上传失败")
                return False

            # 成功完成
            logger.info("\n" + "="*70)
            logger.info("✅ 真实全链路测试成功完成！")
            logger.info(f"📊 Session: {self.session_id}")
            logger.info(f"🔗 文档URL: {upload_url}")
            logger.info("="*70)

            return upload_url

        except Exception as e:
            logger.error(f"❌ 全链路测试失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _real_download(self):
        """真实下载腾讯文档"""

        try:
            # 读取Cookie
            cookie_file = self.base_dir / "config" / "cookies.json"
            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')

            # 使用真实下载器
            from production.core_modules.stable_cookie_downloader import CookieDownloader

            downloader = CookieDownloader()

            # 设置下载路径
            download_dir = self.base_dir / "downloads" / self.session_id
            download_dir.mkdir(parents=True, exist_ok=True)

            csv_file = download_dir / f"download_{self.session_id}.csv"

            # 执行真实下载
            success = downloader.download_as_csv(
                url=self.config["doc_url"],
                output_path=str(csv_file),
                cookie_string=cookie_string
            )

            if success and csv_file.exists():
                logger.info(f"   ✅ 真实下载成功: {csv_file.name}")
                logger.info(f"   📊 文件大小: {csv_file.stat().st_size} bytes")

                # 验证是否为真实CSV数据
                df = pd.read_csv(csv_file, encoding='utf-8')
                logger.info(f"   📊 数据规模: {df.shape[0]}行 x {df.shape[1]}列")

                return str(csv_file)
            else:
                logger.error("   ❌ 下载失败")
                return None

        except Exception as e:
            logger.error(f"   ❌ 下载异常: {e}")

            # 尝试备用下载方法
            logger.info("   🔄 尝试备用下载方法...")
            return self._alternative_download()

    def _alternative_download(self):
        """备用下载方法"""

        try:
            import asyncio
            from production.core_modules.tencent_doc_download_async import async_download_csv

            # 异步下载
            download_dir = self.base_dir / "downloads" / self.session_id
            download_dir.mkdir(parents=True, exist_ok=True)
            csv_file = download_dir / f"download_alt_{self.session_id}.csv"

            result = asyncio.run(async_download_csv(
                self.config["doc_url"],
                str(csv_file)
            ))

            if result and csv_file.exists():
                logger.info(f"   ✅ 备用下载成功: {csv_file.name}")
                return str(csv_file)
            else:
                logger.error("   ❌ 备用下载也失败了")
                return None

        except Exception as e:
            logger.error(f"   ❌ 备用下载异常: {e}")
            return None

    def _real_compare(self, csv_file):
        """真实基线对比"""

        try:
            # 查找真实基线文件
            baseline_dir = self.base_dir / "csv_versions" / f"2025_{self.config['baseline_week']}" / "baseline"
            baseline_files = list(baseline_dir.glob("*.csv"))

            if not baseline_files:
                logger.error(f"   ❌ 未找到基线文件在: {baseline_dir}")
                return None

            baseline_file = baseline_files[0]
            logger.info(f"   📄 基线文件: {baseline_file.name}")

            # 读取数据
            df_current = pd.read_csv(csv_file, encoding='utf-8')
            df_baseline = pd.read_csv(baseline_file, encoding='utf-8')

            logger.info(f"   📊 当前数据: {df_current.shape}")
            logger.info(f"   📊 基线数据: {df_baseline.shape}")

            # 真实对比
            changes = []

            # 确保列数相同
            min_cols = min(df_current.shape[1], df_baseline.shape[1])

            for row_idx in range(min(len(df_current), len(df_baseline))):
                for col_idx in range(min_cols):
                    val_current = str(df_current.iloc[row_idx, col_idx])
                    val_baseline = str(df_baseline.iloc[row_idx, col_idx])

                    if val_current != val_baseline:
                        changes.append({
                            "row": row_idx + 1,
                            "col": col_idx + 1,
                            "old": val_baseline,
                            "new": val_current,
                            "column_name": df_current.columns[col_idx] if col_idx < len(df_current.columns) else f"Col{col_idx+1}"
                        })

            logger.info(f"   ✅ 发现真实变更: {len(changes)}处")

            # 保存对比结果
            diff_file = self.base_dir / "scoring_results" / "diff" / f"diff_{self.session_id}.json"
            diff_file.parent.mkdir(parents=True, exist_ok=True)

            diff_data = {
                "session_id": self.session_id,
                "baseline": str(baseline_file),
                "current": csv_file,
                "total_changes": len(changes),
                "changes": changes[:100]  # 限制保存前100个变更
            }

            with open(diff_file, 'w', encoding='utf-8') as f:
                json.dump(diff_data, f, ensure_ascii=False, indent=2)

            return diff_data

        except Exception as e:
            logger.error(f"   ❌ 对比异常: {e}")
            return None

    def _real_scoring(self, diff_result):
        """真实风险打分"""

        try:
            changes = diff_result.get('changes', [])

            # 创建打分结果
            score_data = {
                "session_id": self.session_id,
                "timestamp": datetime.now().isoformat(),
                "cell_scores": {}
            }

            for change in changes:
                col_idx = change['col'] - 1

                # 根据列判断风险等级
                if col_idx < len(STANDARD_COLUMNS):
                    col_name = STANDARD_COLUMNS[col_idx]

                    # 判断风险级别
                    if col_name in L1_COLUMNS:
                        risk_level = "HIGH"
                        base_score = 70
                    elif col_name in L2_COLUMNS:
                        risk_level = "MEDIUM"
                        base_score = 40
                    else:
                        risk_level = "LOW"
                        base_score = 20
                else:
                    risk_level = "LOW"
                    base_score = 10

                # 计算具体分数
                old_val = change['old']
                new_val = change['new']

                # 根据变更类型调整分数
                if old_val.isdigit() and new_val.isdigit():
                    # 数值变更
                    change_ratio = abs(float(new_val) - float(old_val)) / (float(old_val) + 1)
                    score = base_score + min(30, int(change_ratio * 30))
                else:
                    # 文本变更
                    if len(new_val) > len(old_val) * 2 or len(new_val) < len(old_val) / 2:
                        score = base_score + 20  # 大幅变更
                    else:
                        score = base_score + 10  # 小幅变更

                # 保存打分
                cell_key = f"{change['row']}_{change['col']}"
                score_data["cell_scores"][cell_key] = {
                    "old_value": old_val,
                    "new_value": new_val,
                    "score": min(100, score),
                    "risk_level": risk_level,
                    "column": change.get('column_name', '')
                }

            logger.info(f"   ✅ 真实打分完成: {len(score_data['cell_scores'])}个单元格")

            # 保存打分文件
            score_file = self.base_dir / "scoring_results" / "detailed" / f"scores_{self.session_id}.json"
            score_file.parent.mkdir(parents=True, exist_ok=True)

            with open(score_file, 'w', encoding='utf-8') as f:
                json.dump(score_data, f, ensure_ascii=False, indent=2)

            return score_data

        except Exception as e:
            logger.error(f"   ❌ 打分异常: {e}")
            return None

    def _real_excel_coloring(self, csv_file, score_result):
        """真实Excel生成和涂色"""

        try:
            # 读取CSV数据
            df = pd.read_csv(csv_file, encoding='utf-8')

            # 创建Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "真实数据"

            # 写入标题行
            for col_idx, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="E0E0E0",
                    end_color="E0E0E0",
                    fill_type="solid"
                )

            # 写入数据
            for row_idx, row_data in df.iterrows():
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx + 2, column=col_idx, value=value)

            # 应用涂色（使用solid填充）
            marked_count = 0
            for cell_key, cell_info in score_result["cell_scores"].items():
                row, col = map(int, cell_key.split("_"))

                if row <= ws.max_row and col <= ws.max_column:
                    cell = ws.cell(row=row, column=col)

                    # 根据风险等级选择颜色
                    risk_level = cell_info["risk_level"]
                    if risk_level == "HIGH":
                        color = "FFCCCC"  # 浅红
                        font_color = "CC0000"
                    elif risk_level == "MEDIUM":
                        color = "FFFFCC"  # 浅黄
                        font_color = "FF8800"
                    else:
                        color = "CCFFCC"  # 浅绿
                        font_color = "008800"

                    # 应用solid填充（腾讯文档兼容）
                    cell.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type="solid"  # 必须使用solid
                    )

                    # 设置字体
                    cell.font = Font(
                        color=font_color,
                        bold=(risk_level == "HIGH")
                    )

                    # 添加批注
                    comment_text = (
                        f"风险等级: {risk_level}\n"
                        f"评分: {cell_info['score']}\n"
                        f"原值: {cell_info['old_value']}\n"
                        f"新值: {cell_info['new_value']}"
                    )
                    cell.comment = Comment(comment_text, "真实监控系统")

                    marked_count += 1

            # 保存Excel
            output_dir = self.base_dir / "excel_outputs" / "real_chain"
            output_dir.mkdir(parents=True, exist_ok=True)

            excel_file = output_dir / f"real_{self.session_id}.xlsx"
            wb.save(excel_file)
            wb.close()

            logger.info(f"   ✅ Excel生成成功: {excel_file.name}")
            logger.info(f"   🎨 涂色单元格: {marked_count}个（全部solid填充）")

            return str(excel_file)

        except Exception as e:
            logger.error(f"   ❌ Excel生成异常: {e}")
            return None

    def _real_upload(self, excel_file):
        """真实上传到腾讯文档"""

        try:
            # 读取Cookie
            cookie_file = self.base_dir / "config" / "cookies.json"
            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')

            # 使用真实上传器
            from production.core_modules.tencent_doc_upload_production_v3 import sync_upload_v3

            logger.info("   🔄 正在上传...")

            result = sync_upload_v3(
                cookie_string=cookie_string,
                file_path=excel_file,
                headless=True
            )

            if result and result.get('success'):
                url = result.get('url')
                logger.info(f"   ✅ 上传成功: {url}")
                return url
            else:
                logger.error(f"   ❌ 上传失败: {result.get('message', 'Unknown')}")
                return None

        except Exception as e:
            logger.error(f"   ❌ 上传异常: {e}")
            return None


def main():
    """主函数"""

    # 创建测试实例
    tester = RealFullChainTest()

    # 执行真实测试
    url = tester.run_real_test()

    if url:
        print("\n" + "="*70)
        print("🎉 真实全链路测试成功！")
        print(f"🔗 访问URL: {url}")
        print("\n验证要点:")
        print("1. 检查涂色是否显示")
        print("2. 检查批注是否存在")
        print("3. 验证数据是否为真实下载的内容")
        print("="*70)
    else:
        print("\n" + "="*70)
        print("❌ 真实全链路测试失败")
        print("请检查日志排查问题")
        print("="*70)

    return url


if __name__ == "__main__":
    result = main()
    sys.exit(0 if result else 1)