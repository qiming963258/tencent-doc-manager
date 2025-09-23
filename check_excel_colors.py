#!/usr/bin/env python3
"""检查Excel文件的涂色情况"""

from openpyxl import load_workbook
import sys

excel_file = "/root/projects/tencent-doc-manager/excel_outputs/marked/tencent_111测试版本-出国销售计划表_20250922_0228_midweek_W39_marked_20250922_022809_W00.xlsx"

wb = load_workbook(excel_file)
ws = wb.active

colored_cells = []
for row_idx, row in enumerate(ws.iter_rows(), 1):
    for col_idx, cell in enumerate(row, 1):
        if cell.fill.start_color.index and cell.fill.start_color.index not in ['00000000', None]:
            colored_cells.append({
                'cell': cell.coordinate,
                'color': cell.fill.start_color.index,
                'value': cell.value,
                'fill_type': cell.fill.fill_type
            })

print(f"文件: {excel_file.split('/')[-1]}")
print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")
print(f"涂色单元格数: {len(colored_cells)}")

if colored_cells:
    print("\n前10个涂色单元格:")
    for i, info in enumerate(colored_cells[:10], 1):
        print(f"{i}. {info['cell']}: 颜色={info['color']}, 填充类型={info['fill_type']}, 值={info['value']}")

# 检查fill_type
fill_types = set(c['fill_type'] for c in colored_cells)
print(f"\n使用的填充类型: {fill_types}")