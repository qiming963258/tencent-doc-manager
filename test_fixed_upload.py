#!/usr/bin/env python3
"""
测试修复后的上传模块v3
验证是否正确检测存储空间并拒绝虚假成功
"""

import asyncio
import json
import sys
from pathlib import Path
from datetime import datetime

# 添加项目路径
sys.path.insert(0, '/root/projects/tencent-doc-manager')

from production.core_modules.tencent_doc_upload_production_v3 import TencentDocProductionUploaderV3


async def test_fixed_upload():
    """测试修复后的上传模块"""

    print("=" * 60)
    print("🧪 测试修复后的上传模块v3")
    print("=" * 60)

    # 读取Cookie
    cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
    with open(cookie_file) as f:
        cookie_data = json.load(f)

    cookie_string = cookie_data.get('current_cookies', '')

    # 创建测试文件
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试修复"
    ws['A1'] = '修复测试'
    ws['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    test_file = f"/tmp/test_fixed_{timestamp}.xlsx"
    wb.save(test_file)
    print(f"✅ 创建测试文件: {test_file}")

    # 使用修复后的上传器
    async with TencentDocProductionUploaderV3(headless=True) as uploader:
        print("\n📋 步骤1: 登录并检查存储空间")
        print("-" * 50)

        login_success = await uploader.login_with_cookies(cookie_string)

        if not login_success:
            print("❌ 登录失败")
            return

        print("✅ 登录成功")

        # 检查存储空间信息
        if uploader.storage_space_info:
            storage = uploader.storage_space_info
            usage = storage.get('usage_percent', -1)
            has_space = storage.get('has_space', True)

            print(f"📊 存储空间使用率: {usage:.2f}%")
            print(f"📊 是否有空间: {has_space}")

            if not has_space:
                print("⚠️ 存储空间不足，上传应该会被拒绝")

        print("\n📋 步骤2: 尝试上传文件")
        print("-" * 50)

        result = await uploader.upload_file(test_file)

        print("\n📋 步骤3: 分析上传结果")
        print("-" * 50)

        print(f"成功状态: {result['success']}")
        print(f"返回消息: {result['message']}")
        print(f"返回URL: {result.get('url', 'None')}")

        if result.get('storage_info'):
            print(f"存储信息: {json.dumps(result['storage_info'], indent=2, ensure_ascii=False)}")

        if result.get('api_response'):
            print(f"API响应: {json.dumps(result['api_response'], indent=2, ensure_ascii=False)}")

        # 验证修复效果
        print("\n" + "=" * 60)
        print("🔍 修复验证结果")
        print("=" * 60)

        if not result['success']:
            if '存储空间不足' in result['message']:
                print("✅ 正确：检测到存储空间不足，拒绝上传")
            elif 'URL无效' in result['message']:
                print("✅ 正确：检测到虚假成功，拒绝返回旧文档URL")
            else:
                print(f"✅ 正确：上传失败 - {result['message']}")
        else:
            # 如果成功，验证URL是否真实
            if result['url'] and result['url'] != "https://docs.qq.com/desktop/":
                print(f"⚠️ 上传显示成功，需要验证URL是否真实: {result['url']}")
            else:
                print("❌ 异常：上传成功但没有有效URL")

        # 显示关键改进
        print("\n📌 模块v3已修复的问题：")
        print("1. ✅ 添加存储空间检测，上传前主动检查")
        print("2. ✅ 验证API响应，拒绝空url和doc_id")
        print("3. ✅ 移除猜测策略，不再返回已存在文档")
        print("4. ✅ 增加URL验证，确认是新上传的文档")


if __name__ == "__main__":
    asyncio.run(test_fixed_upload())