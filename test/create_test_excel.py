#!/usr/bin/env python3
"""
创建测试用Excel文件，用于验证Excel MCP服务器功能
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.chart import LineChart, Reference
import os

def create_test_excel():
    """创建包含多种格式的测试Excel文件"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试数据"
    
    # 添加标题行
    headers = ["序号", "项目名称", "负责人", "状态", "风险等级", "完成进度", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # 添加测试数据
    test_data = [
        [1, "腾讯文档监控系统", "张三", "进行中", "L1", 85, "核心系统正常运行"],
        [2, "热力图优化", "李四", "完成", "L3", 100, "性能提升显著"],
        [3, "AI语义分析", "王五", "暂停", "L2", 60, "等待API接口"],
        [4, "Excel MCP集成", "赵六", "进行中", "L1", 75, "测试阶段"],
        [5, "数据安全审计", "孙七", "计划中", "L2", 10, "下周开始"]
    ]
    
    # 填充数据并设置格式
    for row, data in enumerate(test_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            
            # 根据风险等级设置颜色
            if col == 5:  # 风险等级列
                if value == "L1":
                    cell.fill = PatternFill(start_color="FFD5D5", end_color="FFD5D5", fill_type="solid")
                elif value == "L2":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif value == "L3":
                    cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
            
            # 根据状态设置字体颜色
            if col == 4:  # 状态列
                if value == "完成":
                    cell.font = Font(color="008000")
                elif value == "暂停":
                    cell.font = Font(color="FF0000")
                elif value == "进行中":
                    cell.font = Font(color="0066CC")
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
    
    # 调整列宽
    column_widths = [8, 20, 12, 12, 12, 12, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 创建第二个工作表 - 图表数据
    ws2 = wb.create_sheet("图表数据")
    chart_data = [
        ["月份", "完成项目数", "风险项目数"],
        ["1月", 8, 2],
        ["2月", 12, 3],
        ["3月", 15, 1],
        ["4月", 18, 4],
        ["5月", 22, 2]
    ]
    
    for row, data in enumerate(chart_data, 1):
        for col, value in enumerate(data, 1):
            ws2.cell(row=row, column=col, value=value)
    
    # 创建图表
    chart = LineChart()
    chart.title = "项目进度统计"
    chart.style = 13
    chart.x_axis.title = "月份"
    chart.y_axis.title = "项目数量"
    
    data = Reference(ws2, min_col=2, min_row=1, max_col=3, max_row=6)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws2.add_chart(chart, "E2")
    
    # 保存文件
    output_path = "/root/projects/tencent-doc-manager/excel_test/test_excel_file.xlsx"
    wb.save(output_path)
    print(f"✅ 测试Excel文件已创建: {output_path}")
    
    return output_path

if __name__ == "__main__":
    create_test_excel()