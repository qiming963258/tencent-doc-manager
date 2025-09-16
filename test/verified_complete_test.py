#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
腾讯文档完整流程真实测试
下载 → Excel修改 → 验证
确保每一步都是真实可验证的
"""

import sys
import os
import json
import asyncio
from datetime import datetime
import time
from pathlib import Path

# 添加浏览器自动化工具路径
sys.path.append('/root/projects/tencent-doc-manager/测试版本-性能优化开发-20250811-001430')

async def step1_real_download():
    """步骤1: 真实下载腾讯文档Excel格式"""
    print("🎯 步骤1: 真实下载腾讯文档Excel格式")
    print("=" * 60)
    
    try:
        from tencent_export_automation import TencentDocAutoExporter
        
        # 创建新的测试目录
        timestamp = datetime.now().strftime('%H%M%S')
        download_dir = f"/root/projects/tencent-doc-manager/real_test_results/verified_test_{timestamp}"
        os.makedirs(download_dir, exist_ok=True)
        
        # 读取Cookie
        print("🔐 读取认证信息...")
        with open('/root/projects/参考/cookie', 'r') as f:
            content = f.read()
            # 提取cookie行
            lines = content.strip().split('\n')
            cookies = ""
            for line in lines:
                if line.startswith('fingerprint='):
                    cookies = line
                    break
        
        print(f"📍 Cookie长度: {len(cookies)} 字符")
        
        # 创建导出器
        exporter = TencentDocAutoExporter(download_dir=download_dir)
        
        print("🚀 启动浏览器...")
        await exporter.start_browser(headless=True)
        
        print("🔐 加载认证...")
        await exporter.login_with_cookies(cookies)
        
        # 下载文档 - 小红书部门
        doc_url = "https://docs.qq.com/sheet/DWEVjZndkR2xVSWJN"
        print(f"📥 开始下载Excel格式...")
        print(f"🌐 目标文档: {doc_url}")
        
        start_time = time.time()
        result = await exporter.auto_export_document(doc_url, export_format="excel")
        end_time = time.time()
        
        if result and len(result) > 0:
            file_path = result[0]
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"✅ 下载成功！")
                print(f"📁 文件路径: {file_path}")
                print(f"📏 文件大小: {file_size} bytes")
                print(f"⏱️ 下载用时: {end_time - start_time:.2f} 秒")
                
                # 清理
                if exporter.browser:
                    await exporter.browser.close()
                if exporter.playwright:
                    await exporter.playwright.stop()
                
                return {
                    'success': True,
                    'file_path': file_path,
                    'file_size': file_size,
                    'download_time': end_time - start_time,
                    'timestamp': datetime.now().isoformat()
                }
            else:
                print("❌ 文件不存在")
                return {'success': False, 'error': '文件不存在'}
        else:
            print("❌ 下载失败")
            return {'success': False, 'error': '下载返回空结果'}
            
    except Exception as e:
        print(f"❌ 下载异常: {e}")
        return {'success': False, 'error': str(e)}

def step2_correct_excel_modification(file_path):
    """步骤2: 正确修改Excel文件，保持数据完整性"""
    print(f"\n🛠️ 步骤2: 正确修改Excel文件")
    print("=" * 60)
    
    try:
        # 使用openpyxl库进行正确的Excel修改
        import openpyxl
        
        print(f"📂 原始文件: {file_path}")
        original_size = os.path.getsize(file_path)
        print(f"📏 原始大小: {original_size} bytes")
        
        # 加载工作簿
        print("📖 加载Excel工作簿...")
        wb = openpyxl.load_workbook(file_path)
        
        # 获取活动工作表
        ws = wb.active
        print(f"📊 工作表名称: {ws.title}")
        
        # 检查A1单元格的原始内容
        original_a1 = ws['A1'].value
        print(f"📝 A1原始内容: {original_a1}")
        
        # 修改A1单元格 - 添加测试标识
        if original_a1:
            new_a1_value = f"[已测试]{original_a1}"
        else:
            new_a1_value = "[已测试]"
        
        ws['A1'] = new_a1_value
        print(f"✏️ A1修改后内容: {new_a1_value}")
        
        # 添加修改标记到B1单元格
        ws['B1'] = f"修改时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 保存修改后的文件
        modified_path = file_path.replace('.xlsx', '_真实修改.xlsx')
        wb.save(modified_path)
        wb.close()
        
        # 验证修改后的文件
        modified_size = os.path.getsize(modified_path)
        print(f"💾 修改后文件: {modified_path}")
        print(f"📏 修改后大小: {modified_size} bytes")
        
        # 数据完整性检查
        size_difference = abs(modified_size - original_size)
        size_change_percent = (size_difference / original_size) * 100
        
        if size_change_percent < 5:  # 允许5%的大小变化
            print(f"✅ 数据完整性检查通过 (大小变化: {size_change_percent:.1f}%)")
            integrity_ok = True
        else:
            print(f"⚠️ 数据完整性警告 (大小变化: {size_change_percent:.1f}%)")
            integrity_ok = False
        
        # 验证修改内容
        wb_verify = openpyxl.load_workbook(modified_path)
        ws_verify = wb_verify.active
        verified_a1 = ws_verify['A1'].value
        verified_b1 = ws_verify['B1'].value
        wb_verify.close()
        
        print(f"🔍 验证A1: {verified_a1}")
        print(f"🔍 验证B1: {verified_b1}")
        
        return {
            'success': True,
            'modified_file_path': modified_path,
            'original_size': original_size,
            'modified_size': modified_size,
            'size_change_percent': size_change_percent,
            'data_integrity_ok': integrity_ok,
            'original_a1': original_a1,
            'modified_a1': verified_a1,
            'modification_timestamp': verified_b1,
            'timestamp': datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"❌ Excel修改失败: {e}")
        return {'success': False, 'error': str(e)}

async def main():
    """主测试函数"""
    print("🎯 腾讯文档完整流程真实测试")
    print("=" * 60)
    print("📋 测试计划: 下载 → Excel修改 → 验证")
    print(f"⏰ 测试开始: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # 步骤1: 真实下载
    download_result = await step1_real_download()
    
    if not download_result['success']:
        print("❌ 下载失败，测试终止")
        return
    
    # 等待3秒
    print("\n⏳ 等待3秒后进行修改...")
    await asyncio.sleep(3)
    
    # 步骤2: Excel修改
    modification_result = step2_correct_excel_modification(download_result['file_path'])
    
    # 生成完整测试报告
    print(f"\n{'=' * 60}")
    print("📊 完整测试结果报告")
    print(f"{'=' * 60}")
    
    # 下载结果
    print("📥 下载阶段:")
    if download_result['success']:
        print(f"  ✅ 成功下载 {download_result['file_size']} bytes")
        print(f"  ⏱️ 用时: {download_result['download_time']:.2f} 秒")
    else:
        print(f"  ❌ 下载失败: {download_result.get('error', '未知错误')}")
    
    # 修改结果
    print("🛠️ 修改阶段:")
    if modification_result['success']:
        print(f"  ✅ 成功修改Excel文件")
        print(f"  📏 原始大小: {modification_result['original_size']} bytes")
        print(f"  📏 修改后大小: {modification_result['modified_size']} bytes")
        print(f"  📊 大小变化: {modification_result['size_change_percent']:.1f}%")
        print(f"  🔍 数据完整性: {'✅ 良好' if modification_result['data_integrity_ok'] else '⚠️ 警告'}")
        print(f"  📝 A1修改: {modification_result['original_a1']} → {modification_result['modified_a1']}")
    else:
        print(f"  ❌ 修改失败: {modification_result.get('error', '未知错误')}")
    
    # 最终结论
    overall_success = download_result['success'] and modification_result['success']
    
    print(f"\n🎉 测试总结:")
    if overall_success and modification_result.get('data_integrity_ok', False):
        print("✅ 完整流程测试成功！")
        print("✅ 真实下载腾讯文档Excel文件")
        print("✅ 正确修改文件并保持数据完整性")
        print("✅ 所有修改可追踪可验证")
    elif overall_success:
        print("⚠️ 流程基本成功，但存在数据完整性警告")
    else:
        print("❌ 流程测试失败")
    
    # 保存完整报告
    test_report = {
        'test_timestamp': datetime.now().isoformat(),
        'overall_success': overall_success,
        'download_result': download_result,
        'modification_result': modification_result,
        'test_conclusion': '完整流程成功' if overall_success else '测试失败'
    }
    
    report_path = '/root/projects/tencent-doc-manager/real_test_results/verified_complete_test_report.json'
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(test_report, f, ensure_ascii=False, indent=2)
    
    print(f"💾 完整测试报告已保存: {report_path}")

if __name__ == "__main__":
    asyncio.run(main())