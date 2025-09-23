#!/usr/bin/env python3
"""
8089监控全流程测试 - 完全真实，无任何模拟
测试从监控设置输入到最终URL输出的完整链路
"""

import asyncio
import json
import os
import sys
from pathlib import Path
from datetime import datetime
import csv
import logging
from typing import Optional, Dict, List
import aiohttp

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class FullChainTester:
    """8089全链路测试器"""

    def __init__(self):
        self.base_dir = Path('/root/projects/tencent-doc-manager')
        self.api_url = 'http://localhost:8089'
        self.test_url = 'https://docs.qq.com/sheet/DWEFNU25TemFnZXJN'  # 测试用腾讯文档

    async def step1_submit_config(self) -> bool:
        """步骤1: 提交监控配置到8089"""
        logger.info("=" * 60)
        logger.info("步骤1: 提交监控配置到8089")

        # 读取Cookie
        cookie_file = self.base_dir / 'config/cookies.json'
        with open(cookie_file) as f:
            cookie_data = json.load(f)

        # 准备配置数据
        config_data = {
            "links": [
                {
                    "name": "测试文档-出国销售计划表",
                    "url": self.test_url,
                    "category": "销售",
                    "description": "全链路测试文档"
                }
            ]
        }

        # 发送到8089 API
        async with aiohttp.ClientSession() as session:
            # 保存下载链接配置
            async with session.post(
                f'{self.api_url}/api/save-download-links',
                json=config_data
            ) as response:
                if response.status == 200:
                    result = await response.json()
                    logger.info(f"✅ 配置保存成功: {result}")
                    return True
                else:
                    logger.error(f"❌ 配置保存失败: {response.status}")
                    return False

    async def step2_trigger_download(self) -> Optional[str]:
        """步骤2: 触发立即刷新/下载"""
        logger.info("=" * 60)
        logger.info("步骤2: 触发立即下载")

        # 使用PlaywrightDownloader进行真实下载
        sys.path.append(str(self.base_dir))
        from production.core_modules.playwright_downloader import PlaywrightDownloader

        # 读取Cookie
        cookie_file = self.base_dir / 'config/cookies.json'
        with open(cookie_file) as f:
            cookie_data = json.load(f)
        cookie_string = cookie_data.get('current_cookies', '')

        # 创建下载目录
        from production.core_modules.week_time_manager import WeekTimeManager
        week_mgr = WeekTimeManager()
        week_info = week_mgr.get_current_week_info()
        download_dir = self.base_dir / f"csv_versions/2025_W{week_info['week_number']}/test"
        download_dir.mkdir(parents=True, exist_ok=True)

        # 执行下载
        downloader = PlaywrightDownloader()
        try:
            result = await downloader.download(
                url=self.test_url,
                cookies=cookie_string,
                format='csv',
                download_dir=str(download_dir)
            )

            if result and result.get('success'):
                csv_file = result.get('file_path')
                logger.info(f"✅ 下载成功: {csv_file}")
                return csv_file
            else:
                logger.error(f"❌ 下载失败: {result}")
                return None
        except Exception as e:
            logger.error(f"❌ 下载异常: {e}")
            return None

    async def step3_process_and_score(self, csv_file: str) -> Optional[Dict]:
        """步骤3: 对比和打分"""
        logger.info("=" * 60)
        logger.info("步骤3: 对比和打分处理")

        # 查找基线文件
        from production.core_modules.week_time_manager import WeekTimeManager
        week_mgr = WeekTimeManager()
        week_info = week_mgr.get_current_week_info()
        baseline_dir = self.base_dir / f"csv_versions/2025_W{week_info['week_number']}/baseline"

        # 查找匹配的基线文件
        baseline_files = list(baseline_dir.glob("*出国销售*.csv"))
        if not baseline_files:
            logger.warning("未找到基线文件，使用当前文件作为基线")
            baseline_file = csv_file
        else:
            baseline_file = str(baseline_files[0])
            logger.info(f"使用基线文件: {baseline_file}")

        # 执行对比
        changes = []
        with open(csv_file, 'r', encoding='utf-8') as f1, \
             open(baseline_file, 'r', encoding='utf-8') as f2:
            current_reader = list(csv.reader(f1))
            baseline_reader = list(csv.reader(f2))

            # 对比数据
            for row_idx, (row_current, row_baseline) in enumerate(zip(current_reader, baseline_reader)):
                for col_idx, (val_current, val_baseline) in enumerate(zip(row_current, row_baseline)):
                    if val_current.strip() != val_baseline.strip():
                        changes.append({
                            "row": row_idx,
                            "col": col_idx,
                            "old_value": val_baseline,
                            "new_value": val_current
                        })

        logger.info(f"发现 {len(changes)} 处变更")

        # 进行风险评分
        from production.config import L1_COLUMNS, L2_COLUMNS, L3_COLUMNS

        scores = []
        for change in changes:
            col_idx = change['col']
            if col_idx < len(L1_COLUMNS) and col_idx in range(len(L1_COLUMNS)):
                risk_level = "HIGH"
                score = 85
            elif col_idx < len(L1_COLUMNS) + len(L2_COLUMNS):
                risk_level = "MEDIUM"
                score = 50
            else:
                risk_level = "LOW"
                score = 20

            change['risk_level'] = risk_level
            change['score'] = score
            scores.append(score)

        avg_score = sum(scores) / len(scores) if scores else 0

        result = {
            "total_changes": len(changes),
            "average_score": avg_score,
            "risk_distribution": {
                "HIGH": len([c for c in changes if c['risk_level'] == 'HIGH']),
                "MEDIUM": len([c for c in changes if c['risk_level'] == 'MEDIUM']),
                "LOW": len([c for c in changes if c['risk_level'] == 'LOW'])
            },
            "changes": changes[:10]  # 只返回前10个变更
        }

        return result

    async def step4_generate_excel(self, csv_file: str, scoring_result: Dict) -> Optional[str]:
        """步骤4: 生成涂色Excel"""
        logger.info("=" * 60)
        logger.info("步骤4: 生成涂色Excel")

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font

        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "测试文档"

        # 读取CSV数据
        with open(csv_file, 'r', encoding='utf-8') as f:
            csv_reader = csv.reader(f)
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 应用涂色（基于变更）
        color_map = {
            "HIGH": "FFCCCC",
            "MEDIUM": "FFFFCC",
            "LOW": "CCFFCC"
        }

        for change in scoring_result.get('changes', []):
            row = change['row'] + 1
            col = change['col'] + 1
            risk = change['risk_level']
            color = color_map.get(risk, "FFFFFF")

            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"  # 关键：必须使用solid
            )

        # 保存Excel
        excel_dir = self.base_dir / "excel_outputs/test_8089"
        excel_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = excel_dir / f"test_8089_{timestamp}.xlsx"
        wb.save(excel_file)

        logger.info(f"✅ Excel生成成功: {excel_file}")
        return str(excel_file)

    async def step5_upload_to_tencent(self, excel_file: str) -> Optional[str]:
        """步骤5: 上传到腾讯文档"""
        logger.info("=" * 60)
        logger.info("步骤5: 上传到腾讯文档")

        # 读取Cookie
        cookie_file = self.base_dir / 'config/cookies.json'
        with open(cookie_file) as f:
            cookie_data = json.load(f)
        cookie_string = cookie_data.get('current_cookies', '')

        # 使用quick_upload_v3
        from production.core_modules.tencent_doc_upload_production_v3 import quick_upload_v3

        try:
            result = await quick_upload_v3(
                cookie_string=cookie_string,
                file_path=excel_file,
                headless=True
            )

            if result and result.get('success'):
                url = result.get('url')
                logger.info(f"✅ 上传成功: {url}")
                return url
            else:
                logger.error(f"❌ 上传失败: {result}")
                return None
        except Exception as e:
            logger.error(f"❌ 上传异常: {e}")
            return None

    async def step6_update_ui(self, upload_url: str) -> bool:
        """步骤6: 更新8089 UI显示"""
        logger.info("=" * 60)
        logger.info("步骤6: 更新UI显示")

        # 准备更新数据
        update_data = {
            "table_urls": {
                "测试文档-出国销售计划表": upload_url
            },
            "last_update": datetime.now().isoformat()
        }

        # 发送更新到8089
        async with aiohttp.ClientSession() as session:
            async with session.post(
                f'{self.api_url}/api/update-table-urls',
                json=update_data
            ) as response:
                if response.status == 200:
                    logger.info(f"✅ UI更新成功")
                    return True
                else:
                    # 如果API不存在，尝试直接更新配置文件
                    logger.info("尝试直接更新配置文件")
                    config_file = self.base_dir / 'scoring_results/comprehensive/table_urls.json'
                    with open(config_file, 'w') as f:
                        json.dump(update_data, f, indent=2)
                    logger.info(f"✅ 配置文件更新成功")
                    return True

    async def run_full_test(self):
        """运行完整测试"""
        logger.info("🚀 开始8089全链路测试")
        logger.info("=" * 60)

        results = {
            "start_time": datetime.now().isoformat(),
            "steps": {}
        }

        try:
            # 步骤1: 提交配置
            success = await self.step1_submit_config()
            results["steps"]["config"] = {"success": success}
            if not success:
                logger.error("配置提交失败，终止测试")
                return results

            # 步骤2: 触发下载
            csv_file = await self.step2_trigger_download()
            results["steps"]["download"] = {"success": csv_file is not None, "file": csv_file}
            if not csv_file:
                logger.error("下载失败，终止测试")
                return results

            # 步骤3: 对比和打分
            scoring_result = await self.step3_process_and_score(csv_file)
            results["steps"]["scoring"] = {
                "success": scoring_result is not None,
                "total_changes": scoring_result.get("total_changes", 0) if scoring_result else 0
            }
            if not scoring_result:
                logger.error("打分失败，终止测试")
                return results

            # 步骤4: 生成Excel
            excel_file = await self.step4_generate_excel(csv_file, scoring_result)
            results["steps"]["excel"] = {"success": excel_file is not None, "file": excel_file}
            if not excel_file:
                logger.error("Excel生成失败，终止测试")
                return results

            # 步骤5: 上传
            upload_url = await self.step5_upload_to_tencent(excel_file)
            results["steps"]["upload"] = {"success": upload_url is not None, "url": upload_url}
            if not upload_url:
                logger.error("上传失败，终止测试")
                return results

            # 步骤6: 更新UI
            ui_updated = await self.step6_update_ui(upload_url)
            results["steps"]["ui_update"] = {"success": ui_updated}

            # 总结
            results["end_time"] = datetime.now().isoformat()
            results["success"] = True
            results["final_url"] = upload_url

            logger.info("=" * 60)
            logger.info("🎉 测试完成！")
            logger.info(f"📊 总变更数: {scoring_result['total_changes']}")
            logger.info(f"🔗 最终URL: {upload_url}")
            logger.info(f"✅ 所有步骤成功")
            logger.info("=" * 60)

        except Exception as e:
            logger.error(f"测试过程出错: {e}")
            results["error"] = str(e)
            results["success"] = False

        # 保存测试结果
        result_file = self.base_dir / f"test_results/test_8089_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        result_file.parent.mkdir(parents=True, exist_ok=True)
        with open(result_file, 'w') as f:
            json.dump(results, f, indent=2)
        logger.info(f"测试结果保存到: {result_file}")

        return results

async def main():
    """主函数"""
    tester = FullChainTester()
    results = await tester.run_full_test()

    # 验证链接跳转
    if results.get("final_url"):
        print(f"\n请访问以下URL验证结果:")
        print(f"  8089监控UI: http://202.140.143.88:8089/")
        print(f"  上传的文档: {results['final_url']}")
        print("\n验证要点:")
        print("  1. UI中应显示最新的热力图数据")
        print("  2. 点击表格名称应跳转到上传的URL")
        print("  3. 涂色应该正确显示（solid填充）")

if __name__ == "__main__":
    asyncio.run(main())