# -*- coding: utf-8 -*-
"""
Excel MCP可视化标记模块
基于MCP (Model Context Protocol) 实现专业的Excel风险标记可视化

⚠️ 重要提示: 
在使用此模块前，请务必阅读 docs/Excel-MCP-AI-使用指南.md
确保正确配置和使用Excel MCP服务，避免常见错误。

推荐使用 mcp__excel-optimized__* 系列函数，而非基础版本。
"""

import os
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Union
import logging
from dataclasses import dataclass, asdict
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.worksheet.table import Table, TableStyleInfo

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class ModificationDetails:
    """修改详情数据结构"""
    row_number: int
    column_index: int
    column_name: str
    original_value: str
    new_value: str
    risk_level: str  # L1, L2, L3
    confidence: float
    ai_analysis: Optional[Dict[str, Any]] = None
    modification_time: Optional[str] = None
    modifier: Optional[str] = None
    business_impact: Optional[str] = None


@dataclass
class VisualizationConfig:
    """可视化配置"""
    enable_diagonal_pattern: bool = True
    enable_detailed_comments: bool = True
    enable_risk_charts: bool = True
    enable_interactive_dashboard: bool = True
    color_scheme: str = "professional"  # professional, vibrant, minimal
    export_format: str = "xlsx"  # xlsx, xlsm


class ExcelMCPVisualizationClient:
    """Excel MCP可视化标记客户端"""
    
    def __init__(self, config: Optional[VisualizationConfig] = None):
        self.config = config or VisualizationConfig()
        
        # 风险等级的专业颜色配置
        self.risk_color_schemes = {
            "professional": {
                "L1": {
                    "fill_color": "DC2626",      # 深红色
                    "pattern_color": "FFFFFF",    # 白色对角线
                    "border_color": "DC2626",
                    "font_color": "DC2626"
                },
                "L2": {
                    "fill_color": "F59E0B",      # 橙色  
                    "pattern_color": "FFFFFF",    # 白色对角线
                    "border_color": "F59E0B",
                    "font_color": "F59E0B"
                },
                "L3": {
                    "fill_color": "10B981",      # 绿色
                    "pattern_color": "FFFFFF",    # 白色对角线  
                    "border_color": "10B981",
                    "font_color": "10B981"
                }
            }
        }
        
        # 获取当前颜色方案
        self.colors = self.risk_color_schemes.get(
            self.config.color_scheme, 
            self.risk_color_schemes["professional"]
        )
        
        # 标准列映射
        self.standard_columns = [
            "序号", "项目类型", "来源", "任务发起时间", "目标对齐",
            "关键KR对齐", "具体计划内容", "邓总指导登记", "负责人",
            "协助人", "监督人", "重要程度", "预计完成时间", "完成进度", 
            "形成计划清单", "复盘时间", "对上汇报", "应用情况", "进度分析总结"
        ]
        
        # 处理统计
        self.processing_stats = {
            "workbook_created": 0,
            "sheets_created": 0,
            "cells_marked": 0,
            "comments_added": 0,
            "charts_created": 0
        }
    
    async def create_comprehensive_risk_report(self, analysis_data: Dict[str, Any], 
                                             output_path: str) -> Dict[str, Any]:
        """
        创建全面的风险分析报告
        
        Args:
            analysis_data: {
                "comparison_results": [...],
                "ai_analysis_results": {...},
                "table_metadata": {...}
            }
            output_path: 输出Excel文件路径
        
        Returns:
            {
                "report_path": "生成的报告路径",
                "marked_cells_count": "标记的单元格数量",  
                "risk_summary": "风险汇总信息"
            }
        """
        
        logger.info(f"开始创建全面风险分析报告: {output_path}")
        start_time = datetime.now()
        
        try:
            # 创建新的工作簿
            workbook = openpyxl.Workbook()
            
            # 移除默认工作表
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            
            # 步骤1: 创建总览工作表
            overview_sheet = await self._create_overview_sheet(workbook, analysis_data)
            
            # 步骤2: 为每个表格创建详细分析工作表
            marked_cells_count = 0
            table_sheets = []
            
            comparison_results = analysis_data.get("comparison_results", [])
            ai_analysis_results = analysis_data.get("ai_analysis_results", {})
            
            for table_result in comparison_results:
                if "error" not in table_result:
                    table_ai_results = ai_analysis_results.get(table_result.get("table_id"), {})
                    
                    table_sheet, cell_count = await self._create_table_analysis_sheet(
                        workbook, table_result, table_ai_results
                    )
                    
                    table_sheets.append(table_sheet)
                    marked_cells_count += cell_count
            
            # 步骤3: 创建风险分布图表工作表
            if self.config.enable_risk_charts:
                charts_sheet = await self._create_risk_charts_sheet(workbook, analysis_data)
            
            # 步骤4: 创建交互式仪表板
            if self.config.enable_interactive_dashboard:
                dashboard_sheet = await self._create_interactive_dashboard(workbook, analysis_data)
            
            # 步骤5: 应用全局格式化
            await self._apply_global_formatting(workbook)
            
            # 保存工作簿
            workbook.save(output_path)
            
            # 生成报告摘要
            processing_time = datetime.now() - start_time
            risk_summary = self._generate_risk_summary(analysis_data, marked_cells_count, processing_time)
            
            # 更新统计
            self.processing_stats["workbook_created"] += 1
            self.processing_stats["sheets_created"] += len(workbook.sheetnames)
            self.processing_stats["cells_marked"] += marked_cells_count
            
            logger.info(f"风险报告创建成功: {output_path}")
            
            return {
                "report_path": output_path,
                "marked_cells_count": marked_cells_count,
                "risk_summary": risk_summary,
                "generation_timestamp": datetime.now().isoformat(),
                "processing_time_seconds": processing_time.total_seconds(),
                "sheets_created": workbook.sheetnames
            }
            
        except Exception as e:
            logger.error(f"创建风险报告失败: {e}")
            raise
    
    async def _create_overview_sheet(self, workbook: openpyxl.Workbook, 
                                   analysis_data: Dict[str, Any]) -> openpyxl.worksheet.worksheet.Worksheet:
        """创建风险总览工作表"""
        
        sheet = workbook.create_sheet(title="风险总览")
        
        # 设置标题
        sheet['A1'] = "腾讯文档风险分析总览"
        sheet.merge_cells('A1:G1')
        
        # 标题样式
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        sheet['A1'].font = title_font
        sheet['A1'].fill = title_fill
        sheet['A1'].alignment = title_alignment
        
        # 设置表头
        headers = ["表格名称", "L1风险数", "L2风险数", "L3风险数", "AI建议", "总体状态", "最后更新"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
        
        # 填充数据
        comparison_results = analysis_data.get("comparison_results", [])
        
        for row_idx, table_result in enumerate(comparison_results, 4):
            if "error" in table_result:
                continue
                
            table_name = table_result.get("table_name", "未知表格")
            
            # 统计风险数量
            change_result = table_result.get("change_detection_result", {})
            risk_distribution = change_result.get("risk_distribution", {"L1": 0, "L2": 0, "L3": 0})
            
            # AI建议汇总
            ai_recommendation = "待分析"
            ai_results = analysis_data.get("ai_analysis_results", {}).get(table_result.get("table_id"), {})
            if ai_results:
                # 简化AI建议显示
                ai_recommendation = "已分析"
            
            # 总体状态评估
            total_l1 = risk_distribution.get("L1", 0)
            total_l2 = risk_distribution.get("L2", 0)
            
            if total_l1 > 0:
                status = "🔴 严重"
                status_color = "DC2626"
            elif total_l2 > 3:
                status = "🟡 注意"
                status_color = "F59E0B"
            else:
                status = "🟢 正常"
                status_color = "10B981"
            
            # 填充行数据
            row_data = [
                table_name,
                risk_distribution.get("L1", 0),
                risk_distribution.get("L2", 0),
                risk_distribution.get("L3", 0),
                ai_recommendation,
                status,
                datetime.now().strftime("%Y-%m-%d %H:%M")
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 状态列特殊颜色
                if col == 6:  # 总体状态列
                    cell.font = Font(color=status_color, bold=True)
        
        # 自动调整列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("风险总览工作表创建完成")
        return sheet
    
    async def _create_table_analysis_sheet(self, workbook: openpyxl.Workbook, 
                                         table_result: Dict[str, Any], 
                                         ai_results: Dict[str, Any]) -> Tuple[openpyxl.worksheet.worksheet.Worksheet, int]:
        """为单个表格创建详细分析工作表"""
        
        table_name = table_result.get("table_name", "未知表格")
        safe_sheet_name = table_name[:28] + "..." if len(table_name) > 31 else table_name
        
        # 创建工作表
        sheet = workbook.create_sheet(title=safe_sheet_name)
        
        # 获取修改详情
        modification_details = self._extract_modification_details(table_result, ai_results)
        
        # 创建分析矩阵表头
        sheet['A1'] = f"{table_name} - 详细风险分析"
        sheet.merge_cells('A1:S1')
        
        # 表头样式
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        sheet['A1'].font = header_font
        sheet['A1'].fill = header_fill
        sheet['A1'].alignment = Alignment(horizontal="center")
        
        # 设置列标题（标准19列）
        for col_idx, col_name in enumerate(self.standard_columns, 1):
            cell = sheet.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 填充数据并应用标记
        marked_cells_count = 0
        standardized_data = table_result.get("standardization_result", {}).get("standardized_data", [])
        
        for row_idx, row_data in enumerate(standardized_data[:30], 4):  # 限制30行
            for col_idx, col_name in enumerate(self.standard_columns, 1):
                value = row_data.get(col_name, "")
                cell = sheet.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                
                # 基础边框
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 检查是否需要标记
                modification = self._find_modification_for_cell(modification_details, row_idx-4, col_idx-1)
                
                if modification and modification.risk_level in ["L1", "L2"]:
                    # 应用半填充对角线标记
                    await self._apply_diagonal_marking(cell, modification)
                    
                    # 添加详细批注
                    if self.config.enable_detailed_comments:
                        comment_text = self._generate_modification_comment(modification)
                        cell.comment = Comment(comment_text, "System")
                        self.processing_stats["comments_added"] += 1
                    
                    marked_cells_count += 1
        
        # 设置列宽
        for col_idx in range(1, len(self.standard_columns) + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 12
        
        # 添加图表统计（如果启用）
        if self.config.enable_risk_charts and modification_details:
            await self._add_table_statistics_chart(sheet, modification_details)
        
        logger.info(f"表格分析工作表创建完成: {safe_sheet_name}, 标记单元格: {marked_cells_count}")
        return sheet, marked_cells_count
    
    def _extract_modification_details(self, table_result: Dict[str, Any], 
                                    ai_results: Dict[str, Any]) -> List[ModificationDetails]:
        """提取修改详情"""
        modifications = []
        
        change_result = table_result.get("change_detection_result")
        if not change_result or "changes" not in change_result:
            return modifications
        
        changes = change_result["changes"]
        
        for change in changes:
            # 查找对应的AI分析结果
            ai_analysis = None
            change_id = f"{change.get('row_index', 0)}_{change.get('column_name', '')}"
            if change_id in ai_results:
                ai_analysis = ai_results[change_id]
            
            modification = ModificationDetails(
                row_number=change.get("row_index", 0) + 4,  # 加上表头偏移
                column_index=self._get_column_index(change.get("column_name", "")),
                column_name=change.get("column_name", ""),
                original_value=str(change.get("original_value", "")),
                new_value=str(change.get("new_value", "")),
                risk_level=change.get("risk_level", "L3"),
                confidence=ai_analysis.get("confidence", 0.5) if ai_analysis else 0.5,
                ai_analysis=ai_analysis,
                modification_time=change.get("modification_time"),
                modifier=change.get("modifier"),
                business_impact=ai_analysis.get("business_impact") if ai_analysis else None
            )
            
            modifications.append(modification)
        
        return modifications
    
    def _get_column_index(self, column_name: str) -> int:
        """获取列索引"""
        try:
            return self.standard_columns.index(column_name)
        except ValueError:
            return 0
    
    def _find_modification_for_cell(self, modifications: List[ModificationDetails], 
                                  row_idx: int, col_idx: int) -> Optional[ModificationDetails]:
        """查找特定单元格的修改记录"""
        for mod in modifications:
            if mod.row_number == row_idx + 4 and mod.column_index == col_idx:
                return mod
        return None
    
    async def _apply_diagonal_marking(self, cell: openpyxl.cell.cell.Cell, 
                                    modification: ModificationDetails):
        """应用半填充对角线标记"""
        
        risk_level = modification.risk_level
        color_config = self.colors.get(risk_level, self.colors["L3"])
        
        # 应用填充和边框
        if self.config.enable_diagonal_pattern:
            # 对角线图案填充（在openpyxl中使用渐变近似）
            cell.fill = PatternFill(
                start_color=color_config["fill_color"],
                end_color=color_config["pattern_color"],
                fill_type="lightUp"  # 使用lightUp图案
            )
        else:
            # 纯色填充
            cell.fill = PatternFill(
                start_color=color_config["fill_color"],
                end_color=color_config["fill_color"],
                fill_type="solid"
            )
        
        # 边框样式
        border_style = "thick" if risk_level == "L1" else "medium"
        cell.border = Border(
            left=Side(style=border_style, color=color_config["border_color"]),
            right=Side(style=border_style, color=color_config["border_color"]),
            top=Side(style=border_style, color=color_config["border_color"]),
            bottom=Side(style=border_style, color=color_config["border_color"])
        )
        
        # 字体样式
        cell.font = Font(
            bold=True,
            color=color_config["font_color"]
        )
    
    def _generate_modification_comment(self, modification: ModificationDetails) -> str:
        """生成修改说明批注"""
        
        ai_analysis = modification.ai_analysis or {}
        
        comment_template = f"""🚨 {modification.risk_level}级别修改检测

📝 修改详情:
• 字段: {modification.column_name}
• 原值: {modification.original_value}
• 新值: {modification.new_value}
• 修改时间: {modification.modification_time or '未知'}
• 修改人: {modification.modifier or '未知'}

⚠️ 风险评估:
• 风险等级: {modification.risk_level}
• 影响程度: {modification.business_impact or '中等'}
• 置信度: {modification.confidence:.1%}

🤖 AI分析结果:
{ai_analysis.get('reasoning', '暂无AI分析')[:100]}...

📋 建议操作:
{self._format_suggested_actions(ai_analysis)}

🔍 审批要求:
{self._format_approval_requirements(modification)}"""
        
        return comment_template.strip()
    
    def _format_suggested_actions(self, ai_analysis: Dict[str, Any]) -> str:
        """格式化建议操作"""
        if not ai_analysis:
            return "• 需要人工审核确认"
        
        actions = []
        recommendation = ai_analysis.get("recommendation", "REVIEW")
        
        if recommendation == "APPROVE":
            actions.append("• ✅ AI建议：批准此修改")
        elif recommendation == "REJECT":
            actions.append("• ❌ AI建议：拒绝此修改")
        else:
            actions.append("• 🔍 AI建议：需要进一步审核")
        
        suggested_action = ai_analysis.get("suggested_action")
        if suggested_action:
            actions.append(f"• 📋 具体建议：{suggested_action}")
        
        return "\n".join(actions) if actions else "• 需要人工审核确认"
    
    def _format_approval_requirements(self, modification: ModificationDetails) -> str:
        """格式化审批要求"""
        risk_level = modification.risk_level
        
        if risk_level == "L1":
            return "• 🔴 L1级别：需要部门经理和总监双重审批\n• ⏱️ 紧急处理：24小时内必须审批完成"
        elif risk_level == "L2":
            return "• 🟠 L2级别：需要项目经理审批\n• ⏱️ 标准处理：72小时内完成审批"
        else:
            return "• 🟢 L3级别：可自动批准或团队负责人确认"
    
    async def _create_risk_charts_sheet(self, workbook: openpyxl.Workbook, 
                                      analysis_data: Dict[str, Any]) -> openpyxl.worksheet.worksheet.Worksheet:
        """创建风险分布图表工作表"""
        
        sheet = workbook.create_sheet(title="风险分布图表")
        
        # 标题
        sheet['A1'] = "风险分布统计图表"
        sheet.merge_cells('A1:H1')
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal="center")
        
        # 统计数据
        risk_stats = self._calculate_risk_statistics(analysis_data)
        
        # 创建数据表
        sheet['A3'] = "风险等级"
        sheet['B3'] = "修改数量"
        sheet['C3'] = "占比"
        
        for col in ['A3', 'B3', 'C3']:
            sheet[col].font = Font(bold=True)
            sheet[col].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        
        # 填充统计数据
        total_changes = sum(risk_stats.values())
        
        row = 4
        for risk_level in ['L1', 'L2', 'L3']:
            count = risk_stats.get(risk_level, 0)
            percentage = (count / total_changes * 100) if total_changes > 0 else 0
            
            sheet[f'A{row}'] = risk_level
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = f"{percentage:.1f}%"
            
            # 颜色编码
            color = self.colors[risk_level]["fill_color"]
            sheet[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            row += 1
        
        # 创建图表
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "风险等级分布"
        chart.y_axis.title = '修改数量'
        chart.x_axis.title = '风险等级'
        
        data = Reference(sheet, min_col=2, min_row=3, max_row=6, max_col=2)
        cats = Reference(sheet, min_col=1, min_row=4, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        sheet.add_chart(chart, "E3")
        self.processing_stats["charts_created"] += 1
        
        logger.info("风险分布图表工作表创建完成")
        return sheet
    
    def _calculate_risk_statistics(self, analysis_data: Dict[str, Any]) -> Dict[str, int]:
        """计算风险统计"""
        risk_stats = {"L1": 0, "L2": 0, "L3": 0}
        
        comparison_results = analysis_data.get("comparison_results", [])
        
        for table_result in comparison_results:
            change_result = table_result.get("change_detection_result", {})
            risk_distribution = change_result.get("risk_distribution", {})
            
            for risk_level, count in risk_distribution.items():
                if risk_level in risk_stats:
                    risk_stats[risk_level] += count
        
        return risk_stats
    
    async def _create_interactive_dashboard(self, workbook: openpyxl.Workbook, 
                                          analysis_data: Dict[str, Any]) -> openpyxl.worksheet.worksheet.Worksheet:
        """创建交互式仪表板"""
        
        sheet = workbook.create_sheet(title="交互式仪表板")
        
        # 仪表板标题
        sheet['A1'] = "🎯 风险分析仪表板"
        sheet.merge_cells('A1:F1')
        sheet['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        sheet['A1'].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        sheet['A1'].alignment = Alignment(horizontal="center")
        
        # 更新时间
        sheet['G1'] = f"📅 更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['G1'].font = Font(size=10)
        sheet['G1'].alignment = Alignment(horizontal="right")
        
        # 仪表板数据
        dashboard_data = self._prepare_dashboard_data(analysis_data)
        
        # 总体统计区域
        sheet['A3'] = "📊 总体统计"
        sheet['A3'].font = Font(size=14, bold=True)
        
        stats_data = [
            ("总表格数", dashboard_data.get('total_tables', 0)),
            ("总修改数", dashboard_data.get('total_modifications', 0)),
            ("L1违规", dashboard_data.get('l1_count', 0)),
            ("L2修改", dashboard_data.get('l2_count', 0)),
            ("L3修改", dashboard_data.get('l3_count', 0))
        ]
        
        for row, (label, value) in enumerate(stats_data, 4):
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            
            # 格式化
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'].font = Font(size=12, color="2563EB")
        
        # 快速导航区域
        sheet['D3'] = "🚀 快速导航"
        sheet['D3'].font = Font(size=14, bold=True)
        
        navigation_links = [
            ("查看风险总览", "风险总览!A1"),
            ("查看风险图表", "风险分布图表!A1"),
        ]
        
        for row, (link_text, link_target) in enumerate(navigation_links, 4):
            cell = sheet[f'D{row}']
            cell.value = f"🔗 {link_text}"
            cell.font = Font(color="2563EB", underline="single")
            
            # 创建超链接（注意：openpyxl的超链接语法）
            cell.hyperlink = f"#{link_target}"
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['D'].width = 20
        
        logger.info("交互式仪表板创建完成")
        return sheet
    
    def _prepare_dashboard_data(self, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """准备仪表板数据"""
        comparison_results = analysis_data.get("comparison_results", [])
        
        total_tables = len([r for r in comparison_results if "error" not in r])
        risk_stats = self._calculate_risk_statistics(analysis_data)
        
        return {
            "total_tables": total_tables,
            "total_modifications": sum(risk_stats.values()),
            "l1_count": risk_stats.get("L1", 0),
            "l2_count": risk_stats.get("L2", 0),
            "l3_count": risk_stats.get("L3", 0),
            "hotspot_table": "待分析",  # TODO: 实现热点表格识别
            "high_risk_count": risk_stats.get("L1", 0) + risk_stats.get("L2", 0)
        }
    
    async def _add_table_statistics_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                                        modifications: List[ModificationDetails]):
        """添加表格统计图表"""
        
        # 在右侧区域添加小型统计图表
        start_col = len(self.standard_columns) + 2  # 在数据表右侧
        
        # 标题
        title_cell = sheet.cell(row=3, column=start_col, value="修改统计")
        title_cell.font = Font(size=12, bold=True)
        
        # 统计修改类型
        risk_counts = {"L1": 0, "L2": 0, "L3": 0}
        for mod in modifications:
            risk_counts[mod.risk_level] = risk_counts.get(mod.risk_level, 0) + 1
        
        # 创建简单的统计显示
        row = 4
        for risk_level, count in risk_counts.items():
            if count > 0:
                label_cell = sheet.cell(row=row, column=start_col, value=f"{risk_level}:")
                count_cell = sheet.cell(row=row, column=start_col+1, value=count)
                
                # 应用颜色
                color = self.colors[risk_level]["fill_color"]
                label_cell.font = Font(color=color, bold=True)
                count_cell.font = Font(color=color, bold=True)
                
                row += 1
    
    async def _apply_global_formatting(self, workbook: openpyxl.Workbook):
        """应用全局格式化"""
        
        # 为所有工作表设置默认字体
        default_font = Font(name="Calibri", size=10)
        
        for sheet in workbook.worksheets:
            # 设置默认行高
            sheet.row_dimensions[1].height = 25
            
            # 设置网格线显示
            sheet.sheet_view.showGridLines = True
            
            # 冻结窗格（前3行和第1列）
            if sheet.max_row > 3:
                sheet.freeze_panes = "B4"
        
        logger.info("全局格式化应用完成")
    
    def _generate_risk_summary(self, analysis_data: Dict[str, Any], 
                             marked_cells_count: int, processing_time) -> Dict[str, Any]:
        """生成风险汇总信息"""
        
        risk_stats = self._calculate_risk_statistics(analysis_data)
        comparison_results = analysis_data.get("comparison_results", [])
        
        summary = {
            "total_tables_analyzed": len([r for r in comparison_results if "error" not in r]),
            "total_modifications_detected": sum(risk_stats.values()),
            "risk_distribution": risk_stats,
            "marked_cells_count": marked_cells_count,
            "processing_time_seconds": processing_time.total_seconds(),
            "high_risk_tables": [
                result.get("table_name", "未知")
                for result in comparison_results
                if result.get("change_detection_result", {}).get("risk_distribution", {}).get("L1", 0) > 0
            ][:5],
            "processing_stats": self.processing_stats.copy()
        }
        
        return summary


# 使用示例和测试
def test_excel_mcp_visualization():
    """Excel MCP可视化测试"""
    
    # 模拟分析数据
    test_analysis_data = {
        "comparison_results": [
            {
                "table_id": "table_0",
                "table_name": "小红书部门-工作表2",
                "change_detection_result": {
                    "changes": [
                        {
                            "row_index": 5,
                            "column_name": "负责人",
                            "original_value": "张三",
                            "new_value": "李四",
                            "risk_level": "L2",
                            "change_type": "modification"
                        },
                        {
                            "row_index": 8,
                            "column_name": "目标对齐",
                            "original_value": "品牌提升",
                            "new_value": "用户增长",
                            "risk_level": "L1",
                            "change_type": "modification"
                        }
                    ],
                    "risk_distribution": {"L1": 1, "L2": 1, "L3": 0}
                },
                "standardization_result": {
                    "standardized_data": [
                        {col: f"数据_{i}_{j}" for j, col in enumerate(["序号", "项目类型", "来源", "任务发起时间", "目标对齐", "关键KR对齐", "具体计划内容", "邓总指导登记", "负责人", "协助人", "监督人", "重要程度", "预计完成时间", "完成进度", "形成计划清单", "复盘时间", "对上汇报", "应用情况", "进度分析总结"])}
                        for i in range(10)
                    ]
                }
            }
        ],
        "ai_analysis_results": {
            "table_0": {
                "5_负责人": {
                    "recommendation": "APPROVE",
                    "confidence": 0.85,
                    "reasoning": "负责人变更符合项目需求，新负责人具备相应能力",
                    "business_impact": "MEDIUM",
                    "suggested_action": "建议与原负责人做好工作交接"
                }
            }
        },
        "table_metadata": {}
    }
    
    # 创建可视化客户端
    config = VisualizationConfig(
        enable_diagonal_pattern=True,
        enable_detailed_comments=True,
        enable_risk_charts=True,
        enable_interactive_dashboard=True
    )
    
    client = ExcelMCPVisualizationClient(config)
    
    # 同步包装器（因为这是测试函数）
    import asyncio
    
    async def run_test():
        output_path = "/root/projects/tencent-doc-manager/风险分析报告.xlsx"
        
        try:
            result = await client.create_comprehensive_risk_report(
                test_analysis_data, output_path
            )
            
            print("=== Excel MCP可视化测试结果 ===")
            print(f"报告路径: {result['report_path']}")
            print(f"标记单元格数: {result['marked_cells_count']}")
            print(f"处理时间: {result['processing_time_seconds']:.2f}秒")
            print(f"创建工作表: {result['sheets_created']}")
            print("✅ Excel可视化报告创建成功")
            
        except Exception as e:
            print(f"❌ Excel可视化测试失败: {e}")
    
    # 运行测试
    asyncio.run(run_test())


if __name__ == "__main__":
    test_excel_mcp_visualization()