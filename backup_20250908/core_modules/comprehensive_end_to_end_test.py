#!/usr/bin/env python3
"""
完整真实端到端测试执行器
执行真实的腾讯文档下载、修改、上传操作

测试流程：
1. 下载指定腾讯文档：https://docs.qq.com/sheet/DWEVjZndkR2xVSWJN?tab=c2p5hs
2. 修改Excel文件，在标题中加入"123123"
3. 上传修改后的文件，文件名为"123123"
4. 获取真实可访问的上传链接
5. 返回完整操作日志

依赖：
- Cookie配置：/root/projects/tencent-doc-manager/config/cookies.json
- 下载器：tencent_export_automation.py
- 上传器：production_upload_manager.py
"""

import os
import sys
import json
import datetime
import time
import pandas as pd
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging
import traceback

# 设置路径
BASE_DIR = '/root/projects/tencent-doc-manager'
sys.path.append(os.path.join(BASE_DIR, 'production/core_modules'))
sys.path.append(os.path.join(BASE_DIR, '测试版本-性能优化开发-20250811-001430'))

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/root/projects/tencent-doc-manager/production/core_modules/end_to_end_test.log'),
        logging.StreamHandler()
    ]
)

class ComprehensiveEndToEndTester:
    def __init__(self):
        self.test_id = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        self.operation_log = {
            'test_id': self.test_id,
            'start_time': datetime.datetime.now().isoformat(),
            'operations': [],
            'results': {},
            'errors': []
        }
        self.logger = logging.getLogger(__name__)
        
        # 目标URL和配置
        self.target_url = 'https://docs.qq.com/sheet/DWEVjZndkR2xVSWJN?tab=c2p5hs'
        self.target_filename = '123123'
        self.modification_marker = '123123'
        
        # 文件路径
        self.config_dir = '/root/projects/tencent-doc-manager/config'
        self.downloads_dir = '/root/projects/tencent-doc-manager/production/core_modules/downloads'
        self.temp_dir = '/root/projects/tencent-doc-manager/temp'
        
        # 确保目录存在
        os.makedirs(self.downloads_dir, exist_ok=True)
        os.makedirs(self.temp_dir, exist_ok=True)
        
        self.logger.info(f"🚀 启动端到端测试 - ID: {self.test_id}")

    def log_operation(self, operation, status, details, duration=None):
        """记录操作日志"""
        log_entry = {
            'timestamp': datetime.datetime.now().isoformat(),
            'operation': operation,
            'status': status,
            'details': details,
            'duration': duration
        }
        self.operation_log['operations'].append(log_entry)
        
        status_icon = "✅" if status == "success" else "❌" if status == "error" else "⚠️"
        self.logger.info(f"{status_icon} {operation}: {details}")

    def load_cookies(self):
        """加载Cookie配置"""
        try:
            cookie_file = os.path.join(self.config_dir, 'cookies.json')
            if not os.path.exists(cookie_file):
                raise FileNotFoundError(f"Cookie文件不存在: {cookie_file}")
            
            with open(cookie_file, 'r', encoding='utf-8') as f:
                cookie_data = json.load(f)
            
            if not cookie_data.get('is_valid', False):
                raise ValueError("Cookie配置无效")
            
            cookies = cookie_data.get('current_cookies', '')
            self.log_operation("加载Cookie", "success", f"Cookie长度: {len(cookies)}")
            return cookies
            
        except Exception as e:
            self.log_operation("加载Cookie", "error", str(e))
            raise

    def download_document(self):
        """下载腾讯文档"""
        start_time = time.time()
        try:
            # 导入下载器
            from tencent_export_automation import TencentDocAutoExporter
            
            cookies = self.load_cookies()
            
            # 初始化下载器
            exporter = TencentDocAutoExporter()
            
            # 从URL中提取文档ID
            # URL格式: https://docs.qq.com/sheet/DWEVjZndkR2xVSWJN?tab=c2p5hs
            doc_id = self.target_url.split('/sheet/')[1].split('?')[0] if '/sheet/' in self.target_url else None
            
            if not doc_id:
                raise ValueError(f"无法从URL提取文档ID: {self.target_url}")
            
            self.log_operation("提取文档ID", "success", f"文档ID: {doc_id}")
            
            # 执行下载
            download_result = exporter.export_to_csv(
                doc_id=doc_id,
                cookies=cookies,
                output_dir=self.downloads_dir,
                filename_prefix="tencent_doc_test"
            )
            
            if not download_result or 'file_path' not in download_result:
                raise Exception("下载失败，未返回文件路径")
            
            downloaded_file = download_result['file_path']
            
            # 验证文件是否存在
            if not os.path.exists(downloaded_file):
                raise FileNotFoundError(f"下载文件不存在: {downloaded_file}")
            
            file_size = os.path.getsize(downloaded_file)
            duration = time.time() - start_time
            
            self.log_operation(
                "下载文档", "success", 
                f"文件: {downloaded_file}, 大小: {file_size} bytes",
                duration
            )
            
            self.operation_log['results']['downloaded_file'] = downloaded_file
            self.operation_log['results']['download_size'] = file_size
            
            return downloaded_file
            
        except Exception as e:
            duration = time.time() - start_time
            error_msg = f"下载失败: {str(e)}"
            self.log_operation("下载文档", "error", error_msg, duration)
            self.operation_log['errors'].append(error_msg)
            raise

    def modify_excel_file(self, input_file):
        """修改Excel文件，在标题中添加123123"""
        start_time = time.time()
        try:
            # 确定输出文件路径
            output_file = os.path.join(self.temp_dir, f'modified_{self.modification_marker}.xlsx')
            
            # 读取原始文件
            if input_file.endswith('.csv'):
                # 如果是CSV，先转换为Excel
                df = pd.read_csv(input_file, encoding='utf-8')
                df.to_excel(output_file, index=False, engine='openpyxl')
            else:
                # 如果已经是Excel，直接复制
                import shutil
                shutil.copy2(input_file, output_file)
            
            # 使用openpyxl修改Excel文件
            wb = load_workbook(output_file)
            
            modifications_made = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 修改工作表名称
                original_sheet_name = sheet_name
                if self.modification_marker not in sheet_name:
                    new_sheet_name = f"{self.modification_marker}_{sheet_name}"
                    ws.title = new_sheet_name
                    modifications_made.append(f"工作表名: {original_sheet_name} -> {new_sheet_name}")
                
                # 修改第一行标题（表头）
                if ws.max_row > 0:
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col)
                        if cell.value and isinstance(cell.value, str):
                            original_value = cell.value
                            if self.modification_marker not in original_value:
                                new_value = f"{self.modification_marker}_{original_value}"
                                cell.value = new_value
                                
                                # 添加样式标记修改
                                cell.font = Font(bold=True, color="FF0000")  # 红色粗体
                                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色背景
                                cell.alignment = Alignment(horizontal="center")
                                
                                modifications_made.append(f"标题修改: {original_value} -> {new_value}")
                
                # 在A1单元格前插入标识
                if ws['A1'].value:
                    ws.insert_rows(1)
                    ws['A1'] = f"[{self.modification_marker}] 修改标识"
                    ws['A1'].font = Font(bold=True, size=14, color="FF0000")
                    ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    modifications_made.append("插入修改标识行")
            
            # 保存修改后的文件
            wb.save(output_file)
            
            file_size = os.path.getsize(output_file)
            duration = time.time() - start_time
            
            self.log_operation(
                "修改Excel文件", "success",
                f"文件: {output_file}, 修改数量: {len(modifications_made)}, 大小: {file_size} bytes",
                duration
            )
            
            self.operation_log['results']['modified_file'] = output_file
            self.operation_log['results']['modifications'] = modifications_made
            self.operation_log['results']['modified_size'] = file_size
            
            return output_file
            
        except Exception as e:
            duration = time.time() - start_time
            error_msg = f"修改文件失败: {str(e)}"
            self.log_operation("修改Excel文件", "error", error_msg, duration)
            self.operation_log['errors'].append(error_msg)
            raise

    def upload_document(self, file_path):
        """上传修改后的文档到腾讯文档"""
        start_time = time.time()
        try:
            # 导入上传管理器
            from production_upload_manager import ProductionUploadManager
            
            cookies = self.load_cookies()
            
            # 初始化上传管理器
            upload_manager = ProductionUploadManager()
            
            # 执行上传
            upload_result = upload_manager.upload_excel_to_tencent(
                file_path=file_path,
                document_name=self.target_filename,
                cookies=cookies
            )
            
            if not upload_result or not upload_result.get('success', False):
                error_msg = upload_result.get('error', '上传失败，未知错误') if upload_result else '上传失败'
                raise Exception(error_msg)
            
            upload_url = upload_result.get('document_url', '')
            document_id = upload_result.get('document_id', '')
            
            duration = time.time() - start_time
            
            self.log_operation(
                "上传文档", "success",
                f"URL: {upload_url}, ID: {document_id}",
                duration
            )
            
            self.operation_log['results']['upload_url'] = upload_url
            self.operation_log['results']['document_id'] = document_id
            self.operation_log['results']['upload_success'] = True
            
            return upload_result
            
        except Exception as e:
            duration = time.time() - start_time
            error_msg = f"上传失败: {str(e)}"
            self.log_operation("上传文档", "error", error_msg, duration)
            self.operation_log['errors'].append(error_msg)
            raise

    def verify_upload_accessibility(self, upload_url):
        """验证上传链接的可访问性"""
        start_time = time.time()
        try:
            if not upload_url:
                raise ValueError("上传URL为空")
            
            # 发送HTTP请求验证链接
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(upload_url, headers=headers, timeout=30)
            
            is_accessible = response.status_code == 200
            response_size = len(response.content)
            
            duration = time.time() - start_time
            
            if is_accessible:
                self.log_operation(
                    "验证链接可访问性", "success",
                    f"HTTP {response.status_code}, 响应大小: {response_size} bytes",
                    duration
                )
            else:
                self.log_operation(
                    "验证链接可访问性", "warning",
                    f"HTTP {response.status_code}, 可能需要登录访问",
                    duration
                )
            
            self.operation_log['results']['url_accessible'] = is_accessible
            self.operation_log['results']['http_status'] = response.status_code
            self.operation_log['results']['response_size'] = response_size
            
            return is_accessible
            
        except Exception as e:
            duration = time.time() - start_time
            error_msg = f"链接验证失败: {str(e)}"
            self.log_operation("验证链接可访问性", "error", error_msg, duration)
            self.operation_log['errors'].append(error_msg)
            return False

    def save_test_report(self):
        """保存测试报告"""
        try:
            self.operation_log['end_time'] = datetime.datetime.now().isoformat()
            self.operation_log['total_duration'] = (
                datetime.datetime.fromisoformat(self.operation_log['end_time']) - 
                datetime.datetime.fromisoformat(self.operation_log['start_time'])
            ).total_seconds()
            
            # 保存详细报告
            report_file = os.path.join(
                '/root/projects/tencent-doc-manager/production/test_results',
                f'end_to_end_test_report_{self.test_id}.json'
            )
            
            os.makedirs(os.path.dirname(report_file), exist_ok=True)
            
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(self.operation_log, f, ensure_ascii=False, indent=2)
            
            self.log_operation("保存测试报告", "success", f"报告文件: {report_file}")
            
            return report_file
            
        except Exception as e:
            error_msg = f"保存报告失败: {str(e)}"
            self.log_operation("保存测试报告", "error", error_msg)
            self.operation_log['errors'].append(error_msg)
            return None

    def run_complete_test(self):
        """执行完整的端到端测试"""
        try:
            self.logger.info("=" * 60)
            self.logger.info(f"🎯 开始执行真实端到端测试 - {self.test_id}")
            self.logger.info(f"目标URL: {self.target_url}")
            self.logger.info(f"目标文件名: {self.target_filename}")
            self.logger.info("=" * 60)
            
            # 步骤1: 下载文档
            self.logger.info("📥 步骤1: 下载腾讯文档...")
            downloaded_file = self.download_document()
            
            # 步骤2: 修改文件
            self.logger.info("✏️  步骤2: 修改Excel文件...")
            modified_file = self.modify_excel_file(downloaded_file)
            
            # 步骤3: 上传文档
            self.logger.info("📤 步骤3: 上传修改后的文档...")
            upload_result = self.upload_document(modified_file)
            
            # 步骤4: 验证链接可访问性
            upload_url = upload_result.get('document_url', '')
            if upload_url:
                self.logger.info("🔗 步骤4: 验证上传链接可访问性...")
                self.verify_upload_accessibility(upload_url)
            
            # 步骤5: 保存测试报告
            self.logger.info("💾 步骤5: 保存测试报告...")
            report_file = self.save_test_report()
            
            # 输出最终结果
            self.logger.info("=" * 60)
            self.logger.info("🎉 端到端测试完成!")
            self.logger.info(f"📊 测试ID: {self.test_id}")
            self.logger.info(f"📄 下载文件: {self.operation_log['results'].get('downloaded_file', 'N/A')}")
            self.logger.info(f"✏️  修改文件: {self.operation_log['results'].get('modified_file', 'N/A')}")
            self.logger.info(f"🔗 上传链接: {self.operation_log['results'].get('upload_url', 'N/A')}")
            self.logger.info(f"✅ 链接可访问: {self.operation_log['results'].get('url_accessible', False)}")
            self.logger.info(f"📋 测试报告: {report_file}")
            self.logger.info(f"🕒 总耗时: {self.operation_log.get('total_duration', 0):.2f}秒")
            
            if self.operation_log['errors']:
                self.logger.warning(f"⚠️  错误数量: {len(self.operation_log['errors'])}")
                for error in self.operation_log['errors']:
                    self.logger.warning(f"   - {error}")
            
            self.logger.info("=" * 60)
            
            return self.operation_log
            
        except Exception as e:
            error_msg = f"端到端测试执行失败: {str(e)}"
            self.logger.error(error_msg)
            self.logger.error(traceback.format_exc())
            
            self.operation_log['errors'].append(error_msg)
            self.operation_log['test_failed'] = True
            self.operation_log['failure_reason'] = str(e)
            
            # 仍然保存报告，记录失败信息
            self.save_test_report()
            
            raise

def main():
    """主函数"""
    try:
        tester = ComprehensiveEndToEndTester()
        result = tester.run_complete_test()
        
        print("\n" + "=" * 80)
        print("🎯 真实端到端测试执行完毕")
        print("=" * 80)
        print(f"测试ID: {result['test_id']}")
        print(f"开始时间: {result['start_time']}")
        print(f"结束时间: {result.get('end_time', 'N/A')}")
        print(f"总耗时: {result.get('total_duration', 0):.2f}秒")
        print(f"操作数量: {len(result['operations'])}")
        print(f"错误数量: {len(result['errors'])}")
        
        if result.get('results'):
            print("\n📋 主要结果:")
            results = result['results']
            print(f"  下载成功: {'✅' if results.get('downloaded_file') else '❌'}")
            print(f"  修改成功: {'✅' if results.get('modified_file') else '❌'}")
            print(f"  上传成功: {'✅' if results.get('upload_success') else '❌'}")
            print(f"  链接可访问: {'✅' if results.get('url_accessible') else '❌'}")
            
            if results.get('upload_url'):
                print(f"\n🔗 上传链接: {results['upload_url']}")
            
            if results.get('modifications'):
                print(f"\n✏️  文件修改详情:")
                for mod in results['modifications'][:5]:  # 只显示前5个修改
                    print(f"  - {mod}")
                if len(results['modifications']) > 5:
                    print(f"  ... 还有 {len(results['modifications']) - 5} 个修改")
        
        print("=" * 80)
        
        return result
        
    except Exception as e:
        print(f"❌ 测试执行失败: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    main()