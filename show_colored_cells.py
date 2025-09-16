#!/usr/bin/env python3
"""显示具体的涂色单元格内容"""

import openpyxl
from openpyxl.styles import PatternFill
import json

def show_colored_cells(excel_path):
    """显示所有涂色的单元格"""
    
    print(f"检查文件: {excel_path}")
    print("=" * 80)
    
    # 打开工作簿
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['工作表1']
    
    # 收集所有有涂色的单元格
    colored_cells = []
    
    # 检查每个单元格
    for row in range(1, min(ws.max_row + 1, 100)):  # 检查前100行
        for col in range(1, min(ws.max_column + 1, 20)):  # 检查前20列
            cell = ws.cell(row=row, column=col)
            
            # 获取填充样式
            fill = cell.fill
            
            # 检查是否有填充（排除默认填充）
            if fill and fill.patternType and fill.patternType != 'none':
                # 获取前景色和背景色
                fg_color = fill.fgColor
                bg_color = fill.bgColor
                
                # 判断颜色类型
                color_name = "未知"
                if fg_color and fg_color.rgb:
                    rgb = str(fg_color.rgb).upper()
                    if 'FF0000' in rgb or 'FF0101' in rgb or 'FE0000' in rgb:
                        color_name = "🔴红色"
                    elif 'FFFF00' in rgb or 'FFD700' in rgb:
                        color_name = "🟡黄色"
                    elif '00FF00' in rgb or '90EE90' in rgb:
                        color_name = "🟢绿色"
                    elif 'FFFFFF' in rgb:
                        color_name = "⬜白色"
                    elif '000000' in rgb:
                        color_name = "⬛黑色"
                    else:
                        color_name = f"🔵其他({rgb[:6]})"
                
                # 记录单元格信息
                cell_info = {
                    'address': f"{openpyxl.utils.get_column_letter(col)}{row}",
                    'row': row,
                    'col': col,
                    'value': str(cell.value) if cell.value else "空",
                    'pattern': fill.patternType,
                    'color': color_name,
                    'fg_rgb': str(fg_color.rgb) if fg_color and fg_color.rgb else None
                }
                colored_cells.append(cell_info)
    
    # 按行列排序
    colored_cells.sort(key=lambda x: (x['row'], x['col']))
    
    # 打印涂色单元格
    print("\n📊 涂色的单元格（按位置排序）：\n")
    print(f"{'位置':<8} {'行':<4} {'列':<4} {'图案类型':<20} {'颜色':<15} {'内容'}")
    print("-" * 80)
    
    # 只显示条纹图案的单元格（这些是我们关注的）
    stripe_patterns = ['lightHorizontal', 'darkVertical', 'lightUp', 'lightDown', 'darkHorizontal', 'darkUp', 'darkDown']
    
    stripe_cells = [c for c in colored_cells if c['pattern'] in stripe_patterns]
    
    if stripe_cells:
        print("\n✨ 条纹图案涂色单元格（这些是系统标记的变更）：")
        print("-" * 80)
        for cell in stripe_cells:
            value_display = cell['value'][:50] if len(cell['value']) > 50 else cell['value']
            print(f"{cell['address']:<8} {cell['row']:<4} {cell['col']:<4} {cell['pattern']:<20} {cell['color']:<15} {value_display}")
    
    # 显示前20个涂色单元格
    print("\n📋 所有涂色单元格（前30个）：")
    print("-" * 80)
    for cell in colored_cells[:30]:
        value_display = cell['value'][:50] if len(cell['value']) > 50 else cell['value']
        print(f"{cell['address']:<8} {cell['row']:<4} {cell['col']:<4} {cell['pattern']:<20} {cell['color']:<15} {value_display}")
    
    # 统计
    print(f"\n📊 统计：")
    print(f"  总涂色单元格数: {len(colored_cells)}")
    print(f"  条纹图案单元格数: {len(stripe_cells)}")
    
    # 按图案类型统计
    pattern_stats = {}
    for cell in colored_cells:
        pattern = cell['pattern']
        if pattern not in pattern_stats:
            pattern_stats[pattern] = 0
        pattern_stats[pattern] += 1
    
    print(f"\n📊 按图案类型统计：")
    for pattern, count in sorted(pattern_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {pattern:<20}: {count} 个")
    
    return colored_cells


if __name__ == "__main__":
    # 最新的标记文件
    excel_path = "/root/projects/tencent-doc-manager/excel_outputs/marked/tencent_副本-副本-测试版本-出国销售计划表_20250911_2106_midweek_W37_marked_20250911_210610_W37.xlsx"
    
    colored_cells = show_colored_cells(excel_path)