#!/usr/bin/env python3
"""
修复并涂色Excel文件
"""

import os
import sys
import zipfile
import shutil
from datetime import datetime

def fix_excel_file(input_file, output_file):
    """修复腾讯文档Excel的空fill标签问题"""
    print(f"🔧 修复文件: {os.path.basename(input_file)}")

    # 创建临时目录
    temp_dir = f"/tmp/excel_fix_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    os.makedirs(temp_dir, exist_ok=True)

    try:
        # 解压Excel文件
        with zipfile.ZipFile(input_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # 修复styles.xml
        styles_path = os.path.join(temp_dir, 'xl', 'styles.xml')
        if os.path.exists(styles_path):
            with open(styles_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # 替换空的fill标签
            content = content.replace('<fill/>', '<fill><patternFill patternType="none"/></fill>')

            with open(styles_path, 'w', encoding='utf-8') as f:
                f.write(content)

            print("✅ 已修复空fill标签")

        # 重新打包
        with zipfile.ZipFile(output_file, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arc_name = os.path.relpath(file_path, temp_dir)
                    zip_ref.write(file_path, arc_name)

        print(f"✅ 修复完成: {os.path.basename(output_file)}")
        return True

    finally:
        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)

def apply_coloring(excel_file):
    """应用涂色到Excel文件"""
    print(f"\n🎨 涂色文件: {os.path.basename(excel_file)}")

    import openpyxl
    from openpyxl.styles import PatternFill

    # 打开文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    print(f"📊 工作表: {ws.title}")
    print(f"📏 数据范围: {ws.max_row}行 x {ws.max_column}列")

    # 定义涂色规则
    color_rules = [
        {"range": "A2:C4", "color": "FFFF0000", "name": "红色（高风险）"},
        {"range": "D2:F4", "color": "FFFFA500", "name": "橙色（中风险）"},
        {"range": "G2:I4", "color": "FF00FF00", "name": "绿色（低风险）"},
        {"range": "A6:C8", "color": "FFFFCCCC", "name": "浅红色"},
        {"range": "D6:F8", "color": "FFFFE9E8", "name": "浅橙色"},
        {"range": "G6:I8", "color": "FFFFFF00", "name": "黄色"},
    ]

    # 应用涂色
    colored_count = 0
    for rule in color_rules:
        try:
            # 解析范围
            start_cell, end_cell = rule["range"].split(":")

            # 创建填充
            fill = PatternFill(
                patternType="solid",
                fgColor=rule["color"],
                bgColor=rule["color"]
            )

            # 应用到范围
            for row in ws[rule["range"]]:
                for cell in row:
                    cell.fill = fill
                    colored_count += 1

            print(f"✅ 应用{rule['name']}到{rule['range']}")

        except Exception as e:
            print(f"⚠️ 跳过{rule['range']}: {str(e)}")

    # 保存文件
    output_file = excel_file.replace('_fixed.xlsx', '_colored.xlsx')
    if '_fixed' not in excel_file:
        output_file = excel_file.replace('.xlsx', '_colored.xlsx')

    wb.save(output_file)
    wb.close()

    print(f"\n✅ 涂色完成: {os.path.basename(output_file)}")
    print(f"📊 共涂色{colored_count}个单元格")

    return output_file

def main():
    # 原始文件
    original_file = "/root/projects/tencent-doc-manager/csv_versions/2025_W38/midweek/tencent_111测试版本-小红书部门_20250919_2255_midweek_W38.xlsx"

    if not os.path.exists(original_file):
        print(f"❌ 文件不存在: {original_file}")
        return

    print("="*60)
    print("📋 腾讯文档Excel修复并涂色")
    print("="*60)

    # 1. 修复文件
    fixed_file = original_file.replace('.xlsx', '_fixed.xlsx')
    if fix_excel_file(original_file, fixed_file):

        # 2. 涂色文件
        colored_file = apply_coloring(fixed_file)

        if colored_file and os.path.exists(colored_file):
            print("\n" + "="*60)
            print("🎉 处理成功！")
            print(f"📄 最终文件: {colored_file}")
            print(f"📏 文件大小: {os.path.getsize(colored_file):,} bytes")
            print("="*60)

            return colored_file

    return None

if __name__ == "__main__":
    result = main()
    if result:
        print(f"\n🔗 可以上传的文件: {result}")