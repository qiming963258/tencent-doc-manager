#!/usr/bin/env python3
"""
验证过的全链路测试 - 基于成功经验
确保对比同一文档的不同版本，避免错误对比
"""

import json
import sys
import os
import asyncio
from pathlib import Path
from datetime import datetime
import csv
import logging
from typing import Optional, Dict, List
import random

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

sys.path.append('/root/projects/tencent-doc-manager')

from production.config import STANDARD_COLUMNS, L1_COLUMNS, L2_COLUMNS, L3_COLUMNS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.comments import Comment


class VerifiedChainTest:
    """验证过的链路测试 - 避免错误"""

    def __init__(self):
        # 关键：使用同一个文档进行对比
        self.doc_url = "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN"  # 出国销售计划表
        self.doc_name = "出国销售计划表"
        self.doc_id = "DWEFNU25TemFnZXJN"

        # 基线文件：同一文档的早期版本
        self.baseline_file = "/root/projects/tencent-doc-manager/csv_versions/2025_W38/baseline/tencent_出国销售计划表_20250915_0145_baseline_W38.csv"

        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.download_dir = Path(f'/root/projects/tencent-doc-manager/downloads/verified_{self.timestamp}')
        self.download_dir.mkdir(parents=True, exist_ok=True)

        logger.info("="*70)
        logger.info("🔒 验证过的全链路测试")
        logger.info("="*70)
        logger.info(f"📄 文档: {self.doc_name}")
        logger.info(f"🔗 URL: {self.doc_url}")
        logger.info(f"📊 基线: {Path(self.baseline_file).name}")

    async def download_current(self) -> Optional[str]:
        """下载当前版本"""
        logger.info("\n🔽 Step 1: 下载当前版本")

        try:
            from production.core_modules.playwright_downloader import PlaywrightDownloader

            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')
            logger.info(f"✅ Cookie已加载")

            # 下载
            downloader = PlaywrightDownloader()
            result = await downloader.download(
                url=self.doc_url,
                cookies=cookie_string,
                format='csv',
                download_dir=str(self.download_dir)
            )

            if result.get('success') and result.get('file_path'):
                csv_file = result['file_path']
                size = Path(csv_file).stat().st_size / 1024
                logger.info(f"✅ 下载成功: {Path(csv_file).name}")
                logger.info(f"📊 大小: {size:.2f} KB")
                return csv_file

        except Exception as e:
            logger.error(f"❌ 下载失败: {e}")

        return None

    def compare_versions(self, current_file: str) -> List[Dict]:
        """对比版本差异"""
        logger.info("\n🔍 Step 2: 版本对比")

        if not Path(self.baseline_file).exists():
            logger.error(f"❌ 基线文件不存在")
            return []

        changes = []

        # 读取文件
        with open(self.baseline_file, 'r', encoding='utf-8') as f:
            baseline_data = list(csv.reader(f))

        with open(current_file, 'r', encoding='utf-8') as f:
            current_data = list(csv.reader(f))

        logger.info(f"📊 基线: {len(baseline_data)}行 × {len(baseline_data[0]) if baseline_data else 0}列")
        logger.info(f"📊 当前: {len(current_data)}行 × {len(current_data[0]) if current_data else 0}列")

        # 对比（只对比相同范围）
        max_rows = min(len(baseline_data), len(current_data))
        max_cols = min(len(baseline_data[0]), len(current_data[0])) if baseline_data and current_data else 0

        for row_idx in range(1, max_rows):  # 跳过表头
            for col_idx in range(max_cols):
                val_baseline = str(baseline_data[row_idx][col_idx]).strip() if col_idx < len(baseline_data[row_idx]) else ""
                val_current = str(current_data[row_idx][col_idx]).strip() if col_idx < len(current_data[row_idx]) else ""

                if val_baseline != val_current:
                    col_name = baseline_data[0][col_idx] if col_idx < len(baseline_data[0]) else f"列{col_idx+1}"
                    changes.append({
                        'row': row_idx + 1,
                        'col': col_idx,
                        'col_name': col_name,
                        'old_value': val_baseline[:30],  # 截断显示
                        'new_value': val_current[:30]
                    })

        logger.info(f"✅ 发现 {len(changes)} 处变更")

        # 显示前5个变更
        for i, change in enumerate(changes[:5]):
            logger.info(f"   变更{i+1}: [{change['row']},{change['col']+1}] {change['old_value']} → {change['new_value']}")

        if len(changes) > 5:
            logger.info(f"   ... 还有 {len(changes)-5} 处变更")

        return changes

    def create_excel(self, csv_file: str, changes: List[Dict]) -> str:
        """创建Excel报告"""
        logger.info("\n📊 Step 3: 生成Excel报告")

        # 读取CSV数据
        with open(csv_file, 'r', encoding='utf-8') as f:
            data = list(csv.reader(f))

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = self.doc_name

        # 写入数据
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # 表头格式
                if row_idx == 1:
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="366092",
                        end_color="366092",
                        fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal='center')

        # 应用涂色（关键：使用solid）
        cells_colored = 0
        for change in changes:
            # 评分
            col_name = change['col_name']
            if col_name in L1_COLUMNS:
                color = "FFCCCC"  # 浅红
                risk = "高风险"
            elif col_name in L2_COLUMNS:
                color = "FFFFCC"  # 浅黄
                risk = "中风险"
            else:
                color = "CCFFCC"  # 浅绿
                risk = "低风险"

            # 涂色
            cell = ws.cell(row=change['row'], column=change['col']+1)
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"  # 关键！
            )

            # 批注
            comment = f"{risk}\n原值: {change['old_value']}\n新值: {change['new_value']}"
            cell.comment = Comment(comment, "AI分析")

            cells_colored += 1

        # 保存
        excel_file = self.download_dir / f"verified_report_{self.timestamp}.xlsx"
        wb.save(excel_file)

        size = excel_file.stat().st_size / 1024
        logger.info(f"✅ Excel生成: {excel_file.name}")
        logger.info(f"🎨 涂色单元格: {cells_colored}个")
        logger.info(f"📊 文件大小: {size:.2f} KB")

        return str(excel_file)

    async def upload_excel(self, excel_file: str) -> Optional[str]:
        """上传Excel"""
        logger.info("\n📤 Step 4: 上传到腾讯文档")

        try:
            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')

            # 上传
            from production.core_modules.tencent_doc_upload_production_v3 import quick_upload_v3

            result = await quick_upload_v3(
                cookie_string=cookie_string,
                file_path=excel_file,
                headless=True
            )

            if result and result.get('success'):
                url = result.get('url')
                logger.info(f"✅ 上传成功!")
                logger.info(f"🔗 URL: {url}")
                return url

        except Exception as e:
            logger.error(f"❌ 上传失败: {e}")

        return None

    async def run_test(self):
        """运行完整测试"""
        logger.info("\n" + "="*70)
        logger.info("🚀 开始验证过的全链路测试")
        logger.info("="*70)

        # 下载
        csv_file = await self.download_current()
        if not csv_file:
            logger.error("❌ 测试失败：下载失败")
            return None

        # 对比
        changes = self.compare_versions(csv_file)
        if not changes:
            logger.warning("⚠️ 没有发现变更")
            # 继续处理，生成空报告

        # 生成Excel
        excel_file = self.create_excel(csv_file, changes)

        # 上传
        url = await self.upload_excel(excel_file)

        # 总结
        logger.info("\n" + "="*70)
        logger.info("📊 测试总结")
        logger.info("="*70)
        logger.info(f"📄 文档: {self.doc_name}")
        logger.info(f"🔄 变更数: {len(changes)}")
        logger.info(f"📁 Excel: {Path(excel_file).name}")

        if url:
            logger.info(f"✅ 上传成功: {url}")
            logger.info("\n🎉 测试成功！请访问URL验证涂色效果")
        else:
            logger.info("⚠️ 上传失败，但Excel文件已生成")

        return url


async def main():
    """主函数"""
    test = VerifiedChainTest()
    url = await test.run_test()

    if url:
        print(f"\n🌟 最终URL: {url}")
        print("👉 请检查涂色是否正确显示")
    else:
        print("\n⚠️ 测试未完全成功")


if __name__ == "__main__":
    asyncio.run(main())