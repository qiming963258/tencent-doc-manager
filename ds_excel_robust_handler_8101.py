#!/usr/bin/env python3
"""
DS Excel处理器 - 强化版
支持处理各种格式的Excel文件，包括腾讯文档导出的文件
"""

import os
import sys
import logging
import shutil
from flask import Flask, request, jsonify, send_file, render_template_string
from werkzeug.utils import secure_filename
from datetime import datetime
import traceback

# 导入多个Excel处理库
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("正在安装openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

# 尝试导入xlrd用于处理.xls文件
try:
    import xlrd
    import xlwt
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    print("提示: xlrd/xlwt未安装，将只支持.xlsx格式")

# 尝试导入pandas作为备用方案
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("提示: pandas未安装，某些高级功能可能受限")

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Flask应用配置
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
app.config['UPLOAD_FOLDER'] = '/root/projects/tencent-doc-manager/excel_uploads'

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 支持的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm', 'xlsb'}

# 默认测试文件路径
DEFAULT_TEST_FILE = '/root/projects/tencent-doc-manager/test_sales_data.xlsx'

# 全局变量存储当前工作簿
current_workbook = None
current_worksheet = None
current_file_path = None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def safe_load_excel(filepath):
    """
    安全加载Excel文件，支持多种格式和错误恢复
    """
    file_ext = filepath.rsplit('.', 1)[1].lower()
    
    # 方案1: 尝试直接用openpyxl加载
    if file_ext in ['xlsx', 'xlsm', 'xlsb']:
        try:
            # 首先尝试只读模式（更稳定）
            wb = load_workbook(filepath, read_only=True, data_only=True)
            # 如果成功，转换为可写模式
            temp_wb = Workbook()
            source_ws = wb.active
            target_ws = temp_wb.active
            
            # 复制数据
            for row in source_ws.iter_rows(values_only=True):
                target_ws.append(row)
            
            wb.close()
            return temp_wb, None
        except Exception as e1:
            logger.warning(f"只读模式加载失败，尝试标准模式: {e1}")
            try:
                # 尝试标准加载（可能会忽略某些格式错误）
                wb = load_workbook(filepath, data_only=False)
                return wb, None
            except Exception as e2:
                logger.warning(f"标准模式加载失败: {e2}")
    
    # 方案2: 对于.xls文件，使用xlrd转换
    if file_ext == 'xls' and HAS_XLRD:
        try:
            logger.info("检测到.xls文件，尝试转换为.xlsx")
            # 读取.xls文件
            xls_book = xlrd.open_workbook(filepath)
            xls_sheet = xls_book.sheet_by_index(0)
            
            # 创建新的.xlsx工作簿
            new_wb = Workbook()
            new_ws = new_wb.active
            
            # 复制数据
            for row_idx in range(xls_sheet.nrows):
                row_data = []
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    row_data.append(cell_value)
                new_ws.append(row_data)
            
            return new_wb, None
        except Exception as e:
            logger.error(f"xls转换失败: {e}")
    
    # 方案3: 使用pandas作为最后手段
    if HAS_PANDAS:
        try:
            logger.info("尝试使用pandas读取文件")
            # 读取Excel文件
            df = pd.read_excel(filepath, engine=None)
            
            # 创建新工作簿
            new_wb = Workbook()
            new_ws = new_wb.active
            
            # 写入表头
            headers = df.columns.tolist()
            new_ws.append(headers)
            
            # 写入数据
            for _, row in df.iterrows():
                new_ws.append(row.tolist())
            
            return new_wb, None
        except Exception as e:
            logger.error(f"pandas读取失败: {e}")
    
    # 方案4: 尝试修复损坏的xlsx文件
    if file_ext == 'xlsx':
        try:
            logger.info("尝试修复可能损坏的xlsx文件")
            # 创建临时文件
            temp_file = filepath + '.temp.xlsx'
            
            # 使用zip修复（xlsx本质是zip文件）
            import zipfile
            import tempfile
            import shutil
            
            with tempfile.TemporaryDirectory() as temp_dir:
                # 解压文件
                try:
                    with zipfile.ZipFile(filepath, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                    
                    # 重新打包
                    with zipfile.ZipFile(temp_file, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                        for root, dirs, files in os.walk(temp_dir):
                            for file in files:
                                file_path = os.path.join(root, file)
                                arc_name = os.path.relpath(file_path, temp_dir)
                                zip_ref.write(file_path, arc_name)
                    
                    # 尝试加载修复后的文件
                    wb = load_workbook(temp_file, data_only=False)
                    os.remove(temp_file)
                    return wb, None
                except Exception as repair_error:
                    logger.error(f"文件修复失败: {repair_error}")
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
        except Exception as e:
            logger.error(f"修复过程失败: {e}")
    
    return None, "无法加载Excel文件，文件可能已损坏或格式不支持"

def create_test_excel():
    """创建测试Excel文件"""
    if not os.path.exists(DEFAULT_TEST_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "测试数据"
        
        # 添加表头
        headers = ['序号', '产品名称', '类别', '数量', '单价', '总价']
        ws.append(headers)
        
        # 添加示例数据
        for i in range(1, 11):
            ws.append([i, f'产品{i}', '类别A', i*10, i*100, i*1000])
        
        # 设置表头样式
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        wb.save(DEFAULT_TEST_FILE)
        logger.info(f"创建测试文件: {DEFAULT_TEST_FILE}")

# 简化的自然语言处理器
class RobustNLPProcessor:
    def __init__(self):
        self.color_map = {
            '红': 'FF0000', '红色': 'FF0000', 'red': 'FF0000',
            '绿': '00FF00', '绿色': '00FF00', 'green': '00FF00',
            '蓝': '0000FF', '蓝色': '0000FF', 'blue': '0000FF',
            '黄': 'FFFF00', '黄色': 'FFFF00', 'yellow': 'FFFF00',
            '橙': 'FFA500', '橙色': 'FFA500', 'orange': 'FFA500',
            '紫': '800080', '紫色': '800080', 'purple': '800080',
            '灰': '808080', '灰色': '808080', 'gray': '808080', 'grey': '808080',
            '黑': '000000', '黑色': '000000', 'black': '000000',
            '白': 'FFFFFF', '白色': 'FFFFFF', 'white': 'FFFFFF',
            '粉': 'FFC0CB', '粉色': 'FFC0CB', 'pink': 'FFC0CB',
            '青': '00FFFF', '青色': '00FFFF', 'cyan': '00FFFF',
            '棕': 'A52A2A', '棕色': 'A52A2A', 'brown': 'A52A2A'
        }
    
    def parse_instruction(self, instruction):
        """解析自然语言指令"""
        import re
        instruction = instruction.lower().strip()
        
        # 单元格涂色
        patterns = [
            r'([a-z]+\d+)[^\w]*(?:涂|填|标记|改|变|设|染|着)(?:成|为|上)?[^\w]*(.+色|red|green|blue|yellow|orange|purple|gray|grey|black|white|pink|cyan|brown)',
            r'(?:把|将)?([a-z]+\d+)[^\w]*(?:涂|填|标记|改|变)(?:成|为)?[^\w]*(.+色)',
            r'(?:给)?([a-z]+\d+)[^\w]*(?:涂|填|加|添)(?:上)?[^\w]*(.+色)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, instruction)
            if match:
                cell = match.group(1).upper()
                color = match.group(2)
                return {
                    'action': 'color',
                    'type': 'cell',
                    'target': cell,
                    'color': self._get_color_code(color)
                }
        
        return {'action': 'unknown'}
    
    def _get_color_code(self, color_text):
        """获取颜色代码"""
        color_text = color_text.strip().lower()
        for key, value in self.color_map.items():
            if key in color_text:
                return value
        return 'FF0000'  # 默认红色

# 创建处理器实例
nlp_processor = RobustNLPProcessor()

# HTML模板
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DS Excel处理器 - 强化版</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container { max-width: 1200px; margin: 0 auto; }
        .header {
            text-align: center;
            color: white;
            margin-bottom: 30px;
        }
        .card {
            background: white;
            border-radius: 16px;
            padding: 30px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }
        .upload-area {
            border: 3px dashed #ddd;
            border-radius: 12px;
            padding: 40px;
            text-align: center;
            background: #f8f9fa;
            cursor: pointer;
            transition: all 0.3s;
        }
        .upload-area:hover { 
            border-color: #667eea;
            background: #f0f0ff;
        }
        .upload-area.dragover {
            border-color: #667eea;
            background: #e8e8ff;
        }
        input[type="file"] { display: none; }
        .btn {
            padding: 12px 24px;
            border: none;
            border-radius: 8px;
            font-size: 16px;
            cursor: pointer;
            transition: all 0.3s;
            margin: 5px;
        }
        .btn-primary {
            background: linear-gradient(135deg, #667eea, #764ba2);
            color: white;
        }
        .btn-success {
            background: linear-gradient(135deg, #11998e, #38ef7d);
            color: white;
        }
        .btn:hover { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(0,0,0,0.2); }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; }
        .format-info {
            background: #e8f5e9;
            border-left: 4px solid #4caf50;
            padding: 15px;
            margin: 20px 0;
            border-radius: 4px;
        }
        .error { color: #f44336; margin-top: 10px; }
        .success { color: #4caf50; margin-top: 10px; }
        .instruction-input {
            width: 100%;
            padding: 15px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 16px;
            margin: 10px 0;
        }
        .instruction-input:focus {
            outline: none;
            border-color: #667eea;
        }
        .examples {
            background: #f5f5f5;
            padding: 15px;
            border-radius: 8px;
            margin-top: 20px;
        }
        .examples h4 { margin-bottom: 10px; color: #333; }
        .examples ul { list-style: none; }
        .examples li {
            padding: 5px 0;
            color: #666;
        }
        .examples code {
            background: #e0e0e0;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: monospace;
        }
        #result {
            margin-top: 20px;
            padding: 15px;
            background: #f8f9fa;
            border-radius: 8px;
            min-height: 60px;
        }
        .file-info {
            background: #e3f2fd;
            padding: 15px;
            border-radius: 8px;
            margin: 15px 0;
        }
        .path-input-group {
            display: flex;
            gap: 10px;
            margin: 20px 0;
        }
        .path-input {
            flex: 1;
            padding: 12px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 14px;
        }
        .test-files {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            margin: 15px 0;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 DS Excel处理器 - 强化版</h1>
            <p>支持多种Excel格式，智能错误恢复</p>
        </div>
        
        <div class="card">
            <h2>📁 步骤1: 选择Excel文件</h2>
            
            <div class="format-info">
                <strong>✅ 支持格式：</strong>
                <ul style="margin-top: 5px; margin-left: 20px;">
                    <li>• Excel 2007+ (.xlsx) - 推荐</li>
                    <li>• Excel 97-2003 (.xls) - 自动转换</li>
                    <li>• 腾讯文档导出文件 - 自动修复</li>
                    <li>• 带宏的Excel (.xlsm)</li>
                </ul>
            </div>
            
            <div class="upload-area" id="uploadArea">
                <svg width="64" height="64" viewBox="0 0 24 24" fill="none" stroke="#667eea" stroke-width="2">
                    <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path>
                    <polyline points="17 8 12 3 7 8"></polyline>
                    <line x1="12" y1="3" x2="12" y2="15"></line>
                </svg>
                <h3 style="margin: 20px 0; color: #333;">点击或拖拽Excel文件到这里</h3>
                <p style="color: #999;">支持多种Excel格式，最大50MB</p>
                <input type="file" id="fileInput" accept=".xlsx,.xls,.xlsm,.xlsb">
            </div>
            
            <div class="path-input-group">
                <input type="text" class="path-input" id="serverPath" 
                       placeholder="或输入服务器文件路径（如：/root/projects/file.xlsx）">
                <button class="btn btn-primary" onclick="loadServerFile()">加载服务器文件</button>
            </div>
            
            <div class="test-files">
                <strong>快速测试：</strong>
                <button class="btn btn-success" onclick="loadTestFile()">
                    📊 加载默认测试文件
                </button>
            </div>
            
            <div id="fileInfo"></div>
        </div>
        
        <div class="card">
            <h2>✏️ 步骤2: 输入修改指令</h2>
            <input type="text" class="instruction-input" id="instructionInput" 
                   placeholder="输入自然语言指令（如：g5涂红色）" disabled>
            <button class="btn btn-primary" id="executeBtn" onclick="executeInstruction()" disabled>
                执行命令
            </button>
            <button class="btn btn-primary" id="downloadBtn" onclick="downloadFile()" disabled>
                下载修改后的文件
            </button>
            
            <div class="examples">
                <h4>📝 支持的指令示例：</h4>
                <ul>
                    <li>• 单元格涂色：<code>g5涂红色</code>、<code>把b3填成蓝色</code></li>
                    <li>• 使用小写字母+数字格式（如 g5 而不是 G5）</li>
                    <li>• 支持的颜色：红、绿、蓝、黄、橙、紫、灰、黑、粉、青、棕</li>
                </ul>
            </div>
            
            <div id="result"></div>
        </div>
    </div>
    
    <script>
        let currentFile = null;
        let modifiedFilePath = null;
        
        // 拖拽上传
        const uploadArea = document.getElementById('uploadArea');
        const fileInput = document.getElementById('fileInput');
        
        uploadArea.addEventListener('click', () => fileInput.click());
        
        uploadArea.addEventListener('dragover', (e) => {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });
        
        uploadArea.addEventListener('dragleave', () => {
            uploadArea.classList.remove('dragover');
        });
        
        uploadArea.addEventListener('drop', (e) => {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            const files = e.dataTransfer.files;
            if (files.length > 0) {
                handleFile(files[0]);
            }
        });
        
        fileInput.addEventListener('change', (e) => {
            if (e.target.files.length > 0) {
                handleFile(e.target.files[0]);
            }
        });
        
        function handleFile(file) {
            const formData = new FormData();
            formData.append('file', file);
            
            document.getElementById('result').innerHTML = '<div class="success">上传中...</div>';
            
            fetch('/upload', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    currentFile = file.name;
                    document.getElementById('fileInfo').innerHTML = 
                        `<div class="file-info">✅ 已加载文件: <strong>${file.name}</strong> (${(file.size/1024).toFixed(2)} KB)</div>`;
                    document.getElementById('instructionInput').disabled = false;
                    document.getElementById('executeBtn').disabled = false;
                    document.getElementById('result').innerHTML = '<div class="success">文件上传成功！</div>';
                } else {
                    document.getElementById('result').innerHTML = 
                        `<div class="error">文件上传失败：${data.message}</div>`;
                }
            })
            .catch(error => {
                document.getElementById('result').innerHTML = 
                    `<div class="error">上传错误：${error}</div>`;
            });
        }
        
        function loadServerFile() {
            const filepath = document.getElementById('serverPath').value.trim();
            if (!filepath) {
                alert('请输入文件路径');
                return;
            }
            
            fetch('/load_server_file', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({filepath: filepath})
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    currentFile = data.filename;
                    document.getElementById('fileInfo').innerHTML = 
                        `<div class="file-info">✅ 已加载服务器文件: <strong>${data.filename}</strong> (${data.filesize})</div>`;
                    document.getElementById('instructionInput').disabled = false;
                    document.getElementById('executeBtn').disabled = false;
                    document.getElementById('result').innerHTML = '<div class="success">服务器文件加载成功！</div>';
                } else {
                    document.getElementById('result').innerHTML = 
                        `<div class="error">加载失败：${data.message}</div>`;
                }
            });
        }
        
        function loadTestFile() {
            fetch('/load_test_file', {method: 'POST'})
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    currentFile = data.filename;
                    document.getElementById('fileInfo').innerHTML = 
                        `<div class="file-info">✅ 已加载测试文件: <strong>${data.filename}</strong></div>`;
                    document.getElementById('instructionInput').disabled = false;
                    document.getElementById('executeBtn').disabled = false;
                    document.getElementById('result').innerHTML = '<div class="success">测试文件加载成功！</div>';
                }
            });
        }
        
        function executeInstruction() {
            const instruction = document.getElementById('instructionInput').value.trim();
            if (!instruction) {
                alert('请输入指令');
                return;
            }
            
            document.getElementById('result').innerHTML = '<div>处理中...</div>';
            
            fetch('/execute', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({instruction: instruction})
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    modifiedFilePath = data.file_path;
                    document.getElementById('downloadBtn').disabled = false;
                    document.getElementById('result').innerHTML = 
                        `<div class="success">${data.message}</div>`;
                } else {
                    document.getElementById('result').innerHTML = 
                        `<div class="error">${data.message}</div>`;
                }
            });
        }
        
        function downloadFile() {
            if (modifiedFilePath) {
                window.open('/download?filepath=' + encodeURIComponent(modifiedFilePath));
            }
        }
    </script>
</body>
</html>
'''

@app.route('/')
def index():
    create_test_excel()
    return render_template_string(HTML_TEMPLATE)

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有文件'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # 保存文件
            file.save(filepath)
            logger.info(f"文件已保存到: {filepath}")
            
            # 尝试加载文件
            global current_workbook, current_worksheet, current_file_path
            wb, error = safe_load_excel(filepath)
            
            if wb:
                current_workbook = wb
                current_worksheet = wb.active
                current_file_path = filepath
                logger.info(f"Excel文件加载成功: {filepath}")
                return jsonify({'success': True, 'message': '文件上传成功'})
            else:
                # 清理失败的文件
                if os.path.exists(filepath):
                    os.remove(filepath)
                return jsonify({'success': False, 'message': f'文件加载失败: {error}'})
        else:
            return jsonify({'success': False, 'message': f'不支持的文件格式，请使用: {", ".join(ALLOWED_EXTENSIONS)}'})
    except Exception as e:
        logger.error(f"上传处理失败: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'处理失败: {str(e)}'})

@app.route('/load_server_file', methods=['POST'])
def load_server_file():
    try:
        data = request.get_json()
        filepath = data.get('filepath', '').strip()
        
        if not filepath:
            return jsonify({'success': False, 'message': '请提供文件路径'})
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': '文件不存在'})
        
        # 加载文件
        global current_workbook, current_worksheet, current_file_path
        wb, error = safe_load_excel(filepath)
        
        if wb:
            # 复制到上传目录
            filename = os.path.basename(filepath)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{timestamp}_{filename}"
            new_filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
            
            # 保存工作簿
            wb.save(new_filepath)
            
            current_workbook = wb
            current_worksheet = wb.active
            current_file_path = new_filepath
            
            filesize = os.path.getsize(filepath)
            return jsonify({
                'success': True,
                'message': '服务器文件加载成功',
                'filename': filename,
                'filesize': f"{filesize/1024:.2f} KB"
            })
        else:
            return jsonify({'success': False, 'message': f'文件加载失败: {error}'})
    except Exception as e:
        logger.error(f"服务器文件加载失败: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/load_test_file', methods=['POST'])
def load_test_file():
    try:
        create_test_excel()
        
        global current_workbook, current_worksheet, current_file_path
        wb, error = safe_load_excel(DEFAULT_TEST_FILE)
        
        if wb:
            # 复制到上传目录
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{timestamp}_test_data.xlsx"
            new_filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
            
            wb.save(new_filepath)
            
            current_workbook = wb
            current_worksheet = wb.active
            current_file_path = new_filepath
            
            return jsonify({
                'success': True,
                'message': '测试文件加载成功',
                'filename': 'test_sales_data.xlsx'
            })
        else:
            return jsonify({'success': False, 'message': '测试文件加载失败'})
    except Exception as e:
        logger.error(f"测试文件加载失败: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/execute', methods=['POST'])
def execute_instruction():
    try:
        global current_workbook, current_worksheet, current_file_path
        
        if not current_workbook:
            return jsonify({'success': False, 'message': '请先上传文件'})
        
        data = request.get_json()
        instruction = data.get('instruction', '').strip()
        
        if not instruction:
            return jsonify({'success': False, 'message': '请输入指令'})
        
        # 解析指令
        result = nlp_processor.parse_instruction(instruction)
        
        if result['action'] == 'color' and result['type'] == 'cell':
            # 执行单元格涂色
            cell = result['target']
            color = result['color']
            
            try:
                cell_obj = current_worksheet[cell]
                cell_obj.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                # 保存修改后的文件
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                base_name = os.path.basename(current_file_path).rsplit('.', 1)[0]
                new_filename = f"{timestamp}_{base_name}_modified.xlsx"
                new_filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
                
                current_workbook.save(new_filepath)
                
                return jsonify({
                    'success': True,
                    'message': f'✅ 已将{cell}涂成颜色#{color}',
                    'file_path': new_filepath
                })
            except Exception as e:
                return jsonify({'success': False, 'message': f'执行失败: {str(e)}'})
        else:
            return jsonify({'success': False, 'message': f'未能识别指令: {instruction}'})
    except Exception as e:
        logger.error(f"执行指令失败: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/download')
def download_file():
    filepath = request.args.get('filepath')
    if filepath and os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return "文件不存在", 404

if __name__ == '__main__':
    print("="*60)
    print("🚀 DS Excel处理器 - 强化版启动中...")
    print("="*60)
    print(f"📍 访问地址: http://localhost:8101")
    print(f"📍 外网地址: http://202.140.143.88:8101")
    print(f"📁 上传目录: {app.config['UPLOAD_FOLDER']}")
    print(f"📊 默认测试文件: {DEFAULT_TEST_FILE}")
    print("💡 增强功能:")
    print("   - 支持.xls/.xlsx等多种格式")
    print("   - 自动修复腾讯文档导出文件")
    print("   - 智能错误恢复机制")
    print("   - 多重加载策略")
    print("="*60)
    
    app.run(host='0.0.0.0', port=8101, debug=False)