#!/usr/bin/env python3
"""
测试修复后的涂色功能
确保颜色正确应用到Excel文件
"""

import json
import openpyxl
from intelligent_excel_marker import IntelligentExcelMarker
from openpyxl.styles import PatternFill
import os

def test_coloring():
    """测试涂色功能"""
    print("=" * 60)
    print("测试腾讯文档兼容的涂色功能")
    print("=" * 60)
    
    # 1. 找到最新的Excel文件和打分文件
    excel_dir = "/root/projects/tencent-doc-manager/csv_versions/2025_W37/midweek/"
    score_dir = "/root/projects/tencent-doc-manager/scoring_results/detailed/"
    
    # 获取最新的Excel文件
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('_fixed.xlsx')]
    if not excel_files:
        print("❌ 没有找到修复后的Excel文件")
        return False
    
    excel_file = os.path.join(excel_dir, sorted(excel_files)[-1])
    print(f"✅ 找到Excel文件: {excel_file}")
    
    # 获取最新的打分文件
    score_files = [f for f in os.listdir(score_dir) if f.startswith('detailed_score_')]
    if not score_files:
        print("❌ 没有找到打分文件")
        return False
    
    score_file = os.path.join(score_dir, sorted(score_files)[-1])
    print(f"✅ 找到打分文件: {score_file}")
    
    # 2. 加载打分数据查看风险等级
    with open(score_file, 'r', encoding='utf-8') as f:
        score_data = json.load(f)
    
    print("\n风险等级分布:")
    risk_counts = {}
    if 'scores' in score_data:
        for item in score_data['scores']:
            risk_level = item.get('risk_assessment', {}).get('risk_level', 'UNKNOWN')
            risk_counts[risk_level] = risk_counts.get(risk_level, 0) + 1
            print(f"  {item.get('cell')}: {risk_level}")
    
    print(f"\n风险统计: {risk_counts}")
    
    # 3. 应用涂色
    marker = IntelligentExcelMarker()
    output_file = marker.apply_striped_coloring(excel_file, score_file)
    
    if not output_file:
        print("❌ 涂色失败")
        return False
    
    print(f"\n✅ 涂色完成: {output_file}")
    
    # 4. 验证涂色结果
    print("\n验证涂色结果...")
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    colored_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.patternType == 'solid':
                if cell.fill.fgColor and str(cell.fill.fgColor.rgb) != '00000000':
                    colored_cells.append({
                        'cell': f"{cell.column_letter}{cell.row}",
                        'value': cell.value,
                        'color': str(cell.fill.fgColor.rgb)
                    })
    
    wb.close()
    
    print(f"\n找到 {len(colored_cells)} 个涂色单元格:")
    for cell_info in colored_cells[:10]:
        print(f"  {cell_info['cell']}: {cell_info['value']}")
        print(f"    颜色: {cell_info['color']}")
    
    if len(colored_cells) > 0:
        print(f"\n✅ 涂色验证成功！共涂色 {len(colored_cells)} 个单元格")
        return True
    else:
        print("\n❌ 没有找到涂色的单元格")
        return False

if __name__ == "__main__":
    success = test_coloring()
    exit(0 if success else 1)