#!/usr/bin/env python3
"""
终极Excel修复工具 - 专门解决腾讯文档和其他自动导出Excel的兼容性问题
根本原因：腾讯文档导出的Excel文件styles.xml包含不完整的Fill定义
"""

import zipfile
import os
import tempfile
import shutil
from pathlib import Path
import xml.etree.ElementTree as ET

def fix_excel_styles(input_file, output_file=None):
    """
    修复Excel文件的样式问题
    核心思路：重写styles.xml文件，确保所有Fill定义完整
    """
    if not output_file:
        base = Path(input_file).stem
        output_file = Path(input_file).parent / f"{base}_fixed.xlsx"
    
    print(f"🔧 正在修复: {input_file}")
    
    # 创建临时目录
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        
        # 解压Excel文件
        with zipfile.ZipFile(input_file, 'r') as zip_ref:
            zip_ref.extractall(temp_path)
        
        # 修复styles.xml
        styles_path = temp_path / 'xl' / 'styles.xml'
        
        if styles_path.exists():
            print("📝 找到styles.xml，开始修复...")
            
            # 创建最小化但完整的styles.xml
            minimal_styles = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <numFmts count="0"/>
    <fonts count="2">
        <font>
            <sz val="11"/>
            <name val="Calibri"/>
            <family val="2"/>
            <scheme val="minor"/>
        </font>
        <font>
            <sz val="11"/>
            <name val="Calibri"/>
            <family val="2"/>
            <scheme val="minor"/>
            <b/>
        </font>
    </fonts>
    <fills count="3">
        <fill>
            <patternFill patternType="none"/>
        </fill>
        <fill>
            <patternFill patternType="gray125"/>
        </fill>
        <fill>
            <patternFill patternType="solid">
                <fgColor rgb="FFFFFF00"/>
            </patternFill>
        </fill>
    </fills>
    <borders count="1">
        <border>
            <left/>
            <right/>
            <top/>
            <bottom/>
            <diagonal/>
        </border>
    </borders>
    <cellStyleXfs count="1">
        <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
    </cellStyleXfs>
    <cellXfs count="4">
        <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
        <xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>
        <xf numFmtId="0" fontId="0" fillId="2" borderId="0" xfId="0" applyFill="1"/>
        <xf numFmtId="14" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>
    </cellXfs>
    <cellStyles count="1">
        <cellStyle name="Normal" xfId="0" builtinId="0"/>
    </cellStyles>
    <dxfs count="0"/>
    <tableStyles count="0" defaultTableStyle="TableStyleMedium2" defaultPivotStyle="PivotStyleLight16"/>
</styleSheet>'''
            
            # 写入修复后的styles.xml
            with open(styles_path, 'w', encoding='utf-8') as f:
                f.write(minimal_styles)
            
            print("✅ styles.xml已修复")
        
        # 重新打包为xlsx
        with zipfile.ZipFile(output_file, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for root, dirs, files in os.walk(temp_path):
                for file in files:
                    file_path = Path(root) / file
                    arc_name = file_path.relative_to(temp_path)
                    zip_ref.write(file_path, str(arc_name))
    
    print(f"✅ 文件修复完成: {output_file}")
    return output_file

def test_fixed_file(file_path):
    """测试修复后的文件是否能正常加载"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        print(f"✅ 文件可以正常加载")
        print(f"📊 工作表: {wb.sheetnames}")
        print(f"📏 数据范围: {ws.max_row}行 × {ws.max_column}列")
        
        # 读取前几个单元格
        sample_data = []
        for row in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(4, ws.max_column + 1)):
                value = ws.cell(row=row, column=col).value
                row_data.append(str(value)[:20] if value else '')
            sample_data.append(row_data)
        
        print(f"📝 示例数据:")
        for row in sample_data:
            print(f"   {row}")
        
        wb.close()
        return True
    except Exception as e:
        print(f"❌ 加载失败: {e}")
        return False

def batch_fix_excel_files(directory, pattern="*.xlsx"):
    """批量修复目录下的Excel文件"""
    dir_path = Path(directory)
    files = list(dir_path.glob(pattern))
    
    print(f"📁 找到 {len(files)} 个Excel文件")
    
    fixed_files = []
    for file in files:
        if '_fixed' not in file.stem:  # 跳过已修复的文件
            try:
                output = fix_excel_styles(file)
                if test_fixed_file(output):
                    fixed_files.append(output)
            except Exception as e:
                print(f"⚠️ 修复失败 {file}: {e}")
    
    return fixed_files

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("用法:")
        print("  修复单个文件: python ultimate_excel_fixer.py <文件路径>")
        print("  批量修复: python ultimate_excel_fixer.py --batch <目录路径>")
        sys.exit(1)
    
    if sys.argv[1] == "--batch" and len(sys.argv) > 2:
        # 批量修复
        fixed = batch_fix_excel_files(sys.argv[2])
        print(f"\n📊 修复完成: {len(fixed)} 个文件")
    else:
        # 单个文件修复
        input_file = sys.argv[1]
        if not os.path.exists(input_file):
            print(f"错误: 文件不存在 - {input_file}")
            sys.exit(1)
        
        output_file = fix_excel_styles(input_file)
        test_fixed_file(output_file)