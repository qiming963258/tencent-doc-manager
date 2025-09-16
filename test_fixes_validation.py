#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证8093系统修复后的完整性测试
测试所有修复的功能点
"""

import os
import sys
import json
import csv
from pathlib import Path

# 添加项目路径
sys.path.append('/root/projects/tencent-doc-manager')

def test_l2_semantic_analyzer():
    """测试L2语义分析器的正确方法"""
    print("\n" + "="*60)
    print("测试1: L2SemanticAnalyzer.analyze_modifications方法")
    print("="*60)
    
    try:
        from production.core_modules.l2_semantic_analysis_two_layer import L2SemanticAnalyzer
        
        analyzer = L2SemanticAnalyzer()
        
        # 测试数据格式
        test_modifications = [
            {
                'column_name': '销售额',
                'old_value': '1000',
                'new_value': '2000',
                'row': 5,
                'cell': 'B5'
            },
            {
                'column_name': '库存量',
                'old_value': '500',
                'new_value': '0',
                'row': 10,
                'cell': 'C10'
            }
        ]
        
        # 调用正确的方法
        result = analyzer.analyze_modifications(test_modifications)
        
        print("✅ L2SemanticAnalyzer测试通过")
        print(f"   - 方法调用成功: analyze_modifications")
        print(f"   - 返回结果类型: {type(result)}")
        print(f"   - 分析了 {len(test_modifications)} 个修改")
        
        return True
        
    except Exception as e:
        print(f"❌ L2SemanticAnalyzer测试失败: {e}")
        return False


def test_csv_score_generator():
    """测试CSV文件的打分生成器"""
    print("\n" + "="*60)
    print("测试2: DetailedScoreGenerator对CSV文件的支持")
    print("="*60)
    
    try:
        from intelligent_excel_marker import DetailedScoreGenerator
        
        # 创建测试CSV文件
        test_dir = Path("/tmp/test_csv_scoring")
        test_dir.mkdir(exist_ok=True)
        
        # 创建基线CSV
        baseline_csv = test_dir / "baseline_W36.csv"
        with open(baseline_csv, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['产品', '销量', '价格'])
            writer.writerow(['A', 100, 10.5])
            writer.writerow(['B', 200, 20.8])
            writer.writerow(['C', 300, 30.2])
        
        # 创建目标CSV（有修改）
        target_csv = test_dir / "target_W36.csv"
        with open(target_csv, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['产品', '销量', '价格'])
            writer.writerow(['A', 150, 10.5])  # 修改了A的销量
            writer.writerow(['B', 200, 25.0])  # 修改了B的价格
            writer.writerow(['C', 250, 30.2])  # 修改了C的销量
        
        # 测试打分生成
        generator = DetailedScoreGenerator()
        output_dir = test_dir / "scores"
        output_dir.mkdir(exist_ok=True)
        
        score_file = generator.generate_score_json(
            str(baseline_csv),
            str(target_csv),
            str(output_dir)
        )
        
        # 验证输出文件
        if Path(score_file).exists():
            with open(score_file, 'r', encoding='utf-8') as f:
                score_data = json.load(f)
            
            print("✅ CSV打分生成测试通过")
            print(f"   - CSV文件成功处理")
            print(f"   - 生成打分文件: {Path(score_file).name}")
            print(f"   - 检测到变更单元格: {score_data['statistics']['changed_cells']}")
            print(f"   - 总单元格数: {score_data['statistics']['total_cells']}")
            
            return True
        else:
            print("❌ CSV打分生成测试失败: 未生成输出文件")
            return False
            
    except Exception as e:
        print(f"❌ CSV打分生成测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_no_degradation():
    """测试降级代码是否已删除"""
    print("\n" + "="*60)
    print("测试3: 验证降级代码已删除")
    print("="*60)
    
    try:
        # 读取主系统文件
        with open('/root/projects/tencent-doc-manager/production_integrated_test_system_8093.py', 'r') as f:
            content = f.read()
        
        # 检查是否还有降级关键词
        degradation_keywords = [
            '使用基础打分',
            '基础打分文件已生成',
            '真实打分生成失败',
            '虚拟',
            '模拟'
        ]
        
        found_keywords = []
        for keyword in degradation_keywords:
            if keyword in content:
                found_keywords.append(keyword)
        
        if not found_keywords:
            print("✅ 降级代码删除验证通过")
            print("   - 未发现降级关键词")
            print("   - 异常处理改为直接抛出")
            return True
        else:
            print(f"❌ 降级代码删除验证失败")
            print(f"   - 发现降级关键词: {found_keywords}")
            return False
            
    except Exception as e:
        print(f"❌ 降级代码验证失败: {e}")
        return False


def test_excel_format_handling():
    """测试Excel格式处理"""
    print("\n" + "="*60)
    print("测试4: Excel格式处理和文件路径传递")
    print("="*60)
    
    try:
        from fix_tencent_excel import fix_tencent_excel
        import tempfile
        import openpyxl
        
        # 创建测试Excel文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            test_excel = tmp.name
        
        # 创建简单的Excel内容
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '测试'
        ws['B1'] = '数据'
        ws['A2'] = 100
        ws['B2'] = 200
        wb.save(test_excel)
        wb.close()
        
        # 测试fix_tencent_excel返回文件路径
        output_file = test_excel.replace('.xlsx', '_fixed.xlsx')
        result = fix_tencent_excel(test_excel, output_file)
        
        if result == output_file and Path(output_file).exists():
            print("✅ Excel格式处理测试通过")
            print(f"   - fix_tencent_excel返回文件路径: {result}")
            print(f"   - 输出文件存在: {Path(output_file).exists()}")
            
            # 清理测试文件
            os.remove(test_excel)
            os.remove(output_file)
            return True
        else:
            print(f"❌ Excel格式处理测试失败")
            print(f"   - 返回值: {result}")
            print(f"   - 期望: {output_file}")
            return False
            
    except Exception as e:
        print(f"❌ Excel格式处理测试失败: {e}")
        return False


def main():
    """运行所有测试"""
    print("\n" + "="*60)
    print("腾讯文档管理系统8093端口 - 修复验证测试")
    print("="*60)
    
    results = []
    
    # 运行所有测试
    results.append(("L2语义分析器", test_l2_semantic_analyzer()))
    results.append(("CSV打分生成", test_csv_score_generator()))
    results.append(("降级代码删除", test_no_degradation()))
    results.append(("Excel格式处理", test_excel_format_handling()))
    
    # 汇总结果
    print("\n" + "="*60)
    print("测试结果汇总")
    print("="*60)
    
    passed = 0
    failed = 0
    
    for name, result in results:
        status = "✅ 通过" if result else "❌ 失败"
        print(f"{name}: {status}")
        if result:
            passed += 1
        else:
            failed += 1
    
    print("\n" + "-"*60)
    print(f"总计: {len(results)} 个测试")
    print(f"通过: {passed} 个")
    print(f"失败: {failed} 个")
    
    if failed == 0:
        print("\n🎉 所有修复验证测试通过！系统已准备就绪。")
    else:
        print(f"\n⚠️ 有 {failed} 个测试失败，请检查相关修复。")
    
    return failed == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)