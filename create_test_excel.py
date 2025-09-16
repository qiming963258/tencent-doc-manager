#!/usr/bin/env python3
"""
创建测试Excel文件
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import random

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "销售数据"

# 设置表头
headers = ['序号', '产品名称', '类别', '数量', '单价', '总价', '销售日期', '销售员', '状态', '备注']
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# 添加示例数据
products = ['笔记本电脑', 'iPhone 15', 'iPad Pro', '无线耳机', '智能手表', '键盘', '鼠标', '显示器', '打印机', '路由器']
categories = ['电子产品', '办公用品', '数码配件']
sales_persons = ['张三', '李四', '王五', '赵六']
statuses = ['已完成', '处理中', '待发货', '已取消']

data = []
for i in range(1, 21):  # 20行数据
    quantity = random.randint(1, 20)
    unit_price = random.randint(100, 5000)
    total_price = quantity * unit_price
    
    row = [
        i,  # 序号
        random.choice(products),  # 产品名称
        random.choice(categories),  # 类别
        quantity,  # 数量
        unit_price,  # 单价
        total_price,  # 总价
        f"2025-01-{random.randint(1, 31):02d}",  # 销售日期
        random.choice(sales_persons),  # 销售员
        random.choice(statuses),  # 状态
        '重要' if i % 5 == 0 else ''  # 备注
    ]
    data.append(row)
    ws.append(row)

# 调整列宽
column_widths = [8, 15, 12, 8, 10, 12, 12, 10, 10, 10]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = width

# 保存文件
file_path = '/root/projects/tencent-doc-manager/test_sales_data.xlsx'
wb.save(file_path)

print(f"✅ 测试Excel文件已创建: {file_path}")
print(f"📊 数据概况:")
print(f"   - 表格大小: {len(headers)}列 × {len(data)+1}行")
print(f"   - 包含产品: {', '.join(products[:5])}...")
print(f"   - 测试单元格示例:")
print(f"     • G5 - 第5行的销售日期")
print(f"     • B3 - 第3行的产品名称")
print(f"     • A1:J1 - 表头区域")
print(f"     • 第3行 - 完整的数据行")
print(f"     • A列 - 序号列")