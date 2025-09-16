#!/usr/bin/env python3
"""
DeepSeek Excel MCP 测试脚本
使用DeepSeek V3 API (通过硅基流动) 替代Claude进行Excel文件的智能修改
"""

import os
import sys
import json
import requests
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from typing import Dict, List, Any, Optional

# 配置
DEEPSEEK_API_KEY = os.getenv('DEEPSEEK_API_KEY')
API_BASE_URL = "https://api.siliconflow.cn/v1"

# 目录配置
BASE_DIR = Path('/root/projects/tencent-doc-manager')
DOWNLOAD_DIR = BASE_DIR / 'downloads'
OUTPUT_DIR = BASE_DIR / 'excel_outputs'
OUTPUT_DIR.mkdir(exist_ok=True)

class DeepSeekExcelProcessor:
    """使用DeepSeek处理Excel文件的类"""
    
    def __init__(self, api_key: str = None):
        self.api_key = api_key or DEEPSEEK_API_KEY
        if not self.api_key:
            raise ValueError("❌ 未设置DEEPSEEK_API_KEY环境变量")
        
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        
        # Excel样式配置
        self.styles = {
            'high_risk': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            'medium_risk': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
            'low_risk': PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'header_font': Font(bold=True, size=12),
            'risk_font': Font(bold=True, color="FFFFFF")
        }
    
    def analyze_with_deepseek(self, content: str, prompt: str = None) -> Dict[str, Any]:
        """使用DeepSeek V3分析内容"""
        
        if not prompt:
            prompt = """分析这个Excel数据的风险等级和变更合理性。
请返回JSON格式的分析结果，包含：
1. risk_level: 风险等级(high/medium/low)
2. changes: 具体的变更项列表
3. recommendations: 建议
4. confidence: 置信度(0-1)
5. approval_status: 审批建议(approve/review/reject)"""
        
        try:
            response = requests.post(
                f"{API_BASE_URL}/chat/completions",
                headers=self.headers,
                json={
                    "model": "deepseek-v3",
                    "messages": [
                        {"role": "system", "content": "你是一个专业的数据分析专家，擅长识别数据变更的风险。请用JSON格式返回分析结果。"},
                        {"role": "user", "content": f"{prompt}\n\n数据内容：\n{content[:2000]}"}  # 限制内容长度
                    ],
                    "temperature": 0.7,
                    "max_tokens": 1000,
                    "response_format": {"type": "json_object"}
                },
                timeout=30
            )
            
            if response.status_code == 200:
                result = response.json()
                content = result['choices'][0]['message']['content']
                try:
                    return json.loads(content)
                except json.JSONDecodeError:
                    # 如果返回的不是JSON，创建默认结构
                    return {
                        "risk_level": "medium",
                        "analysis": content,
                        "confidence": 0.7,
                        "approval_status": "review"
                    }
            else:
                print(f"❌ DeepSeek API错误: {response.status_code}")
                return self._default_analysis()
                
        except Exception as e:
            print(f"❌ 调用DeepSeek API失败: {e}")
            return self._default_analysis()
    
    def _default_analysis(self) -> Dict[str, Any]:
        """默认分析结果"""
        return {
            "risk_level": "medium",
            "changes": ["无法获取AI分析"],
            "recommendations": "建议人工审核",
            "confidence": 0.5,
            "approval_status": "review"
        }
    
    def process_excel_file(self, input_file: Path, output_file: Path = None) -> Path:
        """处理Excel文件，添加AI分析和风险标记"""
        
        print(f"📊 开始处理Excel文件: {input_file}")
        
        # 如果没有指定输出文件，自动生成
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = OUTPUT_DIR / f"deepseek_marked_{timestamp}_{input_file.name}"
        
        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
            
            # 获取数据内容用于分析
            data_rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell for cell in row):  # 跳过空行
                    data_rows.append(row)
            
            # 转换为字符串进行分析
            data_str = "\n".join([",".join(str(cell) if cell else "" for cell in row) for row in data_rows[:50]])
            
            print("🤖 调用DeepSeek V3进行风险分析...")
            analysis = self.analyze_with_deepseek(data_str)
            
            # 添加分析结果到Excel
            self._apply_risk_marking(ws, analysis)
            
            # 添加AI分析汇总sheet
            self._add_analysis_sheet(wb, analysis)
            
            # 保存文件
            wb.save(output_file)
            print(f"✅ 文件处理完成: {output_file}")
            
            return output_file
            
        except Exception as e:
            print(f"❌ 处理Excel文件失败: {e}")
            raise
    
    def _apply_risk_marking(self, ws, analysis: Dict[str, Any]):
        """应用风险标记到工作表"""
        
        risk_level = analysis.get('risk_level', 'medium')
        fill_style = self.styles.get(f'{risk_level}_risk')
        
        # 添加风险标记列
        max_col = ws.max_column
        ws.cell(row=1, column=max_col + 1, value="风险等级").font = self.styles['header_font']
        ws.cell(row=1, column=max_col + 2, value="AI分析").font = self.styles['header_font']
        
        # 为每行添加风险标记
        for row_idx in range(2, min(ws.max_row + 1, 100)):  # 最多处理100行
            risk_cell = ws.cell(row=row_idx, column=max_col + 1)
            risk_cell.value = risk_level.upper()
            risk_cell.fill = fill_style
            risk_cell.font = self.styles['risk_font']
            risk_cell.border = self.styles['border']
            
            # 添加AI分析批注
            analysis_cell = ws.cell(row=row_idx, column=max_col + 2)
            analysis_cell.value = f"置信度: {analysis.get('confidence', 0):.0%}"
            
            # 添加详细批注
            comment_text = f"DeepSeek分析结果:\n"
            comment_text += f"审批建议: {analysis.get('approval_status', 'review')}\n"
            if 'recommendations' in analysis:
                comment_text += f"建议: {analysis['recommendations']}"
            analysis_cell.comment = Comment(comment_text, "DeepSeek V3")
            
    def _add_analysis_sheet(self, wb, analysis: Dict[str, Any]):
        """添加AI分析汇总sheet"""
        
        # 创建新的sheet
        if "AI分析报告" in wb.sheetnames:
            del wb["AI分析报告"]
        
        ws_analysis = wb.create_sheet("AI分析报告", 0)
        
        # 添加标题
        ws_analysis.merge_cells('A1:D1')
        title_cell = ws_analysis['A1']
        title_cell.value = "DeepSeek V3 智能分析报告"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # 添加分析时间
        ws_analysis['A3'] = "分析时间:"
        ws_analysis['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 添加分析结果
        row = 5
        ws_analysis[f'A{row}'] = "风险等级:"
        ws_analysis[f'B{row}'] = analysis.get('risk_level', 'N/A').upper()
        
        row += 1
        ws_analysis[f'A{row}'] = "置信度:"
        ws_analysis[f'B{row}'] = f"{analysis.get('confidence', 0):.0%}"
        
        row += 1
        ws_analysis[f'A{row}'] = "审批建议:"
        ws_analysis[f'B{row}'] = analysis.get('approval_status', 'review')
        
        # 添加详细变更
        if 'changes' in analysis and isinstance(analysis['changes'], list):
            row += 2
            ws_analysis[f'A{row}'] = "识别的变更:"
            for idx, change in enumerate(analysis['changes'][:10], 1):
                row += 1
                ws_analysis[f'B{row}'] = f"{idx}. {change}"
        
        # 添加建议
        if 'recommendations' in analysis:
            row += 2
            ws_analysis[f'A{row}'] = "处理建议:"
            ws_analysis.merge_cells(f'B{row}:D{row}')
            ws_analysis[f'B{row}'] = str(analysis['recommendations'])
        
        # 设置列宽
        ws_analysis.column_dimensions['A'].width = 20
        ws_analysis.column_dimensions['B'].width = 50
        
        # 应用样式
        for row in ws_analysis.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = self.styles['border']

def test_deepseek_excel_processing():
    """测试DeepSeek Excel处理功能"""
    
    print("=" * 60)
    print("🚀 DeepSeek Excel MCP 测试系统")
    print("=" * 60)
    
    # 检查API密钥
    if not DEEPSEEK_API_KEY:
        print("❌ 请设置DEEPSEEK_API_KEY环境变量")
        print("export DEEPSEEK_API_KEY='your-api-key'")
        return
    
    processor = DeepSeekExcelProcessor()
    
    # 查找下载目录中的Excel文件
    excel_files = list(DOWNLOAD_DIR.glob("*.xlsx")) + list(DOWNLOAD_DIR.glob("*.xls"))
    
    if not excel_files:
        print("⚠️ 未找到Excel文件，创建测试文件...")
        # 创建测试Excel文件
        test_file = DOWNLOAD_DIR / "test_deepseek.xlsx"
        df = pd.DataFrame({
            '产品名称': ['产品A', '产品B', '产品C'],
            '销售数量': [100, 200, 150],
            '单价': [50, 75, 60],
            '风险等级': ['低', '中', '高']
        })
        df.to_excel(test_file, index=False)
        excel_files = [test_file]
    
    # 处理第一个找到的Excel文件
    input_file = excel_files[0]
    print(f"📄 处理文件: {input_file}")
    
    try:
        output_file = processor.process_excel_file(input_file)
        print(f"✅ 处理成功！输出文件: {output_file}")
        
        # 显示分析统计
        print("\n📊 处理统计:")
        print(f"  - 输入文件: {input_file.name}")
        print(f"  - 输出文件: {output_file.name}")
        print(f"  - 文件大小: {output_file.stat().st_size:,} bytes")
        print(f"  - AI模型: DeepSeek V3")
        print(f"  - API提供: 硅基流动")
        
    except Exception as e:
        print(f"❌ 处理失败: {e}")

def main():
    """主函数"""
    
    # 命令行参数处理
    if len(sys.argv) > 1:
        input_file = Path(sys.argv[1])
        if input_file.exists():
            processor = DeepSeekExcelProcessor()
            output_file = processor.process_excel_file(input_file)
            print(f"✅ 完成: {output_file}")
        else:
            print(f"❌ 文件不存在: {input_file}")
    else:
        # 运行测试
        test_deepseek_excel_processing()

if __name__ == "__main__":
    main()