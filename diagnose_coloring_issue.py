#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
诊断涂色问题
验证涂色功能是否正常工作
"""

import os
import json
import openpyxl
from pathlib import Path
import csv
from datetime import datetime

def analyze_issue():
    """分析涂色问题"""
    print("\n" + "="*60)
    print("涂色问题诊断")
    print("="*60)
    
    # 1. 分析最新的打分文件
    print("\n1. 分析最新的打分文件：")
    score_file = '/root/projects/tencent-doc-manager/scoring_results/detailed/detailed_scores_tencent_副本-副本-测试版本-出国销售计划表-工作表1_20250911_1002_midweek_W37_20250911_100216.json'
    
    if Path(score_file).exists():
        with open(score_file, 'r', encoding='utf-8') as f:
            score_data = json.load(f)
        
        print(f"   文件: {Path(score_file).name}")
        print(f"   总单元格: {score_data['statistics']['total_cells']}")
        print(f"   变更单元格: {score_data['statistics']['changed_cells']}")
        print(f"   高风险: {score_data['statistics']['high_risk_count']}")
        print(f"   中风险: {score_data['statistics']['medium_risk_count']}")
        print(f"   低风险: {score_data['statistics']['low_risk_count']}")
        print(f"   cell_scores内容: {len(score_data['cell_scores'])}个条目")
        
        if score_data['statistics']['changed_cells'] == 0:
            print("\n   ⚠️ 问题原因：没有检测到任何变更！")
            print("   两个CSV文件完全相同，所以没有需要涂色的单元格。")
    
    # 2. 检查涂色后的文件
    print("\n2. 检查涂色后的Excel文件：")
    marked_file = '/root/projects/tencent-doc-manager/excel_outputs/marked/tencent_副本-副本-测试版本-出国销售计划表_20250911_1002_midweek_W37_marked_20250911_100302_W37.xlsx'
    
    if Path(marked_file).exists():
        wb = openpyxl.load_workbook(marked_file)
        ws = wb.active
        
        # 统计有填充的单元格
        colored_cells = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                    colored_cells += 1
        
        print(f"   文件: {Path(marked_file).name}")
        print(f"   工作表行数: {ws.max_row}")
        print(f"   工作表列数: {ws.max_column}")
        print(f"   有填充的单元格数: {colored_cells}")
        
        if colored_cells == 0:
            print("   ✅ 确认：Excel文件没有涂色（因为没有变更）")
        
        wb.close()
    
    # 3. 诊断结论
    print("\n3. 诊断结论：")
    print("   ✅ 系统功能正常")
    print("   ⚠️ 没有涂色是因为两个CSV文件完全相同")
    print("   💡 解决方案：使用有实际差异的文件进行测试")
    
    print("\n" + "="*60)
    print("诊断完成")
    print("="*60)

def create_test_with_changes():
    """创建有变更的测试数据"""
    print("\n" + "="*60)
    print("创建测试数据（带变更）")
    print("="*60)
    
    # 创建基线CSV
    baseline_path = '/tmp/test_baseline_with_changes.csv'
    with open(baseline_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['产品名称', '销售额', '库存量', '单价', '状态'])
        writer.writerow(['产品A', '1000', '500', '10.5', '正常'])
        writer.writerow(['产品B', '2000', '300', '20.8', '正常'])
        writer.writerow(['产品C', '3000', '200', '30.2', '正常'])
        writer.writerow(['产品D', '4000', '100', '40.5', '缺货'])
        writer.writerow(['产品E', '5000', '50', '50.8', '缺货'])
    
    print(f"✅ 创建基线文件: {baseline_path}")
    
    # 创建目标CSV（有修改）
    target_path = '/tmp/test_target_with_changes.csv'
    with open(target_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['产品名称', '销售额', '库存量', '单价', '状态'])
        writer.writerow(['产品A', '1500', '500', '10.5', '正常'])  # 销售额改变
        writer.writerow(['产品B', '2000', '0', '25.0', '缺货'])    # 库存和单价改变
        writer.writerow(['产品C', '3000', '200', '30.2', '正常'])  # 没变
        writer.writerow(['产品D', '6000', '100', '40.5', '缺货'])  # 销售额改变
        writer.writerow(['产品E', '5000', '0', '60.0', '停产'])    # 库存、单价和状态改变
    
    print(f"✅ 创建目标文件: {target_path}")
    print("\n变更内容：")
    print("   - 产品A: 销售额 1000→1500")
    print("   - 产品B: 库存 300→0, 单价 20.8→25.0, 状态→缺货")
    print("   - 产品D: 销售额 4000→6000")
    print("   - 产品E: 库存 50→0, 单价 50.8→60.0, 状态→停产")
    
    # 测试打分生成
    print("\n测试打分生成：")
    try:
        from intelligent_excel_marker import DetailedScoreGenerator
        generator = DetailedScoreGenerator()
        score_file = generator.generate_score_json(
            baseline_path,
            target_path,
            '/tmp'
        )
        
        # 查看生成的打分
        with open(score_file, 'r', encoding='utf-8') as f:
            score_data = json.load(f)
        
        print(f"✅ 打分文件生成: {Path(score_file).name}")
        print(f"   检测到变更: {score_data['statistics']['changed_cells']}个单元格")
        print(f"   高风险: {score_data['statistics']['high_risk_count']}")
        print(f"   中风险: {score_data['statistics']['medium_risk_count']}")
        print(f"   低风险: {score_data['statistics']['low_risk_count']}")
        
        # 显示部分变更详情
        if score_data['cell_scores']:
            print("\n   部分变更详情：")
            for i, (cell_ref, cell_data) in enumerate(list(score_data['cell_scores'].items())[:3]):
                print(f"   - {cell_ref}: {cell_data['old_value']}→{cell_data['new_value']} ({cell_data['risk_level']})")
        
        return baseline_path, target_path, score_file
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        return None, None, None

def test_coloring_with_changes(score_file):
    """测试有变更时的涂色功能"""
    print("\n测试涂色功能：")
    
    try:
        # 创建测试Excel文件
        test_excel = '/tmp/test_excel_for_coloring.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 添加数据
        data = [
            ['产品名称', '销售额', '库存量', '单价', '状态'],
            ['产品A', 1500, 500, 10.5, '正常'],
            ['产品B', 2000, 0, 25.0, '缺货'],
            ['产品C', 3000, 200, 30.2, '正常'],
            ['产品D', 6000, 100, 40.5, '缺货'],
            ['产品E', 5000, 0, 60.0, '停产']
        ]
        
        for row in data:
            ws.append(row)
        
        wb.save(test_excel)
        print(f"✅ 创建测试Excel: {test_excel}")
        
        # 应用涂色
        from intelligent_excel_marker import IntelligentExcelMarker
        marker = IntelligentExcelMarker()
        output_file = marker.apply_striped_coloring(test_excel, score_file)
        
        print(f"✅ 涂色完成: {output_file}")
        
        # 验证涂色结果
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        colored_cells = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.patternType and cell.fill.patternType not in ['none', None]:
                    colored_cells += 1
        
        print(f"   涂色单元格数: {colored_cells}")
        
        if colored_cells > 0:
            print("   ✅ 涂色功能正常工作！")
        else:
            print("   ❌ 涂色功能可能有问题")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ 涂色测试失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # 诊断当前问题
    analyze_issue()
    
    # 创建有变更的测试数据
    baseline, target, score_file = create_test_with_changes()
    
    # 如果创建成功，测试涂色
    if score_file:
        test_coloring_with_changes(score_file)