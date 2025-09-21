#!/usr/bin/env python3
"""
XLSX兼容性深度测试
测试不同方法处理腾讯文档XLSX格式
"""

import asyncio
import json
from pathlib import Path
from datetime import datetime
import logging
import traceback

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class XLSXCompatibilityTest:
    """测试XLSX格式兼容性"""

    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.test_dir = Path(f"/root/projects/tencent-doc-manager/xlsx_test_{self.timestamp}")
        self.test_dir.mkdir(parents=True, exist_ok=True)

        # 测试URL
        self.doc_url = "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN"

    async def download_xlsx(self) -> str:
        """下载XLSX格式文件"""
        logger.info("="*60)
        logger.info("📥 下载XLSX格式文件")
        logger.info("="*60)

        from production.core_modules.playwright_downloader import PlaywrightDownloader

        # 读取Cookie
        cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
        with open(cookie_file) as f:
            cookie_data = json.load(f)

        cookie_string = cookie_data.get('current_cookies', '')

        # 下载XLSX
        downloader = PlaywrightDownloader()
        result = await downloader.download(
            url=self.doc_url,
            cookies=cookie_string,
            format='xlsx',  # 关键：请求XLSX格式
            download_dir=str(self.test_dir)
        )

        if result.get('success'):
            xlsx_file = result.get('file_path')
            logger.info(f"✅ 下载成功: {xlsx_file}")
            size = Path(xlsx_file).stat().st_size / 1024
            logger.info(f"📊 文件大小: {size:.2f} KB")
            return xlsx_file
        else:
            logger.error(f"❌ 下载失败")
            return None

    def test_openpyxl_standard(self, xlsx_file: str):
        """测试标准openpyxl读取"""
        logger.info("\n🔬 测试1: 标准openpyxl读取")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            logger.info(f"✅ 成功读取，工作表: {ws.title}")
            logger.info(f"   行数: {ws.max_row}, 列数: {ws.max_column}")
            return True
        except Exception as e:
            logger.error(f"❌ 读取失败: {type(e).__name__}: {str(e)[:100]}")
            return False

    def test_openpyxl_readonly(self, xlsx_file: str):
        """测试只读模式"""
        logger.info("\n🔬 测试2: openpyxl只读模式")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_file, read_only=True, data_only=True)
            ws = wb.active

            # 读取前几行数据
            row_count = 0
            for row in ws.iter_rows(max_row=5, values_only=True):
                row_count += 1
                logger.info(f"   行{row_count}: {str(row[:3])[:50]}...")

            logger.info(f"✅ 只读模式成功")
            return True
        except Exception as e:
            logger.error(f"❌ 只读模式失败: {str(e)[:100]}")
            return False

    def test_pandas_read(self, xlsx_file: str):
        """测试pandas读取"""
        logger.info("\n🔬 测试3: pandas读取")
        try:
            import pandas as pd
            df = pd.read_excel(xlsx_file, engine='openpyxl')
            logger.info(f"✅ pandas读取成功")
            logger.info(f"   数据形状: {df.shape}")
            logger.info(f"   列名: {list(df.columns[:5])}")
            return df
        except Exception as e:
            logger.error(f"❌ pandas读取失败: {str(e)[:100]}")
            return None

    def test_xlrd_read(self, xlsx_file: str):
        """测试xlrd库读取"""
        logger.info("\n🔬 测试4: xlrd读取（如果已安装）")
        try:
            import xlrd
            book = xlrd.open_workbook(xlsx_file)
            sheet = book.sheet_by_index(0)
            logger.info(f"✅ xlrd读取成功")
            logger.info(f"   行数: {sheet.nrows}, 列数: {sheet.ncols}")
            return True
        except ImportError:
            logger.info("⚠️ xlrd未安装，跳过测试")
            return None
        except Exception as e:
            logger.error(f"❌ xlrd读取失败: {str(e)[:100]}")
            return False

    def test_repair_and_read(self, xlsx_file: str):
        """尝试修复并读取"""
        logger.info("\n🔬 测试5: 尝试修复XLSX文件")

        try:
            # 使用pandas读取并重新保存
            import pandas as pd

            # 读取数据（忽略样式）
            df = pd.read_excel(xlsx_file, engine='openpyxl')

            # 保存为新文件
            repaired_file = self.test_dir / f"repaired_{self.timestamp}.xlsx"
            df.to_excel(repaired_file, index=False, engine='openpyxl')

            logger.info(f"✅ 文件修复成功: {repaired_file.name}")

            # 测试修复后的文件
            from openpyxl import load_workbook
            wb = load_workbook(repaired_file)
            ws = wb.active

            logger.info(f"✅ 修复后可以正常打开")
            logger.info(f"   工作表: {ws.title}")

            # 测试涂色
            from openpyxl.styles import PatternFill
            test_cell = ws['B2']
            test_cell.fill = PatternFill(
                start_color="FFCCCC",
                end_color="FFCCCC",
                fill_type="solid"
            )

            # 保存涂色后的文件
            colored_file = self.test_dir / f"colored_{self.timestamp}.xlsx"
            wb.save(colored_file)

            logger.info(f"✅ 涂色测试成功: {colored_file.name}")
            size = colored_file.stat().st_size / 1024
            logger.info(f"   文件大小: {size:.2f} KB")

            return str(colored_file)

        except Exception as e:
            logger.error(f"❌ 修复失败: {str(e)}")
            traceback.print_exc()
            return None

    async def run_all_tests(self):
        """运行所有测试"""
        logger.info("🚀 开始XLSX兼容性测试")
        logger.info(f"📁 测试目录: {self.test_dir}")

        # 下载XLSX
        xlsx_file = await self.download_xlsx()
        if not xlsx_file:
            logger.error("下载失败，无法继续测试")
            return

        # 运行各项测试
        results = {
            'standard_openpyxl': self.test_openpyxl_standard(xlsx_file),
            'readonly_openpyxl': self.test_openpyxl_readonly(xlsx_file),
            'pandas': self.test_pandas_read(xlsx_file) is not None,
            'xlrd': self.test_xlrd_read(xlsx_file),
            'repair': self.test_repair_and_read(xlsx_file) is not None
        }

        # 总结
        logger.info("\n" + "="*60)
        logger.info("📊 测试总结")
        logger.info("="*60)

        for test_name, result in results.items():
            status = "✅" if result else "❌" if result is False else "⚠️"
            logger.info(f"{status} {test_name}: {result}")

        # 找出可行方案
        if results['repair']:
            logger.info("\n💡 可行方案: pandas读取 → 重新保存 → openpyxl处理")
            logger.info("   这种方式可以绕过腾讯XLSX的兼容性问题")
        elif results['pandas']:
            logger.info("\n💡 可行方案: 使用pandas处理数据，但可能无法保留格式")
        else:
            logger.info("\n⚠️ 所有方法都失败，建议继续使用CSV方案")

        return results


async def main():
    """主函数"""
    test = XLSXCompatibilityTest()
    results = await test.run_all_tests()

    if results and results.get('repair'):
        logger.info("\n🎉 找到了XLSX兼容性解决方案！")
    else:
        logger.info("\n⚠️ XLSX直接处理仍有挑战，CSV方案仍是最稳定的")


if __name__ == "__main__":
    asyncio.run(main())