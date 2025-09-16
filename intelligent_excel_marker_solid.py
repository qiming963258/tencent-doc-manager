#!/usr/bin/env python3
"""
智能Excel标记器 - 纯色填充版本
兼容腾讯文档的涂色方案
"""

import json
import logging
import os
from datetime import datetime
from typing import Dict, Optional, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.comments import Comment

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class IntelligentExcelMarkerSolid:
    """
    智能Excel标记器 - 使用纯色填充以兼容腾讯文档
    """
    
    def __init__(self):
        """初始化标记器"""
        self.output_dir = "/root/projects/tencent-doc-manager/excel_outputs/marked"
        
        # 纯色填充映射（兼容腾讯文档）
        self.color_mapping = {
            'high': {
                'color': 'FFCCCC',  # 浅红色
                'font_color': 'CC0000'  # 深红色字体
            },
            'medium': {
                'color': 'FFE5CC',  # 浅橙色
                'font_color': 'FF6600'  # 橙色字体
            },
            'low': {
                'color': 'FFFFCC',  # 浅黄色
                'font_color': 'CC9900'  # 深黄色字体
            }
        }
        
        # 风险等级阈值
        self.risk_thresholds = {
            'high': 0.6,    # >= 0.6 高风险
            'medium': 0.3,  # >= 0.3 中风险
            'low': 0.0      # >= 0.0 低风险
        }
    
    def get_risk_level(self, score: float) -> str:
        """根据分数获取风险等级"""
        if score >= self.risk_thresholds['high']:
            return 'high'
        elif score >= self.risk_thresholds['medium']:
            return 'medium'
        else:
            return 'low'
    
    def apply_solid_coloring(self, excel_file: str, score_file: str, output_file: str = None) -> str:
        """
        应用纯色涂色到Excel文件（兼容腾讯文档）
        
        Args:
            excel_file: 要涂色的Excel文件
            score_file: 详细打分JSON文件
            output_file: 输出文件路径（可选）
            
        Returns:
            输出文件路径
        """
        logger.info(f"开始应用纯色涂色（腾讯文档兼容模式）")
        logger.info(f"  Excel文件: {excel_file}")
        logger.info(f"  打分文件: {score_file}")
        
        # 加载打分数据
        with open(score_file, 'r', encoding='utf-8') as f:
            score_data = json.load(f)
        
        # 加载Excel文件
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 统计涂色情况
        color_stats = {"high": 0, "medium": 0, "low": 0}
        
        # 遍历所有变更的单元格
        for cell_ref, cell_data in score_data.get('cell_scores', {}).items():
            try:
                # 获取单元格
                cell = ws[cell_ref]
                
                # 获取风险等级和分数
                score = cell_data.get('score', 0)
                risk_level = self.get_risk_level(score)
                
                # 获取颜色配置
                colors = self.color_mapping[risk_level]
                
                # 创建纯色填充（腾讯文档支持）
                fill = PatternFill(
                    patternType='solid',  # 纯色填充
                    fgColor=colors['color'],
                    bgColor=colors['color']
                )
                
                # 应用填充
                cell.fill = fill
                
                # 设置字体颜色以增强对比度
                cell.font = Font(
                    color=colors['font_color'],
                    bold=True if risk_level == 'high' else False
                )
                
                # 添加边框以突出显示
                if risk_level in ['high', 'medium']:
                    border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    cell.border = border
                
                # 添加批注说明
                if risk_level == 'high':
                    comment_text = f"高风险变更\n分数: {score:.2f}\n原值: {cell_data.get('old_value', 'N/A')}\n新值: {cell_data.get('new_value', 'N/A')}"
                    cell.comment = Comment(comment_text, "风险监控系统")
                
                color_stats[risk_level] += 1
                
            except Exception as e:
                logger.warning(f"处理单元格 {cell_ref} 时出错: {e}")
        
        # 生成输出文件名
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            week_num = score_data.get('metadata', {}).get('week_number', '00')
            base_name = os.path.basename(excel_file).replace('.xlsx', '')
            output_file = os.path.join(
                self.output_dir, 
                f"{base_name}_solid_marked_{timestamp}_W{week_num}.xlsx"
            )
        
        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        # 保存文件
        wb.save(output_file)
        wb.close()
        
        # 输出统计信息
        logger.info(f"✓ 涂色完成（纯色模式）！")
        logger.info(f"  输出文件: {output_file}")
        logger.info(f"  涂色统计:")
        logger.info(f"    - 高风险（红色）: {color_stats['high']} 个单元格")
        logger.info(f"    - 中风险（橙色）: {color_stats['medium']} 个单元格")
        logger.info(f"    - 低风险（黄色）: {color_stats['low']} 个单元格")
        
        return output_file
    
    def create_legend_sheet(self, wb):
        """创建图例说明工作表"""
        if "图例说明" in wb.sheetnames:
            ws = wb["图例说明"]
        else:
            ws = wb.create_sheet("图例说明", 0)
        
        # 标题
        ws['A1'] = "风险等级涂色说明"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 图例
        legends = [
            ('A3', '高风险', 'FFCCCC', 'CC0000', '分数 >= 0.6'),
            ('A4', '中风险', 'FFE5CC', 'FF6600', '0.3 <= 分数 < 0.6'),
            ('A5', '低风险', 'FFFFCC', 'CC9900', '分数 < 0.3')
        ]
        
        for cell_ref, label, bg_color, font_color, description in legends:
            cell = ws[cell_ref]
            cell.value = label
            cell.fill = PatternFill(patternType='solid', fgColor=bg_color)
            cell.font = Font(color=font_color, bold=True)
            
            # 添加说明
            ws[f'B{cell.row}'] = description
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        
        return ws


def test_solid_coloring():
    """测试纯色涂色功能"""
    import json
    
    # 创建测试打分数据
    test_score_data = {
        "metadata": {
            "week_number": "37",
            "timestamp": datetime.now().isoformat()
        },
        "cell_scores": {
            "A1": {"score": 0.8, "old_value": "高", "new_value": "低"},
            "B2": {"score": 0.5, "old_value": "100", "new_value": "50"},
            "C3": {"score": 0.2, "old_value": "完成", "new_value": "进行中"}
        }
    }
    
    # 保存测试打分文件
    test_score_file = "/root/projects/tencent-doc-manager/test_score_solid.json"
    with open(test_score_file, 'w', encoding='utf-8') as f:
        json.dump(test_score_data, f, ensure_ascii=False, indent=2)
    
    # 创建测试Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '测试高风险'
    ws['B2'] = '测试中风险'
    ws['C3'] = '测试低风险'
    
    test_excel_file = "/root/projects/tencent-doc-manager/test_excel_solid.xlsx"
    wb.save(test_excel_file)
    
    # 应用纯色涂色
    marker = IntelligentExcelMarkerSolid()
    output_file = marker.apply_solid_coloring(test_excel_file, test_score_file)
    
    print(f"\n测试完成！")
    print(f"涂色文件: {output_file}")
    print("\n这个文件使用纯色填充，完全兼容腾讯文档！")
    
    return output_file


if __name__ == "__main__":
    # 运行测试
    test_solid_coloring()