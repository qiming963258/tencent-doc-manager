#!/usr/bin/env python3
"""
修复腾讯文档导出的Excel文件，使其能被openpyxl正确打开
问题：styles.xml中包含空的<fill/>标签，违反OOXML标准
"""

import zipfile
import os
import shutil
from io import BytesIO

def fix_tencent_excel(input_file, output_file=None):
    """
    修复腾讯文档导出的Excel文件
    
    Args:
        input_file: 输入的Excel文件路径
        output_file: 输出的Excel文件路径（如果为None，则生成_fixed后缀的文件）
    
    Returns:
        bool: 修复是否成功
    """
    if output_file is None:
        output_file = input_file.replace('.xlsx', '_fixed.xlsx')
    
    print(f"正在修复: {input_file}")
    
    try:
        # 创建临时文件
        temp_file = input_file + '.tmp'
        
        # 读取原始文件
        with zipfile.ZipFile(input_file, 'r') as zip_in:
            # 创建新的ZIP文件
            with zipfile.ZipFile(temp_file, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                # 遍历所有文件
                for item in zip_in.namelist():
                    # 读取文件内容
                    data = zip_in.read(item)
                    
                    # 如果是styles.xml，修复它
                    if item == 'xl/styles.xml':
                        # 转换为字符串
                        content = data.decode('utf-8')
                        
                        # 修复空的fill标签
                        content = content.replace('<fill/>', 
                            '<fill><patternFill patternType="none"/></fill>')
                        
                        # 转换回字节
                        data = content.encode('utf-8')
                        print("  ✓ 已修复styles.xml中的空fill标签")
                    
                    # 写入到新文件
                    zip_out.writestr(item, data)
        
        # 替换原文件
        shutil.move(temp_file, output_file)
        print(f"  ✓ 修复完成: {output_file}")
        # 返回修复后的文件路径而不是布尔值
        return output_file
        
    except Exception as e:
        print(f"  ✗ 修复失败: {e}")
        # 清理临时文件
        if os.path.exists(temp_file):
            os.remove(temp_file)
        # 失败时返回None
        return None

def test_with_openpyxl(file_path):
    """测试文件是否能被openpyxl打开"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
        
        # 尝试打开文件
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"\n✅ 文件可以被openpyxl打开:")
        print(f"  - 工作表名: {ws.title}")
        print(f"  - 数据范围: {ws.max_row}行 x {ws.max_column}列")
        
        # 测试涂色功能
        yellow_fill = PatternFill(start_color='FFFF00', 
                                 end_color='FFFF00', 
                                 fill_type='solid')
        ws['A1'].fill = yellow_fill
        print(f"  - 涂色测试: 成功（A1单元格设为黄色）")
        
        # 保存测试（不实际保存）
        temp_save = BytesIO()
        wb.save(temp_save)
        print(f"  - 保存测试: 成功")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"\n❌ 文件无法被openpyxl打开: {e}")
        return False

if __name__ == "__main__":
    # 修复最近的腾讯文档Excel文件
    test_files = [
        "downloads/副本-副本-测试版本-出国销售计划表.xlsx",
        "downloads/test_excel_file.xlsx"
    ]
    
    for file_path in test_files:
        if os.path.exists(file_path):
            print(f"\n{'='*50}")
            print(f"处理文件: {file_path}")
            print('='*50)
            
            # 测试原始文件
            print("\n1. 测试原始文件:")
            test_with_openpyxl(file_path)
            
            # 修复文件
            print("\n2. 修复文件:")
            output_file = file_path.replace('.xlsx', '_fixed.xlsx')
            if fix_tencent_excel(file_path, output_file):
                # 测试修复后的文件
                print("\n3. 测试修复后的文件:")
                test_with_openpyxl(output_file)
