#!/usr/bin/env python3
"""
自然语言Excel处理器 - DeepSeek + openpyxl结合
支持自然语言指令直接修改Excel文件
"""

import os
import json
import re
import asyncio
import requests
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DeepSeekNLPProcessor:
    """DeepSeek自然语言处理器"""
    
    def __init__(self):
        # 使用项目中的共享API密钥
        self.api_key = "sk-onzowexyblsrgltveejlnajoutrjqwbrqkwzcccskwmzonvb"
        self.base_url = "https://api.siliconflow.cn/v1"
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        
        # 颜色映射字典
        self.color_map = {
            '红': 'FF0000', '红色': 'FF0000', 'red': 'FF0000',
            '绿': '00FF00', '绿色': '00FF00', 'green': '00FF00',
            '蓝': '0000FF', '蓝色': '0000FF', 'blue': '0000FF',
            '黄': 'FFFF00', '黄色': 'FFFF00', 'yellow': 'FFFF00',
            '橙': 'FFA500', '橙色': 'FFA500', 'orange': 'FFA500',
            '紫': '800080', '紫色': '800080', 'purple': '800080',
            '灰': '808080', '灰色': '808080', 'gray': '808080', 'grey': '808080',
            '黑': '000000', '黑色': '000000', 'black': '000000',
            '白': 'FFFFFF', '白色': 'FFFFFF', 'white': 'FFFFFF',
            '粉': 'FFC0CB', '粉色': 'FFC0CB', 'pink': 'FFC0CB',
            '青': '00FFFF', '青色': '00FFFF', 'cyan': '00FFFF',
            '棕': 'A52A2A', '棕色': 'A52A2A', 'brown': 'A52A2A',
        }
    
    def parse_natural_language(self, instruction: str) -> Dict:
        """
        解析自然语言指令
        
        例如：
        - "将G5涂成红色" -> {'action': 'fill', 'cell': 'G5', 'color': 'FF0000'}
        - "把第3行标记为黄色" -> {'action': 'fill_row', 'row': 3, 'color': 'FFFF00'}
        - "A列到C列涂蓝色" -> {'action': 'fill_range', 'start': 'A', 'end': 'C', 'color': '0000FF'}
        """
        result = {'action': None, 'params': {}}
        
        # 清理指令
        instruction = instruction.strip().lower()
        
        # 模式1: 单个单元格涂色 (如 "G5涂红色", "将B3涂成蓝色")
        cell_pattern = r'([a-z]+\d+)[^\w]*(?:涂|填|标记|改|变|设|染|着)(?:成|为|上)?[^\w]*(.+色|red|green|blue|yellow|orange|purple|gray|grey|black|white|pink|cyan|brown)'
        match = re.search(cell_pattern, instruction)
        if match:
            cell = match.group(1).upper()
            color_text = match.group(2)
            color_code = self._get_color_code(color_text)
            return {
                'action': 'fill_cell',
                'cell': cell,
                'color': color_code
            }
        
        # 模式2: 整行涂色 (如 "第3行涂红色", "把第5行标记为黄色")
        row_pattern = r'第?(\d+)行[^\w]*(?:涂|填|标记|改|变|设|染|着)(?:成|为|上)?[^\w]*(.+色|red|green|blue|yellow)'
        match = re.search(row_pattern, instruction)
        if match:
            row = int(match.group(1))
            color_text = match.group(2)
            color_code = self._get_color_code(color_text)
            return {
                'action': 'fill_row',
                'row': row,
                'color': color_code
            }
        
        # 模式3: 整列涂色 (如 "A列涂蓝色", "把C列标记为绿色")
        col_pattern = r'([a-z]+)列[^\w]*(?:涂|填|标记|改|变|设|染|着)(?:成|为|上)?[^\w]*(.+色|red|green|blue|yellow)'
        match = re.search(col_pattern, instruction)
        if match:
            col = match.group(1).upper()
            color_text = match.group(2)
            color_code = self._get_color_code(color_text)
            return {
                'action': 'fill_column',
                'column': col,
                'color': color_code
            }
        
        # 模式4: 范围涂色 (如 "A1到C3涂绿色", "B2:D5涂黄色")
        range_pattern1 = r'([a-z]+\d+)[^\w]*(?:到|至|~|－|-)[^\w]*([a-z]+\d+)[^\w]*(?:涂|填|标记|改|变|设|染|着)(?:成|为|上)?[^\w]*(.+色|red|green|blue)'
        range_pattern2 = r'([a-z]+\d+):([a-z]+\d+)[^\w]*(?:涂|填|标记|改|变|设|染|着)(?:成|为|上)?[^\w]*(.+色|red|green|blue)'
        
        for pattern in [range_pattern1, range_pattern2]:
            match = re.search(pattern, instruction)
            if match:
                start_cell = match.group(1).upper()
                end_cell = match.group(2).upper()
                color_text = match.group(3)
                color_code = self._get_color_code(color_text)
                return {
                    'action': 'fill_range',
                    'start_cell': start_cell,
                    'end_cell': end_cell,
                    'color': color_code
                }
        
        # 模式5: 总结表格 (如 "总结表格", "分析数据")
        if any(word in instruction for word in ['总结', '分析', '统计', '汇总', 'summary', 'analyze']):
            return {'action': 'summarize', 'params': {}}
        
        # 模式6: 添加批注 (如 "G5添加批注：重要数据")
        comment_pattern = r'([a-z]+\d+)[^\w]*(?:添加|加|写|备注)[^\w]*(?:批注|注释|备注|说明)[^\w]*[:：]?\s*(.+)'
        match = re.search(comment_pattern, instruction)
        if match:
            cell = match.group(1).upper()
            comment_text = match.group(2).strip()
            return {
                'action': 'add_comment',
                'cell': cell,
                'comment': comment_text
            }
        
        # 如果无法解析，调用DS API进行智能解析
        return self._call_ds_for_parsing(instruction)
    
    def _get_color_code(self, color_text: str) -> str:
        """获取颜色代码"""
        color_text = color_text.strip().lower()
        
        # 查找颜色映射
        for key, value in self.color_map.items():
            if key in color_text:
                return value
        
        # 默认红色
        return 'FF0000'
    
    def _call_ds_for_parsing(self, instruction: str) -> Dict:
        """调用DeepSeek API进行智能解析"""
        prompt = f"""
        请将以下自然语言指令转换为JSON格式的操作指令：
        
        指令：{instruction}
        
        返回格式示例：
        {{"action": "fill_cell", "cell": "G5", "color": "FF0000"}}
        {{"action": "fill_row", "row": 3, "color": "FFFF00"}}
        {{"action": "fill_column", "column": "A", "color": "0000FF"}}
        {{"action": "fill_range", "start_cell": "A1", "end_cell": "C3", "color": "00FF00"}}
        {{"action": "summarize"}}
        {{"action": "add_comment", "cell": "B2", "comment": "重要数据"}}
        
        只返回JSON，不要其他内容。
        """
        
        try:
            endpoint = f"{self.base_url}/chat/completions"
            payload = {
                "model": "deepseek-ai/DeepSeek-V3",
                "messages": [
                    {"role": "system", "content": "你是一个Excel操作指令解析器。"},
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 200,
                "temperature": 0.3
            }
            
            response = requests.post(endpoint, headers=self.headers, json=payload, timeout=10)
            if response.status_code == 200:
                result = response.json()
                content = result['choices'][0]['message']['content']
                # 提取JSON
                json_match = re.search(r'\{.*\}', content, re.DOTALL)
                if json_match:
                    return json.loads(json_match.group())
        except Exception as e:
            logger.error(f"DS API调用失败: {e}")
        
        return {'action': 'unknown', 'params': {}}
    
    async def analyze_table(self, data: List[List]) -> str:
        """使用DS分析表格数据"""
        prompt = f"""
        请分析以下Excel表格数据并提供总结：
        
        数据（前10行）：
        {json.dumps(data[:10], ensure_ascii=False, indent=2)}
        
        请提供：
        1. 数据概况（行数、列数）
        2. 数据特征
        3. 关键发现
        4. 建议
        
        保持简洁，200字以内。
        """
        
        try:
            endpoint = f"{self.base_url}/chat/completions"
            payload = {
                "model": "deepseek-ai/DeepSeek-V3",
                "messages": [
                    {"role": "system", "content": "你是一个数据分析专家。"},
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 500,
                "temperature": 0.7
            }
            
            response = requests.post(endpoint, headers=self.headers, json=payload, timeout=15)
            if response.status_code == 200:
                result = response.json()
                return result['choices'][0]['message']['content']
        except Exception as e:
            logger.error(f"表格分析失败: {e}")
        
        return "分析失败，请检查数据格式。"


class NaturalLanguageExcelProcessor:
    """自然语言Excel处理器主类"""
    
    def __init__(self):
        self.nlp = DeepSeekNLPProcessor()
        self.current_workbook = None
        self.current_worksheet = None
        self.file_path = None
    
    def load_excel(self, file_path: str) -> bool:
        """加载Excel文件"""
        try:
            self.file_path = file_path
            self.current_workbook = load_workbook(file_path)
            self.current_worksheet = self.current_workbook.active
            logger.info(f"✅ 成功加载Excel文件: {file_path}")
            return True
        except Exception as e:
            logger.error(f"❌ 加载Excel失败: {e}")
            return False
    
    async def process_instruction(self, instruction: str) -> Dict:
        """处理自然语言指令"""
        if not self.current_workbook:
            return {'success': False, 'message': '请先加载Excel文件'}
        
        # 解析指令
        parsed = self.nlp.parse_natural_language(instruction)
        logger.info(f"📝 解析结果: {parsed}")
        
        # 执行操作
        result = await self._execute_action(parsed)
        
        # 保存文件
        if result['success'] and parsed.get('action') != 'summarize':
            self._save_file()
        
        return result
    
    async def _execute_action(self, parsed: Dict) -> Dict:
        """执行解析后的操作"""
        action = parsed.get('action')
        ws = self.current_worksheet
        
        try:
            if action == 'fill_cell':
                # 单个单元格涂色
                cell = parsed['cell']
                color = parsed['color']
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws[cell].fill = fill
                return {'success': True, 'message': f'✅ 已将{cell}涂成颜色#{color}'}
            
            elif action == 'fill_row':
                # 整行涂色
                row = parsed['row']
                color = parsed['color']
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
                return {'success': True, 'message': f'✅ 已将第{row}行涂成颜色#{color}'}
            
            elif action == 'fill_column':
                # 整列涂色
                column = parsed['column']
                color = parsed['color']
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                col_idx = column_index_from_string(column)
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).fill = fill
                return {'success': True, 'message': f'✅ 已将{column}列涂成颜色#{color}'}
            
            elif action == 'fill_range':
                # 范围涂色
                start_cell = parsed['start_cell']
                end_cell = parsed['end_cell']
                color = parsed['color']
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                # 解析起止单元格
                start_col = column_index_from_string(re.match(r'([A-Z]+)', start_cell).group(1))
                start_row = int(re.match(r'[A-Z]+(\d+)', start_cell).group(1))
                end_col = column_index_from_string(re.match(r'([A-Z]+)', end_cell).group(1))
                end_row = int(re.match(r'[A-Z]+(\d+)', end_cell).group(1))
                
                # 涂色
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        ws.cell(row=row, column=col).fill = fill
                
                return {'success': True, 'message': f'✅ 已将{start_cell}到{end_cell}涂成颜色#{color}'}
            
            elif action == 'add_comment':
                # 添加批注
                cell = parsed['cell']
                comment_text = parsed['comment']
                ws[cell].comment = Comment(comment_text, "DeepSeek AI")
                return {'success': True, 'message': f'✅ 已为{cell}添加批注: {comment_text}'}
            
            elif action == 'summarize':
                # 总结表格
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
                
                summary = await self.nlp.analyze_table(data)
                return {
                    'success': True,
                    'message': '📊 表格分析完成',
                    'summary': summary,
                    'stats': {
                        'rows': ws.max_row,
                        'columns': ws.max_column,
                        'cells': ws.max_row * ws.max_column
                    }
                }
            
            else:
                return {'success': False, 'message': f'❌ 未知操作: {action}'}
                
        except Exception as e:
            return {'success': False, 'message': f'❌ 执行失败: {str(e)}'}
    
    def _save_file(self):
        """保存文件"""
        try:
            output_path = self.file_path.replace('.xlsx', '_modified.xlsx')
            self.current_workbook.save(output_path)
            logger.info(f"💾 文件已保存: {output_path}")
        except Exception as e:
            logger.error(f"保存失败: {e}")
    
    def get_preview(self, max_rows: int = 10, max_cols: int = 10) -> List[List]:
        """获取表格预览数据"""
        if not self.current_worksheet:
            return []
        
        ws = self.current_worksheet
        preview = []
        
        for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                                min_col=1, max_col=min(max_cols, ws.max_column),
                                values_only=True):
            preview.append(list(row))
        
        return preview


# 测试函数
async def test_processor():
    """测试自然语言处理器"""
    
    # 创建测试Excel文件
    wb = Workbook()
    ws = wb.active
    
    # 添加测试数据
    headers = ['序号', '产品', '数量', '价格', '总额', '状态']
    ws.append(headers)
    
    data = [
        [1, '产品A', 10, 100, 1000, '正常'],
        [2, '产品B', 5, 200, 1000, '缺货'],
        [3, '产品C', 15, 150, 2250, '正常'],
        [4, '产品D', 8, 300, 2400, '预订'],
        [5, '产品E', 20, 50, 1000, '正常'],
    ]
    
    for row in data:
        ws.append(row)
    
    test_file = 'test_natural_language.xlsx'
    wb.save(test_file)
    
    # 测试处理器
    processor = NaturalLanguageExcelProcessor()
    processor.load_excel(test_file)
    
    # 测试各种指令
    test_instructions = [
        "将G5涂成红色",
        "把第3行标记为黄色",
        "A列涂蓝色",
        "B2到D4涂绿色",
        "G2添加批注：这是重要数据",
        "总结表格"
    ]
    
    for instruction in test_instructions:
        print(f"\n📝 指令: {instruction}")
        result = await processor.process_instruction(instruction)
        print(f"📤 结果: {result}")
    
    print("\n✅ 测试完成！")


if __name__ == "__main__":
    asyncio.run(test_processor())