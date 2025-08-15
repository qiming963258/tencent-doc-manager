#!/usr/bin/env python3
import openpyxl
import sys

def check_excel_half_fill(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print('📊 半填充Excel文件内容预览:')
        print('=' * 60)
        print(f'工作表尺寸: {ws.max_row}行 x {ws.max_column}列')
        print()
        
        # 显示所有内容
        for row in range(1, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                value = str(cell.value) if cell.value is not None else ''
                
                # 检查单元格样式
                has_fill = False
                try:
                    if cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                        has_fill = True
                except:
                    pass
                
                if has_fill:
                    value = f'🟨{value}'
                    
                row_data.append(value[:25] if value else '')
            
            print(f'行{row:2d}: {" | ".join(row_data)}')
            
    except Exception as e:
        print(f'❌ 检查失败: {e}')

if __name__ == '__main__':
    file_path = '/root/projects/tencent-doc-manager/uploads/half_filled_result_1755067386.xlsx'
    check_excel_half_fill(file_path)