#!/usr/bin/env python3
"""应用新的颜色配置到Excel文件"""

import sys
sys.path.append('/root/projects/tencent-doc-manager')

import openpyxl
from openpyxl.styles import PatternFill
import json
from datetime import datetime
import os

def apply_new_colors(excel_file, score_file, output_file=None):
    """应用新的更明显的颜色到Excel文件"""

    print("🎨 应用新的颜色配置")
    print("="*60)

    # 新的颜色配置（更明显的颜色）
    column_level_colors = {
        "L1": "FF6666",  # 明显红色
        "L2": "FFB366",  # 明显橙色
        "L3": "66FF66"   # 明显绿色
    }

    risk_level_colors = {
        "EXTREME_HIGH": "FF3333",  # 深红色
        "HIGH": "FF6666",          # 红色
        "MEDIUM": "FFB366",        # 橙色
        "LOW": "66FF66",           # 绿色
        "EXTREME_LOW": "00CC00"    # 深绿色
    }

    # 加载打分数据
    with open(score_file, 'r', encoding='utf-8') as f:
        score_data = json.load(f)

    print(f"📊 打分文件: {os.path.basename(score_file)}")
    print(f"  总变更: {score_data['metadata']['total_modifications']} 处")

    # 加载Excel文件（使用原始未涂色的文件）
    # 先下载一个干净的Excel文件
    original_excel = "/root/projects/tencent-doc-manager/downloads/111测试版本-出国销售计划表_fixed.xlsx"

    if not os.path.exists(original_excel):
        # 使用最新的Excel文件作为基础
        original_excel = excel_file
        print(f"  使用现有文件: {os.path.basename(original_excel)}")
    else:
        print(f"  使用干净文件: {os.path.basename(original_excel)}")

    wb = openpyxl.load_workbook(original_excel)
    ws = wb.active

    # 统计涂色
    color_stats = {
        "L1": 0, "L2": 0, "L3": 0,
        "HIGH": 0, "MEDIUM": 0, "LOW": 0
    }

    # 遍历所有变更并涂色
    for score in score_data['scores']:
        cell_ref = score.get('cell')
        if not cell_ref:
            continue

        try:
            cell = ws[cell_ref]

            # 优先使用列级别涂色
            column_level = score.get('column_level')
            if column_level and column_level in column_level_colors:
                color = column_level_colors[column_level]
                stat_key = column_level
                print(f"  涂色 {cell_ref}: {column_level} -> {color}")
            else:
                # 使用风险级别涂色
                risk_level = score.get('risk_assessment', {}).get('risk_level', 'LOW')
                color = risk_level_colors.get(risk_level, "CCCCCC")
                stat_key = risk_level
                print(f"  涂色 {cell_ref}: {risk_level} -> {color}")

            # 应用solid填充（腾讯文档兼容）
            fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )
            cell.fill = fill

            if stat_key in color_stats:
                color_stats[stat_key] += 1

        except Exception as e:
            print(f"  ⚠️ 无法涂色 {cell_ref}: {e}")

    # 生成输出文件名
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"/root/projects/tencent-doc-manager/excel_outputs/marked/new_colors_{timestamp}.xlsx"

    # 保存文件
    wb.save(output_file)
    wb.close()

    # 输出统计
    print(f"\n📊 涂色统计:")
    for key, count in color_stats.items():
        if count > 0:
            print(f"  {key}: {count} 个单元格")

    print(f"\n✅ 涂色完成: {output_file}")

    # 验证颜色
    verify_colors(output_file)

    return output_file

def verify_colors(excel_file):
    """验证涂色的颜色值"""
    print(f"\n🔍 验证涂色颜色:")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    color_counts = {}
    samples = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.patternType == 'solid':
                fg_color = cell.fill.fgColor
                if fg_color and fg_color.rgb:
                    color = fg_color.rgb
                    if color not in ['FFFFFFFF', '00000000']:
                        color_counts[color] = color_counts.get(color, 0) + 1
                        if len(samples) < 5:
                            samples.append((cell.coordinate, color))

    print(f"  使用的颜色:")
    for color, count in sorted(color_counts.items(), key=lambda x: x[1], reverse=True):
        # 判断颜色类型
        if color in ['FF6666', 'FF3333']:
            color_type = '红色'
        elif color in ['FFB366']:
            color_type = '橙色'
        elif color in ['66FF66', '00CC00']:
            color_type = '绿色'
        else:
            color_type = '其他'
        print(f"    {color} ({color_type}): {count} 个单元格")

    print(f"\n  颜色样本:")
    for cell, color in samples:
        print(f"    {cell}: {color}")

    wb.close()

if __name__ == "__main__":
    # 使用有实际变更的打分文件
    score_file = "/root/projects/tencent-doc-manager/scoring_results/detailed/detailed_score_tmpewdh0_p5_20250929_011949.json"
    excel_file = "/root/projects/tencent-doc-manager/downloads/111测试版本-出国销售计划表_fixed.xlsx"

    # 如果fixed文件不存在，使用marked文件
    if not os.path.exists(excel_file):
        excel_file = "/root/projects/tencent-doc-manager/excel_outputs/marked/tencent_副本-副本-测试版本-出国销售计划表_20250929_0121_midweek_W40_marked_20250929_012144_W00.xlsx"

    result = apply_new_colors(excel_file, score_file)
    if result:
        print(f"\n🎉 成功生成新涂色文件: {result}")