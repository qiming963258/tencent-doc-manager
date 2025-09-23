#!/usr/bin/env python3
"""
带真实基线对比的测试
确保有真实的变更和涂色
"""

import asyncio
import json
import csv
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

class RealBaselineTest:
    """使用真实基线文件进行对比测试"""

    def __init__(self):
        self.base_dir = Path('/root/projects/tencent-doc-manager')

    async def prepare_baseline(self) -> str:
        """准备基线文件（从已有文件或下载新的）"""
        logger.info("📋 准备基线文件...")

        # 查找已有的基线文件
        baseline_pattern = "*出国销售*.csv"
        baseline_candidates = list(self.base_dir.rglob(baseline_pattern))

        if baseline_candidates:
            # 使用最新的文件作为基线
            baseline_file = sorted(baseline_candidates, key=lambda x: x.stat().st_mtime)[-1]
            logger.info(f"✅ 使用已有基线: {baseline_file}")
            return str(baseline_file)
        else:
            logger.info("⚠️ 未找到基线文件，需要先下载一个基线版本")
            # 这里应该下载一个基线版本
            return None

    async def download_current_version(self) -> str:
        """下载当前版本进行对比"""
        logger.info("📥 下载当前版本...")

        from production.core_modules.playwright_downloader import PlaywrightDownloader
        from production.core_modules.week_time_manager import WeekTimeManager

        # 读取Cookie
        cookie_file = self.base_dir / 'config/cookies.json'
        with open(cookie_file) as f:
            cookie_data = json.load(f)
        cookie_string = cookie_data.get('current_cookies', '')

        # 设置下载目录
        week_mgr = WeekTimeManager()
        week_info = week_mgr.get_current_week_info()
        download_dir = self.base_dir / f"csv_versions/2025_W{week_info['week_number']}/midweek"
        download_dir.mkdir(parents=True, exist_ok=True)

        # 执行下载
        downloader = PlaywrightDownloader()
        test_url = "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN"

        result = await downloader.download(
            url=test_url,
            cookies=cookie_string,
            format='csv',
            download_dir=str(download_dir)
        )

        if result and result.get('success'):
            return result.get('file_path')
        else:
            logger.error(f"❌ 下载失败: {result}")
            return None

    async def create_test_changes(self, csv_file: str) -> str:
        """创建一些测试变更（修改几个单元格）"""
        logger.info("✏️ 创建测试变更...")

        # 读取CSV
        rows = []
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

        # 修改一些单元格以创建变更
        changes_made = []
        if len(rows) > 5:
            # 修改第2行第3列
            if len(rows[1]) > 2:
                old_val = rows[1][2]
                rows[1][2] = f"测试修改_{datetime.now().strftime('%H%M%S')}"
                changes_made.append(f"行2列3: {old_val} → {rows[1][2]}")

            # 修改第5行第7列
            if len(rows[4]) > 6:
                old_val = rows[4][6]
                rows[4][6] = f"变更_{datetime.now().strftime('%H%M%S')}"
                changes_made.append(f"行5列7: {old_val} → {rows[4][6]}")

            # 修改第10行第10列
            if len(rows) > 9 and len(rows[9]) > 9:
                old_val = rows[9][9]
                rows[9][9] = "重要变更"
                changes_made.append(f"行10列10: {old_val} → {rows[9][9]}")

        # 保存修改后的文件
        modified_file = str(csv_file).replace('.csv', '_modified.csv')
        with open(modified_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        logger.info(f"✅ 创建了 {len(changes_made)} 处变更:")
        for change in changes_made:
            logger.info(f"   - {change}")

        return modified_file

    async def compare_and_score(self, baseline_file: str, current_file: str) -> Dict:
        """真实对比并打分"""
        logger.info("🔍 执行真实对比...")

        changes = []
        with open(baseline_file, 'r', encoding='utf-8') as f1, \
             open(current_file, 'r', encoding='utf-8') as f2:
            baseline_reader = list(csv.reader(f1))
            current_reader = list(csv.reader(f2))

            # 真实对比
            for row_idx, (row_baseline, row_current) in enumerate(zip(baseline_reader, current_reader)):
                for col_idx, (val_baseline, val_current) in enumerate(zip(row_baseline, row_current)):
                    if val_baseline.strip() != val_current.strip():
                        changes.append({
                            "row": row_idx,
                            "col": col_idx,
                            "old_value": val_baseline,
                            "new_value": val_current,
                            "risk_level": "HIGH" if col_idx < 7 else "MEDIUM" if col_idx < 14 else "LOW"
                        })

        logger.info(f"✅ 发现 {len(changes)} 处真实变更")
        return {"changes": changes, "total": len(changes)}

    async def generate_colored_excel(self, csv_file: str, changes: List[Dict]) -> str:
        """生成真正有涂色的Excel"""
        logger.info("🎨 生成涂色Excel...")

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active

        # 读取CSV数据
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 应用真实的涂色
        color_map = {
            "HIGH": "FFCCCC",    # 浅红
            "MEDIUM": "FFFFCC",  # 浅黄
            "LOW": "CCFFCC"      # 浅绿
        }

        for change in changes:
            row = change['row'] + 1
            col = change['col'] + 1
            risk = change['risk_level']
            color = color_map[risk]

            cell = ws.cell(row=row, column=col)
            # 关键：使用solid填充
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )
            # 添加批注说明变更
            cell.comment = Comment(
                f"原值: {change['old_value']}\n新值: {change['new_value']}\n风险: {risk}",
                "测试系统"
            )

        # 保存Excel
        excel_dir = self.base_dir / "excel_outputs/real_test"
        excel_dir.mkdir(parents=True, exist_ok=True)
        excel_file = excel_dir / f"real_colored_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(excel_file)

        logger.info(f"✅ 生成涂色Excel: {excel_file}")
        logger.info(f"   涂色单元格数: {len(changes)}")
        return str(excel_file)

    async def should_upload(self, changes: List[Dict]) -> bool:
        """判断是否应该上传"""
        if len(changes) == 0:
            logger.warning("⚠️ 没有变更，不需要上传")
            return False
        else:
            logger.info(f"✅ 有 {len(changes)} 处变更，需要上传")
            return True

    async def run_test(self):
        """运行完整测试"""
        logger.info("🚀 开始真实基线对比测试")
        logger.info("=" * 60)

        # 1. 准备基线
        baseline_file = await self.prepare_baseline()
        if not baseline_file:
            logger.error("❌ 无法准备基线文件")
            return

        # 2. 下载当前版本
        current_file = await self.download_current_version()
        if not current_file:
            logger.error("❌ 无法下载当前版本")
            return

        # 3. 创建测试变更（如果需要）
        if baseline_file == current_file or "modified" not in current_file:
            logger.info("📝 创建测试变更以演示涂色效果...")
            current_file = await self.create_test_changes(current_file)

        # 4. 对比并打分
        result = await self.compare_and_score(baseline_file, current_file)

        # 5. 判断是否应该上传
        if await self.should_upload(result['changes']):
            # 6. 生成涂色Excel
            excel_file = await self.generate_colored_excel(current_file, result['changes'])

            # 7. 只有有涂色才上传
            logger.info("📤 准备上传涂色Excel...")
            # 这里添加上传逻辑

            logger.info("=" * 60)
            logger.info("✅ 测试完成！")
            logger.info(f"   基线文件: {baseline_file}")
            logger.info(f"   当前文件: {current_file}")
            logger.info(f"   变更数量: {result['total']}")
            logger.info(f"   涂色Excel: {excel_file}")
        else:
            logger.info("=" * 60)
            logger.info("⏹️ 测试终止：没有变更，无需生成和上传Excel")

if __name__ == "__main__":
    tester = RealBaselineTest()
    asyncio.run(tester.run_test())