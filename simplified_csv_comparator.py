#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化的CSV对比算法
只输出必要的信息：修改列的Excel列号、列名，以及修改的单元格内容
"""

import csv
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Set
import openpyxl

class SimplifiedCSVComparator:
    """简化的CSV对比器 - 只输出核心信息"""
    
    def get_column_letter(self, col_index: int) -> str:
        """将列索引转换为Excel列字母（A, B, C...AA, AB等）"""
        result = ""
        col_num = col_index + 1
        
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
            
        return result
    
    def get_cell_address(self, row: int, col: int) -> str:
        """获取单元格地址（如A1, B2, C3等）"""
        return f"{self.get_column_letter(col)}{row + 1}"

    def _read_file(self, file_path: str) -> List[List[str]]:
        """读取CSV或XLSX文件，返回二维数组"""
        file_path = Path(file_path)

        # 根据文件扩展名判断格式
        if file_path.suffix.lower() == '.xlsx':
            # 读取XLSX文件
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True):
                    # 将None转换为空字符串，确保与CSV格式一致
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    data.append(row_data)
                workbook.close()
                return data
            except Exception as e:
                print(f"读取XLSX文件失败 {file_path}: {str(e)}")
                raise
        else:
            # 读取CSV文件
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return list(csv.reader(f))
            except UnicodeDecodeError:
                # 尝试其他编码
                with open(file_path, 'r', encoding='gbk') as f:
                    return list(csv.reader(f))
    
    def compare(self, baseline_path: str, target_path: str, 
                output_dir: str = None) -> Dict[str, Any]:
        """
        执行简化的CSV对比，只输出必要信息
        
        Returns:
            {
                "modified_columns": {
                    "C": "项目类型",
                    "D": "来源",
                    "E": "任务发起时间"
                },
                "modifications": [
                    {
                        "cell": "C4",
                        "column_name": "项目类型",
                        "old": "目标管理", 
                        "new": "固定计划"
                    }
                ],
                "statistics": {
                    "total_modifications": 63,
                    "similarity": 0.978
                }
            }
        """
        # 读取文件（支持CSV和XLSX）
        baseline_data = self._read_file(baseline_path)
        target_data = self._read_file(target_path)
        
        # 腾讯文档CSV格式：
        # 第0行：标题行（如"2025年项目计划与安排表"）
        # 第1行：实际的列名
        # 第2行及以后：数据行
        # 获取列名（使用第二行，索引1）
        column_names = baseline_data[1] if len(baseline_data) > 1 else []
        
        # 收集所有修改的列和单元格
        modified_columns = {}  # {列号: 列名}
        modifications = []  # 修改的单元格列表
        modified_column_indices = set()  # 用于去重的列索引集合
        
        # 从第3行开始比较数据（跳过标题行和列名行）
        start_row = 2  # 从索引2开始（第3行）
        max_rows = max(len(baseline_data), len(target_data))
        for row_idx in range(start_row, max_rows):
            baseline_row = baseline_data[row_idx] if row_idx < len(baseline_data) else []
            target_row = target_data[row_idx] if row_idx < len(target_data) else []
            
            max_cols = max(len(baseline_row), len(target_row))
            for col_idx in range(max_cols):
                baseline_value = str(baseline_row[col_idx]) if col_idx < len(baseline_row) else ''
                target_value = str(target_row[col_idx]) if col_idx < len(target_row) else ''
                
                # 如果单元格值不同
                if baseline_value.strip() != target_value.strip():
                    # 获取列信息
                    column_letter = self.get_column_letter(col_idx)
                    column_name = column_names[col_idx] if col_idx < len(column_names) else ''
                    
                    # 记录修改的列（用于去重汇总）
                    if col_idx not in modified_column_indices:
                        modified_column_indices.add(col_idx)
                        modified_columns[column_letter] = column_name
                    
                    # 记录修改的单元格（包含列名）
                    cell_address = self.get_cell_address(row_idx, col_idx)
                    modifications.append({
                        'cell': cell_address,
                        'column_name': column_name,  # 添加列名到每个修改块
                        'old': baseline_value,
                        'new': target_value
                    })
        
        # 计算相似度（排除标题行和列名行）
        data_rows_baseline = max(0, len(baseline_data) - 2)  # 减去标题行和列名行
        data_rows_target = max(0, len(target_data) - 2)
        
        total_cells_baseline = data_rows_baseline * (len(baseline_data[0]) if baseline_data else 0)
        total_cells_target = data_rows_target * (len(target_data[0]) if target_data else 0)
        max_cells = max(total_cells_baseline, total_cells_target)
        
        similarity = 1 - (len(modifications) / max_cells) if max_cells > 0 else 1.0
        
        # 构建简化的结果
        result = {
            'modified_columns': modified_columns,  # 去重的修改列及其Excel列号
            'modifications': modifications,  # 所有修改的单元格
            'statistics': {
                'total_modifications': len(modifications),
                'similarity': round(similarity, 3)
            }
        }
        
        # 如果指定了输出目录，保存简化的参数文件
        if output_dir:
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
            
            # 生成简化的文件名
            baseline_name = Path(baseline_path).stem.split('_')[1] if '_' in Path(baseline_path).stem else Path(baseline_path).stem
            target_name = Path(target_path).stem.split('_')[1] if '_' in Path(target_path).stem else Path(target_path).stem
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            param_filename = f"simplified_{baseline_name}_vs_{target_name}_{timestamp}.json"
            param_filepath = output_path / param_filename
            
            # 保存参数文件
            with open(param_filepath, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            
            print(f"✅ 简化参数文件已保存: {param_filepath}")
        
        return result


def test_simplified_comparator():
    """测试简化对比器"""
    comparator = SimplifiedCSVComparator()
    
    # 测试文件路径
    baseline = "/root/projects/tencent-doc-manager/csv_versions/2025_W36/midweek/tencent_副本-测试版本-出国销售计划表-工作表1_csv_20250905_1137_midweek_W36.csv"
    target = "/root/projects/tencent-doc-manager/csv_versions/2025_W36/midweek/tencent_副本-副本-测试版本-出国销售计划表-工作表1_csv_20250905_1137_midweek_W36.csv"
    output_dir = "/root/projects/tencent-doc-manager/comparison_results"
    
    from pathlib import Path
    if Path(baseline).exists() and Path(target).exists():
        result = comparator.compare(baseline, target, output_dir)
        
        print("\n" + "="*60)
        print("简化CSV对比结果")
        print("="*60)
        print(f"\n📊 统计信息:")
        print(f"  - 相似度: {result['statistics']['similarity'] * 100:.1f}%")
        print(f"  - 总修改数: {result['statistics']['total_modifications']}")
        
        print(f"\n📝 修改的列（去重）:")
        for col_letter, col_name in result['modified_columns'].items():
            print(f"  - {col_letter}: {col_name}")
        
        print(f"\n🔍 前10个修改示例:")
        for mod in result['modifications'][:10]:
            print(f"  - {mod['cell']}: '{mod['old']}' → '{mod['new']}'")
    else:
        print("❌ 测试文件不存在")


if __name__ == "__main__":
    test_simplified_comparator()