#!/usr/bin/env python3
"""
Excel涂色调试脚本
深度检查涂色问题的根本原因
"""

import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment


def check_colored_file():
    """检查涂色后的文件"""

    print("="*60)
    print("🔍 Excel涂色问题诊断")
    print("="*60)

    # 1. 检查涂色文件
    colored_file = "/root/projects/tencent-doc-manager/excel_outputs/marked/colored_WF_20250921_180701_95de839b.xlsx"

    if not Path(colored_file).exists():
        print("❌ 涂色文件不存在!")
        return

    print(f"\n1. 检查涂色文件: {colored_file}")
    print(f"   文件大小: {Path(colored_file).stat().st_size} 字节")

    # 2. 加载并检查文件内容
    wb = load_workbook(colored_file)
    ws = wb.active

    print(f"\n2. 工作表信息:")
    print(f"   工作表名: {ws.title}")
    print(f"   行数: {ws.max_row}")
    print(f"   列数: {ws.max_column}")

    # 3. 检查是否有填充样式
    print(f"\n3. 检查单元格填充:")

    has_fill = False
    fill_count = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.patternType:
                has_fill = True
                fill_count += 1
                print(f"   ✅ 发现填充: 单元格{cell.coordinate}")
                print(f"      填充类型: {cell.fill.patternType}")
                print(f"      前景色: {cell.fill.fgColor.rgb if cell.fill.fgColor else 'None'}")
                print(f"      背景色: {cell.fill.bgColor.rgb if cell.fill.bgColor else 'None'}")

                # 只显示前5个
                if fill_count >= 5:
                    break
        if fill_count >= 5:
            break

    if not has_fill:
        print("   ❌ 没有发现任何填充样式!")
    else:
        print(f"   总共发现 {fill_count} 个有填充的单元格")

    # 4. 检查批注
    print(f"\n4. 检查批注:")
    has_comment = False
    comment_count = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                has_comment = True
                comment_count += 1
                print(f"   ✅ 发现批注: 单元格{cell.coordinate}")
                print(f"      内容: {cell.comment.text[:50]}...")

                if comment_count >= 3:
                    break
        if comment_count >= 3:
            break

    if not has_comment:
        print("   ❌ 没有发现任何批注!")
    else:
        print(f"   总共发现 {comment_count} 个批注")


def test_correct_coloring():
    """测试正确的涂色方法"""

    print("\n" + "="*60)
    print("🎨 测试正确的涂色方法")
    print("="*60)

    # 创建测试文件
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "涂色测试"

    # 添加标题
    headers = ["测试列1", "测试列2", "测试列3", "测试列4", "测试列5"]
    ws.append(headers)

    # 添加数据和不同的涂色方式
    test_data = [
        ["数据1", "数据2", "数据3", "数据4", "数据5"],
        ["数据6", "数据7", "数据8", "数据9", "数据10"],
        ["数据11", "数据12", "数据13", "数据14", "数据15"],
    ]

    for row_data in test_data:
        ws.append(row_data)

    print("\n测试不同的填充方式:")

    # 方法1: solid填充（最可靠）
    fill1 = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws['A2'].fill = fill1
    print("✅ A2: solid红色填充")

    # 方法2: lightUp填充
    fill2 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="lightUp")
    ws['B2'].fill = fill2
    print("✅ B2: lightUp绿色填充")

    # 方法3: 带透明度的solid填充
    fill3 = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    ws['C2'].fill = fill3
    print("✅ C2: 带透明度的红色solid填充")

    # 方法4: 使用RGB值
    fill4 = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws['D2'].fill = fill4
    print("✅ D2: 浅红色solid填充")

    # 方法5: 使用内置颜色索引
    fill5 = PatternFill(patternType="solid", fgColor="FFFFCC00")
    ws['E2'].fill = fill5
    print("✅ E2: 黄色solid填充（使用fgColor）")

    # 添加批注测试
    from openpyxl.comments import Comment
    comment = Comment("这是一个测试批注\n包含多行内容", "测试用户")
    ws['A3'].comment = comment
    print("✅ A3: 添加批注")

    # 添加边框测试
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws['B3'].border = thin_border
    print("✅ B3: 添加边框")

    # 保存测试文件
    test_file = "/root/projects/tencent-doc-manager/excel_outputs/test_coloring.xlsx"
    wb.save(test_file)
    print(f"\n测试文件已保存: {test_file}")

    return test_file


def fix_coloring_method():
    """修复涂色方法，使用solid填充"""

    print("\n" + "="*60)
    print("🔧 使用solid填充修复涂色")
    print("="*60)

    # 读取分数文件
    scores_file = "/root/projects/tencent-doc-manager/scoring_results/detailed/scores_WF_20250921_180701_95de839b.json"
    scores_data = json.loads(Path(scores_file).read_text())

    # 读取原始Excel
    original_file = "/root/projects/tencent-doc-manager/excel_outputs/export_WF_20250921_180701_95de839b.xlsx"

    wb = load_workbook(original_file)
    ws = wb.active

    print(f"正在处理 {len(scores_data['cell_scores'])} 个单元格...")

    # 应用solid填充
    for cell_key, cell_info in scores_data["cell_scores"].items():
        row, col = map(int, cell_key.split("_"))
        cell = ws.cell(row=row, column=col)

        # 使用solid填充而不是lightUp
        if cell_info["score"] >= 70:
            # 高风险 - 红色
            fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            print(f"  🔴 {cell.coordinate}: 高风险 (分数={cell_info['score']})")
        elif cell_info["score"] >= 40:
            # 中风险 - 黄色
            fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            print(f"  🟡 {cell.coordinate}: 中风险 (分数={cell_info['score']})")
        else:
            # 低风险 - 绿色
            fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            print(f"  🟢 {cell.coordinate}: 低风险 (分数={cell_info['score']})")

        cell.fill = fill

        # 添加批注
        comment = Comment(
            f"风险分数: {cell_info['score']}\n"
            f"风险等级: {cell_info['level']}\n"
            f"权重: {cell_info['weight']}",
            "AI分析系统"
        )
        cell.comment = comment

        # 可选：添加字体颜色
        if cell_info["score"] >= 70:
            cell.font = Font(color="FF0000", bold=True)  # 红色加粗
        elif cell_info["score"] >= 40:
            cell.font = Font(color="FF8800")  # 橙色

        # 可选：添加边框
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = border

    # 保存修复后的文件
    fixed_file = "/root/projects/tencent-doc-manager/excel_outputs/marked/fixed_colored_WF_20250921.xlsx"
    wb.save(fixed_file)
    print(f"\n✅ 修复后的文件已保存: {fixed_file}")

    # 验证修复
    print("\n验证修复后的文件:")
    wb_check = load_workbook(fixed_file)
    ws_check = wb_check.active

    fill_count = 0
    for row in ws_check.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.patternType == 'solid':
                fill_count += 1

    print(f"✅ 发现 {fill_count} 个solid填充的单元格")

    return fixed_file


def main():
    """主函数"""

    # 1. 检查当前涂色文件
    check_colored_file()

    # 2. 测试正确的涂色方法
    test_file = test_correct_coloring()

    # 3. 修复涂色
    fixed_file = fix_coloring_method()

    print("\n" + "="*60)
    print("📊 诊断完成")
    print("="*60)
    print("\n建议:")
    print("1. 使用solid填充而不是lightUp")
    print("2. 确保颜色代码格式正确（6位或8位十六进制）")
    print("3. 添加字体颜色和边框增强视觉效果")
    print("4. 测试文件在不同Excel查看器中的兼容性")

    print(f"\n测试文件:")
    print(f"  - 测试涂色: {test_file}")
    print(f"  - 修复涂色: {fixed_file}")


if __name__ == "__main__":
    main()