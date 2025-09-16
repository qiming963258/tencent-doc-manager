#!/usr/bin/env python3
"""
DS+openpyxl自然语言Excel修改测试服务器（修复版）
端口：8101
功能：通过自然语言指令修改Excel文件
"""

from flask import Flask, render_template_string, request, jsonify, send_file
import os
import json
import re
from pathlib import Path
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string
import logging
import traceback

# 配置
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = '/root/projects/tencent-doc-manager/excel_uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 确保上传文件夹存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 配置日志
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# 全局变量存储当前文件路径
current_file_path = None
current_workbook = None
current_worksheet = None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 简化的自然语言处理器
class SimpleNLPProcessor:
    """简化的自然语言处理器（不依赖异步）"""
    
    def __init__(self):
        self.color_map = {
            '红': 'FF0000', '红色': 'FF0000', 'red': 'FF0000',
            '绿': '00FF00', '绿色': '00FF00', 'green': '00FF00',
            '蓝': '0000FF', '蓝色': '0000FF', 'blue': '0000FF',
            '黄': 'FFFF00', '黄色': 'FFFF00', 'yellow': 'FFFF00',
            '橙': 'FFA500', '橙色': 'FFA500', 'orange': 'FFA500',
            '紫': '800080', '紫色': '800080', 'purple': '800080',
            '灰': '808080', '灰色': '808080', 'gray': '808080',
            '粉': 'FFC0CB', '粉色': 'FFC0CB', 'pink': 'FFC0CB',
        }
    
    def parse_instruction(self, instruction):
        """解析自然语言指令"""
        instruction = instruction.strip().lower()
        
        # 单元格涂色
        cell_pattern = r'([a-z]+\d+)[^\w]*(?:涂|填|标记|改|变|设|染|着)(?:成|为|上)?[^\w]*(.+色)'
        match = re.search(cell_pattern, instruction)
        if match:
            cell = match.group(1).upper()
            color_text = match.group(2)
            color_code = self._get_color_code(color_text)
            return {
                'action': 'fill_cell',
                'cell': cell,
                'color': color_code
            }
        
        # 整行涂色
        row_pattern = r'第?(\d+)行[^\w]*(?:涂|填|标记|改|变)(?:成|为)?[^\w]*(.+色)'
        match = re.search(row_pattern, instruction)
        if match:
            row = int(match.group(1))
            color_text = match.group(2)
            color_code = self._get_color_code(color_text)
            return {
                'action': 'fill_row',
                'row': row,
                'color': color_code
            }
        
        # 整列涂色
        col_pattern = r'([a-z]+)列[^\w]*(?:涂|填|标记)(?:成|为)?[^\w]*(.+色)'
        match = re.search(col_pattern, instruction)
        if match:
            col = match.group(1).upper()
            color_text = match.group(2)
            color_code = self._get_color_code(color_text)
            return {
                'action': 'fill_column',
                'column': col,
                'color': color_code
            }
        
        # 范围涂色
        range_pattern = r'([a-z]+\d+)[^\w]*(?:到|至|:)[^\w]*([a-z]+\d+)[^\w]*(?:涂|填)(?:成)?[^\w]*(.+色)'
        match = re.search(range_pattern, instruction)
        if match:
            start_cell = match.group(1).upper()
            end_cell = match.group(2).upper()
            color_text = match.group(3)
            color_code = self._get_color_code(color_text)
            return {
                'action': 'fill_range',
                'start_cell': start_cell,
                'end_cell': end_cell,
                'color': color_code
            }
        
        # 总结表格
        if any(word in instruction for word in ['总结', '分析', '统计']):
            return {'action': 'summarize'}
        
        # 添加批注
        comment_pattern = r'([a-z]+\d+)[^\w]*(?:添加|加)[^\w]*(?:批注|注释)[^\w]*[:：]?\s*(.+)'
        match = re.search(comment_pattern, instruction)
        if match:
            cell = match.group(1).upper()
            comment_text = match.group(2).strip()
            return {
                'action': 'add_comment',
                'cell': cell,
                'comment': comment_text
            }
        
        return {'action': 'unknown'}
    
    def _get_color_code(self, color_text):
        """获取颜色代码"""
        color_text = color_text.strip().lower()
        for key, value in self.color_map.items():
            if key in color_text:
                return value
        return 'FF0000'  # 默认红色

# 创建处理器实例
nlp_processor = SimpleNLPProcessor()

# HTML模板（保持不变，只改端口显示）
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DS自然语言Excel处理器 - 8101端口</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
        }
        
        .header {
            text-align: center;
            color: white;
            margin-bottom: 30px;
        }
        
        .header h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }
        
        .header p {
            font-size: 1.1em;
            opacity: 0.95;
        }
        
        .main-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
            margin-bottom: 20px;
        }
        
        @media (max-width: 1024px) {
            .main-grid {
                grid-template-columns: 1fr;
            }
        }
        
        .card {
            background: white;
            border-radius: 15px;
            padding: 25px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.1);
        }
        
        .card h2 {
            color: #333;
            margin-bottom: 20px;
            font-size: 1.3em;
            border-bottom: 2px solid #667eea;
            padding-bottom: 10px;
        }
        
        .upload-area {
            border: 2px dashed #667eea;
            border-radius: 10px;
            padding: 40px;
            text-align: center;
            transition: all 0.3s;
            cursor: pointer;
            background: #f8f9ff;
        }
        
        .upload-area:hover {
            border-color: #764ba2;
            background: #f0f2ff;
        }
        
        .upload-area.dragover {
            background: #e8ebff;
            border-color: #764ba2;
        }
        
        .upload-icon {
            font-size: 3em;
            color: #667eea;
            margin-bottom: 10px;
        }
        
        .file-input {
            display: none;
        }
        
        .file-info {
            margin-top: 20px;
            padding: 15px;
            background: #f0f2ff;
            border-radius: 8px;
            display: none;
        }
        
        .file-info.active {
            display: block;
        }
        
        .instruction-input {
            width: 100%;
            padding: 15px;
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            font-size: 1.1em;
            transition: border-color 0.3s;
            margin-bottom: 15px;
        }
        
        .instruction-input:focus {
            outline: none;
            border-color: #667eea;
        }
        
        .examples {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-bottom: 15px;
        }
        
        .example-chip {
            padding: 6px 12px;
            background: #f0f2ff;
            border: 1px solid #667eea;
            border-radius: 20px;
            color: #667eea;
            cursor: pointer;
            font-size: 0.9em;
            transition: all 0.3s;
        }
        
        .example-chip:hover {
            background: #667eea;
            color: white;
        }
        
        .btn {
            padding: 12px 30px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 1.1em;
            cursor: pointer;
            transition: all 0.3s;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
        }
        
        .btn:hover:not(:disabled) {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.4);
        }
        
        .btn:disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }
        
        .btn-secondary {
            background: #6c757d;
            box-shadow: 0 4px 15px rgba(108, 117, 125, 0.3);
            margin-left: 10px;
        }
        
        .preview-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 0.9em;
        }
        
        .preview-table th,
        .preview-table td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        
        .preview-table th {
            background: #667eea;
            color: white;
            font-weight: 600;
        }
        
        .preview-table tr:nth-child(even) {
            background: #f9f9f9;
        }
        
        .result-box {
            margin-top: 20px;
            padding: 15px;
            border-radius: 8px;
            display: none;
        }
        
        .result-box.success {
            background: #d4edda;
            border: 1px solid #c3e6cb;
            color: #155724;
        }
        
        .result-box.error {
            background: #f8d7da;
            border: 1px solid #f5c6cb;
            color: #721c24;
        }
        
        .result-box.active {
            display: block;
        }
        
        .history-list {
            max-height: 300px;
            overflow-y: auto;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            padding: 10px;
        }
        
        .history-item {
            padding: 10px;
            margin-bottom: 8px;
            background: #f8f9ff;
            border-radius: 6px;
            border-left: 3px solid #667eea;
        }
        
        .loading {
            display: none;
            text-align: center;
            padding: 20px;
        }
        
        .loading.active {
            display: block;
        }
        
        .spinner {
            border: 3px solid #f3f3f3;
            border-top: 3px solid #667eea;
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
            margin: 0 auto;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .download-btn {
            background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
            box-shadow: 0 4px 15px rgba(40, 167, 69, 0.3);
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 DS自然语言Excel处理器</h1>
            <p>使用自然语言指令快速修改Excel文件 - 端口8101</p>
        </div>
        
        <div class="main-grid">
            <!-- 左侧：文件上传和预览 -->
            <div class="card">
                <h2>📁 文件管理</h2>
                
                <div class="upload-area" onclick="document.getElementById('fileInput').click();" 
                     ondrop="dropHandler(event);" ondragover="dragOverHandler(event);" 
                     ondragleave="dragLeaveHandler(event);">
                    <div class="upload-icon">📤</div>
                    <p>点击或拖拽Excel文件到此处</p>
                    <p style="color: #999; font-size: 0.9em; margin-top: 10px;">支持 .xlsx 格式</p>
                    <input type="file" id="fileInput" class="file-input" accept=".xlsx,.xls" onchange="handleFileSelect(event)">
                </div>
                
                <div id="fileInfo" class="file-info">
                    <strong>当前文件：</strong> <span id="fileName"></span><br>
                    <strong>文件大小：</strong> <span id="fileSize"></span><br>
                    <strong>上传时间：</strong> <span id="uploadTime"></span>
                </div>
                
                <h2 style="margin-top: 30px;">📊 表格预览</h2>
                <div id="previewArea">
                    <p style="color: #999;">请先上传Excel文件</p>
                </div>
            </div>
            
            <!-- 右侧：指令输入和结果 -->
            <div class="card">
                <h2>✏️ 自然语言指令</h2>
                
                <div class="examples">
                    <div class="example-chip" onclick="setInstruction('将G5涂成红色')">将G5涂成红色</div>
                    <div class="example-chip" onclick="setInstruction('把第3行标记为黄色')">第3行涂黄色</div>
                    <div class="example-chip" onclick="setInstruction('A列涂蓝色')">A列涂蓝色</div>
                    <div class="example-chip" onclick="setInstruction('B2到D4涂绿色')">B2:D4涂绿色</div>
                    <div class="example-chip" onclick="setInstruction('G2添加批注：重要数据')">添加批注</div>
                    <div class="example-chip" onclick="setInstruction('总结表格')">总结表格</div>
                </div>
                
                <input type="text" id="instructionInput" class="instruction-input" 
                       placeholder="输入指令，如：将G5涂成红色" 
                       onkeypress="if(event.key==='Enter') executeInstruction()">
                
                <div style="display: flex; gap: 10px;">
                    <button class="btn" onclick="executeInstruction()" id="executeBtn">
                        🎯 执行指令
                    </button>
                    <button class="btn btn-secondary" onclick="clearHistory()">
                        🗑️ 清空历史
                    </button>
                    <button class="btn download-btn" onclick="downloadModified()" id="downloadBtn" style="display:none;">
                        📥 下载文件
                    </button>
                </div>
                
                <div id="loading" class="loading">
                    <div class="spinner"></div>
                    <p style="margin-top: 10px;">处理中...</p>
                </div>
                
                <div id="resultBox" class="result-box"></div>
                
                <h2 style="margin-top: 30px;">📝 操作历史</h2>
                <div id="historyList" class="history-list">
                    <p style="color: #999;">暂无操作记录</p>
                </div>
            </div>
        </div>
    </div>
    
    <script>
        let currentFile = null;
        let modifiedFilePath = null;
        let operationHistory = [];
        let fileUploaded = false;
        
        function dragOverHandler(ev) {
            ev.preventDefault();
            document.querySelector('.upload-area').classList.add('dragover');
        }
        
        function dragLeaveHandler(ev) {
            document.querySelector('.upload-area').classList.remove('dragover');
        }
        
        function dropHandler(ev) {
            ev.preventDefault();
            document.querySelector('.upload-area').classList.remove('dragover');
            
            if (ev.dataTransfer.items) {
                for (let i = 0; i < ev.dataTransfer.items.length; i++) {
                    if (ev.dataTransfer.items[i].kind === 'file') {
                        const file = ev.dataTransfer.items[i].getAsFile();
                        processFile(file);
                        break;
                    }
                }
            }
        }
        
        function handleFileSelect(event) {
            const file = event.target.files[0];
            if (file) {
                processFile(file);
            }
        }
        
        function processFile(file) {
            if (!file.name.match(/\.(xlsx|xls)$/)) {
                alert('请上传Excel文件（.xlsx或.xls）');
                return;
            }
            
            currentFile = file;
            
            // 显示文件信息
            document.getElementById('fileName').textContent = file.name;
            document.getElementById('fileSize').textContent = formatFileSize(file.size);
            document.getElementById('uploadTime').textContent = new Date().toLocaleString('zh-CN');
            document.getElementById('fileInfo').classList.add('active');
            
            // 上传文件
            uploadFile(file);
        }
        
        function formatFileSize(bytes) {
            if (bytes === 0) return '0 Bytes';
            const k = 1024;
            const sizes = ['Bytes', 'KB', 'MB', 'GB'];
            const i = Math.floor(Math.log(bytes) / Math.log(k));
            return Math.round(bytes / Math.pow(k, i) * 100) / 100 + ' ' + sizes[i];
        }
        
        async function uploadFile(file) {
            const formData = new FormData();
            formData.append('file', file);
            
            try {
                const response = await fetch('/upload', {
                    method: 'POST',
                    body: formData
                });
                
                const result = await response.json();
                
                if (result.success) {
                    showResult('文件上传成功！', 'success');
                    fileUploaded = true;
                    document.getElementById('executeBtn').disabled = false;
                    loadPreview();
                } else {
                    showResult('文件上传失败：' + result.message, 'error');
                    fileUploaded = false;
                    document.getElementById('executeBtn').disabled = true;
                }
            } catch (error) {
                showResult('上传错误：' + error.message, 'error');
                fileUploaded = false;
                document.getElementById('executeBtn').disabled = true;
            }
        }
        
        async function loadPreview() {
            try {
                const response = await fetch('/preview');
                const result = await response.json();
                
                if (result.success) {
                    displayPreview(result.data);
                }
            } catch (error) {
                console.error('加载预览失败:', error);
            }
        }
        
        function displayPreview(data) {
            if (!data || data.length === 0) {
                document.getElementById('previewArea').innerHTML = '<p style="color: #999;">表格为空</p>';
                return;
            }
            
            let html = '<div style="overflow-x: auto;"><table class="preview-table">';
            
            // 表头
            html += '<thead><tr>';
            for (let i = 0; i < data[0].length; i++) {
                html += `<th>列${String.fromCharCode(65 + i)}</th>`;
            }
            html += '</tr></thead>';
            
            // 数据行
            html += '<tbody>';
            data.forEach((row, rowIndex) => {
                html += '<tr>';
                row.forEach((cell, colIndex) => {
                    const cellId = String.fromCharCode(65 + colIndex) + (rowIndex + 1);
                    html += `<td id="cell_${cellId}">${cell || ''}</td>`;
                });
                html += '</tr>';
            });
            html += '</tbody></table></div>';
            
            document.getElementById('previewArea').innerHTML = html;
        }
        
        function setInstruction(text) {
            document.getElementById('instructionInput').value = text;
        }
        
        async function executeInstruction() {
            const instruction = document.getElementById('instructionInput').value.trim();
            
            if (!instruction) {
                alert('请输入指令');
                return;
            }
            
            if (!fileUploaded) {
                alert('请先上传Excel文件');
                return;
            }
            
            document.getElementById('loading').classList.add('active');
            document.getElementById('executeBtn').disabled = true;
            
            try {
                const response = await fetch('/execute', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({ instruction: instruction })
                });
                
                const result = await response.json();
                
                if (result.success) {
                    showResult(result.message, 'success');
                    addToHistory(instruction, result.message);
                    
                    if (result.file_path) {
                        modifiedFilePath = result.file_path;
                        document.getElementById('downloadBtn').style.display = 'inline-block';
                    }
                    
                    loadPreview();
                    document.getElementById('instructionInput').value = '';
                } else {
                    showResult(result.message, 'error');
                }
            } catch (error) {
                showResult('执行错误：' + error.message, 'error');
            } finally {
                document.getElementById('loading').classList.remove('active');
                document.getElementById('executeBtn').disabled = false;
            }
        }
        
        function showResult(message, type) {
            const resultBox = document.getElementById('resultBox');
            resultBox.textContent = message;
            resultBox.className = `result-box ${type} active`;
            
            setTimeout(() => {
                resultBox.classList.remove('active');
            }, 5000);
        }
        
        function addToHistory(instruction, result) {
            const time = new Date().toLocaleTimeString('zh-CN');
            operationHistory.unshift({
                instruction: instruction,
                result: result,
                time: time
            });
            
            updateHistoryDisplay();
        }
        
        function updateHistoryDisplay() {
            const historyList = document.getElementById('historyList');
            
            if (operationHistory.length === 0) {
                historyList.innerHTML = '<p style="color: #999;">暂无操作记录</p>';
                return;
            }
            
            let html = '';
            operationHistory.slice(0, 10).forEach(item => {
                html += `
                    <div class="history-item">
                        <div><strong>指令：</strong> ${item.instruction}</div>
                        <div><strong>结果：</strong> ${item.result}</div>
                        <div style="color: #666; font-size: 0.8em;">${item.time}</div>
                    </div>
                `;
            });
            
            historyList.innerHTML = html;
        }
        
        function clearHistory() {
            operationHistory = [];
            updateHistoryDisplay();
        }
        
        function downloadModified() {
            if (modifiedFilePath) {
                window.location.href = '/download?path=' + encodeURIComponent(modifiedFilePath);
            }
        }
        
        // 页面加载时禁用执行按钮
        document.addEventListener('DOMContentLoaded', function() {
            document.getElementById('executeBtn').disabled = true;
        });
    </script>
</body>
</html>
'''

@app.route('/')
def index():
    """主页"""
    return render_template_string(HTML_TEMPLATE)

@app.route('/upload', methods=['POST'])
def upload_file():
    """上传文件"""
    global current_file_path, current_workbook, current_worksheet
    
    try:
        if 'file' not in request.files:
            logger.error("没有文件在请求中")
            return jsonify({'success': False, 'message': '没有文件'})
        
        file = request.files['file']
        if file.filename == '':
            logger.error("文件名为空")
            return jsonify({'success': False, 'message': '文件名为空'})
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # 保存文件
            file.save(filepath)
            logger.info(f"文件已保存到: {filepath}")
            
            # 尝试加载文件
            try:
                current_workbook = load_workbook(filepath)
                current_worksheet = current_workbook.active
                current_file_path = filepath
                logger.info(f"Excel文件加载成功: {filepath}")
                return jsonify({'success': True, 'message': '文件上传成功'})
            except Exception as load_error:
                logger.error(f"加载Excel文件失败: {load_error}")
                return jsonify({'success': False, 'message': f'文件加载失败: {str(load_error)}'})
        else:
            return jsonify({'success': False, 'message': '不支持的文件格式'})
            
    except Exception as e:
        logger.error(f"上传处理失败: {e}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'上传失败: {str(e)}'})

@app.route('/preview')
def get_preview():
    """获取表格预览"""
    global current_worksheet
    
    if not current_worksheet:
        return jsonify({'success': False, 'message': '请先上传文件'})
    
    try:
        preview_data = []
        max_rows = min(10, current_worksheet.max_row)
        max_cols = min(10, current_worksheet.max_column)
        
        for row in current_worksheet.iter_rows(min_row=1, max_row=max_rows,
                                              min_col=1, max_col=max_cols,
                                              values_only=True):
            preview_data.append(list(row))
        
        return jsonify({'success': True, 'data': preview_data})
    except Exception as e:
        logger.error(f"获取预览失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/execute', methods=['POST'])
def execute_instruction():
    """执行自然语言指令"""
    global current_workbook, current_worksheet, current_file_path
    
    if not current_workbook:
        return jsonify({'success': False, 'message': '请先上传文件'})
    
    data = request.json
    instruction = data.get('instruction', '')
    
    if not instruction:
        return jsonify({'success': False, 'message': '指令不能为空'})
    
    try:
        # 解析指令
        parsed = nlp_processor.parse_instruction(instruction)
        logger.info(f"解析结果: {parsed}")
        
        # 执行操作
        action = parsed.get('action')
        ws = current_worksheet
        
        if action == 'fill_cell':
            # 单元格涂色
            cell = parsed['cell']
            color = parsed['color']
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[cell].fill = fill
            message = f'✅ 已将{cell}涂成颜色#{color}'
            
        elif action == 'fill_row':
            # 整行涂色
            row = parsed['row']
            color = parsed['color']
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill
            message = f'✅ 已将第{row}行涂成颜色#{color}'
            
        elif action == 'fill_column':
            # 整列涂色
            column = parsed['column']
            color = parsed['color']
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            col_idx = column_index_from_string(column)
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).fill = fill
            message = f'✅ 已将{column}列涂成颜色#{color}'
            
        elif action == 'fill_range':
            # 范围涂色
            start_cell = parsed['start_cell']
            end_cell = parsed['end_cell']
            color = parsed['color']
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            # 解析起止单元格
            start_col = column_index_from_string(re.match(r'([A-Z]+)', start_cell).group(1))
            start_row = int(re.match(r'[A-Z]+(\d+)', start_cell).group(1))
            end_col = column_index_from_string(re.match(r'([A-Z]+)', end_cell).group(1))
            end_row = int(re.match(r'[A-Z]+(\d+)', end_cell).group(1))
            
            # 涂色
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).fill = fill
            
            message = f'✅ 已将{start_cell}到{end_cell}涂成颜色#{color}'
            
        elif action == 'add_comment':
            # 添加批注
            cell = parsed['cell']
            comment_text = parsed['comment']
            from openpyxl.comments import Comment
            ws[cell].comment = Comment(comment_text, "DS Excel")
            message = f'✅ 已为{cell}添加批注: {comment_text}'
            
        elif action == 'summarize':
            # 总结表格
            stats = {
                'rows': ws.max_row,
                'columns': ws.max_column,
                'cells': ws.max_row * ws.max_column
            }
            message = f'📊 表格概况：{stats["rows"]}行 × {stats["columns"]}列，共{stats["cells"]}个单元格'
            
        else:
            return jsonify({'success': False, 'message': f'未能识别指令: {instruction}'})
        
        # 保存文件
        if action != 'summarize':
            output_path = current_file_path.replace('.xlsx', '_modified.xlsx')
            current_workbook.save(output_path)
            logger.info(f"文件已保存: {output_path}")
            
            return jsonify({
                'success': True,
                'message': message,
                'file_path': output_path
            })
        else:
            return jsonify({
                'success': True,
                'message': message
            })
            
    except Exception as e:
        logger.error(f"执行指令失败: {e}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'执行失败: {str(e)}'})

@app.route('/download')
def download_file():
    """下载修改后的文件"""
    file_path = request.args.get('path')
    if file_path and os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, 
                        download_name=os.path.basename(file_path))
    return "文件不存在", 404

if __name__ == '__main__':
    print("="*60)
    print("🚀 DS自然语言Excel处理器启动中...")
    print("="*60)
    print(f"📍 访问地址: http://localhost:8101")
    print(f"📍 外网地址: http://202.140.143.88:8101")
    print(f"📁 上传目录: {app.config['UPLOAD_FOLDER']}")
    print("💡 功能特性:")
    print("   - 自然语言指令修改Excel")
    print("   - 单元格/行/列/范围涂色")
    print("   - 添加批注和表格总结")
    print("="*60)
    
    app.run(host='0.0.0.0', port=8101, debug=False)