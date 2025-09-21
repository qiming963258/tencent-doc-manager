#!/usr/bin/env python3
"""
XLSX真实测试 - 使用已有基线文件对比
下载当前XLSX，与基线对比，涂色后上传
"""

import json
import sys
import os
import asyncio
from pathlib import Path
from datetime import datetime
import logging
import csv
from typing import Optional, Dict, List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

sys.path.append('/root/projects/tencent-doc-manager')

from production.config import STANDARD_COLUMNS, L1_COLUMNS, L2_COLUMNS, L3_COLUMNS


class XLSXRealTestWithBaseline:
    """XLSX真实测试 - 使用基线文件"""

    def __init__(self):
        self.doc_url = "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN"
        self.doc_id = "DWEFNU25TemFnZXJN"
        # 使用已有的基线CSV文件
        self.baseline_csv = "/root/projects/tencent-doc-manager/csv_versions/2025_W38/baseline/tencent_出国销售计划表_20250915_0145_baseline_W38.csv"
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.download_dir = Path(f'/root/projects/tencent-doc-manager/downloads/xlsx_real_{self.timestamp}')
        self.download_dir.mkdir(parents=True, exist_ok=True)

    def convert_baseline_csv_to_xlsx(self) -> Optional[str]:
        """将基线CSV转换为XLSX格式用于对比"""

        logger.info("="*70)
        logger.info("准备基线文件：CSV转XLSX")
        logger.info("="*70)

        if not Path(self.baseline_csv).exists():
            logger.error(f"❌ 基线CSV文件不存在: {self.baseline_csv}")
            return None

        try:
            # 创建XLSX文件
            wb = Workbook()
            ws = wb.active
            ws.title = "基线数据"

            # 读取CSV并写入XLSX
            with open(self.baseline_csv, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # 保存基线XLSX
            baseline_xlsx = self.download_dir / f'baseline_{self.timestamp}.xlsx'
            wb.save(baseline_xlsx)
            wb.close()

            logger.info(f"✅ 基线CSV成功转换为XLSX: {baseline_xlsx.name}")
            logger.info(f"   文件大小: {baseline_xlsx.stat().st_size:,} bytes")
            return str(baseline_xlsx)

        except Exception as e:
            logger.error(f"❌ 转换失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def step1_download_current_xlsx(self) -> Optional[str]:
        """Step 1: 下载当前XLSX格式文件"""

        logger.info("\n" + "="*70)
        logger.info("Step 1: 下载当前XLSX文件")
        logger.info("="*70)
        logger.info(f"目标URL: {self.doc_url}")

        try:
            from production.core_modules.playwright_downloader import PlaywrightDownloader

            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            if not cookie_file.exists():
                logger.error("❌ Cookie文件不存在")
                return None

            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')
            if not cookie_string:
                logger.error("❌ Cookie为空")
                return None

            logger.info(f"✅ Cookie已加载 (最后更新: {cookie_data.get('last_update', 'Unknown')})")

            # 创建下载器
            downloader = PlaywrightDownloader()

            # 尝试下载XLSX格式
            logger.info("🔄 尝试下载XLSX格式...")
            result = await downloader.download(
                url=self.doc_url,
                cookies=cookie_string,
                format='xlsx',  # 请求XLSX格式
                download_dir=str(self.download_dir)
            )

            if result.get('success'):
                # 获取下载的文件
                if result.get('file_path'):
                    downloaded_file = result['file_path']
                elif result.get('files'):
                    downloaded_file = result['files'][0]
                else:
                    logger.error("❌ 下载结果中没有文件")
                    return None

                if Path(downloaded_file).exists():
                    # 检查文件类型
                    if downloaded_file.endswith('.csv'):
                        logger.warning("⚠️ 腾讯文档仅返回CSV格式，正在转换为XLSX...")
                        return await self._convert_csv_to_xlsx(downloaded_file)
                    else:
                        # 是XLSX格式，重命名
                        xlsx_file = self.download_dir / f'current_{self.timestamp}.xlsx'
                        import shutil
                        shutil.copy2(downloaded_file, xlsx_file)
                        logger.info(f"✅ XLSX下载成功: {xlsx_file.name}")
                        logger.info(f"   文件大小: {xlsx_file.stat().st_size:,} bytes")
                        return str(xlsx_file)
                else:
                    logger.error(f"❌ 文件不存在: {downloaded_file}")
                    return None
            else:
                logger.error(f"❌ 下载失败: {result.get('error', 'Unknown error')}")
                return None

        except Exception as e:
            logger.error(f"❌ 下载异常: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def _convert_csv_to_xlsx(self, csv_file: str) -> Optional[str]:
        """将下载的CSV转换为XLSX格式"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "当前数据"

            with open(csv_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            xlsx_file = self.download_dir / f'current_{self.timestamp}.xlsx'
            wb.save(xlsx_file)
            wb.close()

            logger.info(f"✅ CSV成功转换为XLSX: {xlsx_file.name}")
            logger.info(f"   文件大小: {xlsx_file.stat().st_size:,} bytes")
            return str(xlsx_file)

        except Exception as e:
            logger.error(f"❌ CSV转换失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def step2_compare_xlsx_files(self, baseline_xlsx: str, current_xlsx: str) -> List[Dict]:
        """Step 2: 对比两个XLSX文件"""

        logger.info("\n" + "="*70)
        logger.info("Step 2: XLSX文件对比分析")
        logger.info("="*70)

        try:
            # 加载两个XLSX文件
            wb_baseline = load_workbook(baseline_xlsx, data_only=True)
            wb_current = load_workbook(current_xlsx, data_only=True)

            ws_baseline = wb_baseline.active
            ws_current = wb_current.active

            logger.info(f"基线XLSX: {ws_baseline.max_row}行 x {ws_baseline.max_column}列")
            logger.info(f"当前XLSX: {ws_current.max_row}行 x {ws_current.max_column}列")

            changes = []

            # 获取标题行
            headers = []
            for col in range(1, ws_current.max_column + 1):
                cell_value = ws_current.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Col{col}")

            # 对比数据
            max_rows = min(ws_baseline.max_row, ws_current.max_row)
            max_cols = min(ws_baseline.max_column, ws_current.max_column)

            for row in range(2, max_rows + 1):  # 跳过标题行
                for col in range(1, max_cols + 1):
                    val_baseline = ws_baseline.cell(row=row, column=col).value
                    val_current = ws_current.cell(row=row, column=col).value

                    # 转换为字符串比较
                    str_baseline = str(val_baseline).strip() if val_baseline is not None else ""
                    str_current = str(val_current).strip() if val_current is not None else ""

                    if str_baseline != str_current:
                        changes.append({
                            "row": row,
                            "col": col,
                            "old": str_baseline,
                            "new": str_current,
                            "column_name": headers[col-1] if col-1 < len(headers) else f"Col{col}",
                            "cell_ref": f"{get_column_letter(col)}{row}"
                        })

            logger.info(f"✅ 发现 {len(changes)} 处变更")

            # 显示前5个变更
            for i, change in enumerate(changes[:5]):
                logger.info(f"   变更{i+1}: [{change['cell_ref']}] "
                          f"'{change['old'][:30]}' → '{change['new'][:30]}'")

            if len(changes) > 5:
                logger.info(f"   ... 还有 {len(changes)-5} 处变更")

            # 关闭工作簿
            wb_baseline.close()
            wb_current.close()

            return changes

        except Exception as e:
            logger.error(f"❌ 对比失败: {e}")
            import traceback
            traceback.print_exc()
            return []

    def step3_apply_coloring(self, xlsx_file: str, changes: List[Dict]) -> Optional[str]:
        """Step 3: 在XLSX文件上应用涂色"""

        logger.info("\n" + "="*70)
        logger.info("Step 3: 应用智能涂色标记")
        logger.info("="*70)

        if not changes:
            logger.warning("⚠️ 没有变更需要标记")
            # 即使没有变更也返回文件，用于演示
            return xlsx_file

        try:
            # 加载XLSX文件
            wb = load_workbook(xlsx_file)
            ws = wb.active

            marked_count = 0
            risk_summary = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}

            logger.info("开始应用涂色...")

            for change in changes:
                row = change["row"]
                col = change["col"]

                # 判断风险等级
                col_idx = col - 1
                if col_idx < len(STANDARD_COLUMNS):
                    col_name = STANDARD_COLUMNS[col_idx]

                    if col_name in L1_COLUMNS:
                        risk_level = "HIGH"
                        color = "FFCCCC"  # 浅红
                        font_color = "CC0000"
                    elif col_name in L2_COLUMNS:
                        risk_level = "MEDIUM"
                        color = "FFFFCC"  # 浅黄
                        font_color = "FF8800"
                    else:
                        risk_level = "LOW"
                        color = "CCFFCC"  # 浅绿
                        font_color = "008800"
                else:
                    risk_level = "LOW"
                    color = "CCFFCC"
                    font_color = "008800"

                risk_summary[risk_level] += 1

                # 获取单元格
                cell = ws.cell(row=row, column=col)

                # 应用solid填充（腾讯文档兼容）
                cell.fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid"  # 必须是solid
                )

                # 应用字体样式
                cell.font = Font(
                    color=font_color,
                    bold=(risk_level == "HIGH")
                )

                # 添加批注
                comment_text = (
                    f"风险等级: {risk_level}\n"
                    f"原值: {change['old'][:50]}\n"
                    f"新值: {change['new'][:50]}\n"
                    f"列名: {change['column_name']}"
                )
                cell.comment = Comment(comment_text, "XLSX智能分析")

                marked_count += 1

            # 保存涂色后的文件
            output_dir = Path('/root/projects/tencent-doc-manager/excel_outputs/xlsx_real')
            output_dir.mkdir(parents=True, exist_ok=True)

            output_file = output_dir / f'xlsx_real_colored_{self.timestamp}.xlsx'
            wb.save(output_file)
            wb.close()

            logger.info(f"✅ 涂色完成: {output_file.name}")
            logger.info(f"   标记单元格: {marked_count}个")
            logger.info(f"   高风险: {risk_summary['HIGH']}")
            logger.info(f"   中风险: {risk_summary['MEDIUM']}")
            logger.info(f"   低风险: {risk_summary['LOW']}")
            logger.info(f"   文件大小: {output_file.stat().st_size:,} bytes")

            return str(output_file)

        except Exception as e:
            logger.error(f"❌ 涂色失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def step4_upload_to_tencent(self, xlsx_file: str) -> Optional[str]:
        """Step 4: 上传到腾讯文档"""

        logger.info("\n" + "="*70)
        logger.info("Step 4: 上传XLSX到腾讯文档")
        logger.info("="*70)

        try:
            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')

            # 使用quick_upload_v3
            from production.core_modules.tencent_doc_upload_production_v3 import quick_upload_v3

            logger.info(f"📄 准备上传: {Path(xlsx_file).name}")
            logger.info(f"   文件大小: {Path(xlsx_file).stat().st_size:,} bytes")
            logger.info("🔄 正在连接腾讯文档...")

            result = await quick_upload_v3(
                cookie_string=cookie_string,
                file_path=xlsx_file,
                headless=True
            )

            if result and result.get('success'):
                url = result.get('url')
                logger.info(f"✅ 上传成功!")
                logger.info(f"🔗 文档URL: {url}")
                return url
            else:
                logger.error(f"❌ 上传失败: {result.get('message', 'Unknown error') if result else 'No result'}")
                return None

        except Exception as e:
            logger.error(f"❌ 上传异常: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def run_complete_test(self) -> Dict:
        """执行完整的XLSX测试流程"""

        logger.info("\n" + "="*80)
        logger.info("🚀 XLSX格式完整测试 - 使用真实基线文件")
        logger.info("="*80)
        logger.info(f"基线文件: {self.baseline_csv}")
        logger.info(f"目标URL: {self.doc_url}")
        logger.info(f"时间戳: {self.timestamp}")

        test_result = {
            "timestamp": self.timestamp,
            "success": False,
            "steps": {},
            "summary": "",
            "final_url": None
        }

        # 准备基线XLSX
        baseline_xlsx = self.convert_baseline_csv_to_xlsx()
        if not baseline_xlsx:
            test_result["summary"] = "基线文件准备失败"
            return test_result

        # Step 1: 下载当前XLSX
        current_xlsx = await self.step1_download_current_xlsx()
        test_result["steps"]["download"] = {
            "success": bool(current_xlsx),
            "file": current_xlsx
        }

        if not current_xlsx:
            test_result["summary"] = "下载失败"
            return test_result

        # Step 2: 对比
        changes = self.step2_compare_xlsx_files(baseline_xlsx, current_xlsx)
        test_result["steps"]["compare"] = {
            "success": True,
            "changes_count": len(changes)
        }

        # Step 3: 涂色
        colored_xlsx = self.step3_apply_coloring(current_xlsx, changes)
        test_result["steps"]["coloring"] = {
            "success": bool(colored_xlsx),
            "file": colored_xlsx
        }

        if not colored_xlsx:
            test_result["summary"] = "涂色失败"
            return test_result

        # Step 4: 上传
        upload_url = await self.step4_upload_to_tencent(colored_xlsx)
        test_result["steps"]["upload"] = {
            "success": bool(upload_url),
            "url": upload_url
        }

        # 总结
        if upload_url:
            test_result["success"] = True
            test_result["final_url"] = upload_url
            test_result["summary"] = f"XLSX测试成功，发现{len(changes)}处变更"

            logger.info("\n" + "="*80)
            logger.info("✅ XLSX格式完整测试成功！")
            logger.info(f"   基线文件: {Path(baseline_xlsx).name}")
            logger.info(f"   当前文件: {Path(current_xlsx).name}")
            logger.info(f"   发现变更: {len(changes)}处")
            logger.info(f"   涂色文件: {Path(colored_xlsx).name if colored_xlsx else 'N/A'}")
            logger.info(f"   上传成功: {upload_url}")
            logger.info("\n" + "🔗 请访问以下URL验证XLSX涂色效果：")
            logger.info(f"   {upload_url}")
            logger.info("="*80)
        else:
            test_result["summary"] = "上传失败"

        return test_result


async def main():
    """主函数"""

    # 创建测试实例
    tester = XLSXRealTestWithBaseline()

    # 执行测试
    result = await tester.run_complete_test()

    # 保存结果
    result_file = Path(f'/root/projects/tencent-doc-manager/test_results/xlsx_real_test_{tester.timestamp}.json')
    result_file.parent.mkdir(exist_ok=True)

    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    # 输出结果
    if result.get("final_url"):
        print(f"\n\n" + "="*80)
        print(f"🎉 XLSX测试最终URL：")
        print(f"   {result['final_url']}")
        print("="*80)
    else:
        print("\n\n❌ 测试失败，未能获得URL")

    return 0 if result.get("success") else 1


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)