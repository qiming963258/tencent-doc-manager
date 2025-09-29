#!/usr/bin/env python3
"""
严格的上传测试 - 不使用任何猜测策略
只依赖真实的API响应
"""

import asyncio
import json
import logging
from pathlib import Path
from datetime import datetime
from playwright.async_api import async_playwright, Response
import time

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class StrictUploadTester:
    """严格的上传测试器 - 不猜测，只看真实响应"""

    def __init__(self, headless: bool = False):
        self.headless = headless
        self.browser = None
        self.context = None
        self.page = None
        self.playwright = None
        self.api_responses = []
        self.import_api_response = None

    async def __aenter__(self):
        await self.setup()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.cleanup()

    async def setup(self):
        """初始化浏览器"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=['--disable-blink-features=AutomationControlled']
        )
        self.context = await self.browser.new_context(
            viewport={'width': 1920, 'height': 1080}
        )
        self.page = await self.context.new_page()

        # 监听所有网络响应
        self.page.on('response', self.handle_response)

        logger.info("✅ 浏览器启动成功")

    async def handle_response(self, response: Response):
        """捕获所有API响应"""
        url = response.url

        # 重点监控导入相关API
        if '/v1/import/' in url or '/import/secretKey' in url:
            try:
                if response.status == 200:
                    content_type = response.headers.get('content-type', '')
                    if 'json' in content_type:
                        data = await response.json()
                        self.import_api_response = data
                        logger.info(f"📡 捕获导入API响应:")
                        logger.info(f"   URL: {url}")
                        logger.info(f"   数据: {json.dumps(data, ensure_ascii=False)}")

                        # 分析响应
                        if 'url' in data:
                            if data['url']:
                                logger.info(f"   ✅ 返回URL: {data['url']}")
                            else:
                                logger.warning(f"   ⚠️ URL字段为空！")

                        if 'doc_id' in data:
                            if data['doc_id']:
                                logger.info(f"   ✅ 文档ID: {data['doc_id']}")
                            else:
                                logger.warning(f"   ⚠️ 文档ID为空！")

                        if 'error' in data:
                            if data['error'] != 0:
                                logger.error(f"   ❌ 错误代码: {data['error']}")

                        if 'msg' in data:
                            logger.info(f"   消息: {data['msg']}")
            except Exception as e:
                logger.debug(f"解析响应失败: {e}")

        # 记录所有API调用
        if 'docs.qq.com' in url and response.status == 200:
            self.api_responses.append({
                'url': url,
                'status': response.status,
                'time': datetime.now().isoformat()
            })

    def parse_cookie_string(self, cookie_string: str) -> list:
        """解析Cookie字符串"""
        cookies = []
        for cookie_pair in cookie_string.split('; '):
            if '=' in cookie_pair:
                name, value = cookie_pair.strip().split('=', 1)
                cookies.append({
                    'name': name.strip(),
                    'value': value.strip(),
                    'domain': '.qq.com',
                    'path': '/'
                })
        return cookies

    async def check_storage_space(self) -> dict:
        """检查存储空间"""
        try:
            # 查找存储空间元素
            storage_elem = await self.page.query_selector('.desktop-storage-panel')
            if storage_elem:
                storage_text = await storage_elem.inner_text()
                logger.info(f"📊 存储空间信息: {storage_text}")

                # 获取使用率
                storage_bar = await self.page.query_selector('.desktop-storage-bar')
                if storage_bar:
                    style = await storage_bar.get_attribute('style')
                    if '--size:' in style:
                        size = style.split('--size:')[1].split('%')[0].strip()
                        usage = float(size)

                        # 检查是否有critical类
                        classes = await storage_bar.get_attribute('class')
                        is_critical = 'critical' in classes

                        result = {
                            'usage_percent': usage,
                            'is_critical': is_critical,
                            'has_space': usage < 95,
                            'text': storage_text
                        }

                        if usage >= 95:
                            logger.error(f"❌ 存储空间不足: {usage:.2f}% 已使用")
                        else:
                            logger.info(f"✅ 存储空间充足: {usage:.2f}% 已使用")

                        return result

        except Exception as e:
            logger.error(f"检查存储空间失败: {e}")

        return {'usage_percent': -1, 'is_critical': False, 'has_space': True}

    async def login_and_check(self, cookie_string: str) -> dict:
        """登录并检查状态"""
        result = {
            'login_success': False,
            'storage_space': {},
            'can_upload': False
        }

        try:
            # 添加Cookie
            cookies = self.parse_cookie_string(cookie_string)
            await self.context.add_cookies(cookies)

            # 访问首页
            await self.page.goto('https://docs.qq.com/desktop/', wait_until='networkidle')
            await self.page.wait_for_timeout(3000)

            # 检查登录状态
            login_btn = await self.page.query_selector('button:has-text("登录")')
            import_btn = await self.page.query_selector('button.desktop-import-button-pc')

            if login_btn:
                logger.error("❌ Cookie已失效，需要重新登录")
                result['login_success'] = False
            elif import_btn:
                logger.info("✅ 登录成功")
                result['login_success'] = True

                # 检查存储空间
                storage_info = await self.check_storage_space()
                result['storage_space'] = storage_info
                result['can_upload'] = storage_info.get('has_space', False)
            else:
                logger.warning("⚠️ 登录状态不确定")

        except Exception as e:
            logger.error(f"登录检查失败: {e}")

        return result

    async def create_minimal_test_file(self) -> str:
        """创建最小测试文件"""
        import openpyxl
        from openpyxl.styles import PatternFill

        # 创建极小的Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试"

        # 只写入一个单元格
        ws['A1'] = 'Test'
        ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 保存
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"/tmp/minimal_test_{timestamp}.xlsx"
        wb.save(filename)

        # 获取文件大小
        file_size = Path(filename).stat().st_size
        logger.info(f"✅ 创建最小测试文件: {filename} ({file_size} bytes)")

        return filename

    async def strict_upload_test(self, file_path: str) -> dict:
        """严格的上传测试 - 不使用任何猜测"""
        result = {
            'upload_started': False,
            'api_response': None,
            'real_success': False,
            'url': None,
            'error': None
        }

        try:
            file_name = Path(file_path).name
            logger.info(f"📤 开始严格上传测试: {file_name}")

            # 点击导入按钮
            import_btn = await self.page.query_selector('button.desktop-import-button-pc')
            if not import_btn:
                result['error'] = "找不到导入按钮"
                return result

            # 选择文件
            async with self.page.expect_file_chooser() as fc_info:
                await import_btn.click()
            file_chooser = await fc_info.value
            await file_chooser.set_files(file_path)

            result['upload_started'] = True
            logger.info("✅ 文件已选择，等待对话框...")

            # 等待并点击确定
            await self.page.wait_for_timeout(2000)
            confirm_btn = await self.page.wait_for_selector(
                'button.dui-button-type-primary:has-text("确定")',
                timeout=5000
            )
            await confirm_btn.click()
            logger.info("✅ 点击确定按钮")

            # 等待上传完成（最多30秒）
            logger.info("⏳ 等待API响应...")
            start_time = time.time()
            while time.time() - start_time < 30:
                await self.page.wait_for_timeout(1000)

                # 检查是否收到API响应
                if self.import_api_response:
                    result['api_response'] = self.import_api_response

                    # 分析响应
                    if self.import_api_response.get('url'):
                        result['real_success'] = True
                        result['url'] = self.import_api_response['url']
                        logger.info(f"✅ 真实上传成功: {result['url']}")
                    else:
                        result['real_success'] = False
                        result['error'] = "API返回的URL为空"
                        logger.error("❌ API返回的URL为空，上传失败")
                    break

                # 检查对话框是否关闭
                import_dialog = await self.page.query_selector('.import-kit-import-file')
                if not import_dialog:
                    logger.info("对话框已关闭")
                    break

            if not result['api_response']:
                result['error'] = "未收到API响应"
                logger.error("❌ 30秒内未收到API响应")

        except Exception as e:
            result['error'] = str(e)
            logger.error(f"❌ 上传测试失败: {e}")

        return result

    async def cleanup(self):
        """清理资源"""
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()


async def main():
    """主测试流程"""

    print("="*60)
    print("🔬 严格上传测试 - 深度分析版")
    print("="*60)

    # 读取Cookie
    cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
    with open(cookie_file) as f:
        cookie_data = json.load(f)

    cookie_string = cookie_data.get('current_cookies', '')

    async with StrictUploadTester(headless=True) as tester:
        # 步骤1：登录并检查
        print("\n📋 步骤1: 登录并检查账户状态")
        print("-"*50)

        status = await tester.login_and_check(cookie_string)

        if not status['login_success']:
            print("❌ 登录失败，Cookie需要更新")
            return

        # 显示存储空间状态
        storage = status['storage_space']
        if storage.get('usage_percent', -1) > 0:
            print(f"📊 存储空间使用: {storage['usage_percent']:.2f}%")
            if storage['is_critical']:
                print("⚠️ 存储空间标记为critical（严重）")
            if not storage['has_space']:
                print("❌ 存储空间不足，可能无法上传")

        # 步骤2：创建最小测试文件
        print("\n📋 步骤2: 创建最小测试文件")
        print("-"*50)

        test_file = await tester.create_minimal_test_file()
        print(f"✅ 测试文件: {test_file}")

        # 步骤3：严格上传测试
        print("\n📋 步骤3: 执行严格上传测试")
        print("-"*50)

        result = await tester.strict_upload_test(test_file)

        # 步骤4：分析结果
        print("\n📋 步骤4: 分析测试结果")
        print("-"*50)

        if result['api_response']:
            print("📡 API响应内容:")
            print(json.dumps(result['api_response'], indent=2, ensure_ascii=False))

        if result['real_success']:
            print(f"✅ 真实上传成功")
            print(f"📄 文档URL: {result['url']}")
        else:
            print(f"❌ 上传失败")
            print(f"原因: {result['error']}")

        # 最终诊断
        print("\n" + "="*60)
        print("🎯 诊断结果")
        print("="*60)

        if not status['storage_space'].get('has_space', True):
            print("问题根因: 存储空间不足")
            print("解决方案:")
            print("1. 清理腾讯文档回收站")
            print("2. 删除不需要的文档")
            print("3. 升级存储空间")
        elif not result['real_success']:
            print("问题根因: 上传权限或API问题")
            print("解决方案:")
            print("1. 更新Cookie")
            print("2. 检查账户权限")
            print("3. 联系腾讯文档支持")
        else:
            print("✅ 上传功能正常")

        # 显示所有API调用
        if tester.api_responses:
            print(f"\n📊 共记录 {len(tester.api_responses)} 个API调用")


if __name__ == "__main__":
    asyncio.run(main())