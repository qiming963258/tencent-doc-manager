#!/usr/bin/env python3
"""
完整真实全链路测试 - 包含上传步骤
绝对无虚拟行为，包含从下载到上传的完整链路
"""

import json
import sys
import os
import asyncio
from pathlib import Path
from datetime import datetime
import csv
import logging
from typing import Optional, Dict, List

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

sys.path.append('/root/projects/tencent-doc-manager')

from production.config import STANDARD_COLUMNS, L1_COLUMNS, L2_COLUMNS, L3_COLUMNS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.comments import Comment


class CompleteRealChainTest:
    """完整的真实测试类 - 包含上传"""

    def __init__(self):
        self.doc_url = "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN"
        self.doc_id = "DWEFNU25TemFnZXJN"
        self.baseline_file = "/root/projects/tencent-doc-manager/csv_versions/2025_W38/baseline/tencent_出国销售计划表_20250915_0145_baseline_W38.csv"
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.download_dir = Path(f'/root/projects/tencent-doc-manager/downloads/complete_real_{self.timestamp}')
        self.download_dir.mkdir(parents=True, exist_ok=True)

    async def step1_real_download(self) -> Optional[str]:
        """Step 1: 真实下载 - 无任何模拟"""

        logger.info("="*70)
        logger.info("Step 1: 真实下载腾讯文档")
        logger.info("="*70)
        logger.info(f"目标URL: {self.doc_url}")

        try:
            from production.core_modules.playwright_downloader import PlaywrightDownloader

            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            if not cookie_file.exists():
                logger.error("❌ Cookie文件不存在")
                return None

            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')
            if not cookie_string:
                logger.error("❌ Cookie为空")
                return None

            logger.info(f"✅ Cookie已加载 (最后更新: {cookie_data.get('last_update', 'Unknown')})")

            # 创建下载器
            downloader = PlaywrightDownloader()

            # 执行异步下载
            logger.info("🔄 正在从腾讯文档下载CSV...")
            result = await downloader.download(
                url=self.doc_url,
                cookies=cookie_string,
                format='csv',
                download_dir=str(self.download_dir)
            )

            if result.get('success') and result.get('file_path'):
                csv_file = result['file_path']
                if Path(csv_file).exists():
                    logger.info(f"✅ 下载成功: {csv_file}")
                    logger.info(f"   文件大小: {Path(csv_file).stat().st_size} bytes")
                    return csv_file
                else:
                    logger.error(f"❌ 文件不存在: {csv_file}")
                    return None
            else:
                logger.error(f"❌ 下载失败: {result.get('error', 'Unknown error')}")
                if result.get('files'):
                    for file in result.get('files', []):
                        if Path(file).exists():
                            logger.info(f"✅ 找到下载文件: {file}")
                            return file
                return None

        except Exception as e:
            logger.error(f"❌ 下载异常: {e}")
            import traceback
            traceback.print_exc()
            return None

    def step2_real_compare(self, downloaded_file: str) -> List[Dict]:
        """Step 2: 真实对比 - 不伪造数据"""

        logger.info("\n" + "="*70)
        logger.info("Step 2: 真实对比分析")
        logger.info("="*70)

        if not Path(self.baseline_file).exists():
            logger.error(f"❌ 基线文件不存在: {self.baseline_file}")
            return []

        if not Path(downloaded_file).exists():
            logger.error(f"❌ 下载文件不存在: {downloaded_file}")
            return []

        try:
            # 读取下载的文件
            with open(downloaded_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows_current = list(reader)

            # 读取基线文件
            with open(self.baseline_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows_baseline = list(reader)

            logger.info(f"下载文件: {len(rows_current)}行 x {len(rows_current[0]) if rows_current else 0}列")
            logger.info(f"基线文件: {len(rows_baseline)}行 x {len(rows_baseline[0]) if rows_baseline else 0}列")

            # 真实对比 - 不人为添加任何变更
            changes = []
            headers = rows_current[0] if rows_current else []

            max_rows = min(len(rows_current), len(rows_baseline))
            for row_idx in range(1, max_rows):
                row_current = rows_current[row_idx]
                row_baseline = rows_baseline[row_idx]

                max_cols = min(len(row_current), len(row_baseline))
                for col_idx in range(max_cols):
                    val_current = str(row_current[col_idx]).strip()
                    val_baseline = str(row_baseline[col_idx]).strip()

                    if val_current != val_baseline:
                        changes.append({
                            "row": row_idx + 1,
                            "col": col_idx + 1,
                            "old": val_baseline,
                            "new": val_current,
                            "column_name": headers[col_idx] if col_idx < len(headers) else f"Col{col_idx+1}"
                        })

            logger.info(f"✅ 发现 {len(changes)} 处真实变更")

            for i, change in enumerate(changes[:5]):
                logger.info(f"   变更{i+1}: [{change['row']},{change['col']}] "
                          f"'{change['old'][:30]}' → '{change['new'][:30]}'")

            if len(changes) > 5:
                logger.info(f"   ... 还有 {len(changes)-5} 处变更")

            return changes

        except Exception as e:
            logger.error(f"❌ 对比失败: {e}")
            import traceback
            traceback.print_exc()
            return []

    def step3_real_scoring(self, changes: List[Dict]) -> Dict:
        """Step 3: 真实打分 - 基于真实数据"""

        logger.info("\n" + "="*70)
        logger.info("Step 3: 真实风险评分")
        logger.info("="*70)

        if not changes:
            logger.warning("⚠️ 没有变更需要打分")
            return {"timestamp": datetime.now().isoformat(), "cell_scores": {}}

        score_data = {
            "timestamp": datetime.now().isoformat(),
            "cell_scores": {},
            "summary": {
                "total_changes": len(changes),
                "high_risk": 0,
                "medium_risk": 0,
                "low_risk": 0
            }
        }

        for change in changes:
            col_idx = change['col'] - 1

            if col_idx < len(STANDARD_COLUMNS):
                col_name = STANDARD_COLUMNS[col_idx]

                if col_name in L1_COLUMNS:
                    risk_level = "HIGH"
                    score = 85
                    score_data["summary"]["high_risk"] += 1
                elif col_name in L2_COLUMNS:
                    risk_level = "MEDIUM"
                    score = 50
                    score_data["summary"]["medium_risk"] += 1
                else:
                    risk_level = "LOW"
                    score = 20
                    score_data["summary"]["low_risk"] += 1
            else:
                risk_level = "LOW"
                score = 15
                score_data["summary"]["low_risk"] += 1

            cell_key = f"{change['row']}_{change['col']}"
            score_data["cell_scores"][cell_key] = {
                "old_value": change['old'],
                "new_value": change['new'],
                "score": score,
                "risk_level": risk_level,
                "column": change.get('column_name', '')
            }

        logger.info(f"✅ 完成 {len(score_data['cell_scores'])} 个单元格评分")
        logger.info(f"   高风险: {score_data['summary']['high_risk']}")
        logger.info(f"   中风险: {score_data['summary']['medium_risk']}")
        logger.info(f"   低风险: {score_data['summary']['low_risk']}")

        return score_data

    def step4_real_excel(self, downloaded_file: str, score_data: Dict) -> Optional[str]:
        """Step 4: 真实Excel生成 - 包含涂色"""

        logger.info("\n" + "="*70)
        logger.info("Step 4: 生成Excel分析报告（带涂色）")
        logger.info("="*70)

        if not Path(downloaded_file).exists():
            logger.error("❌ 下载文件不存在")
            return None

        try:
            # 读取CSV数据
            with open(downloaded_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)

            # 创建Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "真实变更分析"

            # 写入数据
            for row_idx, row_data in enumerate(rows, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # 标题行样式
                    if row_idx == 1:
                        cell.font = Font(bold=True, color="000000")
                        cell.fill = PatternFill(
                            start_color="E0E0E0",
                            end_color="E0E0E0",
                            fill_type="solid"
                        )

            # 应用涂色（基于真实评分）
            marked_count = 0
            logger.info("开始应用涂色...")

            for cell_key, cell_info in score_data.get("cell_scores", {}).items():
                row, col = map(int, cell_key.split("_"))

                if row <= ws.max_row and col <= ws.max_column:
                    cell = ws.cell(row=row, column=col)

                    # 使用solid填充（腾讯文档兼容）
                    risk_level = cell_info["risk_level"]
                    if risk_level == "HIGH":
                        color = "FFCCCC"  # 浅红
                        font_color = "CC0000"
                        logger.debug(f"   涂色[{row},{col}]: 高风险(浅红)")
                    elif risk_level == "MEDIUM":
                        color = "FFFFCC"  # 浅黄
                        font_color = "FF8800"
                        logger.debug(f"   涂色[{row},{col}]: 中风险(浅黄)")
                    else:
                        color = "CCFFCC"  # 浅绿
                        font_color = "008800"
                        logger.debug(f"   涂色[{row},{col}]: 低风险(浅绿)")

                    # 必须使用solid填充类型
                    cell.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type="solid"  # 关键：必须是solid
                    )

                    cell.font = Font(
                        color=font_color,
                        bold=(risk_level == "HIGH")
                    )

                    # 添加批注
                    comment_text = (
                        f"风险等级: {risk_level}\n"
                        f"风险分数: {cell_info['score']}\n"
                        f"原值: {cell_info['old_value'][:50]}\n"
                        f"新值: {cell_info['new_value'][:50]}"
                    )
                    cell.comment = Comment(comment_text, "真实分析")

                    marked_count += 1

            # 保存Excel
            output_dir = Path('/root/projects/tencent-doc-manager/excel_outputs/complete_with_upload')
            output_dir.mkdir(parents=True, exist_ok=True)

            excel_file = output_dir / f'complete_real_analysis_{self.timestamp}.xlsx'
            wb.save(excel_file)
            wb.close()

            logger.info(f"✅ Excel生成成功: {excel_file.name}")
            logger.info(f"   标记单元格: {marked_count}个")
            logger.info(f"   文件大小: {excel_file.stat().st_size:,} bytes")
            logger.info(f"   涂色类型: solid填充（腾讯文档兼容）")

            return str(excel_file)

        except Exception as e:
            logger.error(f"❌ Excel生成失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def step5_real_upload(self, excel_file: str) -> Optional[str]:
        """Step 5: 真实上传到腾讯文档"""

        logger.info("\n" + "="*70)
        logger.info("Step 5: 真实上传到腾讯文档")
        logger.info("="*70)

        if not Path(excel_file).exists():
            logger.error(f"❌ Excel文件不存在: {excel_file}")
            return None

        try:
            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            if not cookie_file.exists():
                logger.error("❌ Cookie文件不存在")
                return None

            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')
            if not cookie_string:
                logger.error("❌ Cookie为空")
                return None

            # 使用V3上传器的异步版本（直接使用quick_upload_v3）
            from production.core_modules.tencent_doc_upload_production_v3 import quick_upload_v3

            logger.info(f"📄 准备上传文件: {Path(excel_file).name}")
            logger.info(f"   文件大小: {Path(excel_file).stat().st_size:,} bytes")
            logger.info("🔄 正在连接腾讯文档...")

            # 直接调用快速上传函数
            result = await quick_upload_v3(
                cookie_string=cookie_string,
                file_path=excel_file,
                headless=True
            )

            if result and result.get('success'):
                url = result.get('url')
                logger.info(f"✅ 上传成功!")
                logger.info(f"🔗 文档URL: {url}")
                logger.info(f"📄 文档名称: {result.get('doc_name', 'Unknown')}")
                return url
            else:
                logger.error(f"❌ 上传失败: {result.get('message', 'Unknown error') if result else 'No result'}")
                return None

        except ImportError as e:
            logger.error(f"❌ 无法导入上传模块: {e}")
            return None

        except Exception as e:
            logger.error(f"❌ 上传异常: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def run_complete_test(self) -> Dict:
        """执行完整测试流程（包含上传）"""

        logger.info("\n" + "="*80)
        logger.info("🚀 完整真实全链路测试 - 从下载到上传")
        logger.info("="*80)
        logger.info(f"基线文件: {self.baseline_file}")
        logger.info(f"目标URL: {self.doc_url}")
        logger.info(f"时间戳: {self.timestamp}")

        test_result = {
            "timestamp": self.timestamp,
            "success": False,
            "steps": {},
            "summary": "",
            "final_url": None  # 关键：返回最终URL
        }

        # Step 1: 真实下载
        downloaded_file = await self.step1_real_download()
        test_result["steps"]["download"] = {
            "success": bool(downloaded_file),
            "file": downloaded_file
        }

        if not downloaded_file:
            test_result["summary"] = "下载失败 - 无法继续测试"
            logger.error("\n❌ 测试失败：下载步骤未能完成")
            return test_result

        # Step 2: 真实对比
        changes = self.step2_real_compare(downloaded_file)
        test_result["steps"]["compare"] = {
            "success": True,
            "changes_count": len(changes)
        }

        # Step 3: 真实打分
        score_data = self.step3_real_scoring(changes)
        test_result["steps"]["scoring"] = {
            "success": True,
            "total_scored": len(score_data.get("cell_scores", {})),
            "summary": score_data.get("summary", {})
        }

        # Step 4: 生成Excel（带涂色）
        excel_file = self.step4_real_excel(downloaded_file, score_data)
        test_result["steps"]["excel"] = {
            "success": bool(excel_file),
            "file": excel_file
        }

        if not excel_file:
            test_result["summary"] = "Excel生成失败"
            logger.error("\n❌ 测试失败：Excel生成失败")
            return test_result

        # Step 5: 上传到腾讯文档
        upload_url = await self.step5_real_upload(excel_file)
        test_result["steps"]["upload"] = {
            "success": bool(upload_url),
            "url": upload_url
        }

        # 总结
        if upload_url:
            test_result["success"] = True
            test_result["final_url"] = upload_url
            test_result["summary"] = f"完整测试成功，发现{len(changes)}处真实变更，已上传到腾讯文档"

            logger.info("\n" + "="*80)
            logger.info("✅ 完整全链路测试成功完成！")
            logger.info(f"   下载文件: {Path(downloaded_file).name}")
            logger.info(f"   发现变更: {len(changes)}处")
            logger.info(f"   生成Excel: {Path(excel_file).name}")
            logger.info(f"   上传成功: {upload_url}")
            logger.info("\n" + "🔗 请访问以下URL验证涂色效果：")
            logger.info(f"   {upload_url}")
            logger.info("\n验证要点：")
            logger.info("   1. 检查是否有浅红、浅黄、浅绿色背景")
            logger.info("   2. 检查批注是否显示（鼠标悬停）")
            logger.info("   3. 确认变更位置与报告一致")
            logger.info("="*80)
        else:
            test_result["summary"] = "上传失败"
            logger.error("\n❌ 测试未能完全成功：上传失败")

        return test_result


async def main():
    """主函数"""

    # 创建测试实例
    tester = CompleteRealChainTest()

    # 执行完整测试
    result = await tester.run_complete_test()

    # 保存测试结果
    result_file = Path(f'/root/projects/tencent-doc-manager/test_results/complete_real_test_{tester.timestamp}.json')
    result_file.parent.mkdir(exist_ok=True)

    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    logger.info(f"\n📊 测试结果已保存: {result_file}")

    # 返回最终URL
    if result.get("final_url"):
        print(f"\n\n" + "="*80)
        print(f"🎉 最终腾讯文档URL：")
        print(f"   {result['final_url']}")
        print("="*80)
    else:
        print("\n\n❌ 未能获得最终URL")

    # 返回成功状态
    return 0 if result.get("success") else 1


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)