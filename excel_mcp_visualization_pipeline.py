# -*- coding: utf-8 -*-
"""
Excel MCP可视化管道 - 从风险检测到标记输出的完整流程
集成所有分析结果，生成专业的Excel可视化标记文件
"""

import os
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Union
import logging
from dataclasses import dataclass, asdict
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, NamedStyle
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule
import uuid
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class VisualizationModification:
    """可视化修改项数据结构"""
    row_number: int
    column_name: str
    column_index: int
    original_value: str
    new_value: str
    risk_level: str  # L1, L2, L3
    confidence: float
    ai_analysis: Optional[Dict[str, Any]] = None
    business_impact: str = "MEDIUM"
    modification_time: str = None
    comments: List[str] = None
    
    def __post_init__(self):
        if self.modification_time is None:
            self.modification_time = datetime.now().isoformat()
        if self.comments is None:
            self.comments = []


@dataclass
class VisualizationConfig:
    """可视化配置参数"""
    # 主题配置
    theme: str = "professional"  # professional, vibrant, minimal
    language: str = "zh-CN"  # zh-CN, en-US
    
    # 功能开关
    enable_risk_heatmap: bool = True
    enable_detailed_comments: bool = True
    enable_ai_insights: bool = True
    enable_summary_dashboard: bool = True
    enable_data_validation: bool = True
    enable_interactive_features: bool = True
    
    # 样式配置
    highlight_intensity: float = 0.8  # 0.1-1.0
    font_size: int = 11
    show_grid_lines: bool = True
    freeze_header_rows: int = 1
    
    # 输出配置
    output_format: str = "xlsx"  # xlsx, xlsm
    include_analysis_sheet: bool = True
    include_charts: bool = True
    file_name_template: str = "分析结果_{timestamp}_{batch_id}"


class ExcelMCPVisualizationPipeline:
    """Excel MCP可视化管道 - 核心处理引擎"""
    
    def __init__(self, config: VisualizationConfig = None):
        self.config = config or VisualizationConfig()
        
        # 主题颜色配置
        self.color_themes = {
            "professional": {
                "L1": {"fill": "DC2626", "font": "FFFFFF", "border": "991B1B"},  # 深红
                "L2": {"fill": "F59E0B", "font": "FFFFFF", "border": "D97706"},  # 橙色
                "L3": {"fill": "10B981", "font": "FFFFFF", "border": "047857"},  # 绿色
                "header": {"fill": "1F2937", "font": "FFFFFF", "border": "374151"},
                "summary": {"fill": "F3F4F6", "font": "1F2937", "border": "D1D5DB"}
            },
            "vibrant": {
                "L1": {"fill": "EF4444", "font": "FFFFFF", "border": "DC2626"},
                "L2": {"fill": "F97316", "font": "FFFFFF", "border": "EA580C"},
                "L3": {"fill": "22C55E", "font": "FFFFFF", "border": "16A34A"},
                "header": {"fill": "3B82F6", "font": "FFFFFF", "border": "2563EB"},
                "summary": {"fill": "E0E7FF", "font": "1E40AF", "border": "C7D2FE"}
            },
            "minimal": {
                "L1": {"fill": "FEE2E2", "font": "991B1B", "border": "FCA5A5"},
                "L2": {"fill": "FEF3C7", "font": "92400E", "border": "FCD34D"},
                "L3": {"fill": "D1FAE5", "font": "065F46", "border": "A7F3D0"},
                "header": {"fill": "F9FAFB", "font": "374151", "border": "E5E7EB"},
                "summary": {"fill": "FFFFFF", "font": "6B7280", "border": "D1D5DB"}
            }
        }
        
        # 当前主题
        self.current_theme = self.color_themes.get(self.config.theme, self.color_themes["professional"])
        
        # 文本模板
        self.text_templates = {
            "zh-CN": {
                "risk_levels": {"L1": "高风险", "L2": "中风险", "L3": "低风险"},
                "ai_decisions": {
                    "APPROVE": "AI建议：通过",
                    "REJECT": "AI建议：拒绝",
                    "REVIEW": "AI建议：人工审核",
                    "CONDITIONAL": "AI建议：条件通过"
                },
                "sheet_names": {
                    "data": "数据分析结果",
                    "summary": "分析汇总",
                    "charts": "图表分析",
                    "ai_insights": "AI洞察"
                },
                "comment_templates": {
                    "modification": "修改项：{original} → {new}",
                    "risk": "风险等级：{level}",
                    "confidence": "置信度：{confidence:.1%}",
                    "ai_analysis": "AI分析：{reasoning}",
                    "timestamp": "分析时间：{time}"
                }
            }
        }
        
        # 当前语言包
        self.texts = self.text_templates.get(self.config.language, self.text_templates["zh-CN"])
        
        # 处理统计
        self.processing_stats = {
            "workbooks_created": 0,
            "modifications_processed": 0,
            "comments_added": 0,
            "charts_created": 0,
            "ai_insights_integrated": 0
        }
    
    def create_comprehensive_visualization(self, 
                                         source_data: pd.DataFrame,
                                         modifications: List[VisualizationModification],
                                         analysis_summary: Dict[str, Any],
                                         output_path: str = None) -> Dict[str, Any]:
        """
        创建综合可视化Excel文件
        
        Args:
            source_data: 源数据DataFrame
            modifications: 修改项列表
            analysis_summary: 分析汇总信息
            output_path: 输出路径
            
        Returns:
            创建结果信息
        """
        logger.info(f"开始创建综合可视化，修改项数量: {len(modifications)}")
        
        try:
            # 生成输出路径
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                batch_id = str(uuid.uuid4())[:8]
                filename = self.config.file_name_template.format(
                    timestamp=timestamp,
                    batch_id=batch_id
                ) + f".{self.config.output_format}"
                output_path = os.path.join(tempfile.gettempdir(), filename)
            
            # 创建工作簿
            workbook = openpyxl.Workbook()
            
            # 删除默认工作表
            workbook.remove(workbook.active)
            
            # 创建主要工作表
            results = {}
            
            # 1. 数据分析结果表
            data_sheet = self._create_data_analysis_sheet(workbook, source_data, modifications)
            results["data_sheet"] = data_sheet
            
            # 2. 分析汇总表
            if self.config.include_analysis_sheet:
                summary_sheet = self._create_summary_sheet(workbook, analysis_summary, modifications)
                results["summary_sheet"] = summary_sheet
            
            # 3. AI洞察表
            if self.config.enable_ai_insights:
                ai_insights_sheet = self._create_ai_insights_sheet(workbook, modifications)
                results["ai_insights_sheet"] = ai_insights_sheet
            
            # 4. 图表分析表
            if self.config.include_charts:
                charts_sheet = self._create_charts_sheet(workbook, modifications, analysis_summary)
                results["charts_sheet"] = charts_sheet
            
            # 5. 应用全局样式
            self._apply_global_styles(workbook)
            
            # 6. 设置工作表保护（可选）
            if self.config.enable_interactive_features:
                self._setup_interactive_features(workbook)
            
            # 保存工作簿
            workbook.save(output_path)
            
            # 更新统计
            self.processing_stats["workbooks_created"] += 1
            self.processing_stats["modifications_processed"] += len(modifications)
            
            logger.info(f"可视化文件创建成功: {output_path}")
            
            return {
                "success": True,
                "output_path": output_path,
                "file_size_bytes": os.path.getsize(output_path),
                "sheets_created": list(results.keys()),
                "modifications_count": len(modifications),
                "metadata": {
                    "created_at": datetime.now().isoformat(),
                    "config": asdict(self.config),
                    "processing_stats": self.processing_stats.copy()
                }
            }
            
        except Exception as e:
            logger.error(f"可视化创建失败: {e}")
            return {
                "success": False,
                "error": str(e),
                "output_path": output_path
            }
    
    def _create_data_analysis_sheet(self, workbook: openpyxl.Workbook, 
                                  source_data: pd.DataFrame,
                                  modifications: List[VisualizationModification]) -> Dict[str, Any]:
        """创建数据分析结果工作表"""
        sheet_name = self.texts["sheet_names"]["data"]
        sheet = workbook.create_sheet(title=sheet_name, index=0)
        
        # 设置为活跃工作表
        workbook.active = sheet
        
        # 写入数据
        self._write_dataframe_to_sheet(sheet, source_data)
        
        # 应用修改标记
        modification_results = self._apply_modification_markers(sheet, modifications)
        
        # 添加表头样式
        self._style_header_row(sheet, len(source_data.columns))
        
        # 冻结窗格
        if self.config.freeze_header_rows > 0:
            sheet.freeze_panes = f"A{self.config.freeze_header_rows + 1}"
        
        # 自动调整列宽
        self._auto_adjust_column_widths(sheet)
        
        # 添加筛选器
        if len(source_data) > 0:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(source_data.columns))}{len(source_data) + 1}"
        
        return {
            "sheet_name": sheet_name,
            "data_rows": len(source_data),
            "data_columns": len(source_data.columns),
            "modifications_applied": modification_results["applied_count"],
            "comments_added": modification_results["comments_count"]
        }
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook,
                            analysis_summary: Dict[str, Any],
                            modifications: List[VisualizationModification]) -> Dict[str, Any]:
        """创建分析汇总工作表"""
        sheet_name = self.texts["sheet_names"]["summary"]
        sheet = workbook.create_sheet(title=sheet_name)
        
        # 汇总统计
        summary_data = self._prepare_summary_data(analysis_summary, modifications)
        
        # 写入汇总信息
        row = 1
        
        # 标题
        sheet[f"A{row}"] = "腾讯文档分析汇总报告"
        sheet[f"A{row}"].font = Font(size=16, bold=True, color=self.current_theme["header"]["font"])
        sheet.merge_cells(f"A{row}:D{row}")
        row += 2
        
        # 基本信息
        sheet[f"A{row}"] = "分析时间："
        sheet[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        
        sheet[f"A{row}"] = "分析文件数量："
        sheet[f"B{row}"] = summary_data.get("total_files", 1)
        row += 1
        
        sheet[f"A{row}"] = "检测到的修改项："
        sheet[f"B{row}"] = len(modifications)
        row += 2
        
        # 风险等级分布
        sheet[f"A{row}"] = "风险等级分布"
        sheet[f"A{row}"].font = Font(bold=True)
        row += 1
        
        risk_distribution = self._calculate_risk_distribution(modifications)
        for risk_level, count in risk_distribution.items():
            risk_text = self.texts["risk_levels"][risk_level]
            sheet[f"A{row}"] = f"{risk_text} ({risk_level})"
            sheet[f"B{row}"] = count
            
            # 应用颜色
            fill_color = self.current_theme[risk_level]["fill"]
            font_color = self.current_theme[risk_level]["font"]
            sheet[f"A{row}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            sheet[f"A{row}"].font = Font(color=font_color, bold=True)
            
            row += 1
        
        row += 1
        
        # AI分析结果汇总
        if any(mod.ai_analysis for mod in modifications):
            sheet[f"A{row}"] = "AI分析决策分布"
            sheet[f"A{row}"].font = Font(bold=True)
            row += 1
            
            ai_decisions = self._calculate_ai_decision_distribution(modifications)
            for decision, count in ai_decisions.items():
                decision_text = self.texts["ai_decisions"].get(decision, decision)
                sheet[f"A{row}"] = decision_text
                sheet[f"B{row}"] = count
                row += 1
        
        # 自动调整列宽
        self._auto_adjust_column_widths(sheet)
        
        return {
            "sheet_name": sheet_name,
            "summary_rows": row - 1,
            "risk_distribution": risk_distribution
        }
    
    def _create_ai_insights_sheet(self, workbook: openpyxl.Workbook,
                                modifications: List[VisualizationModification]) -> Dict[str, Any]:
        """创建AI洞察工作表"""
        sheet_name = self.texts["sheet_names"]["ai_insights"]
        sheet = workbook.create_sheet(title=sheet_name)
        
        ai_modifications = [mod for mod in modifications if mod.ai_analysis]
        
        if not ai_modifications:
            sheet["A1"] = "本次分析未包含AI语义分析结果"
            return {"sheet_name": sheet_name, "ai_insights_count": 0}
        
        # 表头
        headers = ["修改项ID", "列名", "原始值", "修改值", "AI决策", "置信度", "分析推理", "业务影响", "建议措施"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color=self.current_theme["header"]["font"])
            cell.fill = PatternFill(
                start_color=self.current_theme["header"]["fill"],
                end_color=self.current_theme["header"]["fill"],
                fill_type="solid"
            )
        
        # 数据行
        for row_idx, mod in enumerate(ai_modifications, 2):
            ai_data = mod.ai_analysis or {}
            
            sheet.cell(row=row_idx, column=1, value=f"MOD_{row_idx-1}")
            sheet.cell(row=row_idx, column=2, value=mod.column_name)
            sheet.cell(row=row_idx, column=3, value=mod.original_value[:50] + "..." if len(mod.original_value) > 50 else mod.original_value)
            sheet.cell(row=row_idx, column=4, value=mod.new_value[:50] + "..." if len(mod.new_value) > 50 else mod.new_value)
            sheet.cell(row=row_idx, column=5, value=ai_data.get("decision", ""))
            sheet.cell(row=row_idx, column=6, value=f"{ai_data.get('confidence', 0):.1%}")
            sheet.cell(row=row_idx, column=7, value=ai_data.get("reasoning", "")[:100] + "..." if len(ai_data.get("reasoning", "")) > 100 else ai_data.get("reasoning", ""))
            sheet.cell(row=row_idx, column=8, value=ai_data.get("business_impact", ""))
            sheet.cell(row=row_idx, column=9, value="; ".join(ai_data.get("suggested_actions", [])))
            
            # 根据AI决策着色
            decision = ai_data.get("decision", "")
            if decision == "REJECT":
                fill_color = self.current_theme["L1"]["fill"]
            elif decision == "REVIEW":
                fill_color = self.current_theme["L2"]["fill"]
            elif decision == "APPROVE":
                fill_color = self.current_theme["L3"]["fill"]
            else:
                fill_color = "F3F4F6"
            
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )
        
        # 自动调整列宽
        self._auto_adjust_column_widths(sheet)
        
        # 添加数据表格样式
        table_ref = f"A1:{get_column_letter(len(headers))}{len(ai_modifications) + 1}"
        table = Table(displayName="AIInsightsTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        sheet.add_table(table)
        
        self.processing_stats["ai_insights_integrated"] += len(ai_modifications)
        
        return {
            "sheet_name": sheet_name,
            "ai_insights_count": len(ai_modifications),
            "table_range": table_ref
        }
    
    def _create_charts_sheet(self, workbook: openpyxl.Workbook,
                           modifications: List[VisualizationModification],
                           analysis_summary: Dict[str, Any]) -> Dict[str, Any]:
        """创建图表分析工作表"""
        sheet_name = self.texts["sheet_names"]["charts"]
        sheet = workbook.create_sheet(title=sheet_name)
        
        charts_created = []
        
        # 1. 风险等级分布饼图
        risk_chart_result = self._create_risk_distribution_pie_chart(sheet, modifications)
        if risk_chart_result:
            charts_created.append("risk_distribution_pie")
        
        # 2. AI决策分布条形图
        if any(mod.ai_analysis for mod in modifications):
            ai_chart_result = self._create_ai_decision_bar_chart(sheet, modifications)
            if ai_chart_result:
                charts_created.append("ai_decision_bar")
        
        # 3. 置信度分布图
        confidence_chart_result = self._create_confidence_distribution_chart(sheet, modifications)
        if confidence_chart_result:
            charts_created.append("confidence_distribution")
        
        self.processing_stats["charts_created"] += len(charts_created)
        
        return {
            "sheet_name": sheet_name,
            "charts_created": charts_created,
            "charts_count": len(charts_created)
        }
    
    def _apply_modification_markers(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                                  modifications: List[VisualizationModification]) -> Dict[str, Any]:
        """应用修改标记到工作表"""
        applied_count = 0
        comments_count = 0
        
        for mod in modifications:
            try:
                # 计算单元格位置（注意Excel从1开始计数，且有表头行）
                row = mod.row_number + 2  # +1 for header, +1 for 0-based to 1-based
                col = mod.column_index + 1  # 0-based to 1-based
                
                cell = sheet.cell(row=row, column=col)
                
                # 应用填充颜色
                risk_colors = self.current_theme[mod.risk_level]
                cell.fill = PatternFill(
                    start_color=risk_colors["fill"],
                    end_color=risk_colors["fill"],
                    fill_type="solid"
                )
                
                # 应用字体颜色
                cell.font = Font(
                    color=risk_colors["font"],
                    bold=True,
                    size=self.config.font_size
                )
                
                # 应用边框
                border_color = risk_colors["border"]
                thick_border = Border(
                    left=Side(style="thick", color=border_color),
                    right=Side(style="thick", color=border_color),
                    top=Side(style="thick", color=border_color),
                    bottom=Side(style="thick", color=border_color)
                )
                cell.border = thick_border
                
                # 添加详细注释
                if self.config.enable_detailed_comments:
                    comment_text = self._generate_modification_comment(mod)
                    comment = Comment(comment_text, "系统分析")
                    comment.width = 300
                    comment.height = 150
                    cell.comment = comment
                    comments_count += 1
                
                applied_count += 1
                
            except Exception as e:
                logger.warning(f"应用修改标记失败 {mod.row_number}, {mod.column_index}: {e}")
        
        self.processing_stats["comments_added"] += comments_count
        
        return {
            "applied_count": applied_count,
            "comments_count": comments_count
        }
    
    def _generate_modification_comment(self, mod: VisualizationModification) -> str:
        """生成修改项的详细注释"""
        template = self.texts["comment_templates"]
        
        comment_lines = [
            template["modification"].format(
                original=mod.original_value[:30] + "..." if len(mod.original_value) > 30 else mod.original_value,
                new=mod.new_value[:30] + "..." if len(mod.new_value) > 30 else mod.new_value
            ),
            template["risk"].format(level=self.texts["risk_levels"][mod.risk_level]),
            template["confidence"].format(confidence=mod.confidence)
        ]
        
        if mod.ai_analysis:
            ai_reasoning = mod.ai_analysis.get("reasoning", "")[:50]
            comment_lines.append(template["ai_analysis"].format(reasoning=ai_reasoning))
        
        comment_lines.append(template["timestamp"].format(time=mod.modification_time[:16]))
        
        return "\n".join(comment_lines)
    
    def _create_risk_distribution_pie_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                                          modifications: List[VisualizationModification]) -> bool:
        """创建风险等级分布饼图"""
        try:
            # 准备数据
            risk_distribution = self._calculate_risk_distribution(modifications)
            
            # 写入图表数据
            start_row = 2
            sheet["A1"] = "风险等级"
            sheet["B1"] = "数量"
            
            row = start_row
            for risk_level, count in risk_distribution.items():
                sheet[f"A{row}"] = self.texts["risk_levels"][risk_level]
                sheet[f"B{row}"] = count
                row += 1
            
            # 创建饼图
            pie_chart = PieChart()
            labels = Reference(sheet, min_col=1, min_row=start_row, max_row=row-1)
            data = Reference(sheet, min_col=2, min_row=start_row, max_row=row-1)
            
            pie_chart.add_data(data)
            pie_chart.set_categories(labels)
            pie_chart.title = "修改项风险等级分布"
            
            # 设置图表位置
            sheet.add_chart(pie_chart, "D2")
            
            return True
            
        except Exception as e:
            logger.error(f"创建风险分布饼图失败: {e}")
            return False
    
    def _create_ai_decision_bar_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                                    modifications: List[VisualizationModification]) -> bool:
        """创建AI决策分布条形图"""
        try:
            ai_decisions = self._calculate_ai_decision_distribution(modifications)
            
            # 写入图表数据（在饼图数据下方）
            start_row = 8
            sheet[f"A{start_row}"] = "AI决策"
            sheet[f"B{start_row}"] = "数量"
            
            row = start_row + 1
            for decision, count in ai_decisions.items():
                sheet[f"A{row}"] = self.texts["ai_decisions"].get(decision, decision)
                sheet[f"B{row}"] = count
                row += 1
            
            # 创建条形图
            bar_chart = BarChart()
            labels = Reference(sheet, min_col=1, min_row=start_row+1, max_row=row-1)
            data = Reference(sheet, min_col=2, min_row=start_row+1, max_row=row-1)
            
            bar_chart.add_data(data)
            bar_chart.set_categories(labels)
            bar_chart.title = "AI分析决策分布"
            
            # 设置图表位置
            sheet.add_chart(bar_chart, "D12")
            
            return True
            
        except Exception as e:
            logger.error(f"创建AI决策条形图失败: {e}")
            return False
    
    def _create_confidence_distribution_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                                            modifications: List[VisualizationModification]) -> bool:
        """创建置信度分布图"""
        try:
            # 计算置信度区间分布
            confidence_ranges = {
                "高置信度 (>80%)": 0,
                "中等置信度 (60%-80%)": 0,
                "低置信度 (<60%)": 0
            }
            
            for mod in modifications:
                if mod.confidence > 0.8:
                    confidence_ranges["高置信度 (>80%)"] += 1
                elif mod.confidence > 0.6:
                    confidence_ranges["中等置信度 (60%-80%)"] += 1
                else:
                    confidence_ranges["低置信度 (<60%)"] += 1
            
            # 写入图表数据
            start_row = 15
            sheet[f"A{start_row}"] = "置信度范围"
            sheet[f"B{start_row}"] = "数量"
            
            row = start_row + 1
            for confidence_range, count in confidence_ranges.items():
                sheet[f"A{row}"] = confidence_range
                sheet[f"B{row}"] = count
                row += 1
            
            # 创建条形图
            bar_chart = BarChart()
            labels = Reference(sheet, min_col=1, min_row=start_row+1, max_row=row-1)
            data = Reference(sheet, min_col=2, min_row=start_row+1, max_row=row-1)
            
            bar_chart.add_data(data)
            bar_chart.set_categories(labels)
            bar_chart.title = "分析置信度分布"
            
            # 设置图表位置
            sheet.add_chart(bar_chart, "D22")
            
            return True
            
        except Exception as e:
            logger.error(f"创建置信度分布图失败: {e}")
            return False
    
    def _write_dataframe_to_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
        """将DataFrame写入工作表"""
        # 写入列名
        for col_idx, column_name in enumerate(df.columns, 1):
            sheet.cell(row=1, column=col_idx, value=str(column_name))
        
        # 写入数据
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell_value = value if pd.notna(value) else ""
                sheet.cell(row=row_idx, column=col_idx, value=str(cell_value))
    
    def _style_header_row(self, sheet: openpyxl.worksheet.worksheet.Worksheet, num_columns: int):
        """设置表头行样式"""
        header_fill = PatternFill(
            start_color=self.current_theme["header"]["fill"],
            end_color=self.current_theme["header"]["fill"],
            fill_type="solid"
        )
        header_font = Font(
            color=self.current_theme["header"]["font"],
            bold=True,
            size=self.config.font_size + 1
        )
        
        for col in range(1, num_columns + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _auto_adjust_column_widths(self, sheet: openpyxl.worksheet.worksheet.Worksheet):
        """自动调整列宽"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_global_styles(self, workbook: openpyxl.Workbook):
        """应用全局样式"""
        # 设置默认字体
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.font == Font():  # 如果是默认字体
                        cell.font = Font(size=self.config.font_size)
    
    def _setup_interactive_features(self, workbook: openpyxl.Workbook):
        """设置交互功能"""
        # 这里可以添加数据验证、下拉菜单等交互功能
        pass
    
    def _calculate_risk_distribution(self, modifications: List[VisualizationModification]) -> Dict[str, int]:
        """计算风险等级分布"""
        distribution = {"L1": 0, "L2": 0, "L3": 0}
        for mod in modifications:
            distribution[mod.risk_level] = distribution.get(mod.risk_level, 0) + 1
        return distribution
    
    def _calculate_ai_decision_distribution(self, modifications: List[VisualizationModification]) -> Dict[str, int]:
        """计算AI决策分布"""
        distribution = {}
        for mod in modifications:
            if mod.ai_analysis:
                decision = mod.ai_analysis.get("decision", "UNKNOWN")
                distribution[decision] = distribution.get(decision, 0) + 1
        return distribution
    
    def _prepare_summary_data(self, analysis_summary: Dict[str, Any], 
                            modifications: List[VisualizationModification]) -> Dict[str, Any]:
        """准备汇总数据"""
        return {
            "total_files": analysis_summary.get("total_files", 1),
            "total_modifications": len(modifications),
            "risk_distribution": self._calculate_risk_distribution(modifications),
            "ai_decision_distribution": self._calculate_ai_decision_distribution(modifications),
            "processing_timestamp": datetime.now().isoformat()
        }
    
    def get_processing_statistics(self) -> Dict[str, Any]:
        """获取处理统计信息"""
        return {
            "processing_stats": self.processing_stats.copy(),
            "configuration": {
                "theme": self.config.theme,
                "language": self.config.language,
                "features_enabled": {
                    "risk_heatmap": self.config.enable_risk_heatmap,
                    "detailed_comments": self.config.enable_detailed_comments,
                    "ai_insights": self.config.enable_ai_insights,
                    "summary_dashboard": self.config.enable_summary_dashboard,
                    "charts": self.config.include_charts
                }
            }
        }