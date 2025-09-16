#!/usr/bin/env python3
"""检查Excel文件的样式和涂色"""

import openpyxl
from openpyxl.styles import PatternFill
import json
import sys

def check_excel_styles(excel_path):
    """检查Excel文件中的样式"""
    
    print(f"检查文件: {excel_path}")
    print("=" * 60)
    
    # 打开工作簿
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['工作表1']
    
    # 统计涂色情况
    colored_cells = {
        'red': [],
        'yellow': [],
        'green': [],
        'other': [],
        'none': []
    }
    
    # 检查每个单元格的样式
    for row in range(1, min(ws.max_row + 1, 50)):  # 检查前50行
        for col in range(1, min(ws.max_column + 1, 20)):  # 检查前20列
            cell = ws.cell(row=row, column=col)
            
            # 获取填充样式
            fill = cell.fill
            
            # 检查是否有填充
            if fill and fill.patternType:
                # 获取前景色和背景色
                fg_color = fill.fgColor
                bg_color = fill.bgColor
                
                # 构建单元格信息
                cell_info = {
                    'cell': f"{openpyxl.utils.get_column_letter(col)}{row}",
                    'value': cell.value,
                    'pattern': fill.patternType,
                    'fg_color': str(fg_color.rgb) if fg_color and fg_color.rgb else None,
                    'bg_color': str(bg_color.rgb) if bg_color and bg_color.rgb else None
                }
                
                # 判断颜色类型
                if fg_color and fg_color.rgb:
                    rgb = str(fg_color.rgb)
                    if 'FF0000' in rgb or 'ff0000' in rgb:  # 红色
                        colored_cells['red'].append(cell_info)
                    elif 'FFFF00' in rgb or 'ffff00' in rgb:  # 黄色
                        colored_cells['yellow'].append(cell_info)
                    elif '00FF00' in rgb or '00ff00' in rgb:  # 绿色
                        colored_cells['green'].append(cell_info)
                    elif rgb != '00000000':  # 其他颜色
                        colored_cells['other'].append(cell_info)
                    else:
                        colored_cells['none'].append(cell_info)
    
    # 打印统计结果
    print("\n📊 涂色统计:")
    print(f"  🔴 红色单元格: {len(colored_cells['red'])} 个")
    print(f"  🟡 黄色单元格: {len(colored_cells['yellow'])} 个")
    print(f"  🟢 绿色单元格: {len(colored_cells['green'])} 个")
    print(f"  🔵 其他颜色: {len(colored_cells['other'])} 个")
    print(f"  ⬜ 无填充: {len(colored_cells['none'])} 个")
    
    # 显示具体的涂色单元格
    if colored_cells['red']:
        print("\n🔴 红色单元格详情:")
        for cell in colored_cells['red'][:5]:  # 显示前5个
            print(f"  {cell['cell']}: {cell['value']} (pattern={cell['pattern']})")
    
    if colored_cells['yellow']:
        print("\n🟡 黄色单元格详情:")
        for cell in colored_cells['yellow'][:5]:  # 显示前5个
            print(f"  {cell['cell']}: {cell['value']} (pattern={cell['pattern']})")
    
    if colored_cells['green']:
        print("\n🟢 绿色单元格详情:")
        for cell in colored_cells['green'][:5]:  # 显示前5个
            print(f"  {cell['cell']}: {cell['value']} (pattern={cell['pattern']})")
    
    if colored_cells['other']:
        print("\n🔵 其他颜色单元格详情:")
        for cell in colored_cells['other'][:10]:  # 显示前10个
            print(f"  {cell['cell']}: pattern={cell['pattern']}, fg={cell['fg_color']}, bg={cell['bg_color']}")
    
    # 检查是否有lightUp图案
    lightup_cells = []
    for row in range(1, min(ws.max_row + 1, 50)):
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.patternType == 'lightUp':
                lightup_cells.append(f"{openpyxl.utils.get_column_letter(col)}{row}")
    
    if lightup_cells:
        print(f"\n✨ 发现lightUp图案单元格: {len(lightup_cells)} 个")
        print(f"  位置: {', '.join(lightup_cells[:10])}")
    else:
        print("\n⚠️ 没有发现lightUp图案单元格")
    
    # 总结
    total_colored = sum(len(cells) for key, cells in colored_cells.items() if key != 'none')
    print(f"\n📊 总结: 共有 {total_colored} 个单元格有颜色填充")
    
    return colored_cells


if __name__ == "__main__":
    # 最新的标记文件
    excel_path = "/root/projects/tencent-doc-manager/excel_outputs/marked/tencent_副本-副本-测试版本-出国销售计划表_20250911_2106_midweek_W37_marked_20250911_210610_W37.xlsx"
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    
    check_excel_styles(excel_path)