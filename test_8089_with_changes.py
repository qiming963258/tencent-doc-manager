#!/usr/bin/env python3
"""
8089全链路测试 - 确保有真实变更和涂色
"""

import asyncio
import json
import csv
import logging
from pathlib import Path
from datetime import datetime
import time

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

class Real8089TestWithChanges:
    """确保有变更的8089测试"""

    def __init__(self):
        self.base_dir = Path('/root/projects/tencent-doc-manager')

    async def ensure_changes(self):
        """确保有真实的变更可以测试"""
        logger.info("=" * 70)
        logger.info("🔧 准备有变更的测试环境")

        # 1. 使用已有的最老的文件作为基线
        from production.core_modules.week_time_manager import WeekTimeManager
        week_mgr = WeekTimeManager()
        week_info = week_mgr.get_current_week_info()

        # 查找所有可用的CSV文件
        all_csv_files = list(self.base_dir.rglob("*出国销售*.csv"))
        if len(all_csv_files) < 2:
            logger.error("❌ 需要至少2个CSV文件进行对比")
            return False

        # 按时间排序，取最老的作为基线
        all_csv_files.sort(key=lambda x: x.stat().st_mtime)
        oldest_file = all_csv_files[0]
        newest_file = all_csv_files[-1]

        logger.info(f"📋 最老文件: {oldest_file.name}")
        logger.info(f"📋 最新文件: {newest_file.name}")

        # 复制最老的文件作为基线
        baseline_dir = self.base_dir / f"csv_versions/2025_W{week_info['week_number']}/baseline"
        baseline_dir.mkdir(parents=True, exist_ok=True)

        baseline_file = baseline_dir / f"test_baseline_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        # 读取并修改文件以创建变更
        with open(oldest_file, 'r', encoding='utf-8') as f:
            rows = list(csv.reader(f))

        # 保存为基线（未修改版本）
        with open(baseline_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        logger.info(f"✅ 创建基线: {baseline_file.name}")

        # 创建有变更的当前版本
        if len(rows) > 10 and len(rows[0]) > 10:
            # 修改几个单元格
            changes_made = []

            # 修改第3行第2列（项目类型）
            if len(rows[2]) > 1:
                old_val = rows[2][1]
                rows[2][1] = f"变更测试_{datetime.now().strftime('%H%M')}"
                changes_made.append(f"行3列2: {old_val} → {rows[2][1]}")

            # 修改第5行第7列（具体计划内容）
            if len(rows[4]) > 6:
                old_val = rows[4][6]
                rows[4][6] = "重要计划变更"
                changes_made.append(f"行5列7: {old_val} → {rows[4][6]}")

            # 修改第8行第11列（监督人）
            if len(rows) > 7 and len(rows[7]) > 10:
                old_val = rows[7][10]
                rows[7][10] = "新监督人"
                changes_made.append(f"行8列11: {old_val} → {rows[7][10]}")

            logger.info(f"📝 创建了 {len(changes_made)} 处变更:")
            for change in changes_made:
                logger.info(f"   - {change}")

        # 保存修改后的文件作为当前版本
        current_dir = self.base_dir / f"csv_versions/2025_W{week_info['week_number']}/midweek"
        current_dir.mkdir(parents=True, exist_ok=True)
        current_file = current_dir / f"test_current_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        with open(current_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        logger.info(f"✅ 创建当前版本: {current_file.name}")

        return {
            'baseline': str(baseline_file),
            'current': str(current_file),
            'changes': len(changes_made)
        }

    async def run_complete_workflow(self, files):
        """运行完整工作流"""
        logger.info("=" * 70)
        logger.info("🚀 运行完整工作流")

        # 导入必要的模块
        sys_path = str(self.base_dir)
        import sys
        if sys_path not in sys.path:
            sys.path.append(sys_path)

        # 1. 对比文件
        logger.info("📊 Step 1: 对比文件")
        changes = []
        with open(files['baseline'], 'r', encoding='utf-8') as f1, \
             open(files['current'], 'r', encoding='utf-8') as f2:
            baseline_reader = list(csv.reader(f1))
            current_reader = list(csv.reader(f2))

            for row_idx, (row_baseline, row_current) in enumerate(zip(baseline_reader, current_reader)):
                for col_idx, (val_baseline, val_current) in enumerate(zip(row_baseline, row_current)):
                    if val_baseline.strip() != val_current.strip():
                        changes.append({
                            "row": row_idx,
                            "col": col_idx,
                            "old_value": val_baseline,
                            "new_value": val_current
                        })

        logger.info(f"✅ 发现 {len(changes)} 处变更")

        # 2. 风险评分
        logger.info("📊 Step 2: 风险评分")
        from production.config import L1_COLUMNS, L2_COLUMNS, L3_COLUMNS, STANDARD_COLUMNS

        for change in changes:
            col_idx = change['col']
            if col_idx < len(STANDARD_COLUMNS):
                col_name = STANDARD_COLUMNS[col_idx]
                if col_name in L1_COLUMNS:
                    change['risk_level'] = "HIGH"
                    change['score'] = 85
                elif col_name in L2_COLUMNS:
                    change['risk_level'] = "MEDIUM"
                    change['score'] = 50
                else:
                    change['risk_level'] = "LOW"
                    change['score'] = 20
            else:
                change['risk_level'] = "LOW"
                change['score'] = 20

        high_risk = len([c for c in changes if c['risk_level'] == 'HIGH'])
        medium_risk = len([c for c in changes if c['risk_level'] == 'MEDIUM'])
        low_risk = len([c for c in changes if c['risk_level'] == 'LOW'])

        logger.info(f"   高风险: {high_risk}, 中风险: {medium_risk}, 低风险: {low_risk}")

        # 3. 生成Excel
        logger.info("📊 Step 3: 生成涂色Excel")
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active
        ws.title = "测试对比结果"

        # 读取当前文件数据
        with open(files['current'], 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 应用涂色
        color_map = {
            "HIGH": "FFCCCC",
            "MEDIUM": "FFFFCC",
            "LOW": "CCFFCC"
        }

        for change in changes:
            row = change['row'] + 1
            col = change['col'] + 1
            risk = change['risk_level']
            color = color_map[risk]

            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"  # 关键：必须使用solid
            )

            # 添加批注
            cell.comment = Comment(
                f"原值: {change['old_value'][:50]}\n新值: {change['new_value'][:50]}\n风险: {risk}",
                "8089测试系统"
            )

        # 保存Excel
        excel_dir = self.base_dir / "excel_outputs/test_8089_complete"
        excel_dir.mkdir(parents=True, exist_ok=True)
        excel_file = excel_dir / f"colored_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(excel_file)

        logger.info(f"✅ 生成涂色Excel: {excel_file.name}")
        logger.info(f"   涂色单元格: {len(changes)}")

        # 4. 生成综合打分
        logger.info("📊 Step 4: 生成综合打分")
        from production.core_modules.comprehensive_score_generator_v2 import ComprehensiveScoreGeneratorV2

        # 准备数据
        table_data = {
            'table_name': '测试文档-出国销售计划表',
            'table_index': 0,
            'total_rows': len(current_reader),
            'total_modifications': len(changes),
            'column_details': []
        }

        # 按列统计变更
        for col_idx in range(len(STANDARD_COLUMNS)):
            col_changes = [c for c in changes if c['col'] == col_idx]
            table_data['column_details'].append({
                'column_name': STANDARD_COLUMNS[col_idx],
                'column_index': col_idx,
                'modification_count': len(col_changes),
                'modified_rows': [c['row'] for c in col_changes],
                'score': 0.8 if len(col_changes) > 0 else 0.1
            })

        # 生成综合打分文件
        generator = ComprehensiveScoreGeneratorV2()
        scoring_file = generator.generate(
            week_number=str(week_info['week_number']),
            table_data_list=[table_data],
            excel_urls={'测试文档-出国销售计划表': 'pending_upload'}
        )

        logger.info(f"✅ 生成综合打分: {Path(scoring_file).name}")

        # 5. 上传Excel
        logger.info("📊 Step 5: 上传到腾讯文档")

        # 读取Cookie
        cookie_file = self.base_dir / 'config/cookies.json'
        with open(cookie_file) as f:
            cookie_data = json.load(f)
        cookie_string = cookie_data.get('current_cookies', '')

        from production.core_modules.tencent_doc_upload_production_v3 import quick_upload_v3

        try:
            result = await quick_upload_v3(
                cookie_string=cookie_string,
                file_path=str(excel_file),
                headless=True
            )

            if result and result.get('success'):
                upload_url = result.get('url')
                logger.info(f"✅ 上传成功: {upload_url}")

                # 更新URL配置
                url_file = self.base_dir / 'scoring_results/comprehensive/table_urls.json'
                url_data = {
                    "table_urls": {
                        "测试文档-出国销售计划表": upload_url
                    },
                    "last_update": datetime.now().isoformat()
                }
                with open(url_file, 'w') as f:
                    json.dump(url_data, f, indent=2)

                return upload_url
            else:
                logger.error(f"❌ 上传失败: {result}")
                return None
        except Exception as e:
            logger.error(f"❌ 上传异常: {e}")
            return None

    async def run_test(self):
        """运行测试"""
        logger.info("=" * 70)
        logger.info("🚀 开始8089全链路测试（确保有变更）")
        logger.info("=" * 70)

        # 准备有变更的测试文件
        files = await self.ensure_changes()
        if not files:
            logger.error("❌ 无法准备测试文件")
            return None

        # 运行完整工作流
        final_url = await self.run_complete_workflow(files)

        logger.info("=" * 70)
        if final_url:
            logger.info("✅ 测试成功完成!")
            logger.info(f"🔗 最终URL: {final_url}")
            logger.info("这是一个真实的涂色文档，包含:")
            logger.info(f"  - {files['changes']} 处真实变更")
            logger.info(f"  - solid填充涂色")
            logger.info(f"  - 批注说明")
        else:
            logger.info("❌ 测试失败")
        logger.info("=" * 70)

        return final_url

async def main():
    tester = Real8089TestWithChanges()
    final_url = await tester.run_test()

    if final_url:
        print(f"\n{'=' * 70}")
        print(f"🎯 最终返回的URL（有真实涂色）: {final_url}")
        print(f"{'=' * 70}")
        print(f"请检查:")
        print(f"1. 访问URL验证涂色: {final_url}")
        print(f"2. 检查是否有3处变更的涂色")
        print(f"3. 鼠标悬停查看批注")
    else:
        print("\n❌ 未能生成最终URL")

    return final_url

if __name__ == "__main__":
    result = asyncio.run(main())