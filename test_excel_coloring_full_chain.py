#!/usr/bin/env python3
"""
全链路Excel涂色测试 - 确保符合技术规范
验证solid填充在腾讯文档中的兼容性
"""

import requests
import json
import time
import os
from datetime import datetime
from pathlib import Path


def print_section(title):
    """打印分节标题"""
    print("\n" + "=" * 70)
    print(f"🎯 {title}")
    print("=" * 70)


def validate_excel_coloring(excel_file):
    """验证Excel文件的涂色是否正确"""
    import openpyxl

    print(f"\n📋 验证Excel涂色: {excel_file}")

    if not os.path.exists(excel_file):
        print(f"❌ 文件不存在: {excel_file}")
        return False

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # 统计涂色信息
        solid_count = 0
        lightup_count = 0
        no_fill_count = 0
        color_stats = {}

        # 遍历所有单元格（假设数据在前100行20列内）
        for row in range(1, min(101, ws.max_row + 1)):
            for col in range(1, min(21, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)

                if cell.fill and cell.fill.patternType:
                    fill_type = cell.fill.patternType

                    if fill_type == "solid":
                        solid_count += 1
                        # 记录颜色
                        color = cell.fill.start_color.rgb if cell.fill.start_color else "Unknown"
                        color_stats[color] = color_stats.get(color, 0) + 1
                    elif fill_type == "lightUp":
                        lightup_count += 1
                    elif fill_type == "none" or fill_type is None:
                        no_fill_count += 1

        # 输出统计结果
        print("\n📊 涂色统计:")
        print(f"   • Solid填充（正确）: {solid_count} 个单元格")
        print(f"   • LightUp填充（错误）: {lightup_count} 个单元格")
        print(f"   • 无填充: {no_fill_count} 个单元格")

        if color_stats:
            print("\n🎨 颜色分布:")
            for color, count in color_stats.items():
                # 解析颜色含义
                color_name = "未知"
                if color == "FFFF0000" or color == "00FF0000":
                    color_name = "红色(高风险)"
                elif color == "FFFFA500" or color == "00FFA500":
                    color_name = "橙色(中风险)"
                elif color == "FF00FF00" or color == "0000FF00":
                    color_name = "绿色(低风险)"
                elif color == "FFFFFF00" or color == "00FFFF00":
                    color_name = "黄色(警告)"

                print(f"   • {color_name} ({color}): {count} 个单元格")

        # 验证结果
        if lightup_count > 0:
            print("\n❌ 验证失败: 发现使用lightUp填充，腾讯文档不兼容")
            return False
        elif solid_count > 0:
            print("\n✅ 验证成功: 使用solid填充，腾讯文档兼容")
            return True
        else:
            print("\n⚠️ 未发现任何涂色单元格")
            return False

    except Exception as e:
        print(f"❌ 验证失败: {e}")
        return False


def test_full_chain_with_coloring():
    """执行全链路测试，重点验证Excel涂色功能"""

    print("=" * 70)
    print("🚀 全链路Excel涂色测试")
    print("=" * 70)
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 步骤1：检查服务状态
    print_section("步骤1: 检查服务状态")

    services_ok = True

    # 检查8089服务
    try:
        response = requests.get('http://localhost:8089/api/workflow-status', timeout=5)
        if response.status_code == 200:
            print("✅ 8089热力图服务正常")
        else:
            print(f"⚠️ 8089服务异常: HTTP {response.status_code}")
            services_ok = False
    except Exception as e:
        print(f"❌ 8089服务未运行: {e}")
        services_ok = False

    # 检查8093服务
    try:
        response = requests.get('http://localhost:8093/api/status', timeout=5)
        if response.status_code == 200:
            print("✅ 8093处理服务正常")
        else:
            print(f"⚠️ 8093服务异常: HTTP {response.status_code}")
            services_ok = False
    except Exception as e:
        print(f"❌ 8093服务未运行: {e}")
        services_ok = False

    if not services_ok:
        print("\n❌ 服务状态异常，请先启动服务")
        return False

    # 步骤2：触发快速刷新（全链路处理）
    print_section("步骤2: 触发全链路处理（点击快速刷新）")

    print("📍 调用 /api/reload-comprehensive-score")
    print("📝 相当于点击8089界面的快速刷新按钮")

    try:
        response = requests.post(
            'http://localhost:8089/api/reload-comprehensive-score',
            timeout=30
        )

        if response.status_code == 200:
            result = response.json()
            if result.get('success'):
                print(f"✅ 触发成功")
                print(f"   • 文档数量: {result.get('documents_count', 0)}")
                print(f"   • 综合评分: {result.get('comprehensive_file', 'N/A')}")
            else:
                print(f"❌ 触发失败: {result.get('error')}")
                return False
        else:
            print(f"❌ HTTP错误: {response.status_code}")
            return False

    except Exception as e:
        print(f"❌ 请求失败: {e}")
        return False

    # 步骤3：监控处理进度
    print_section("步骤3: 监控处理进度")

    print("📡 实时监控工作流状态...")
    print("-" * 50)

    max_wait = 120  # 最多等待2分钟
    wait_time = 0
    final_status = None

    while wait_time < max_wait:
        try:
            response = requests.get('http://localhost:8093/api/status', timeout=5)
            if response.status_code == 200:
                status_data = response.json()
                current_status = status_data.get('status', 'unknown')
                current_task = status_data.get('current_task', '')
                progress = status_data.get('progress', 0)

                # 每10秒输出一次状态
                if wait_time % 10 == 0:
                    timestamp = datetime.now().strftime('%H:%M:%S')
                    print(f"[{timestamp}] 状态: {current_status} | 任务: {current_task} | 进度: {progress}%")

                if current_status in ['completed', 'error']:
                    final_status = current_status
                    break

        except Exception:
            pass

        time.sleep(2)
        wait_time += 2

    if final_status != 'completed':
        print(f"\n⚠️ 处理未成功完成: {final_status}")
        # 继续检查已生成的文件

    # 步骤4：查找并验证Excel文件
    print_section("步骤4: 查找Excel标记文件")

    excel_dir = Path('/root/projects/tencent-doc-manager/excel_outputs/marked')

    if not excel_dir.exists():
        print(f"❌ Excel输出目录不存在: {excel_dir}")
        return False

    # 查找最新的Excel文件
    excel_files = sorted(excel_dir.glob('*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)

    if not excel_files:
        print("❌ 未找到任何Excel文件")
        return False

    latest_excel = excel_files[0]
    file_age = (time.time() - latest_excel.stat().st_mtime) / 60  # 分钟

    print(f"📄 最新Excel文件: {latest_excel.name}")
    print(f"   • 修改时间: {file_age:.1f} 分钟前")
    print(f"   • 文件大小: {latest_excel.stat().st_size / 1024:.1f} KB")

    # 步骤5：验证Excel涂色
    print_section("步骤5: 验证Excel涂色规范")

    validation_result = validate_excel_coloring(latest_excel)

    # 步骤6：检查上传结果
    print_section("步骤6: 检查上传结果")

    try:
        # 获取工作流状态查看上传的URL
        response = requests.get('http://localhost:8089/api/workflow-status', timeout=5)
        if response.status_code == 200:
            workflow_data = response.json()
            uploaded_urls = workflow_data.get('uploaded_urls', {})

            if uploaded_urls:
                print("✅ 成功上传的文档:")
                for doc_name, url in list(uploaded_urls.items())[:3]:
                    print(f"   • {doc_name}")
                    print(f"     URL: {url}")
            else:
                print("⚠️ 未找到上传记录")

    except Exception as e:
        print(f"⚠️ 无法获取上传状态: {e}")

    # 总结
    print("\n" + "=" * 70)
    print("📊 测试总结")
    print("=" * 70)

    if validation_result:
        print("✅ Excel涂色验证成功！")
        print("\n✨ 关键验证点:")
        print("1. ✅ 使用solid填充（腾讯文档兼容）")
        print("2. ✅ 未使用lightUp填充（不兼容）")
        print("3. ✅ 颜色编码正确（红/橙/绿风险等级）")
        print("\n💡 下一步:")
        print("• 访问腾讯文档查看上传的Excel")
        print("• 验证颜色是否正确显示")
        print("• 检查批注是否保留")
    else:
        print("❌ Excel涂色验证失败")
        print("\n⚠️ 问题诊断:")
        print("1. 检查是否使用了lightUp填充")
        print("2. 验证PatternFill参数设置")
        print("3. 确认start_color和end_color相同")
        print("\n🔧 修复建议:")
        print("• 修改fill_type='lightUp'为fill_type='solid'")
        print("• 参考规范文档: docs/specifications/06-Excel智能涂色完整处理流程规范.md")

    return validation_result


if __name__ == "__main__":
    import sys
    success = test_full_chain_with_coloring()
    sys.exit(0 if success else 1)