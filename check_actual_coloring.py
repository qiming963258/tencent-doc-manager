#!/usr/bin/env python3
"""检查Excel文件中实际涂色的单元格"""

import openpyxl
from openpyxl.styles import PatternFill
import sys
from pathlib import Path

def check_colored_cells(excel_file):
    """检查Excel文件中实际涂色的单元格"""

    print(f"📊 检查文件: {excel_file}")
    print("="*60)

    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    print(f"工作表: {ws.title}")
    print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
    print()

    # 统计涂色情况
    colored_cells = []
    solid_count = 0
    other_count = 0
    no_fill_count = 0

    # 遍历所有单元格
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.patternType:
                fill_type = cell.fill.patternType

                if fill_type == 'solid':
                    solid_count += 1
                    # 获取颜色
                    fg_color = cell.fill.fgColor
                    if fg_color and fg_color.rgb:
                        color_rgb = fg_color.rgb
                        # 跳过白色填充
                        if color_rgb != 'FFFFFFFF' and color_rgb != '00000000':
                            colored_cells.append({
                                'cell': cell.coordinate,
                                'value': cell.value,
                                'fill_type': fill_type,
                                'color': color_rgb
                            })
                elif fill_type and fill_type != 'none':
                    other_count += 1
                    colored_cells.append({
                        'cell': cell.coordinate,
                        'value': cell.value,
                        'fill_type': fill_type,
                        'color': cell.fill.fgColor.rgb if cell.fill.fgColor else None
                    })
            else:
                no_fill_count += 1

    # 输出统计结果
    print(f"📊 单元格填充统计:")
    print(f"  Solid填充: {solid_count} 个")
    print(f"  其他填充: {other_count} 个")
    print(f"  无填充: {no_fill_count} 个")
    print()

    # 显示前20个涂色单元格
    print(f"🎨 实际涂色的单元格（前20个）:")
    if colored_cells:
        for i, cell_info in enumerate(colored_cells[:20]):
            print(f"  {i+1}. {cell_info['cell']}: "
                  f"类型={cell_info['fill_type']}, "
                  f"颜色={cell_info['color']}, "
                  f"值={str(cell_info['value'])[:20]}")
    else:
        print("  ❌ 没有找到涂色的单元格！")

    print()
    print("="*60)

    # 颜色分析
    if colored_cells:
        color_counts = {}
        for cell in colored_cells:
            color = cell.get('color')
            if color:
                color_counts[color] = color_counts.get(color, 0) + 1

        print(f"🎨 颜色使用统计:")
        for color, count in sorted(color_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {color}: {count} 个单元格")

    # 判断结果
    print()
    if colored_cells:
        if other_count > 0:
            print("⚠️ 警告：发现非solid填充，可能在腾讯文档中不显示")
        else:
            print("✅ 所有涂色都使用solid填充，应该在腾讯文档中正常显示")

        print(f"\n💡 提示：")
        print(f"  - 共找到 {len(colored_cells)} 个涂色单元格")
        print(f"  - 如果腾讯文档中仍然不显示，可能是：")
        print(f"    1. 颜色太浅（如FFFFFFFF接近白色）")
        print(f"    2. 腾讯文档渲染问题")
        print(f"    3. 上传过程中格式被转换")
    else:
        print("❌ 文件中没有任何涂色单元格！")
        print("   可能原因：")
        print("   1. 涂色步骤失败")
        print("   2. 打分数据为空")
        print("   3. 单元格匹配失败")

    return len(colored_cells) > 0

if __name__ == "__main__":
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        # 默认检查最新的文件
        excel_file = "/root/projects/tencent-doc-manager/excel_outputs/marked/tencent_111测试版本-出国销售计划表_20250929_0157_midweek_W40_marked_20250929_015729_W00.xlsx"

    if not Path(excel_file).exists():
        print(f"❌ 文件不存在: {excel_file}")
        sys.exit(1)

    has_coloring = check_colored_cells(excel_file)
    sys.exit(0 if has_coloring else 1)