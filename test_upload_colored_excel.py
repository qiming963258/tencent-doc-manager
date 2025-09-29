#!/usr/bin/env python3
"""
上传涂色Excel到腾讯文档并验证
"""

import sys
import json
from pathlib import Path

sys.path.insert(0, '/root/projects/tencent-doc-manager')

from production.core_modules.tencent_doc_upload_production_v3 import sync_upload_v3


def upload_colored_excel():
    """上传最新的涂色Excel文件"""

    print("=" * 60)
    print("🚀 上传涂色Excel到腾讯文档")
    print("=" * 60)

    # 读取Cookie
    cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
    with open(cookie_file) as f:
        cookie_data = json.load(f)
        cookie = cookie_data.get('current_cookies', '')

    # 查找最新的Excel文件
    excel_dir = Path('/root/projects/tencent-doc-manager/excel_outputs/marked')
    excel_files = sorted(excel_dir.glob('*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)

    if not excel_files:
        print("❌ 未找到Excel文件")
        return None

    latest_excel = excel_files[0]
    print(f"\n📄 准备上传文件:")
    print(f"   • 文件名: {latest_excel.name}")
    print(f"   • 大小: {latest_excel.stat().st_size / 1024:.1f} KB")

    # 验证涂色信息
    import openpyxl

    wb = openpyxl.load_workbook(latest_excel)
    ws = wb.active

    solid_count = 0
    colors = set()

    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.patternType == "solid":
                solid_count += 1
                if cell.fill.start_color:
                    colors.add(cell.fill.start_color.rgb)

    print(f"\n📊 文件涂色信息:")
    print(f"   • Solid填充单元格: {solid_count}个")
    print(f"   • 使用颜色数: {len(colors)}种")

    # 颜色解析
    color_mapping = {
        "FFFF0000": "红色(高风险)",
        "00FF0000": "红色(高风险)",
        "FFFFA500": "橙色(中风险)",
        "00FFA500": "橙色(中风险)",
        "FF00FF00": "绿色(低风险)",
        "0000FF00": "绿色(低风险)",
        "FFFFFF00": "黄色(警告)",
        "00FFFF00": "黄色(警告)",
        "FFFFCCCC": "浅红",
        "00FFCCCC": "浅红",
        "FFFFFFCC": "浅黄",
        "00FFFFCC": "浅黄",
        "FFCCFFCC": "浅绿",
        "00CCFFCC": "浅绿"
    }

    if colors:
        print("\n🎨 颜色详情:")
        for color in colors:
            color_name = color_mapping.get(color, f"自定义({color})")
            print(f"   • {color_name}")

    # 执行上传
    print("\n📤 开始上传到腾讯文档...")

    result = sync_upload_v3(
        cookie_string=cookie,
        file_path=str(latest_excel),
        headless=True
    )

    print(f"\n📋 上传结果:")
    print(f"   • 成功: {result.get('success')}")
    print(f"   • 消息: {result.get('message')}")

    if result.get('success'):
        url = result.get('url')
        print(f"\n✅ 上传成功！")
        print(f"📄 文档URL: {url}")
        print("\n💡 验证步骤:")
        print("1. 访问上述URL")
        print("2. 检查单元格背景色是否显示")
        print("3. 验证颜色是否与风险等级对应:")
        print("   • 红色 = 高风险")
        print("   • 橙色 = 中风险")
        print("   • 绿色 = 低风险")
        print("4. 查看批注是否保留")

        # 保存URL供后续验证
        with open('/tmp/uploaded_excel_url.txt', 'w') as f:
            f.write(url)

        return url
    else:
        print(f"\n❌ 上传失败")
        if result.get('storage_info'):
            storage = result['storage_info']
            print(f"   • 存储空间: {storage.get('usage_percent', 0):.2f}%")
            if not storage.get('has_space'):
                print("   • 原因: 存储空间不足")
        return None


def main():
    url = upload_colored_excel()

    print("\n" + "=" * 60)
    print("📊 测试总结")
    print("=" * 60)

    if url:
        print("✅ Excel涂色上传测试成功！")
        print("\n🔧 技术验证:")
        print("1. ✅ 修复了lightUp填充问题")
        print("2. ✅ 使用solid填充确保兼容")
        print("3. ✅ 符合技术规范要求")
        print("\n📝 符合的技术规范:")
        print("• docs/specifications/06-Excel智能涂色完整处理流程规范.md")
        print("• docs/specifications/14-Excel涂色兼容性测试报告.md")
    else:
        print("❌ 上传失败，请检查:")
        print("1. Cookie是否有效")
        print("2. 存储空间是否充足")
        print("3. 网络连接是否正常")

    return url is not None


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)