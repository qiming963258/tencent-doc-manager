#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
上传诊断脚本 - 找出上传失败的真实原因
"""

import asyncio
import json
import os
from pathlib import Path
from datetime import datetime
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

async def diagnose_upload():
    """诊断上传问题"""
    
    print("\n" + "="*60)
    print("腾讯文档上传诊断")
    print("="*60)
    
    # 1. 检查Cookie配置
    print("\n1. 检查Cookie配置...")
    cookie_path = Path('/root/projects/tencent-doc-manager/config/cookies.json')
    if not cookie_path.exists():
        print("❌ Cookie配置文件不存在")
        return
    
    with open(cookie_path) as f:
        config = json.load(f)
        cookie_string = config.get('cookie_string', '')
    
    print(f"✅ Cookie长度: {len(cookie_string)}")
    print(f"✅ 包含分号空格: {'; ' in cookie_string}")
    print(f"✅ 包含DOC_SID: {'DOC_SID' in cookie_string}")
    print(f"✅ 包含SID: {'SID' in cookie_string}")
    
    # 2. 检查上传模块
    print("\n2. 检查上传模块...")
    try:
        # 优先尝试修复版
        from tencent_doc_uploader_fixed import TencentDocUploader
        print("✅ 使用修复版上传模块")
        module_type = "fixed"
    except ImportError:
        try:
            from tencent_doc_uploader import TencentDocUploader
            print("⚠️ 使用原版上传模块")
            module_type = "original"
        except ImportError as e:
            print(f"❌ 无法导入上传模块: {e}")
            return
    
    # 3. 创建测试文件
    print("\n3. 创建测试文件...")
    test_file = Path('/tmp/test_upload_diagnostic.xlsx')
    
    # 使用openpyxl创建简单的Excel文件
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '测试上传'
        ws['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A2'] = '诊断脚本'
        ws['B2'] = '验证上传功能'
        wb.save(test_file)
        print(f"✅ 创建测试文件: {test_file}")
    except Exception as e:
        print(f"❌ 创建测试文件失败: {e}")
        return
    
    # 4. 测试上传
    print("\n4. 测试上传功能...")
    uploader = TencentDocUploader()
    
    try:
        # 初始化浏览器（无头模式）
        print("   初始化浏览器...")
        await uploader.init_browser(headless=True)  # 使用无头模式
        
        # 登录
        print("   执行Cookie登录...")
        login_success = await uploader.login_with_cookies(cookie_string)
        if not login_success:
            print("❌ 登录失败")
            await uploader.close()
            return
        print("✅ 登录成功")
        
        # 等待观察
        print("\n   等待5秒观察页面状态...")
        await asyncio.sleep(5)
        
        # 获取页面信息
        page_url = uploader.page.url
        page_title = await uploader.page.title()
        print(f"   当前URL: {page_url}")
        print(f"   页面标题: {page_title}")
        
        # 检查导入按钮
        import_btn = await uploader.page.query_selector('button.desktop-import-button-pc')
        if import_btn:
            print("✅ 找到导入按钮")
        else:
            print("❌ 未找到导入按钮")
            # 尝试其他选择器
            alt_btn = await uploader.page.query_selector('button:has-text("导入")')
            if alt_btn:
                print("✅ 找到备用导入按钮")
        
        # 尝试上传
        print("\n5. 执行上传测试...")
        result = await uploader.create_new_sheet(str(test_file))
        
        if result.get('success'):
            print(f"✅ 上传成功！")
            print(f"   URL: {result.get('url')}")
            print(f"   消息: {result.get('message')}")
        else:
            print(f"❌ 上传失败")
            print(f"   错误: {result.get('error')}")
            print(f"   消息: {result.get('message')}")
        
        # 等待观察结果
        print("\n   等待10秒观察最终结果...")
        await asyncio.sleep(10)
        
        # 获取最终页面状态
        final_url = uploader.page.url
        print(f"   最终URL: {final_url}")
        
        # 检查是否真的创建了新文档
        if '/sheet/' in final_url or '/doc/' in final_url:
            print("✅ 确认：新文档已创建")
        else:
            print("⚠️ 警告：URL未跳转到新文档")
        
    except Exception as e:
        print(f"❌ 测试过程出错: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        print("\n6. 清理...")
        await uploader.close()
        if test_file.exists():
            os.remove(test_file)
            print("✅ 删除测试文件")
    
    print("\n" + "="*60)
    print("诊断完成")
    print("="*60)

if __name__ == "__main__":
    asyncio.run(diagnose_upload())