#!/usr/bin/env python3
"""
XLSX原生格式全链路测试 - 直接下载和修改XLSX格式
完全真实，无任何虚拟行为
"""

import json
import sys
import os
import asyncio
from pathlib import Path
from datetime import datetime
import logging
from typing import Optional, Dict, List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

sys.path.append('/root/projects/tencent-doc-manager')

from production.config import STANDARD_COLUMNS, L1_COLUMNS, L2_COLUMNS, L3_COLUMNS


class XLSXNativeChainTest:
    """XLSX原生格式测试类 - 直接处理XLSX"""

    def __init__(self):
        self.doc_url = "https://docs.qq.com/sheet/DWEFNU25TemFnZXJN"
        self.doc_id = "DWEFNU25TemFnZXJN"
        self.baseline_xlsx = None  # 将通过下载获得基线XLSX
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.download_dir = Path(f'/root/projects/tencent-doc-manager/downloads/xlsx_native_{self.timestamp}')
        self.download_dir.mkdir(parents=True, exist_ok=True)

    async def step1_download_xlsx(self, is_baseline=False) -> Optional[str]:
        """Step 1: 下载XLSX格式文件"""

        logger.info("="*70)
        logger.info(f"Step 1: 下载{'基线' if is_baseline else '当前'}XLSX文件")
        logger.info("="*70)
        logger.info(f"目标URL: {self.doc_url}")

        try:
            from production.core_modules.playwright_downloader import PlaywrightDownloader

            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            if not cookie_file.exists():
                logger.error("❌ Cookie文件不存在")
                return None

            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')
            if not cookie_string:
                logger.error("❌ Cookie为空")
                return None

            logger.info(f"✅ Cookie已加载 (最后更新: {cookie_data.get('last_update', 'Unknown')})")

            # 创建下载器
            downloader = PlaywrightDownloader()

            # 执行XLSX格式下载
            logger.info("🔄 正在从腾讯文档下载XLSX格式...")
            result = await downloader.download(
                url=self.doc_url,
                cookies=cookie_string,
                format='xlsx',  # 关键：使用XLSX格式
                download_dir=str(self.download_dir)
            )

            if result.get('success'):
                # 查找下载的文件
                if result.get('file_path'):
                    xlsx_file = result['file_path']
                elif result.get('files'):
                    xlsx_file = result['files'][0]
                else:
                    logger.error("❌ 下载结果中没有文件")
                    return None

                if Path(xlsx_file).exists():
                    # 重命名为明确的名称
                    if is_baseline:
                        new_name = self.download_dir / f'baseline_{self.timestamp}.xlsx'
                    else:
                        new_name = self.download_dir / f'current_{self.timestamp}.xlsx'

                    # 如果是xlsx结尾，直接使用；如果是csv，需要处理
                    if xlsx_file.endswith('.csv'):
                        logger.warning("⚠️ 下载的是CSV格式，尝试转换为XLSX")
                        return await self._convert_csv_to_xlsx(xlsx_file, str(new_name))
                    else:
                        # 复制或移动文件
                        import shutil
                        shutil.copy2(xlsx_file, new_name)

                    logger.info(f"✅ XLSX下载成功: {new_name.name}")
                    logger.info(f"   文件大小: {new_name.stat().st_size} bytes")
                    return str(new_name)
                else:
                    logger.error(f"❌ 文件不存在: {xlsx_file}")
                    return None
            else:
                logger.error(f"❌ 下载失败: {result.get('error', 'Unknown error')}")
                return None

        except Exception as e:
            logger.error(f"❌ 下载异常: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def _convert_csv_to_xlsx(self, csv_file: str, xlsx_file: str) -> Optional[str]:
        """将CSV转换为XLSX格式"""
        try:
            import csv
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active

            with open(csv_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)

            wb.save(xlsx_file)
            wb.close()

            logger.info(f"✅ CSV成功转换为XLSX: {Path(xlsx_file).name}")
            return xlsx_file

        except Exception as e:
            logger.error(f"❌ CSV转换失败: {e}")
            return None

    def step2_compare_xlsx(self, baseline_xlsx: str, current_xlsx: str) -> List[Dict]:
        """Step 2: 直接对比两个XLSX文件"""

        logger.info("\n" + "="*70)
        logger.info("Step 2: XLSX原生格式对比")
        logger.info("="*70)

        if not Path(baseline_xlsx).exists():
            logger.error(f"❌ 基线XLSX不存在: {baseline_xlsx}")
            return []

        if not Path(current_xlsx).exists():
            logger.error(f"❌ 当前XLSX不存在: {current_xlsx}")
            return []

        try:
            # 加载两个XLSX文件
            wb_baseline = load_workbook(baseline_xlsx, data_only=True)
            wb_current = load_workbook(current_xlsx, data_only=True)

            ws_baseline = wb_baseline.active
            ws_current = wb_current.active

            logger.info(f"基线XLSX: {ws_baseline.max_row}行 x {ws_baseline.max_column}列")
            logger.info(f"当前XLSX: {ws_current.max_row}行 x {ws_current.max_column}列")

            # 对比每个单元格
            changes = []

            # 获取标题行
            headers = []
            for col in range(1, ws_current.max_column + 1):
                cell_value = ws_current.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Col{col}")

            # 对比数据行
            max_rows = min(ws_baseline.max_row, ws_current.max_row)
            max_cols = min(ws_baseline.max_column, ws_current.max_column)

            for row in range(2, max_rows + 1):  # 跳过标题行
                for col in range(1, max_cols + 1):
                    val_baseline = ws_baseline.cell(row=row, column=col).value
                    val_current = ws_current.cell(row=row, column=col).value

                    # 转换为字符串比较
                    str_baseline = str(val_baseline) if val_baseline is not None else ""
                    str_current = str(val_current) if val_current is not None else ""

                    if str_baseline.strip() != str_current.strip():
                        changes.append({
                            "row": row,
                            "col": col,
                            "old": str_baseline,
                            "new": str_current,
                            "column_name": headers[col-1] if col-1 < len(headers) else f"Col{col}",
                            "cell_ref": f"{get_column_letter(col)}{row}"
                        })

            logger.info(f"✅ 发现 {len(changes)} 处真实变更")

            for i, change in enumerate(changes[:5]):
                logger.info(f"   变更{i+1}: [{change['cell_ref']}] "
                          f"'{change['old'][:30]}' → '{change['new'][:30]}'")

            if len(changes) > 5:
                logger.info(f"   ... 还有 {len(changes)-5} 处变更")

            # 关闭工作簿
            wb_baseline.close()
            wb_current.close()

            return changes

        except Exception as e:
            logger.error(f"❌ 对比失败: {e}")
            import traceback
            traceback.print_exc()
            return []

    def step3_apply_coloring_to_xlsx(self, xlsx_file: str, changes: List[Dict]) -> Optional[str]:
        """Step 3: 直接在XLSX文件上应用涂色"""

        logger.info("\n" + "="*70)
        logger.info("Step 3: 在XLSX原文件上应用涂色")
        logger.info("="*70)

        if not Path(xlsx_file).exists():
            logger.error(f"❌ XLSX文件不存在: {xlsx_file}")
            return None

        if not changes:
            logger.warning("⚠️ 没有变更需要标记")
            return xlsx_file

        try:
            # 加载XLSX文件
            wb = load_workbook(xlsx_file)
            ws = wb.active

            # 统计
            marked_count = 0
            risk_summary = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}

            logger.info("开始在原XLSX文件上应用涂色...")

            for change in changes:
                row = change["row"]
                col = change["col"]

                # 判断风险等级
                col_idx = col - 1
                if col_idx < len(STANDARD_COLUMNS):
                    col_name = STANDARD_COLUMNS[col_idx]

                    if col_name in L1_COLUMNS:
                        risk_level = "HIGH"
                        color = "FFCCCC"  # 浅红
                        font_color = "CC0000"
                    elif col_name in L2_COLUMNS:
                        risk_level = "MEDIUM"
                        color = "FFFFCC"  # 浅黄
                        font_color = "FF8800"
                    else:
                        risk_level = "LOW"
                        color = "CCFFCC"  # 浅绿
                        font_color = "008800"
                else:
                    risk_level = "LOW"
                    color = "CCFFCC"
                    font_color = "008800"

                risk_summary[risk_level] += 1

                # 应用涂色
                cell = ws.cell(row=row, column=col)

                # 使用solid填充（腾讯文档兼容）
                cell.fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid"  # 必须是solid
                )

                cell.font = Font(
                    color=font_color,
                    bold=(risk_level == "HIGH")
                )

                # 添加批注
                comment_text = (
                    f"风险等级: {risk_level}\n"
                    f"原值: {change['old'][:50]}\n"
                    f"新值: {change['new'][:50]}\n"
                    f"列名: {change['column_name']}"
                )
                cell.comment = Comment(comment_text, "XLSX原生分析")

                marked_count += 1
                logger.debug(f"   涂色[{change['cell_ref']}]: {risk_level}风险")

            # 保存修改后的文件
            output_dir = Path('/root/projects/tencent-doc-manager/excel_outputs/xlsx_native')
            output_dir.mkdir(parents=True, exist_ok=True)

            output_file = output_dir / f'xlsx_native_colored_{self.timestamp}.xlsx'
            wb.save(output_file)
            wb.close()

            logger.info(f"✅ XLSX涂色完成: {output_file.name}")
            logger.info(f"   标记单元格: {marked_count}个")
            logger.info(f"   高风险: {risk_summary['HIGH']}")
            logger.info(f"   中风险: {risk_summary['MEDIUM']}")
            logger.info(f"   低风险: {risk_summary['LOW']}")
            logger.info(f"   文件大小: {output_file.stat().st_size:,} bytes")

            return str(output_file)

        except Exception as e:
            logger.error(f"❌ 涂色失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def step4_upload_xlsx(self, xlsx_file: str) -> Optional[str]:
        """Step 4: 上传修改后的XLSX到腾讯文档"""

        logger.info("\n" + "="*70)
        logger.info("Step 4: 上传XLSX到腾讯文档")
        logger.info("="*70)

        if not Path(xlsx_file).exists():
            logger.error(f"❌ XLSX文件不存在: {xlsx_file}")
            return None

        try:
            # 读取Cookie
            cookie_file = Path('/root/projects/tencent-doc-manager/config/cookies.json')
            with open(cookie_file) as f:
                cookie_data = json.load(f)

            cookie_string = cookie_data.get('current_cookies', '')

            # 使用quick_upload_v3上传
            from production.core_modules.tencent_doc_upload_production_v3 import quick_upload_v3

            logger.info(f"📄 准备上传XLSX文件: {Path(xlsx_file).name}")
            logger.info(f"   文件大小: {Path(xlsx_file).stat().st_size:,} bytes")
            logger.info("🔄 正在连接腾讯文档...")

            result = await quick_upload_v3(
                cookie_string=cookie_string,
                file_path=xlsx_file,
                headless=True
            )

            if result and result.get('success'):
                url = result.get('url')
                logger.info(f"✅ 上传成功!")
                logger.info(f"🔗 文档URL: {url}")
                logger.info(f"📄 文档名称: {result.get('doc_name', 'Unknown')}")
                return url
            else:
                logger.error(f"❌ 上传失败: {result.get('message', 'Unknown error') if result else 'No result'}")
                return None

        except Exception as e:
            logger.error(f"❌ 上传异常: {e}")
            import traceback
            traceback.print_exc()
            return None

    async def run_xlsx_native_test(self) -> Dict:
        """执行XLSX原生格式全链路测试"""

        logger.info("\n" + "="*80)
        logger.info("🚀 XLSX原生格式全链路测试 - 直接处理XLSX文件")
        logger.info("="*80)
        logger.info(f"目标URL: {self.doc_url}")
        logger.info(f"时间戳: {self.timestamp}")

        test_result = {
            "timestamp": self.timestamp,
            "success": False,
            "steps": {},
            "summary": "",
            "final_url": None
        }

        # Step 1a: 下载基线XLSX（第一次下载作为基线）
        logger.info("\n📥 下载基线XLSX文件...")
        baseline_xlsx = await self.step1_download_xlsx(is_baseline=True)
        test_result["steps"]["download_baseline"] = {
            "success": bool(baseline_xlsx),
            "file": baseline_xlsx
        }

        if not baseline_xlsx:
            test_result["summary"] = "基线XLSX下载失败"
            logger.error("\n❌ 测试失败：基线下载失败")
            return test_result

        # 模拟一些时间过去，让文档有变化（实际应用中这是两个不同时间点的文档）
        logger.info("\n⏳ 等待文档可能的变化...")
        await asyncio.sleep(2)

        # Step 1b: 下载当前XLSX
        logger.info("\n📥 下载当前XLSX文件...")
        current_xlsx = await self.step1_download_xlsx(is_baseline=False)
        test_result["steps"]["download_current"] = {
            "success": bool(current_xlsx),
            "file": current_xlsx
        }

        if not current_xlsx:
            test_result["summary"] = "当前XLSX下载失败"
            logger.error("\n❌ 测试失败：当前文件下载失败")
            return test_result

        # Step 2: 对比两个XLSX文件
        changes = self.step2_compare_xlsx(baseline_xlsx, current_xlsx)
        test_result["steps"]["compare"] = {
            "success": True,
            "changes_count": len(changes)
        }

        if not changes:
            logger.warning("⚠️ 没有发现变更（可能两次下载的是同一版本）")
            # 即使没有变更也继续，展示流程

        # Step 3: 在XLSX上应用涂色
        colored_xlsx = self.step3_apply_coloring_to_xlsx(current_xlsx, changes)
        test_result["steps"]["coloring"] = {
            "success": bool(colored_xlsx),
            "file": colored_xlsx
        }

        if not colored_xlsx:
            test_result["summary"] = "XLSX涂色失败"
            logger.error("\n❌ 测试失败：涂色失败")
            return test_result

        # Step 4: 上传到腾讯文档
        upload_url = await self.step4_upload_xlsx(colored_xlsx)
        test_result["steps"]["upload"] = {
            "success": bool(upload_url),
            "url": upload_url
        }

        # 总结
        if upload_url:
            test_result["success"] = True
            test_result["final_url"] = upload_url
            test_result["summary"] = f"XLSX原生格式测试成功，发现{len(changes)}处变更"

            logger.info("\n" + "="*80)
            logger.info("✅ XLSX原生格式全链路测试成功完成！")
            logger.info(f"   基线文件: {Path(baseline_xlsx).name}")
            logger.info(f"   当前文件: {Path(current_xlsx).name}")
            logger.info(f"   发现变更: {len(changes)}处")
            logger.info(f"   涂色文件: {Path(colored_xlsx).name if colored_xlsx else 'N/A'}")
            logger.info(f"   上传成功: {upload_url}")
            logger.info("\n" + "🔗 请访问以下URL验证XLSX原生格式涂色效果：")
            logger.info(f"   {upload_url}")
            logger.info("\n验证要点：")
            logger.info("   1. 确认是XLSX格式（非CSV转换）")
            logger.info("   2. 检查涂色是否正确显示")
            logger.info("   3. 验证批注信息")
            logger.info("   4. 确认格式保持原样")
            logger.info("="*80)
        else:
            test_result["summary"] = "上传失败"
            logger.error("\n❌ 测试未能完全成功：上传失败")

        return test_result


async def main():
    """主函数"""

    # 创建测试实例
    tester = XLSXNativeChainTest()

    # 执行XLSX原生格式测试
    result = await tester.run_xlsx_native_test()

    # 保存测试结果
    result_file = Path(f'/root/projects/tencent-doc-manager/test_results/xlsx_native_test_{tester.timestamp}.json')
    result_file.parent.mkdir(exist_ok=True)

    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    logger.info(f"\n📊 测试结果已保存: {result_file}")

    # 返回最终URL
    if result.get("final_url"):
        print(f"\n\n" + "="*80)
        print(f"🎉 XLSX原生格式测试最终URL：")
        print(f"   {result['final_url']}")
        print("="*80)
    else:
        print("\n\n❌ 未能获得最终URL")

    return 0 if result.get("success") else 1


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)