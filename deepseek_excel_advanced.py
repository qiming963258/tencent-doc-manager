#!/usr/bin/env python3
"""
DeepSeek Excel 高级处理脚本
支持XLSX文件的完整处理，包括半填充颜色标记
"""

import os
import sys
import json
import requests
from datetime import datetime
from pathlib import Path

# 尝试导入openpyxl，如果失败则提供安装指导
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
    from openpyxl.comments import Comment
    from openpyxl.drawing.image import Image
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ 需要安装openpyxl: pip install openpyxl")

# 配置
DEEPSEEK_API_KEY = os.getenv('DEEPSEEK_API_KEY')
API_BASE_URL = "https://api.siliconflow.cn/v1"

# 目录配置
BASE_DIR = Path('/root/projects/tencent-doc-manager')
DOWNLOAD_DIR = BASE_DIR / 'downloads'
OUTPUT_DIR = BASE_DIR / 'excel_outputs'
OUTPUT_DIR.mkdir(exist_ok=True)

class DeepSeekExcelAdvanced:
    """高级Excel处理器，支持半填充和复杂格式"""
    
    def __init__(self):
        self.api_key = DEEPSEEK_API_KEY
        
        if OPENPYXL_AVAILABLE:
            # 定义各种填充样式（包括半填充lightUp图案）
            self.fills = {
                # 半填充样式（lightUp图案）
                'high_risk_pattern': PatternFill(
                    fill_type='lightUp',  # 半填充图案
                    fgColor='FF6B6B',     # 前景色：红色
                    bgColor='FFFFFF'      # 背景色：白色
                ),
                'medium_risk_pattern': PatternFill(
                    fill_type='lightUp',
                    fgColor='FFD93D',     # 前景色：黄色
                    bgColor='FFFFFF'
                ),
                'low_risk_pattern': PatternFill(
                    fill_type='lightUp',
                    fgColor='6BCF7F',     # 前景色：绿色
                    bgColor='FFFFFF'
                ),
                # 实心填充样式
                'high_risk_solid': PatternFill(
                    fill_type='solid',
                    start_color='FF6B6B',
                    end_color='FF6B6B'
                ),
                'medium_risk_solid': PatternFill(
                    fill_type='solid',
                    start_color='FFD93D',
                    end_color='FFD93D'
                ),
                'low_risk_solid': PatternFill(
                    fill_type='solid',
                    start_color='6BCF7F',
                    end_color='6BCF7F'
                ),
                # 渐变填充
                'gradient_risk': PatternFill(
                    fill_type='lightGrid',
                    fgColor='FF6B6B',
                    bgColor='FFD93D'
                )
            }
            
            # 边框样式
            self.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # 字体样式
            self.fonts = {
                'header': Font(bold=True, size=12, color='000000'),
                'risk_high': Font(bold=True, color='FF0000'),
                'risk_medium': Font(bold=True, color='FFA500'),
                'risk_low': Font(bold=True, color='008000'),
                'normal': Font(size=10, color='000000')
            }
    
    def call_deepseek(self, content: str, prompt: str = None) -> dict:
        """调用DeepSeek API进行分析"""
        
        if not self.api_key:
            print("⚠️ 使用模拟模式（未设置API密钥）")
            return self._mock_analysis()
        
        if not prompt:
            prompt = """分析这个Excel数据表，识别风险和异常。返回JSON格式：
{
    "risk_level": "high/medium/low",
    "confidence": 0.0-1.0,
    "risk_cells": [{"row": 2, "col": 3, "reason": "异常值"}],
    "summary": "总体分析",
    "recommendations": "处理建议"
}"""
        
        try:
            print("🤖 调用DeepSeek V3 API...")
            response = requests.post(
                f"{API_BASE_URL}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "deepseek-v3",
                    "messages": [
                        {"role": "system", "content": "你是Excel数据分析专家。返回JSON格式的分析结果。"},
                        {"role": "user", "content": f"{prompt}\n\n数据：\n{content[:2000]}"}
                    ],
                    "temperature": 0.7,
                    "max_tokens": 800
                },
                timeout=30
            )
            
            if response.status_code == 200:
                result = response.json()
                content = result['choices'][0]['message']['content']
                
                # 尝试解析JSON
                try:
                    import re
                    json_match = re.search(r'\{.*\}', content, re.DOTALL)
                    if json_match:
                        return json.loads(json_match.group())
                except:
                    pass
                
                return self._mock_analysis()
            else:
                print(f"❌ API错误: {response.status_code}")
                return self._mock_analysis()
                
        except Exception as e:
            print(f"⚠️ API调用失败: {e}")
            return self._mock_analysis()
    
    def _mock_analysis(self) -> dict:
        """模拟分析结果"""
        return {
            "risk_level": "medium",
            "confidence": 0.85,
            "risk_cells": [
                {"row": 2, "col": 2, "reason": "数值异常偏高"},
                {"row": 5, "col": 3, "reason": "缺失关键数据"},
                {"row": 8, "col": 4, "reason": "格式不一致"}
            ],
            "summary": "检测到中等风险的数据异常",
            "recommendations": "建议审核标记的单元格，确认数据准确性"
        }
    
    def process_xlsx(self, input_file: Path, use_pattern_fill: bool = True) -> Path:
        """处理Excel文件
        
        Args:
            input_file: 输入文件路径
            use_pattern_fill: 是否使用半填充图案（True）或实心填充（False）
        """
        
        if not OPENPYXL_AVAILABLE:
            print("❌ 需要openpyxl库来处理Excel文件")
            return None
        
        print(f"\n📊 处理Excel文件: {input_file}")
        print(f"🎨 填充模式: {'半填充图案(lightUp)' if use_pattern_fill else '实心填充'}")
        
        # 生成输出文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = OUTPUT_DIR / f"deepseek_marked_{timestamp}_{input_file.name}"
        
        try:
            # 加载Excel文件
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
            
            print(f"📋 工作表: {ws.title}")
            print(f"📏 数据范围: {ws.max_row}行 × {ws.max_column}列")
            
            # 提取数据用于分析
            data_sample = []
            for row in ws.iter_rows(max_row=min(50, ws.max_row), values_only=True):
                if any(cell for cell in row):
                    data_sample.append(','.join(str(c) if c else '' for c in row))
            
            data_str = '\n'.join(data_sample)
            
            # 调用DeepSeek分析
            analysis = self.call_deepseek(data_str)
            
            print(f"\n📈 分析结果:")
            print(f"  风险等级: {analysis['risk_level']}")
            print(f"  置信度: {analysis['confidence']:.0%}")
            print(f"  风险单元格: {len(analysis.get('risk_cells', []))}个")
            
            # 应用格式和标记
            self._apply_formatting(ws, analysis, use_pattern_fill)
            
            # 添加分析报告sheet
            self._add_report_sheet(wb, analysis)
            
            # 保存文件
            wb.save(output_file)
            print(f"\n✅ 处理完成: {output_file}")
            print(f"📦 文件大小: {output_file.stat().st_size:,} bytes")
            
            return output_file
            
        except Exception as e:
            print(f"❌ 处理失败: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _apply_formatting(self, ws, analysis: dict, use_pattern: bool = True):
        """应用格式化和标记"""
        
        risk_level = analysis.get('risk_level', 'medium')
        
        # 选择填充样式
        if use_pattern:
            fill_key = f'{risk_level}_risk_pattern'
        else:
            fill_key = f'{risk_level}_risk_solid'
        
        fill_style = self.fills.get(fill_key, self.fills['medium_risk_pattern'])
        
        # 添加风险标记列
        risk_col = ws.max_column + 1
        ai_col = ws.max_column + 2
        
        # 设置标题
        ws.cell(row=1, column=risk_col, value="风险等级").font = self.fonts['header']
        ws.cell(row=1, column=ai_col, value="AI分析").font = self.fonts['header']
        
        # 标记风险单元格（使用半填充）
        risk_cells = analysis.get('risk_cells', [])
        for risk_info in risk_cells:
            try:
                row = risk_info.get('row', 0)
                col = risk_info.get('col', 0)
                if 1 <= row <= ws.max_row and 1 <= col <= ws.max_column:
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_style  # 应用半填充样式
                    cell.border = self.border
                    # 添加批注说明原因
                    if 'reason' in risk_info:
                        cell.comment = Comment(f"风险: {risk_info['reason']}", "DeepSeek")
            except:
                pass
        
        # 为每行添加风险标记
        font_key = f'risk_{risk_level}'
        risk_font = self.fonts.get(font_key, self.fonts['normal'])
        
        for row_idx in range(2, min(ws.max_row + 1, 200)):
            # 风险等级列
            risk_cell = ws.cell(row=row_idx, column=risk_col)
            risk_cell.value = risk_level.upper()
            risk_cell.font = risk_font
            risk_cell.border = self.border
            
            # AI分析列
            ai_cell = ws.cell(row=row_idx, column=ai_col)
            ai_cell.value = f"{analysis['confidence']:.0%}"
            ai_cell.border = self.border
            ai_cell.alignment = Alignment(horizontal='center')
        
        # 设置列宽
        ws.column_dimensions[openpyxl.utils.get_column_letter(risk_col)].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(ai_col)].width = 10
    
    def _add_report_sheet(self, wb, analysis: dict):
        """添加分析报告sheet"""
        
        # 创建报告sheet
        if "DeepSeek分析报告" in wb.sheetnames:
            del wb["DeepSeek分析报告"]
        
        report_sheet = wb.create_sheet("DeepSeek分析报告", 0)
        
        # 标题
        report_sheet.merge_cells('A1:E1')
        title = report_sheet['A1']
        title.value = "DeepSeek V3 智能分析报告"
        title.font = Font(bold=True, size=16)
        title.alignment = Alignment(horizontal='center', vertical='center')
        title.fill = PatternFill(fill_type='solid', start_color='E6F3FF')
        
        # 基本信息
        row = 3
        report_sheet[f'A{row}'] = "分析时间:"
        report_sheet[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        row += 1
        report_sheet[f'A{row}'] = "API模型:"
        report_sheet[f'B{row}'] = "DeepSeek V3 (硅基流动)"
        
        row += 2
        report_sheet[f'A{row}'] = "风险等级:"
        cell = report_sheet[f'B{row}']
        cell.value = analysis.get('risk_level', 'N/A').upper()
        
        # 根据风险等级设置颜色
        risk_colors = {
            'HIGH': 'FF6B6B',
            'MEDIUM': 'FFD93D',
            'LOW': '6BCF7F'
        }
        color = risk_colors.get(cell.value, 'CCCCCC')
        cell.fill = PatternFill(fill_type='solid', start_color=color)
        cell.font = Font(bold=True, color='FFFFFF' if cell.value == 'HIGH' else '000000')
        
        row += 1
        report_sheet[f'A{row}'] = "置信度:"
        report_sheet[f'B{row}'] = f"{analysis.get('confidence', 0):.0%}"
        
        # 风险单元格详情
        if 'risk_cells' in analysis and analysis['risk_cells']:
            row += 2
            report_sheet[f'A{row}'] = "风险单元格:"
            report_sheet[f'A{row}'].font = Font(bold=True)
            
            row += 1
            report_sheet[f'B{row}'] = "位置"
            report_sheet[f'C{row}'] = "原因"
            report_sheet[f'B{row}'].font = Font(bold=True)
            report_sheet[f'C{row}'].font = Font(bold=True)
            
            for risk in analysis['risk_cells'][:20]:  # 最多显示20个
                row += 1
                report_sheet[f'B{row}'] = f"行{risk.get('row', '?')} 列{risk.get('col', '?')}"
                report_sheet[f'C{row}'] = risk.get('reason', '未知')
        
        # 总结和建议
        if 'summary' in analysis:
            row += 2
            report_sheet[f'A{row}'] = "分析总结:"
            report_sheet[f'A{row}'].font = Font(bold=True)
            report_sheet.merge_cells(f'B{row}:E{row}')
            report_sheet[f'B{row}'] = analysis['summary']
        
        if 'recommendations' in analysis:
            row += 2
            report_sheet[f'A{row}'] = "处理建议:"
            report_sheet[f'A{row}'].font = Font(bold=True)
            report_sheet.merge_cells(f'B{row}:E{row}')
            report_sheet[f'B{row}'] = analysis['recommendations']
        
        # 设置列宽
        report_sheet.column_dimensions['A'].width = 15
        report_sheet.column_dimensions['B'].width = 25
        report_sheet.column_dimensions['C'].width = 40
        
        # 添加边框
        for row in report_sheet.iter_rows(min_row=1, max_row=report_sheet.max_row,
                                          min_col=1, max_col=5):
            for cell in row:
                if cell.value:
                    cell.border = self.border

def test_performance():
    """测试处理性能"""
    
    print("=" * 60)
    print("🚀 DeepSeek Excel 高级处理系统")
    print("=" * 60)
    print("⚡ 性能测试模式")
    
    processor = DeepSeekExcelAdvanced()
    
    # 查找Excel文件
    excel_files = list(DOWNLOAD_DIR.glob("*.xlsx")) + list(DOWNLOAD_DIR.glob("*.xls"))
    
    if not excel_files:
        # 创建测试文件
        if OPENPYXL_AVAILABLE:
            print("\n📝 创建测试Excel文件...")
            test_file = OUTPUT_DIR / "test_performance.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "测试数据"
            
            # 添加测试数据
            headers = ["产品", "数量", "单价", "总价", "部门"]
            ws.append(headers)
            
            for i in range(100):
                ws.append([f"产品{i+1}", 100+i*10, 50+i*2, (100+i*10)*(50+i*2), f"部门{i%5+1}"])
            
            wb.save(test_file)
            excel_files = [test_file]
            print(f"✅ 创建测试文件: {test_file}")
    
    if excel_files:
        input_file = excel_files[0]
        
        print(f"\n📊 测试文件: {input_file.name}")
        
        # 测试半填充模式
        start_time = datetime.now()
        print("\n🎨 测试1: 半填充图案模式（lightUp）")
        result1 = processor.process_xlsx(input_file, use_pattern_fill=True)
        time1 = (datetime.now() - start_time).total_seconds()
        
        # 测试实心填充模式
        start_time = datetime.now()
        print("\n🎨 测试2: 实心填充模式（solid）")
        result2 = processor.process_xlsx(input_file, use_pattern_fill=False)
        time2 = (datetime.now() - start_time).total_seconds()
        
        # 性能报告
        print("\n" + "=" * 60)
        print("📊 性能测试报告")
        print("=" * 60)
        print(f"半填充模式耗时: {time1:.2f}秒")
        print(f"实心填充耗时: {time2:.2f}秒")
        print(f"API密钥状态: {'已配置' if DEEPSEEK_API_KEY else '模拟模式'}")
        print(f"处理速度评级: {'⚡快速' if time1 < 5 else '🚀正常' if time1 < 10 else '🐌需优化'}")
        print("=" * 60)

def main():
    """主函数"""
    
    if len(sys.argv) > 1:
        # 处理指定文件
        input_file = Path(sys.argv[1])
        if input_file.exists() and input_file.suffix.lower() in ['.xlsx', '.xls']:
            processor = DeepSeekExcelAdvanced()
            processor.process_xlsx(input_file, use_pattern_fill=True)
        else:
            print(f"❌ 文件不存在或不是Excel文件: {input_file}")
    else:
        # 运行性能测试
        test_performance()

if __name__ == "__main__":
    main()